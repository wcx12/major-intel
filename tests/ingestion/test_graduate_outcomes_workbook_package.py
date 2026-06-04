from __future__ import annotations

import csv
import importlib.util
from pathlib import Path
import tempfile
import zipfile
import unittest
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
CLEAN_DIR = ROOT / "data" / "cleaned" / "graduate_outcomes"
OUTPUT_DIR = ROOT / "outputs" / "graduate_outcomes"
WORKBOOK = ROOT / "outputs" / "graduate_outcomes" / "graduate_outcomes_clean_data_package.xlsx"


def load_streaming_workbook_builder():
    builder_path = ROOT / "outputs" / "graduate_outcomes" / "build_workbook_streaming.py"
    spec = importlib.util.spec_from_file_location("graduate_outcome_streaming_workbook_builder", builder_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"could not load workbook builder from {builder_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def worksheet_data_row_count(workbook, sheet_name: str) -> int:
    return max(sum(1 for _ in workbook[sheet_name].iter_rows(values_only=True)) - 1, 0)


class GraduateOutcomeWorkbookPackageTest(unittest.TestCase):
    def test_add_worksheet_dimensions_retries_transient_replace_lock(self) -> None:
        builder = load_streaming_workbook_builder()
        with tempfile.TemporaryDirectory() as temp_dir:
            xlsx_path = Path(temp_dir) / "book.xlsx"
            with zipfile.ZipFile(xlsx_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                archive.writestr(
                    "xl/worksheets/sheet1.xml",
                    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                    b"<sheetData/></worksheet>",
                )

            original_replace = Path.replace
            replace_failures = 0

            def replace_once_after_transient_lock(self, target):
                nonlocal replace_failures
                if self.name == "book.dimensioned.tmp.xlsx" and Path(target) == xlsx_path and replace_failures == 0:
                    replace_failures += 1
                    raise PermissionError("simulated transient Windows file lock")
                return original_replace(self, target)

            with mock.patch.object(Path, "replace", replace_once_after_transient_lock):
                builder.add_worksheet_dimensions(xlsx_path, [(1, 1)])

            self.assertEqual(replace_failures, 1)
            with zipfile.ZipFile(xlsx_path) as archive:
                worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
            self.assertIn(b'<dimension ref="A1"/>', worksheet_xml)

    def test_workbook_worksheets_have_dimensions_for_read_only_loading(self) -> None:
        with zipfile.ZipFile(WORKBOOK) as archive:
            worksheet_paths = [
                name
                for name in archive.namelist()
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
            ]
            missing_dimensions = []
            for name in worksheet_paths:
                head = archive.open(name).read(4096)
                if b"<dimension " not in head:
                    missing_dimensions.append(name)

        self.assertFalse(
            missing_dimensions,
            f"worksheet XML should include dimension refs for read-only openpyxl loading: {missing_dimensions}",
        )

    def test_employment_report_tables_are_available_in_cleaned_package_and_workbook(self) -> None:
        sources_csv = CLEAN_DIR / "official_employment_report_sources.csv"
        metrics_csv = CLEAN_DIR / "official_employment_report_metrics.csv"
        self.assertTrue(sources_csv.exists(), "employment report source CSV should be in cleaned package")
        self.assertTrue(metrics_csv.exists(), "employment metric CSV should be in cleaned package")

        with sources_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            source_row_count = max(sum(1 for _ in csv.reader(handle)) - 1, 0)
        with metrics_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            metric_row_count = max(sum(1 for _ in csv.reader(handle)) - 1, 0)

        workbook = load_workbook(WORKBOOK, read_only=True)
        self.assertIn("Employment_Report_Sources", workbook.sheetnames)
        self.assertIn("Employment_Metrics", workbook.sheetnames)
        self.assertEqual(worksheet_data_row_count(workbook, "Employment_Report_Sources"), source_row_count)
        self.assertEqual(worksheet_data_row_count(workbook, "Employment_Metrics"), metric_row_count)

    def test_undergraduate_source_outcome_summary_is_available_in_cleaned_package_and_workbook(self) -> None:
        summary_csv = CLEAN_DIR / "undergraduate_school_outcome_summary.csv"
        self.assertTrue(summary_csv.exists(), "undergraduate source outcome summary CSV should be in cleaned package")

        with summary_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))

        self.assertGreater(len(rows), 0)
        self.assertIn("undergraduate_school", rows[0])
        self.assertIn("destination_school", rows[0])
        self.assertIn("record_count", rows[0])
        self.assertTrue(
            any(row["undergraduate_school"] == "北京林业大学" for row in rows),
            "summary should include schools that appear as undergraduate source schools",
        )

        workbook = load_workbook(WORKBOOK, read_only=True)
        self.assertIn("Undergrad_Source_Outcomes", workbook.sheetnames)
        self.assertEqual(worksheet_data_row_count(workbook, "Undergrad_Source_Outcomes"), len(rows))

    def test_remaining_official_recommendation_source_attempts_are_available_in_cleaned_package_and_workbook(self) -> None:
        attempts_csv = CLEAN_DIR / "official_recommendation_source_attempts.csv"
        coverage_csv = CLEAN_DIR / "official_recommendation_school_coverage.csv"
        self.assertTrue(attempts_csv.exists(), "remaining source-attempt CSV should be in cleaned package")

        with coverage_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            missing_schools = {
                row["school_name"]
                for row in csv.DictReader(handle)
                if row["has_official_recommendation_records"].lower() != "true"
            }
        with attempts_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))

        self.assertGreaterEqual(len(rows), len(missing_schools))
        self.assertTrue(
            missing_schools.issubset({row["school_name"] for row in rows}),
            "every remaining uncovered school should have at least one official-source attempt recorded",
        )
        resolved_attempts = [
            row for row in rows if row["school_name"] not in missing_schools and row["decision"] == "ingested"
        ]
        self.assertGreaterEqual(
            len(resolved_attempts),
            1,
            "source-attempt table may retain resolved original-gap rows marked as ingested",
        )
        self.assertIn("source_url", rows[0])
        self.assertIn("blocker_type", rows[0])
        self.assertIn("decision", rows[0])

        workbook = load_workbook(WORKBOOK, read_only=True)
        self.assertIn("Recommendation_Source_Attempts", workbook.sheetnames)
        self.assertEqual(worksheet_data_row_count(workbook, "Recommendation_Source_Attempts"), len(rows))

    def test_remaining_uncovered_recheck_log_is_available_in_package_and_workbook(self) -> None:
        expected_recheck_date = "2026-06-04"
        recheck_csv = OUTPUT_DIR / f"remaining_uncovered_recheck_{expected_recheck_date}.csv"
        coverage_csv = CLEAN_DIR / "official_recommendation_school_coverage.csv"
        self.assertTrue(recheck_csv.exists(), "remaining uncovered recheck CSV should be in output package")

        with coverage_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            missing_schools = [
                row["school_name"]
                for row in csv.DictReader(handle)
                if row["has_official_recommendation_records"].lower() != "true"
            ]
        with recheck_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)

        self.assertEqual(len(rows), len(missing_schools))
        self.assertFalse(
            any(None in row for row in rows),
            "remaining uncovered recheck CSV rows should not contain extra unnamed columns",
        )
        self.assertEqual([row["school_name"] for row in rows], missing_schools)
        self.assertEqual({row["recheck_date"] for row in rows}, {expected_recheck_date})
        self.assertIn("official_evidence_urls", rows[0])
        self.assertIn("ingestion_decision", rows[0])

        workbook = load_workbook(WORKBOOK, read_only=True)
        self.assertIn("Remaining_Uncovered_Recheck", workbook.sheetnames)
        self.assertEqual(worksheet_data_row_count(workbook, "Remaining_Uncovered_Recheck"), len(rows))
        workbook_dates = {
            row[0]
            for row in workbook["Remaining_Uncovered_Recheck"].iter_rows(
                min_row=2,
                values_only=True,
            )
            if row and row[0]
        }
        self.assertEqual(workbook_dates, {expected_recheck_date})

    def test_workbook_includes_data_dictionary_for_key_downstream_fields(self) -> None:
        workbook = load_workbook(WORKBOOK, read_only=True)
        self.assertIn("Data_Dictionary", workbook.sheetnames)

        dictionary_rows = list(workbook["Data_Dictionary"].iter_rows(values_only=True))
        self.assertGreater(len(dictionary_rows), 1)
        header = dictionary_rows[0]
        self.assertEqual(header[:4], ("sheet_name", "column_name", "data_type", "description"))
        keys = {
            (row[0], row[1])
            for row in dictionary_rows[1:]
            if row and len(row) >= 2 and row[0] and row[1]
        }

        self.assertIn(("Public_Records", "public_record_id"), keys)
        self.assertIn(("Coverage", "has_official_recommendation_records"), keys)
        self.assertIn(("Remaining_Uncovered_Recheck", "ingestion_decision"), keys)

    def test_workbook_includes_quality_checks_for_key_package_invariants(self) -> None:
        workbook = load_workbook(WORKBOOK, read_only=True)
        self.assertIn("Quality_Checks", workbook.sheetnames)

        rows = list(workbook["Quality_Checks"].iter_rows(values_only=True))
        self.assertGreater(len(rows), 1)
        self.assertEqual(rows[0][:5], ("check_name", "actual", "expected", "status", "notes"))
        checks = {row[0]: row for row in rows[1:] if row and row[0]}

        self.assertEqual(checks["coverage_target_school_count"][3], "PASS")
        self.assertEqual(checks["public_records_row_count"][3], "PASS")
        self.assertEqual(checks["remaining_uncovered_school_count"][3], "PASS")

    def test_cupl_web_visible_recommendation_records_are_in_cleaned_coverage(self) -> None:
        coverage_csv = CLEAN_DIR / "official_recommendation_school_coverage.csv"
        master_csv = CLEAN_DIR / "master_records_clean.csv"

        with coverage_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            coverage_by_school = {row["school_name"]: row for row in csv.DictReader(handle)}
        cupl = coverage_by_school["中国政法大学"]
        self.assertEqual(cupl["has_official_recommendation_records"], "True")
        self.assertGreaterEqual(int(cupl["official_recommendation_record_count"]), 315)

        with master_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            cupl_rows = [
                row
                for row in csv.DictReader(handle)
                if row["school_name"] == "中国政法大学"
                and row["source_url"].startswith("https://yjsy.cupl.edu.cn/__local/F/A7/C1/")
            ]
        self.assertEqual(len(cupl_rows), 315)
        self.assertTrue(
            all("official_web_visible_pdf_text" in row["remarks"] for row in cupl_rows)
        )

    def test_nufe_doctor_supplement_image_records_are_in_cleaned_coverage(self) -> None:
        coverage_csv = CLEAN_DIR / "official_recommendation_school_coverage.csv"
        master_csv = CLEAN_DIR / "master_records_clean.csv"

        with coverage_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            coverage_by_school = {row["school_name"]: row for row in csv.DictReader(handle)}
        nufe = coverage_by_school["南京财经大学"]
        self.assertEqual(nufe["has_official_recommendation_records"], "True")
        self.assertGreaterEqual(int(nufe["official_recommendation_record_count"]), 1)

        with master_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            nufe_rows = [
                row
                for row in csv.DictReader(handle)
                if row["school_name"] == "南京财经大学"
                and row["source_url"] == "https://yjsc.nufe.edu.cn/__local/A/C7/31/38635DF087259669397CAD93EEC_C1C13246_111C3.png"
            ]
        self.assertEqual(len(nufe_rows), 1)
        self.assertEqual(nufe_rows[0]["person_name"], "周*帆")
        self.assertIn("source_image_transcribed true", nufe_rows[0]["remarks"])
        self.assertNotIn("放弃拟录取", nufe_rows[0]["remarks"])

    def test_ymu_web_visible_recommendation_records_are_in_cleaned_coverage(self) -> None:
        coverage_csv = CLEAN_DIR / "official_recommendation_school_coverage.csv"
        master_csv = CLEAN_DIR / "master_records_clean.csv"

        with coverage_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            coverage_by_school = {row["school_name"]: row for row in csv.DictReader(handle)}
        ymu = coverage_by_school["云南民族大学"]
        self.assertEqual(ymu["has_official_recommendation_records"], "True")
        self.assertGreaterEqual(int(ymu["official_recommendation_record_count"]), 8)

        with master_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            ymu_rows = [
                row
                for row in csv.DictReader(handle)
                if row["school_name"] == "云南民族大学"
                and row["source_url"] == "https://web.ymu.edu.cn/__local/5/86/BA/EC4D50E9EABCAC4B3283EF800D2_ED4BFA46_151C7.pdf"
            ]
        self.assertEqual(len(ymu_rows), 8)
        self.assertTrue(all("official_web_visible_pdf_text true" in row["remarks"] for row in ymu_rows))
        self.assertTrue(all("local_pdf_fetch HTTP 521 __jsl_clearance_s" in row["remarks"] for row in ymu_rows))

    def test_dlou_web_visible_adjustment_records_are_in_cleaned_coverage(self) -> None:
        coverage_csv = CLEAN_DIR / "official_recommendation_school_coverage.csv"
        master_csv = CLEAN_DIR / "master_records_clean.csv"

        with coverage_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            coverage_by_school = {row["school_name"]: row for row in csv.DictReader(handle)}
        dlou = coverage_by_school["大连海洋大学"]
        self.assertEqual(dlou["has_official_recommendation_records"], "True")
        self.assertGreaterEqual(int(dlou["official_recommendation_record_count"]), 51)

        with master_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            dlou_rows = [
                row
                for row in csv.DictReader(handle)
                if row["school_name"] == "大连海洋大学"
                and row["source_url"] == "https://www.dlou.edu.cn/_upload/article/files/e7/fb/e8bdbb4040f4a9987ac2d09f5ee4/d048fe33-6d3d-463f-8c74-203b3176890b.pdf"
            ]
        self.assertEqual(len(dlou_rows), 51)
        self.assertTrue(all("official_web_visible_pdf_text true" in row["remarks"] for row in dlou_rows))
        self.assertTrue(all("local_pdf_fetch HTTP 404 196_byte_html" in row["remarks"] for row in dlou_rows))
        self.assertTrue(any("note 少数民族骨干计划" in row["remarks"] for row in dlou_rows))

    def test_zcmu_official_xlsx_master_admission_records_are_in_cleaned_coverage(self) -> None:
        coverage_csv = CLEAN_DIR / "official_recommendation_school_coverage.csv"
        master_csv = CLEAN_DIR / "master_records_clean.csv"
        source_url = "https://yjsgl.zcmu.edu.cn/storage/uploads/file/20251020/1760944454983762.xlsx"

        with coverage_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            coverage_by_school = {row["school_name"]: row for row in csv.DictReader(handle)}
        zcmu = coverage_by_school["浙江中医药大学"]
        self.assertEqual(zcmu["has_official_recommendation_records"], "True")
        self.assertGreaterEqual(int(zcmu["official_recommendation_record_count"]), 1715)

        with master_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            zcmu_rows = [
                row
                for row in csv.DictReader(handle)
                if row["school_name"] == "浙江中医药大学"
                and row["source_url"] == source_url
            ]
        self.assertEqual(len(zcmu_rows), 1715)
        self.assertTrue(all(row["document_type"] == "postgraduate_admission_list" for row in zcmu_rows))
        self.assertTrue(all("official_xlsx_download true" in row["remarks"] for row in zcmu_rows))

    def test_nwnu_jykxxy_official_pdf_records_are_in_cleaned_coverage(self) -> None:
        coverage_csv = CLEAN_DIR / "official_recommendation_school_coverage.csv"
        master_csv = CLEAN_DIR / "master_records_clean.csv"
        source_url = (
            "https://jykxxy.nwnu.edu.cn/_upload/article/files/bd/c5/"
            "289f95b04bfe86e89495ed3239ef/1369c3c3-8eeb-42df-b75b-b36687ef3d13.pdf"
        )

        with master_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            nwnu_rows = [
                row
                for row in csv.DictReader(handle)
                if row["source_url"] == source_url
            ]
        self.assertEqual(len(nwnu_rows), 106)
        school_name = nwnu_rows[0]["school_name"]

        with coverage_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            coverage_by_school = {row["school_name"]: row for row in csv.DictReader(handle)}
        nwnu = coverage_by_school[school_name]
        self.assertEqual(nwnu["has_official_recommendation_records"], "True")
        self.assertGreaterEqual(int(nwnu["official_recommendation_record_count"]), 106)

        self.assertTrue(all(row["document_type"] == "postgraduate_admission_list" for row in nwnu_rows))
        self.assertTrue(all(row["college"] == "教育科学学院" for row in nwnu_rows))
        self.assertTrue(all("official_pdf_download true" in row["remarks"] for row in nwnu_rows))


    def test_cmu_doctor_minority_official_html_records_are_in_cleaned_coverage(self) -> None:
        coverage_csv = CLEAN_DIR / "official_recommendation_school_coverage.csv"
        master_csv = CLEAN_DIR / "master_records_clean.csv"
        source_url = "https://www.cmu.edu.cn/cmuyjs/info/1901/9841.htm"

        with master_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            cmu_rows = [
                row
                for row in csv.DictReader(handle)
                if row["source_url"] == source_url
            ]
        self.assertEqual(len(cmu_rows), 3)
        school_name = cmu_rows[0]["school_name"]

        with coverage_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            coverage_by_school = {row["school_name"]: row for row in csv.DictReader(handle)}
        cmu = coverage_by_school[school_name]
        self.assertEqual(cmu["has_official_recommendation_records"], "True")
        self.assertGreaterEqual(int(cmu["official_recommendation_record_count"]), 3)

        self.assertTrue(all(row["document_type"] == "postgraduate_admission_list" for row in cmu_rows))
        self.assertTrue(all("official_html_table true" in row["remarks"] for row in cmu_rows))
        self.assertTrue(all("minority_backbone_plan true" in row["remarks"] for row in cmu_rows))

    def test_bift_doctor_admission_first_official_pdf_records_are_in_cleaned_coverage(self) -> None:
        coverage_csv = CLEAN_DIR / "official_recommendation_school_coverage.csv"
        master_csv = CLEAN_DIR / "master_records_clean.csv"
        source_url = "https://yjs.bift.edu.cn/docs//2026-06/85b36649ec7b43d79acd889705f02a2d.pdf"

        with master_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            bift_rows = [
                row
                for row in csv.DictReader(handle)
                if row["source_url"] == source_url
            ]
        self.assertEqual(len(bift_rows), 35)
        school_name = bift_rows[0]["school_name"]

        with coverage_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            coverage_by_school = {row["school_name"]: row for row in csv.DictReader(handle)}
        bift = coverage_by_school[school_name]
        self.assertEqual(bift["has_official_recommendation_records"], "True")
        self.assertGreaterEqual(int(bift["official_recommendation_record_count"]), 35)

        self.assertTrue(all(row["document_type"] == "postgraduate_admission_list" for row in bift_rows))
        self.assertTrue(all("official_pdf_download true" in row["remarks"] for row in bift_rows))
        self.assertTrue(all("admission_status 拟录取" in row["remarks"] for row in bift_rows))


if __name__ == "__main__":
    unittest.main()
