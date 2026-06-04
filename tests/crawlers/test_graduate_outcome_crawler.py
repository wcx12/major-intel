import csv
import json
import os
import ssl
import sys
import threading
import types
import unittest
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch
from urllib.error import URLError
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SRC_ROOT = PROJECT_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

import major_intel.crawlers.graduate_outcome_crawler as crawler
from major_intel.crawlers.graduate_outcome_crawler import (
    FetchResponse,
    build_discovery_tasks,
    classify_document,
    clean_records_to_outputs,
    collect_chsi_bulletin_seeds,
    collect_chsi_school_index,
    collect_official_site_seeds,
    collect_school_site_index,
    collect_search_results,
    crawl_seed_documents,
    extract_candidate_links,
    official_site_candidate_portal_urls,
    parse_official_site_portal_links,
    parse_school_site_index_rows,
    parse_chsi_bulletin_seed_rows,
    parse_chsi_school_index,
    parse_chsi_school_info_bulletin_urls,
    parse_bing_search_results,
    parse_bing_rss_search_results,
    parse_duckduckgo_search_results,
    parse_excel_records,
    parse_html_records,
    read_official_site_csv,
    read_seed_csv,
    select_seed_rows_from_search_results,
    slice_seed_rows,
    write_discovery_tasks_csv,
    write_seed_rows_csv,
    write_records_csv,
)


HTML_TABLE = """
<html>
  <head><title>2026年接收推免生拟录取名单公示</title></head>
  <body>
    <table>
      <tr>
        <th>姓名</th><th>本科毕业院校</th><th>拟录取学院</th>
        <th>拟录取专业</th><th>备注</th>
      </tr>
      <tr>
        <td>张三</td><td>东北农业大学</td><td>计算机学院</td>
        <td>计算机科学与技术</td><td>推免</td>
      </tr>
    </table>
  </body>
</html>
"""


class GraduateOutcomeCrawlerTests(unittest.TestCase):
    def test_classify_document_distinguishes_recommendation_and_admission_lists(self):
        recommendation = classify_document(
            "关于2026届推荐免试攻读研究生名单的公示",
            "https://jwc.example.edu.cn/info/1001/1.htm",
            "",
        )
        admission = classify_document(
            "2025年硕士研究生拟录取名单公示",
            "https://yz.example.edu.cn/info/1001/2.htm",
            "附件含本科毕业院校",
        )
        exempt_admission = classify_document(
            "长江大学2026年免试攻读研究生拟录取名单公示（一）",
            "https://gs.yangtzeu.edu.cn/list.pdf",
            "",
        )

        self.assertEqual(recommendation["document_type"], "recommendation_exemption_list")
        self.assertEqual(admission["document_type"], "postgraduate_admission_list")
        self.assertEqual(exempt_admission["document_type"], "recommendation_exemption_list")
        self.assertIn("推荐免试", recommendation["matched_keywords"])
        self.assertIn("拟录取名单", admission["matched_keywords"])

    def test_classify_document_treats_candidate_admission_list_as_postgraduate_admission(self):
        classification = classify_document(
            "山东财经大学2026年硕士研究生拟录取考生名单公示",
            "https://yjszs.sdufe.edu.cn/info/1034/2829.htm",
            "",
        )

        self.assertEqual(classification["document_type"], "postgraduate_admission_list")

    def test_classify_document_treats_tuimian_admission_as_incoming_recommendation(self):
        classification = classify_document(
            "河北工业大学2026年推免研究生拟录取名单公示",
            "https://yjs.hebut.edu.cn/info/1.htm",
            "",
        )

        self.assertEqual(classification["document_type"], "incoming_recommendation_admission_list")

    def test_classify_document_treats_bracketed_tuimian_student_admission_as_incoming(self):
        classification = classify_document(
            "山西大学2026年推免生（含直博生）拟录取名单",
            "https://yjszsw.sxu.edu.cn/docs/2025-10/list.pdf",
            "",
        )

        self.assertEqual(classification["document_type"], "incoming_recommendation_admission_list")

    def test_extract_candidate_links_resolves_same_site_relevant_links_and_attachments(self):
        html = """
        <a href="/info/1001/notice.htm">2026年接收推免生拟录取名单</a>
        <a href="../upload/result.xlsx">附件：硕士研究生拟录取名单</a>
        <a href="https://other.example.edu.cn/result.pdf">外站拟录取名单</a>
        <a href="/news/campus.htm">校园新闻</a>
        """

        links = extract_candidate_links(
            html,
            "https://yz.example.edu.cn/path/index.htm",
            allowed_domains={"yz.example.edu.cn"},
        )

        self.assertEqual(
            [link.url for link in links],
            [
                "https://yz.example.edu.cn/info/1001/notice.htm",
                "https://yz.example.edu.cn/upload/result.xlsx",
            ],
        )
        self.assertEqual(links[1].link_kind, "attachment")

    def test_extract_candidate_links_skips_malformed_href_values(self):
        html = """
        <a href="（网址：https://bad.example.edu.cn/result.htm）">硕士研究生拟录取名单</a>
        <a href="/valid/result.htm">硕士研究生拟录取名单</a>
        """

        links = extract_candidate_links(
            html,
            "https://yz.example.edu.cn/path/index.htm",
            allowed_domains={"yz.example.edu.cn"},
        )

        self.assertEqual(
            [link.url for link in links],
            ["https://yz.example.edu.cn/valid/result.htm"],
        )

    def test_extract_candidate_links_can_keep_generic_attachments_from_relevant_page(self):
        html = """
        <a href="/upload/result.xlsx">附件1</a>
        <a href="/news/campus.htm">校园新闻</a>
        """

        links = extract_candidate_links(
            html,
            "https://yz.example.edu.cn/path/index.htm",
            allowed_domains={"yz.example.edu.cn"},
            include_all_attachments=True,
        )

        self.assertEqual(
            [link.url for link in links],
            ["https://yz.example.edu.cn/upload/result.xlsx"],
        )
        self.assertEqual(links[0].document_type, "unknown")

    def test_parse_chsi_school_index_extracts_school_info_links(self):
        html = """
        <a href="/sch/schoolInfo--schId-367878.dhtml"> 北京大学 </a>
        <a href="/sch/schoolInfo--schId-367899.dhtml"> 清华大学 </a>
        """

        schools = parse_chsi_school_index(html, "https://yzst.chsi.com.cn/sch/")

        self.assertEqual(
            schools,
            [
                {
                    "chsi_school_name": "北京大学",
                    "chsi_sch_id": "367878",
                    "chsi_school_url": "https://yzst.chsi.com.cn/sch/schoolInfo--schId-367878.dhtml",
                },
                {
                    "chsi_school_name": "清华大学",
                    "chsi_sch_id": "367899",
                    "chsi_school_url": "https://yzst.chsi.com.cn/sch/schoolInfo--schId-367899.dhtml",
                },
            ],
        )

    def test_parse_chsi_school_info_bulletin_urls_extracts_list_pages(self):
        html = """
        <h4>信息公告</h4>
        <a href="/sch/listBulletin--schId-367878%2CcategoryId-481379.dhtml">更多</a>
        <h4>调剂办法</h4>
        <a href="/sch/listBulletin--schId-367878%2CcategoryId-481383.dhtml">更多</a>
        """

        urls = parse_chsi_school_info_bulletin_urls(
            html,
            "https://yzst.chsi.com.cn/sch/schoolInfo--schId-367878.dhtml",
        )

        self.assertEqual(
            urls,
            [
                "https://yzst.chsi.com.cn/sch/listBulletin--schId-367878%2CcategoryId-481379.dhtml",
                "https://yzst.chsi.com.cn/sch/listBulletin--schId-367878%2CcategoryId-481383.dhtml",
            ],
        )

    def test_parse_chsi_bulletin_seed_rows_filters_relevant_bulletins(self):
        html = """
        <table>
          <tr><td>1</td><td><a href="/sch/viewBulletin--schId-1,infoId-10.dhtml">公示2025年统考硕士研究生拟录取名单</a></td><td>2025-05-08</td></tr>
          <tr><td>2</td><td><a href="/sch/viewBulletin--schId-1,infoId-11.dhtml">招生简章</a></td><td>2025-09-01</td></tr>
          <tr><td>3</td><td><a href="/sch/viewBulletin--schId-1,infoId-12.dhtml">公示2025年拟录取推免生名单</a></td><td>2024-10-14</td></tr>
          <tr><td>4</td><td><a href="/sch/viewBulletin--schId-1,infoId-13.dhtml">2025年接收推免生办法</a></td><td>2024-09-01</td></tr>
        </table>
        """

        seeds = parse_chsi_bulletin_seed_rows(
            html,
            "https://yzst.chsi.com.cn/sch/listBulletin--schId-1%2CcategoryId-2.dhtml",
            school_name="军事科学院",
        )

        self.assertEqual(len(seeds), 2)
        self.assertEqual(seeds[0]["school_name"], "军事科学院")
        self.assertEqual(seeds[0]["source_type"], "postgraduate_admission")
        self.assertEqual(seeds[0]["year"], 2025)
        self.assertEqual(seeds[1]["source_type"], "incoming_recommendation")

    def test_parse_chsi_bulletin_seed_rows_ignores_implausible_year_codes(self):
        html = """
        <a href="/sch/viewBulletin--schId-1,infoId-10.dhtml">单位代码2095 2025年硕士研究生拟录取名单公示</a>
        """

        seeds = parse_chsi_bulletin_seed_rows(
            html,
            "https://yzst.chsi.com.cn/sch/listBulletin--schId-1%2CcategoryId-2.dhtml",
            school_name="样例大学",
        )

        self.assertEqual(seeds[0]["year"], 2025)

    def test_parse_html_records_extracts_major_name_paragraph_lists(self):
        html = """
        <html><body>
          <p>一、物流工程专业</p>
          <p>高婕、倪嘉浩、金哲玉</p>
          <p>二、机器人工程专业</p>
          <p>王雅雯、范青芸</p>
        </body></html>
        """
        document = {
            "school_name": "上海海洋大学",
            "year": 2026,
            "document_type": "recommendation_exemption_list",
            "source_url": "https://gcxy.shou.edu.cn/page.htm",
            "title": "工程学院推免拟推荐名单公示",
        }

        records = parse_html_records(html, document)

        self.assertEqual([record["person_name"] for record in records], ["高婕", "倪嘉浩", "金哲玉", "王雅雯", "范青芸"])
        self.assertEqual(records[0]["major"], "物流工程")
        self.assertEqual(records[-1]["major"], "机器人工程")

    def test_parse_html_records_extracts_gzucm_nested_recommendation_table(self):
        html = """
        <html><body>
          <table>
            <tr>
              <td>
                <table>
                  <tr>
                    <td>院所代码</td><td>院所名称</td><td>考生姓名</td>
                    <td>录取专业代码</td><td>录取专业名称</td>
                    <td>研究方向代码</td><td>研究方向名称</td><td>接收导师</td>
                  </tr>
                  <tr>
                    <td>202</td><td>中药学院</td><td>周顺</td>
                    <td>100700</td><td>药学</td>
                    <td>00</td><td>不区分研究方向</td><td>靳红磊</td>
                  </tr>
                </table>
              </td>
            </tr>
          </table>
        </body></html>
        """
        document = {
            "school_name": "广州中医药大学",
            "year": 2026,
            "document_type": "incoming_recommendation_admission_list",
            "source_url": "https://yjsy.gzucm.edu.cn/info/1004/17006.htm",
            "title": "广州中医药大学2026年推免生拟录取名单公示",
        }

        records = parse_html_records(html, document)

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "周顺")
        self.assertEqual(records[0]["college"], "中药学院")
        self.assertEqual(records[0]["major"], "100700")
        self.assertEqual(records[0]["admission_major"], "100700 药学")
        self.assertIn("direction 00 不区分研究方向", records[0]["remarks"])
        self.assertIn("advisor 靳红磊", records[0]["remarks"])

    def test_parse_html_records_extracts_space_separated_major_name_lists(self):
        html = """
        <html><body>
          <p>经学院推荐免试研究生工作领导小组审核，拟推荐名单如下：</p>
          <p><strong>英语专业：</strong>年雨彤 耿琪茹 周佳怡 苏 晴 邵诗琪</p>
          <p>王铭珺 房小珂 孙一涵 路 笑</p>
          <p><strong>日语专业：</strong>张文静 郭 洁 刘飞燕</p>
          <p>若有名额增加或空缺，递补次序 綦筱涵、戴舒。</p>
        </body></html>
        """
        document = {
            "school_name": "南京农业大学",
            "year": 2026,
            "document_type": "recommendation_exemption_list",
            "source_url": "https://foreign.njau.edu.cn/info/1053/6846.htm",
            "title": "外国语学院2026年推荐优秀应届本科毕业生免试攻读研究生拟推荐名单公示",
        }

        records = parse_html_records(html, document)

        self.assertEqual(
            [(record["major"], record["person_name"]) for record in records],
            [
                ("英语", "年雨彤"),
                ("英语", "耿琪茹"),
                ("英语", "周佳怡"),
                ("英语", "苏晴"),
                ("英语", "邵诗琪"),
                ("英语", "王铭珺"),
                ("英语", "房小珂"),
                ("英语", "孙一涵"),
                ("英语", "路笑"),
                ("日语", "张文静"),
                ("日语", "郭洁"),
                ("日语", "刘飞燕"),
            ],
        )

    def test_parse_html_records_extracts_standalone_names_after_major_heading(self):
        html = """
        <html><body>
          <p>农学专业：</p>
          <p>严婧灵</p>
          <p>余欣蕊</p>
          <p>张 梦</p>
          <p>种子科学与工程专业：</p>
          <p>赵子杰</p>
          <p>任 婧</p>
          <p>其中，国家生命学院直博生专项 3 人，入选情况如下：</p>
          <p>国家生命学院直博生专项</p>
          <p>赵晨捷、王泽鹤、靳思语</p>
        </body></html>
        """
        document = {
            "school_name": "南京农业大学",
            "year": 2026,
            "document_type": "recommendation_exemption_list",
            "source_url": "https://nx.njau.edu.cn/info/1112/10565.htm",
            "title": "农学院2026年推荐优秀应届本科毕业生免试攻读研究生拟推荐名单公示",
        }

        records = parse_html_records(html, document)

        self.assertEqual(
            [(record["major"], record["person_name"]) for record in records],
            [
                ("农学", "严婧灵"),
                ("农学", "余欣蕊"),
                ("农学", "张梦"),
                ("种子科学与工程", "赵子杰"),
                ("种子科学与工程", "任婧"),
            ],
        )

    def test_parse_html_records_does_not_extract_names_from_prose_after_generic_major_heading(self):
        html = """
        <html><body>
          <p>学院专业</p>
          <p>我院遵循全面衡量、择优遴选、宁缺毋滥的原则，现将拟录取名单公示。</p>
        </body></html>
        """
        document = {
            "school_name": "四川农业大学",
            "year": 2026,
            "document_type": "incoming_recommendation_admission_list",
            "source_url": "https://jdxy.sicau.edu.cn/info/1033/3874.htm",
            "title": "机电学院2026年推免研究生招生拟录取名单公示",
        }

        records = parse_html_records(html, document)

        self.assertEqual(records, [])

    def test_parse_html_records_ignores_navigation_items_when_extracting_major_name_paragraphs(self):
        html = """
        <html><body>
          <div class="wp-navi-aside">
            <ul>
              <li>研究生培养</li>
              <li>学科建设</li>
              <li>招生专业</li>
              <li>培养动态</li>
              <li>学术之星</li>
            </ul>
          </div>
          <div class="article">
            <table>
              <tr><td>序号</td><td>姓名</td><td>本科专业名称</td><td>申报类别</td><td>综合名次</td></tr>
              <tr><td>1</td><td>赵俊涛</td><td>化学工程与工艺</td><td>A类计划</td><td>1</td></tr>
              <tr><td>2</td><td>王天怡</td><td>化学工程与工艺</td><td>A类计划</td><td>2</td></tr>
            </table>
          </div>
        </body></html>
        """
        document = {
            "school_name": "武汉工程大学",
            "year": 2026,
            "document_type": "recommendation_exemption_list",
            "source_url": "https://mep.wit.edu.cn/info/1102/52172.htm",
            "title": "化工与制药学院拟推荐2026届优秀应届本科毕业生免试攻读硕士学位研究生名单公示",
        }

        records = parse_html_records(html, document)

        self.assertEqual([record["person_name"] for record in records], ["赵俊涛", "王天怡"])

    def test_parse_html_records_extracts_structured_major_name_lines(self):
        html = """
        <html><body>
          <div class="v_news_content">
            <p>现将推免名单公告如下：</p>
            <p>专业名称</p>
            <p>推免名单（共计65人）</p>
            <p>数字文化试验班</p>
            <p>（3人）</p>
            <p>蒋尚蓉、周柃妍、施雨霏</p>
            <p>图书馆学</p>
            <p>（2人）</p>
            <p>朱紫霖、陈淑杰</p>
            <p>信息管理与信息系统</p>
            <p>（2人）</p>
            <p>袁可（工程硕博专项）、张令</p>
            <p>特此公告</p>
          </div>
        </body></html>
        """
        document = {
            "school_name": "武汉大学",
            "year": 2026,
            "document_type": "recommendation_exemption_list",
            "source_url": "https://sim.whu.edu.cn/info/1776/108742.htm",
            "title": "信息管理学院关于2026届本科毕业生免试攻读硕士学位研究生推荐名单的公告",
        }

        records = parse_html_records(html, document)

        self.assertEqual(
            [(record["major"], record["person_name"]) for record in records],
            [
                ("数字文化试验班", "蒋尚蓉"),
                ("数字文化试验班", "周柃妍"),
                ("数字文化试验班", "施雨霏"),
                ("图书馆学", "朱紫霖"),
                ("图书馆学", "陈淑杰"),
                ("信息管理与信息系统", "袁可"),
                ("信息管理与信息系统", "张令"),
            ],
        )

    def test_parse_html_records_extracts_aufe_replenishment_key_value_notice(self):
        html = """
        <html><body>
          <div class="v_news_content">
            <p>现递补吴韩获得推荐资格，递补学生信息如下：</p>
            <p>姓名：</p>
            <p>吴韩</p>
            <p>专业名称：</p>
            <p>审计学</p>
            <p>综合成绩：63.12</p>
            <p>排名人数：89</p>
          </div>
        </body></html>
        """
        document = {
            "school_name": "安徽财经大学",
            "year": 2026,
            "document_type": "recommendation_exemption_list",
            "source_url": "https://yz.aufe.edu.cn/2025/0918/c13923a238434/page.htm",
            "title": "关于会计学院2026年推荐优秀应届本科毕业生免试攻读硕士学位研究生递补名单公示",
        }

        records = parse_html_records(html, document)

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "吴韩")
        self.assertEqual(records[0]["major"], "审计学")
        self.assertIn("综合成绩 63.12", records[0]["remarks"])

    def test_parse_html_records_skips_recommendation_workgroup_member_lists(self):
        html = """
        <html><body>
          <div class="v_news_content">
            <p>学校成立推荐优秀应届本科毕业生免试攻读硕士学位研究生工作领导小组。</p>
            <p>成员：万光彩、储德银、刘晓光</p>
            <p>副组长：朱红军</p>
            <p>成员：</p>
            <p>周加来、张焕明</p>
            <p>工作职责：全面负责推荐免试研究生工作。</p>
            <p>特此通知。</p>
            <p>各学院应按照通知要求做好推荐工作。</p>
          </div>
        </body></html>
        """
        document = {
            "school_name": "安徽财经大学",
            "year": 2026,
            "document_type": "recommendation_exemption_list",
            "source_url": "https://yz.aufe.edu.cn/notice.htm",
            "title": "安徽财经大学关于做好2026年推荐优秀应届本科毕业生免试攻读硕士学位研究生工作的通知",
        }

        records = parse_html_records(html, document)

        self.assertEqual(records, [])

    def test_parse_html_records_ignores_medical_exam_items_as_candidate_names(self):
        html = """
        <html><body>
          <div class="v_news_content">
            <p>决定递补录取考生符慧菲，现予以公示，具体名单请见附件。</p>
            <p>体检必检项目：</p>
            <p>（一）常规检查：色觉检查、身高、体重、血压、胸片；</p>
            <p>（二）化验检查：血常规、肝功能；</p>
          </div>
        </body></html>
        """
        document = {
            "school_name": "海南医科大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "source_url": "https://www.muhn.edu.cn/zsw/info/1091/10974.htm",
            "title": "海南医科大学2026年硕士研究生招生考试第一志愿递补拟录取名单公示",
        }

        records = parse_html_records(html, document)

        self.assertEqual(records, [])

    def test_parse_html_records_extracts_plain_recommendation_name_list_after_notice_intro(self):
        html = """
        <html><body>
          <div class="nry_main">
            <p>楚才学院推荐2026届优秀本科毕业生免试攻读研究生名单公示</p>
            <p>作者： 编辑：test2 发布时间：2025年09月09日 点击次数：</p>
            <p>根据有关通知，现将拟推荐免试攻读研究生学生名单公示如下：</p>
            <p>史雪怡、熊予康、王忠平、潘子怡</p>
            <p>候补名单：</p>
            <p>1.邓紫浠 2.黄钰惠</p>
            <p>特此公示。</p>
          </div>
        </body></html>
        """
        document = {
            "school_name": "湖北大学",
            "year": 2026,
            "document_type": "recommendation_exemption_list",
            "source_url": "https://ccxy.hubu.edu.cn/info/1009/4691.htm",
            "title": "楚才学院推荐2026届优秀本科毕业生免试攻读研究生名单公示",
        }

        records = parse_html_records(html, document)

        self.assertEqual(
            [(record["person_name"], record["major"]) for record in records],
            [("史雪怡", ""), ("熊予康", ""), ("王忠平", ""), ("潘子怡", "")],
        )

    def test_collect_chsi_school_index_writes_index_csv(self):
        html = '<a href="/sch/schoolInfo--schId-367878.dhtml"> 北京大学 </a>'
        requested_urls = []

        def fake_fetch(url, timeout_seconds=20):
            requested_urls.append(url)
            return FetchResponse(url=url, status_code=200, content_type="text/html; charset=utf-8", content=html.encode("utf-8"))

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "chsi_schools.csv"
            summary = collect_chsi_school_index(
                output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                max_pages=1,
            )
            rows = list(csv.DictReader(output_path.open("r", encoding="utf-8-sig", newline="")))

        self.assertEqual(summary["schools_written"], 1)
        self.assertEqual(rows[0]["chsi_school_name"], "北京大学")
        self.assertEqual(requested_urls, ["https://yzst.chsi.com.cn/sch/"])

    def test_collect_chsi_school_index_supports_start_page(self):
        html = '<a href="/sch/schoolInfo--schId-367899.dhtml"> 清华大学 </a>'
        requested_urls = []

        def fake_fetch(url, timeout_seconds=20):
            requested_urls.append(url)
            return FetchResponse(url=url, status_code=200, content_type="text/html; charset=utf-8", content=html.encode("utf-8"))

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "chsi_schools.csv"
            collect_chsi_school_index(
                output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                start_page=2,
                max_pages=1,
                page_size=20,
            )

        self.assertEqual(requested_urls, ["https://yzst.chsi.com.cn/sch/?start=40"])

    def test_parse_school_site_index_rows_matches_known_recommended_schools(self):
        html = """
        <a href="https://www.pku.edu.cn/">北京大学</a>
        <a href="https://www.example.edu.cn/">普通职业学院</a>
        <a href="/fuwu/other">本页导航</a>
        """
        schools = [
            {
                "id": "1",
                "name": "北京大学",
                "province": "北京",
                "level": "本科",
                "tags": '["保研"]',
            },
            {
                "id": "2",
                "name": "普通职业学院",
                "province": "浙江",
                "level": "专科",
                "tags": "[]",
            },
        ]

        rows = parse_school_site_index_rows(
            html,
            "https://laosheng.top/fuwu/yuanxiao",
            schools,
            recommended_only=True,
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["school_name"], "北京大学")
        self.assertEqual(rows[0]["official_url"], "https://www.pku.edu.cn/")
        self.assertEqual(rows[0]["eligibility_hint"], "recommended")

    def test_parse_official_site_portal_links_keeps_teaching_portals_on_same_school_root(self):
        html = """
        <a href="https://dean.pku.edu.cn/web/notice.php">教务部</a>
        <a href="https://news.other.edu.cn/">校外新闻</a>
        <a href="/about">学校概况</a>
        """

        links = parse_official_site_portal_links(html, "https://www.pku.edu.cn/")

        self.assertEqual(links, ["https://dean.pku.edu.cn/web/notice.php"])

    def test_official_site_candidate_portal_urls_include_common_teaching_subdomains(self):
        urls = official_site_candidate_portal_urls("https://www.pku.edu.cn/")

        self.assertIn("https://dean.pku.edu.cn/", urls)
        self.assertIn("https://jwc.pku.edu.cn/", urls)
        self.assertIn("https://www.pku.edu.cn/jwc/", urls)

    def test_collect_school_site_index_writes_official_site_csv(self):
        html = '<a href="https://www.pku.edu.cn/">北京大学</a>'
        schools = [
            {"id": "1", "name": "北京大学", "province": "北京", "level": "本科", "tags": '["保研"]'}
        ]

        def fake_fetch(url, timeout_seconds=20):
            return FetchResponse(
                url=url,
                status_code=200,
                content_type="text/html; charset=utf-8",
                content=html.encode("utf-8"),
            )

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "sites.csv"
            summary = collect_school_site_index(
                schools,
                output_path,
                source_url="https://laosheng.top/fuwu/yuanxiao",
                fetcher=fake_fetch,
                recommended_only=True,
            )
            with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(summary["schools_considered"], 1)
        self.assertEqual(summary["sites_written"], 1)
        self.assertEqual(rows[0]["school_name"], "北京大学")
        self.assertEqual(rows[0]["official_url"], "https://www.pku.edu.cn/")

    def test_collect_official_site_seeds_follows_teaching_portal_links(self):
        pages = {
            "https://www.pku.edu.cn/": FetchResponse(
                url="https://www.pku.edu.cn/",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content='<a href="https://dean.pku.edu.cn/web/notice.php">教务部</a>'.encode("utf-8"),
            ),
            "https://dean.pku.edu.cn/web/notice.php": FetchResponse(
                url="https://dean.pku.edu.cn/web/notice.php",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content='<a href="/info/1.htm">北京大学2026届推荐免试研究生名单公示</a>'.encode("utf-8"),
            ),
        }

        def fake_fetch(url, timeout_seconds=20):
            return pages[url]

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "seeds.csv"
            summary = collect_official_site_seeds(
                [{"school_name": "北京大学", "official_url": "https://www.pku.edu.cn/"}],
                output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
            )
            with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(summary["seeds_written"], 1)
        self.assertEqual(rows[0]["school_name"], "北京大学")
        self.assertEqual(rows[0]["start_url"], "https://dean.pku.edu.cn/info/1.htm")

    def helper_collect_official_site_seeds_can_probe_common_teaching_subdomains_corrupt_fixture(self):
        pages = {
            "https://www.pku.edu.cn/": FetchResponse(
                url="https://www.pku.edu.cn/",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content=b"<html><body>home</body></html>",
            ),
            "https://dean.pku.edu.cn/": FetchResponse(
                url="https://dean.pku.edu.cn/",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content='<a href="/info/1.htm">鍖椾含澶у2026灞婃帹鑽愬厤璇曠爺绌剁敓鍚嶅崟鍏ず</a>'.encode("utf-8"),
            ),
        }

        def fake_fetch(url, timeout_seconds=20):
            if url not in pages:
                raise crawler.FetchError(f"missing {url}")
            return pages[url]

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "seeds.csv"
            summary = collect_official_site_seeds(
                [{"school_name": "鍖椾含澶у", "official_url": "https://www.pku.edu.cn/"}],
                output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                probe_candidate_portals=True,
                max_candidate_portal_pages_per_site=1,
            )
            with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(summary["seeds_written"], 1)
        self.assertEqual(rows[0]["start_url"], "https://dean.pku.edu.cn/info/1.htm")

    def helper_collect_official_site_seeds_can_probe_common_teaching_subdomains_corrupt_fixture_2(self):
        pages = {
            "https://www.pku.edu.cn/": FetchResponse(
                url="https://www.pku.edu.cn/",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content=b"<html><body>home</body></html>",
            ),
            "https://dean.pku.edu.cn/": FetchResponse(
                url="https://dean.pku.edu.cn/",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content='<a href="/info/1.htm">鍖椾含澶у2026灞婃帹鑽愬厤璇曠爺绌剁敓鍚嶅崟鍏ず</a>'.encode("utf-8"),
            ),
        }

        def fake_fetch(url, timeout_seconds=20):
            if url not in pages:
                raise crawler.FetchError(f"missing {url}")
            return pages[url]

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "seeds.csv"
            summary = collect_official_site_seeds(
                [{"school_name": "鍖椾含澶у", "official_url": "https://www.pku.edu.cn/"}],
                output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                probe_candidate_portals=True,
                max_candidate_portal_pages_per_site=1,
            )
            with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(summary["seeds_written"], 1)
        self.assertEqual(rows[0]["start_url"], "https://dean.pku.edu.cn/info/1.htm")

    def test_collect_official_site_seeds_can_probe_common_teaching_subdomains(self):
        pages = {
            "https://www.pku.edu.cn/": FetchResponse(
                url="https://www.pku.edu.cn/",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content=b"<html><body>home</body></html>",
            ),
            "https://dean.pku.edu.cn/": FetchResponse(
                url="https://dean.pku.edu.cn/",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content=b'<a href="/info/2026%E6%8E%A8%E8%8D%90%E5%85%8D%E8%AF%95%E5%90%8D%E5%8D%95.htm">notice</a>',
            ),
        }

        def fake_fetch(url, timeout_seconds=20):
            if url not in pages:
                raise crawler.FetchError(f"missing {url}")
            return pages[url]

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "seeds.csv"
            summary = collect_official_site_seeds(
                [{"school_name": "school", "official_url": "https://www.pku.edu.cn/"}],
                output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                probe_candidate_portals=True,
                max_candidate_portal_pages_per_site=1,
            )
            with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(summary["seeds_written"], 1)
        self.assertEqual(
            rows[0]["start_url"],
            "https://dean.pku.edu.cn/info/2026%E6%8E%A8%E8%8D%90%E5%85%8D%E8%AF%95%E5%90%8D%E5%8D%95.htm",
        )

    def test_collect_official_site_seeds_can_run_with_workers(self):
        pages = {
            "https://www.alpha.edu.cn/": FetchResponse(
                url="https://www.alpha.edu.cn/",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content='<a href="/notice/1.htm">2026届推荐免试名单公示</a>'.encode("utf-8"),
            ),
            "https://www.beta.edu.cn/": FetchResponse(
                url="https://www.beta.edu.cn/",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content='<a href="/notice/2.htm">2026届推荐免试名单公示</a>'.encode("utf-8"),
            ),
        }
        lock = threading.Lock()
        release = threading.Event()
        active_fetches = 0
        max_active_fetches = 0

        def fake_fetch(url, timeout_seconds=20):
            nonlocal active_fetches, max_active_fetches
            with lock:
                active_fetches += 1
                max_active_fetches = max(max_active_fetches, active_fetches)
                if max_active_fetches >= 2:
                    release.set()
            try:
                if not release.wait(1):
                    raise AssertionError("official-site fetches did not overlap")
                return pages[url]
            finally:
                with lock:
                    active_fetches -= 1

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "seeds.csv"
            summary = collect_official_site_seeds(
                [
                    {"school_name": "Alpha University", "official_url": "https://www.alpha.edu.cn/"},
                    {"school_name": "Beta University", "official_url": "https://www.beta.edu.cn/"},
                ],
                output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                workers=2,
            )
            with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(summary["selected_sites"], 2)
        self.assertEqual(summary["seeds_written"], 2)
        self.assertEqual(max_active_fetches, 2)
        self.assertEqual(
            {row["start_url"] for row in rows},
            {"https://www.alpha.edu.cn/notice/1.htm", "https://www.beta.edu.cn/notice/2.htm"},
        )

    def test_collect_chsi_bulletin_seeds_writes_relevant_seed_csv(self):
        pages = {
            "https://yzst.chsi.com.cn/sch/schoolInfo--schId-1.dhtml": FetchResponse(
                url="https://yzst.chsi.com.cn/sch/schoolInfo--schId-1.dhtml",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content=b'<a href="/sch/listBulletin--schId-1%2CcategoryId-2.dhtml">more</a>',
            ),
            "https://yzst.chsi.com.cn/sch/listBulletin--schId-1%2CcategoryId-2.dhtml": FetchResponse(
                url="https://yzst.chsi.com.cn/sch/listBulletin--schId-1%2CcategoryId-2.dhtml",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content='<a href="/sch/viewBulletin--schId-1,infoId-10.dhtml">2025年硕士研究生拟录取名单</a>'.encode("utf-8"),
            ),
        }

        def fake_fetch(url, timeout_seconds=20):
            return pages[url]

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "seeds.csv"
            summary = collect_chsi_bulletin_seeds(
                [{"chsi_school_name": "样例大学", "chsi_school_url": "https://yzst.chsi.com.cn/sch/schoolInfo--schId-1.dhtml"}],
                output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
            )
            rows = list(csv.DictReader(output_path.open("r", encoding="utf-8-sig", newline="")))

        self.assertEqual(summary["seeds_written"], 1)
        self.assertEqual(rows[0]["school_name"], "样例大学")
        self.assertEqual(rows[0]["source_type"], "postgraduate_admission")

    def test_collect_chsi_bulletin_seeds_clears_stale_failure_log(self):
        pages = {
            "https://yzst.chsi.com.cn/sch/schoolInfo--schId-1.dhtml": FetchResponse(
                url="https://yzst.chsi.com.cn/sch/schoolInfo--schId-1.dhtml",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content=b'<a href="/sch/listBulletin--schId-1%2CcategoryId-2.dhtml">more</a>',
            ),
            "https://yzst.chsi.com.cn/sch/listBulletin--schId-1%2CcategoryId-2.dhtml": FetchResponse(
                url="https://yzst.chsi.com.cn/sch/listBulletin--schId-1%2CcategoryId-2.dhtml",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content='<a href="/sch/viewBulletin--schId-1,infoId-10.dhtml">2025年硕士研究生拟录取名单</a>'.encode("utf-8"),
            ),
        }

        def fake_fetch(url, timeout_seconds=20):
            return pages[url]

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "seeds.csv"
            failures_path = output_path.with_suffix(".failures.jsonl")
            failures_path.write_text('{"error":"stale"}\n', encoding="utf-8")

            collect_chsi_bulletin_seeds(
                [{"chsi_school_name": "Example University", "chsi_school_url": "https://yzst.chsi.com.cn/sch/schoolInfo--schId-1.dhtml"}],
                output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
            )

            failures_text = failures_path.read_text(encoding="utf-8")

        self.assertEqual(failures_text, "")

    def test_collect_chsi_bulletin_seeds_can_run_with_workers(self):
        pages = {}
        for school_id, school_name in [("1", "样例大学A"), ("2", "样例大学B")]:
            pages[f"https://yzst.chsi.com.cn/sch/schoolInfo--schId-{school_id}.dhtml"] = FetchResponse(
                url=f"https://yzst.chsi.com.cn/sch/schoolInfo--schId-{school_id}.dhtml",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content=f'<a href="/sch/listBulletin--schId-{school_id}%2CcategoryId-2.dhtml">more</a>'.encode("utf-8"),
            )
            pages[f"https://yzst.chsi.com.cn/sch/listBulletin--schId-{school_id}%2CcategoryId-2.dhtml"] = FetchResponse(
                url=f"https://yzst.chsi.com.cn/sch/listBulletin--schId-{school_id}%2CcategoryId-2.dhtml",
                status_code=200,
                content_type="text/html; charset=utf-8",
                content=f'<a href="/sch/viewBulletin--schId-{school_id},infoId-10.dhtml">{school_name}2025年硕士研究生拟录取名单公示</a>'.encode("utf-8"),
            )

        def fake_fetch(url, timeout_seconds=20):
            return pages[url]

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "seeds.csv"
            summary = collect_chsi_bulletin_seeds(
                [
                    {"chsi_school_name": "样例大学A", "chsi_school_url": "https://yzst.chsi.com.cn/sch/schoolInfo--schId-1.dhtml"},
                    {"chsi_school_name": "样例大学B", "chsi_school_url": "https://yzst.chsi.com.cn/sch/schoolInfo--schId-2.dhtml"},
                ],
                output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                workers=2,
            )
            rows = list(csv.DictReader(output_path.open("r", encoding="utf-8-sig", newline="")))

        self.assertEqual(summary["selected_schools"], 2)
        self.assertEqual(summary["seeds_written"], 2)
        self.assertEqual({row["school_name"] for row in rows}, {"样例大学A", "样例大学B"})

    def test_parse_html_records_maps_chinese_headers_to_normalized_fields(self):
        records = parse_html_records(
            HTML_TABLE,
            document={
                "school_name": "东北农业大学",
                "document_type": "incoming_recommendation_admission_list",
                "source_url": "https://yz.neau.edu.cn/result.htm",
                "title": "2026年接收推免生拟录取名单公示",
                "year": 2026,
            },
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "张三")
        self.assertEqual(records[0]["undergraduate_school"], "东北农业大学")
        self.assertEqual(records[0]["college"], "计算机学院")
        self.assertEqual(records[0]["admission_major"], "计算机科学与技术")
        self.assertEqual(records[0]["route"], "recommendation_exemption")
        self.assertFalse(records[0]["needs_review"])

    def test_parse_html_records_prefers_table_rows_over_text_fallback_for_sufe_lists(self):
        html = """
        <div class="wp_articlecontent">
          <p>根据复试成绩排名，获得拟录取后备资格的学生名单及候补排序如下：</p>
          <p>国际商务（1人）</p>
          <table>
            <tr>
              <td>候补排序</td><td>姓名</td><td>本科高校</td><td>本科专业</td>
              <td>复试成绩</td><td>申请专业</td>
            </tr>
            <tr>
              <td>1</td><td>寇宁</td><td>内蒙古大学</td><td>金融学</td>
              <td>88.00</td><td>国际商务</td>
            </tr>
          </table>
        </div>
        """

        records = parse_html_records(
            html,
            document={
                "school_name": "上海财经大学",
                "document_type": "incoming_recommendation_admission_list",
                "source_url": "https://cob.sufe.edu.cn/Home/Detail/26771",
                "title": "上海财经大学商学院2026年推荐免试研究生拟录取及候补资格名单",
                "year": 2026,
            },
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "寇宁")
        self.assertEqual(records[0]["undergraduate_school"], "内蒙古大学")
        self.assertEqual(records[0]["undergraduate_major"], "金融学")
        self.assertEqual(records[0]["admission_major"], "国际商务")
        self.assertNotIn("候补排序", {record["person_name"] for record in records})

    def test_parse_html_records_extracts_subject_category_score_table_fields(self):
        html = """
        <table>
          <tr>
            <td>考生编号</td><td>姓名</td><td>一级学科名称</td>
            <td>报考类别</td><td>综合面试成绩</td><td>备注</td>
          </tr>
          <tr>
            <td>102546123100024</td><td>金子涵</td><td>船舶与海洋工程</td>
            <td>非定向就业</td><td>94.67</td><td></td>
          </tr>
          <tr>
            <td>102546123800141</td><td>刘海鑫</td><td>管理科学与工程</td>
            <td>非定向就业</td><td>/</td><td>推免生直接录取</td>
          </tr>
        </table>
        """

        records = parse_html_records(
            html,
            document={
                "school_name": "上海海事大学",
                "document_type": "postgraduate_admission_list",
                "source_url": "https://yz.shmtu.edu.cn/2026/0522/list.htm",
                "title": "上海海事大学2026年博士研究生拟录取名单公示",
                "year": 2026,
            },
        )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "金子涵")
        self.assertEqual(records[0]["student_id"], "102546123100024")
        self.assertEqual(records[0]["admission_major"], "船舶与海洋工程")
        self.assertEqual(records[0]["remarks"], "非定向就业 interview_score 94.67")
        self.assertFalse(records[0]["needs_review"])
        self.assertEqual(records[1]["admission_major"], "管理科学与工程")
        self.assertIn("推免生直接录取", records[1]["remarks"])

    def test_parse_html_records_extracts_cdutcm_recommendation_table_fields(self):
        html = """
        <table>
          <tr><td>2025年现代中药产业学院推荐免试攻读硕士学位研究生拟录取名单</td></tr>
          <tr>
            <td>序号</td><td>考生姓名</td><td>类型</td><td>报考专业代码</td>
            <td>报考专业名称</td><td>报考研究方向代码</td><td>报考研究方向名称</td>
            <td>综合面试成绩</td><td>导师</td>
          </tr>
          <tr>
            <td>1</td><td>税美晴</td><td>专业学位</td><td>105600</td><td>中药学</td>
            <td>04</td><td>中药新制剂、新剂型、新技术应用研究</td><td>81.63</td><td>罗佳</td>
          </tr>
        </table>
        """

        records = parse_html_records(
            html,
            document={
                "school_name": "成都中医药大学",
                "document_type": "incoming_recommendation_admission_list",
                "source_url": "https://www.cdutcm.edu.cn/xdzycyxy/info/1161/1851.htm",
                "title": "成都中医药大学现代中药产业学院2026年推免生拟录取名单（第一批）",
                "year": 2026,
            },
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "税美晴")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertEqual(records[0]["college"], "现代中药产业学院")
        self.assertEqual(records[0]["major"], "105600")
        self.assertEqual(records[0]["admission_major"], "105600 中药学")
        self.assertEqual(
            records[0]["remarks"],
            "专业学位 direction 04 中药新制剂、新剂型、新技术应用研究 interview_score 81.63 advisor 罗佳",
        )

    def test_parse_html_records_extracts_sut_recommendation_table_fields(self):
        html = """
        <table>
          <tr>
            <td>序号</td><td>姓名</td><td>性别</td><td>证件号码</td><td>学院</td>
            <td>专业代码</td><td>专业名称</td><td>学习方式</td><td>招生类型</td><td>复试成绩</td>
          </tr>
          <tr>
            <td>1</td><td>韩 *奇</td><td>女</td><td>210103********6028</td><td>机械工程学院</td>
            <td>080200</td><td>机械工程</td><td>全日制</td><td>硕士</td><td>92.4</td>
          </tr>
        </table>
        """

        records = parse_html_records(
            html,
            document={
                "school_name": "沈阳工业大学",
                "document_type": "incoming_recommendation_admission_list",
                "source_url": "https://yjsxy.sut.edu.cn/info/1311/10141.htm",
                "title": "沈阳工业大学2026年接收推荐免试研究生拟录取名单公示",
                "year": 2026,
            },
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "韩*奇")
        self.assertEqual(records[0]["student_id"], "210103********6028")
        self.assertEqual(records[0]["college"], "机械工程学院")
        self.assertEqual(records[0]["admission_major"], "080200 机械工程")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertIn("gender 女", records[0]["remarks"])
        self.assertIn("study_mode 全日制", records[0]["remarks"])
        self.assertIn("admission_type 硕士", records[0]["remarks"])
        self.assertIn("score 92.4", records[0]["remarks"])

    def test_parse_html_records_extracts_beihua_recommendation_table_fields(self):
        html = """
        <table>
          <tr>
            <td>姓名</td><td>报考专业代码</td><td>报考专业名称</td><td>报考学习形式</td><td>复试成绩</td>
          </tr>
          <tr>
            <td>陈怡彤</td><td>080804</td><td>电力电子与电力传动</td><td>全日制</td><td>88.8</td>
          </tr>
        </table>
        """

        records = parse_html_records(
            html,
            document={
                "school_name": "北华大学",
                "document_type": "incoming_recommendation_admission_list",
                "source_url": "https://grad.beihua.edu.cn/info/1071/3062.htm",
                "title": "北华大学2026年推免生复试拟录取名单公示",
                "year": 2026,
            },
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "陈怡彤")
        self.assertEqual(records[0]["major"], "080804")
        self.assertEqual(records[0]["admission_major"], "080804 电力电子与电力传动")
        self.assertEqual(records[0]["remarks"], "study_mode 全日制; retest_score 88.8")

    def test_parse_html_records_extracts_admitted_major_without_name_suffix(self):
        html = """
        <table>
          <tr>
            <td>序号</td><td>拟录取学院</td><td>考生姓名</td><td>性别</td>
            <td>拟录取专业代码</td><td>拟录取专业</td>
          </tr>
          <tr>
            <td>1</td><td>传媒学院</td><td>丁晓娜</td><td>女</td><td>135400</td><td>戏剧与影视</td>
          </tr>
        </table>
        """

        records = parse_html_records(
            html,
            document={
                "school_name": "贵州民族大学",
                "document_type": "recommendation_exemption_list",
                "source_url": "https://yjsy.gzmu.edu.cn/info/1084/5986.htm",
                "title": "贵州民族大学2026年接收推荐免试攻读硕士研究生拟录取名单公示",
                "year": 2026,
            },
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "丁晓娜")
        self.assertEqual(records[0]["college"], "传媒学院")
        self.assertEqual(records[0]["major"], "135400")
        self.assertEqual(records[0]["admission_major"], "135400 戏剧与影视")
        self.assertIn("gender 女", records[0]["remarks"])

    def test_parse_html_records_skips_aggregate_admission_count_tables(self):
        html = """
        <table>
          <tr><td>学院代码</td><td>学院名称</td><td>专业代码</td><td>专业名称</td><td>研究方向</td><td>录取人数</td></tr>
          <tr><td>011</td><td>理学院</td><td>025200</td><td>应用统计</td><td>所有方向</td><td>2</td></tr>
          <tr><td>012</td><td>机械工程学院</td><td>080200</td><td>机械工程</td><td>所有方向</td><td>26</td></tr>
        </table>
        """

        records = parse_html_records(
            html,
            document={
                "school_name": "河北工业大学",
                "document_type": "incoming_recommendation_admission_list",
                "source_url": "https://yjs.hebut.edu.cn/list.htm",
                "title": "河北工业大学2026年推免研究生拟录取名单公示",
                "year": 2026,
            },
        )

        self.assertEqual(records, [])

    def test_parse_excel_records_reads_header_like_first_row(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "admission.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["序号", "姓名", "本科毕业学校", "录取学院", "录取专业"])
            sheet.append([1, "李四", "杭州电子科技大学", "自动化学院", "控制科学与工程"])
            workbook.save(path)

            records = parse_excel_records(
                path,
                document={
                    "school_name": "浙江大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yz.zju.edu.cn/result.xlsx",
                    "title": "硕士研究生拟录取名单",
                    "year": 2025,
                },
            )

        self.assertEqual(records[0]["person_name"], "李四")
        self.assertEqual(records[0]["undergraduate_school"], "杭州电子科技大学")
        self.assertEqual(records[0]["college"], "自动化学院")
        self.assertEqual(records[0]["admission_major"], "控制科学与工程")
        self.assertEqual(records[0]["route"], "postgraduate_exam_or_admission")

    def test_parse_excel_records_extracts_formal_recommendation_name_columns_only(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "recommendation.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["正式（2人）", "候补"])
            sheet.append(["安梦笛", "蔡瑞"])
            sheet.append(["白屹禾", "曹可心"])
            workbook.save(path)

            records = parse_excel_records(
                path,
                document={
                    "school_name": "北京工商大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://jwc.btbu.edu.cn/list.xlsx",
                    "title": "北京工商大学关于2026届本科生推免资格结果的公示",
                    "year": 2026,
                },
            )

        self.assertEqual([record["person_name"] for record in records], ["安梦笛", "白屹禾"])
        self.assertEqual({record["route"] for record in records}, {"recommendation_exemption"})
        self.assertEqual({record["remarks"] for record in records}, {"正式（2人）"})

    def test_parse_excel_records_clears_header_text_left_in_ranking_field(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "recommendation.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["学院", "专业名称", "学号", "姓名", "专业排名"])
            sheet.append(["经济管理学院", "工程管理", "22377113", "严子一", "专业排名"])
            workbook.save(path)

            records = parse_excel_records(
                path,
                document={
                    "school_name": "北京航空航天大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://sem.buaa.edu.cn/list.pdf",
                    "title": "北京航空航天大学经济管理学院2026年拟推免生名单",
                    "year": 2026,
                },
            )

        self.assertEqual(records[0]["person_name"], "严子一")
        self.assertEqual(records[0]["student_id"], "22377113")
        self.assertEqual(records[0]["admission_major"], "工程管理")
        self.assertEqual(records[0]["ranking"], "")

    def test_parse_excel_records_extracts_recommendation_admission_status_rows(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "recommendation.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["海南医科大学2026年接收推荐免试攻读硕士研究生复试考核成绩及拟录取名单（一）"])
            sheet.append(["序号", "姓名", "身份证号", "报考学院", "报考专业", "学位类型", "学习方式", "复试成绩", "录取状态"])
            sheet.append([1, "李海映", "469006********1666", "001海南医科大学", "105118（麻醉学）", "专业型", "全日制", 84.06, "拟录取"])
            workbook.save(path)

            records = parse_excel_records(
                path,
                document={
                    "school_name": "海南医科大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://www.muhn.edu.cn/download.jsp",
                    "title": "海南医科大学2026年接收推荐免试攻读硕士研究生复试考核成绩及拟录取名单（一）",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "李海映")
        self.assertEqual(records[0]["student_id"], "469006********1666")
        self.assertEqual(records[0]["college"], "001海南医科大学")
        self.assertEqual(records[0]["admission_major"], "105118（麻醉学）")
        self.assertIn("学位类型 专业型", records[0]["remarks"])
        self.assertIn("学习方式 全日制", records[0]["remarks"])
        self.assertIn("复试成绩 84.06", records[0]["remarks"])
        self.assertIn("录取状态 拟录取", records[0]["remarks"])

    def test_classify_document_treats_push免拟录取_as_recommendation(self):
        classification = crawler.classify_document(
            "2026年推免拟录取名单.pdf",
            "https://graduate.nbu.edu.cn/download.jsp",
            "",
        )

        self.assertEqual(
            classification["document_type"],
            "incoming_recommendation_admission_list",
        )

    def test_classify_document_treats_tuimian_master_admission_as_recommendation(self):
        classification = crawler.classify_document(
            "上海中医药大学2025年推免硕士研究生拟录取名单公示",
            "https://yjsy.shutcm.edu.cn/2025/1013/c1143a169399/page.htm",
            "",
        )

        self.assertEqual(
            classification["document_type"],
            "incoming_recommendation_admission_list",
        )

    def test_parse_pdf_records_extracts_nbu_recommendation_rows(self):
        text = """
        姓名    性别 复试成绩     接收学院   接收专业代码    接收专业名称    录取类型
        陈虹    女  82.93    商学院    020200     应用经济学    硕士
        王宇铎    男  87.28      数学与统计学院    070100      数学      直博生
        """

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "宁波大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://graduate.nbu.edu.cn/list.pdf",
                    "title": "2026年推免拟录取名单.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "陈虹")
        self.assertEqual(records[0]["college"], "商学院")
        self.assertEqual(records[0]["major"], "020200")
        self.assertEqual(records[0]["admission_major"], "020200 应用经济学")
        self.assertIn("gender 女", records[0]["remarks"])
        self.assertIn("retest_score 82.93", records[0]["remarks"])
        self.assertIn("admission_type 硕士", records[0]["remarks"])
        self.assertEqual(records[0]["route"], "recommendation_exemption")
        self.assertEqual(records[1]["admission_major"], "070100 数学")
        self.assertIn("admission_type 直博生", records[1]["remarks"])

    def test_parse_pdf_records_extracts_jxnu_recommendation_qualification_rows(self):
        text = """
        江西师范大学
        2026届优秀本科毕业生获得推荐免试攻读研究生资格名单
        学院代码   学院名称   专业代码   专业名称   姓名   性别 综合成绩
        001    政法学院   030101     法学   舒雯    女   67.61
        002    经济与管理学院   120203     会计学   余文静   女   71.29
        """

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "江西师范大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://yz.jxnu.edu.cn/list.pdf",
                    "title": "江西师范大学2026届优秀本科毕业生获得推荐免试攻读研究生资格名单公示.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "舒雯")
        self.assertEqual(records[0]["college"], "001 政法学院")
        self.assertEqual(records[0]["major"], "030101")
        self.assertEqual(records[0]["undergraduate_major"], "030101 法学")
        self.assertIn("gender 女", records[0]["remarks"])
        self.assertIn("composite_score 67.61", records[0]["remarks"])

    def test_parse_pdf_records_extracts_jxnu_incoming_recommendation_rows(self):
        text = """
        江西师范大学2026年推免生接收名单
        院系所码   录取院系所        考生编号           姓名    性别    录取专业代码     录取专业名称     备注
        018   化学与材料学院   104146104070001   张义帅   男    070300     化学       直博生
        001    政法学院     104146104140003   石颜硕   男    035102   法律（法学）     支教团
        """

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "江西师范大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yz.jxnu.edu.cn/list.pdf",
                    "title": "江西师范大学2026年接收推免生（含直博生）名单公示.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "张义帅")
        self.assertEqual(records[0]["student_id"], "104146104070001")
        self.assertEqual(records[0]["college"], "018 化学与材料学院")
        self.assertEqual(records[0]["major"], "070300")
        self.assertEqual(records[0]["admission_major"], "070300 化学")
        self.assertIn("gender 男", records[0]["remarks"])
        self.assertIn("admission_type 直博生", records[0]["remarks"])
        self.assertIn("remark 支教团", records[1]["remarks"])

    def test_parse_pdf_records_extracts_xmu_sectioned_postgraduate_rows(self):
        text = """
        材料学院 2026年硕士研究生拟录取名单
        专业代码和专业名称：080500材料科学与工程
        序号      考生编号           姓名     初试总分 复试成绩      总成绩     学习方式 录取类别   备注
        1    103846214309756   周顺      397    88.2   82.92   全日制   非定向

        专业代码和专业名称：085601材料工程
        序号      考生编号           姓名     初试总分 复试成绩      总成绩     学习方式 录取类别   备注
        41   103846214309757 詹彦逸哲    376   86.6   79.76   全日制   非定向
        """

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "厦门大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://cm.xmu.edu.cn/list.pdf",
                    "title": "材料学院2026年硕士研究生拟录取名单.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "周顺")
        self.assertEqual(records[0]["student_id"], "103846214309756")
        self.assertEqual(records[0]["college"], "材料学院")
        self.assertEqual(records[0]["major"], "080500")
        self.assertEqual(records[0]["admission_major"], "080500 材料科学与工程")
        self.assertIn("initial_score 397", records[0]["remarks"])
        self.assertIn("reexam_score 88.2", records[0]["remarks"])
        self.assertIn("total_score 82.92", records[0]["remarks"])
        self.assertIn("study_mode 全日制", records[0]["remarks"])
        self.assertIn("admission_category 非定向", records[0]["remarks"])
        self.assertFalse(records[0]["needs_review"])
        self.assertEqual(records[1]["person_name"], "詹彦逸哲")
        self.assertEqual(records[1]["admission_major"], "085601 材料工程")

    def test_parse_html_records_extracts_shutcm_meta_description_recommendation_rows(self):
        html = """
        <html><head>
        <title>上海中医药大学2025年推免硕士研究生拟录取名单公示</title>
        <meta name="description" content="序号姓名拟录取学院专业代码专业名称复试成绩推荐学校1王杰A01中医学院100501中医基础理论389甘肃中医药大学2张月阳A01中医学院100502中医临床基础414长江大学3截断A81中西医结合学院100601中西医结合基础(中医方向" />
        </head><body></body></html>
        """

        records = crawler.parse_html_records(
            html,
            {
                "school_name": "上海中医药大学",
                "document_type": "incoming_recommendation_admission_list",
                "source_url": "https://yjsy.shutcm.edu.cn/2025/1013/c1143a169399/page.htm",
                "title": "上海中医药大学2025年推免硕士研究生拟录取名单公示",
                "year": 2025,
            },
        )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "王杰")
        self.assertEqual(records[0]["college"], "A01 中医学院")
        self.assertEqual(records[0]["major"], "100501")
        self.assertEqual(records[0]["admission_major"], "100501 中医基础理论")
        self.assertIn("retest_score 389", records[0]["remarks"])
        self.assertIn("undergraduate_school 甘肃中医药大学", records[0]["remarks"])
        self.assertEqual(records[0]["route"], "recommendation_exemption")
        self.assertEqual(records[1]["person_name"], "张月阳")
        self.assertEqual(records[1]["admission_major"], "100502 中医临床基础")

    def test_parse_legacy_xls_records_uses_converted_workbook(self):
        with TemporaryDirectory() as temp_dir:
            legacy_path = Path(temp_dir) / "recommendation.xls"
            legacy_path.write_bytes(b"legacy-xls")
            converted_path = Path(temp_dir) / "recommendation.converted.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["姓名", "学院", "专业"])
            sheet.append(["严寒", "法治学院、法律硕士教育学院", "法律（法学）"])
            workbook.save(converted_path)

            with patch.object(crawler, "_convert_legacy_xls_to_xlsx", return_value=converted_path):
                records = crawler.parse_legacy_xls_records(
                    legacy_path,
                    document={
                        "school_name": "西北政法大学",
                        "document_type": "incoming_recommendation_admission_list",
                        "source_url": "https://grs.nwupl.edu.cn/list.xls",
                        "title": "西北政法大学2026年推荐免试硕士研究生拟录取名单公示",
                        "year": 2026,
                    },
                )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "严寒")
        self.assertEqual(records[0]["college"], "法治学院、法律硕士教育学院")
        self.assertEqual(records[0]["major"], "法律（法学）")

    def test_extract_pdf_text_falls_back_to_pdftotext_when_pypdf_text_is_garbled(self):
        class FakePage:
            def extract_text(self):
                return "������ѧ2026������Ƽ����Թ����о�"

        class FakePdfReader:
            def __init__(self, path):
                self.pages = [FakePage()]

        fake_pypdf = types.SimpleNamespace(PdfReader=FakePdfReader)

        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "list.pdf"
            path.write_bytes(b"%PDF-1.4\n%fake\n")

            with patch.dict("sys.modules", {"pypdf": fake_pypdf}):
                with patch("subprocess.run") as run:
                    run.return_value.returncode = 0
                    run.return_value.stdout = "辽宁大学2026年接收推荐免试攻读研究生拟录取名单\n"
                    run.return_value.stderr = ""

                    text = crawler.extract_pdf_text(path)

        self.assertIn("辽宁大学2026年接收推荐免试", text)

    def test_extract_pdf_text_prefers_pdftotext_layout_over_single_line_pypdf_table(self):
        class FakePage:
            def extract_text(self):
                return "辽宁大学2026年接收推荐免试攻读研究生拟录取名单拟录取培养单位名称 拟录取专业代码及名称 编号 姓名 经济学院020101政治经济学mpemt****7sicy王*屏"

        class FakePdfReader:
            def __init__(self, path):
                self.pages = [FakePage()]

        fake_pypdf = types.SimpleNamespace(PdfReader=FakePdfReader)

        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "list.pdf"
            path.write_bytes(b"%PDF-1.4\n%fake\n")

            with patch.dict("sys.modules", {"pypdf": fake_pypdf}):
                with patch("subprocess.run") as run:
                    run.return_value.returncode = 0
                    run.return_value.stdout = (
                        "辽宁大学2026年接收推荐免试攻读研究生拟录取名单\n"
                        "拟录取培养单位名称  拟录取专业代码及名称  编号  姓名\n"
                        "经济学院  020101 政治经济学  mpemt****7sicy  王*屏\n"
                    )
                    run.return_value.stderr = ""

                    text = crawler.extract_pdf_text(path)

        self.assertIn("拟录取培养单位名称  拟录取专业代码及名称", text)
        self.assertIn("\n经济学院  020101", text)

    def test_extract_pdf_text_decodes_gbk_pdftotext_output(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "list.pdf"
            path.write_bytes(b"%PDF-1.4\n%fake\n")

            with patch("subprocess.run") as run:
                run.return_value.returncode = 0
                run.return_value.stdout = "南方医科大学2026年招收推荐免试研究生拟录取名单公示\n".encode(
                    "gbk"
                )
                run.return_value.stderr = b""

                text = crawler.extract_pdf_text(path)

        self.assertIn("南方医科大学2026年招收推荐免试研究生拟录取名单公示", text)

    def test_rows_from_text_lines_preserves_pdftotext_spacing_as_columns(self):
        rows = crawler._rows_from_text_lines(
            "拟录取培养单位名称  拟录取专业代码及名称  编号  姓名\n"
            "经济学院  020101 政治经济学  mpemt****7sicy  王*屏\n"
        )

        self.assertEqual(
            rows,
            [
                ["拟录取培养单位名称", "拟录取专业代码及名称", "编号", "姓名"],
                ["经济学院", "020101 政治经济学", "mpemt****7sicy", "王*屏"],
            ],
        )

    def test_records_from_table_aligns_pdf_rows_with_split_code_and_name_columns(self):
        records = crawler._records_from_table(
            [
                ["拟录取培养单位名称", "拟录取专业代码及名称", "拟录取研究方向代码及名称", "编号", "姓名", "拟录取学习方式"],
                ["经济学院", "020101", "政治经济学", "00", "不区分研究方向", "mpemt****7sicy", "王*屏", "全日制"],
            ],
            {
                "school_name": "辽宁大学",
                "document_type": "recommendation_exemption_list",
                "source_url": "https://grs.lnu.edu.cn/26tmgsmd.pdf",
                "title": "辽宁大学2026年接收推荐免试攻读研究生拟录取名单",
                "year": 2026,
            },
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["college"], "经济学院")
        self.assertEqual(records[0]["admission_major"], "020101 政治经济学")
        self.assertEqual(records[0]["student_id"], "mpemt****7sicy")
        self.assertEqual(records[0]["person_name"], "王*屏")

    def test_parse_pdf_records_extracts_college_code_first_admission_rows(self):
        text = (
            "拟录取学院  准考证号  姓名  拟录取专业名称  拟录取研究方向名称  总成绩  录取类别  备注\n"
            "010  管理学院  101656000000784  张*  040102  课程与教学论  00  不区分研究方向  368 136.80  78.88  全日制 非定向\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "辽宁师范大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yjszs.lnnu.edu.cn/list.pdf",
                    "title": "辽宁师范大学2026年硕士研究生拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["college"], "010 管理学院")
        self.assertEqual(records[0]["student_id"], "101656000000784")
        self.assertEqual(records[0]["person_name"], "张*")
        self.assertEqual(records[0]["admission_major"], "040102 课程与教学论")

    def test_parse_pdf_records_extracts_scnu_retest_admission_summary_rows(self):
        text = (
            "华南师范大学 2026 年硕士研究生复试及拟录取情况汇总表\n"
            "招生单位名称（盖章）： 哲学与社会发展学院(001) 制表日期：2026 年 4 月 20 日\n"
            "序号      考生编号      考生姓名      考试方式      调剂类别      专项计划      初试成绩      复试成绩      最终成绩      拟录取专业代码及名称      学习方式      录取类别      是否拟录取      备注\n"
            "1    105746000019577   刘国正   全国统考   一志愿      无    412   91.68   87.04   (010100)哲学   全日制   非定向    是          (01 方向)\n"
            "2    105746105740318   涂天天   推荐免试   一志愿      无           92      92     (030200)政治学   全日制   非定向    是          (01 方向)\n"
            "3    105746000009130   刘美茜   全国统考   一志愿      无    375   85.73   80.37   (045103)学科教学（语文） 非全日制   定向    是\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "华南师范大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://statics.scnu.edu.cn/pics/yz/2026/sample.pdf",
                    "title": "001哲学与社会发展学院2026年硕士研究生复试及拟录取情况汇总表.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 3)
        self.assertEqual(records[0]["college"], "哲学与社会发展学院")
        self.assertEqual(records[0]["student_id"], "105746000019577")
        self.assertEqual(records[0]["person_name"], "刘国正")
        self.assertEqual(records[0]["admission_major"], "(010100)哲学")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertIn("考试方式: 全国统考", records[0]["remarks"])
        self.assertIn("是否拟录取: 是", records[0]["remarks"])
        self.assertEqual(records[1]["route"], "recommendation_exemption")
        self.assertEqual(records[2]["admission_major"], "(045103)学科教学（语文）")
        self.assertIn("学习方式: 非全日制", records[2]["remarks"])
        self.assertIn("是否拟录取: 是", records[2]["remarks"])

    def test_parse_pdf_records_extracts_xupt_recommendation_rows(self):
        text = (
            "西安邮电大学 2026 年推荐免试研究生拟录取名单\n"
            "      身份证号\n"
            "姓名                   拟录取学院   拟录取专业代码             拟录取专业名称\n"
            "      后四位\n"
            "张雨婷   1927   通信与信息工程学院         081000   信息与通信工程\n"
            "王佳维   5028   通信与信息工程学院         085401   新一代电子信息技术（含量子技术等）\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "西安邮电大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://gr.xupt.edu.cn/__local/sample.pdf",
                    "title": "西安邮电大学2026年推荐免试研究生拟录取名单PDF",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "张雨婷")
        self.assertEqual(records[0]["student_id"], "1927")
        self.assertEqual(records[0]["college"], "通信与信息工程学院")
        self.assertEqual(records[0]["major"], "081000")
        self.assertEqual(records[0]["admission_major"], "081000 信息与通信工程")
        self.assertFalse(records[0]["needs_review"])

    def test_parse_pdf_records_extracts_syphu_postgraduate_rows(self):
        text = (
            "沈阳药科大学2025年硕士研究生一志愿复试结果及拟录取名单\n"
            "考生编号      姓名   复试批次       复试学院       复试专业代码、研究方向代码、名称      初试总分  专业考核 综合素质考核 复试总分 总分 录取排名 复试结果 录取意见 录取类别\n"
            "101635000002982 李芳 复试一志愿批次 001【药学院】    070300【化学】                392.00 179.00 84.60 263.60 655.60   1    1复试合格 拟录取 全日制非定向\n"
            "104225510906628    王浩    复试调剂一批次 001【药学院】   100702【药剂学】 00【不区分研究方向】   340.00 162.60 88.60 251.20 591.20   4   拟录取 1复试合格\n"
            "100235411410681 张文祯 复试调剂二批次 002【制药工程学院】 070300【化学】 00【不区分研究方向】 352.00 178.00 89.00 267.00 619.00 1 拟录取 1复试合格\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "沈阳药科大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://grs.syphu.edu.cn/__local/sample.pdf",
                    "title": "沈阳药科大学2025年硕士研究生一志愿复试结果及拟录取名单PDF",
                    "year": 2025,
                },
            )

        self.assertEqual(len(records), 3)
        self.assertEqual(records[0]["student_id"], "101635000002982")
        self.assertEqual(records[0]["person_name"], "李芳")
        self.assertEqual(records[0]["college"], "001【药学院】")
        self.assertEqual(records[0]["admission_major"], "070300【化学】")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertIn("复试批次: 复试一志愿批次", records[0]["remarks"])
        self.assertIn("初试总分: 392.00", records[0]["remarks"])
        self.assertIn("录取意见: 拟录取", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "王浩")
        self.assertEqual(records[1]["admission_major"], "100702【药剂学】 00【不区分研究方向】")
        self.assertEqual(records[2]["college"], "002【制药工程学院】")

    def test_parse_pdf_records_extracts_syphu_wrapped_ninth_batch_row(self):
        text = (
            "沈阳药科大学2025年硕士复试结果及拟录取名单公示（调剂第九批次）\n"
            "复试专业代码、名 研究方向代码、         初试总 专业考 综合素 复试总              录取排 拟录取\n"
            "考生编号           姓名     复试批次    复试学院                                                       总分              备注\n"
            "                        复试调剂第九 004【生命科学与生                01【临床药学研究\n"
            "100625000102925   张亚茹                       105500【药学】               306   172   83   255   561    1   拟录取 复试合格\n"
            "                          批次      物制药学院】                    方向】\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "沈阳药科大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://grs.syphu.edu.cn/__local/sample.pdf",
                    "title": "沈阳药科大学2025年硕士复试结果及拟录取名单公示（调剂第九批次）",
                    "year": 2025,
                },
            )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["student_id"], "100625000102925")
        self.assertEqual(records[0]["person_name"], "张亚茹")
        self.assertEqual(records[0]["college"], "004【生命科学与生物制药学院】")
        self.assertEqual(records[0]["admission_major"], "105500【药学】 01【临床药学研究方向】")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertIn("复试批次: 复试调剂第九批次", records[0]["remarks"])

    def test_parse_pdf_records_extracts_scuec_wrapped_recommendation_major(self):
        text = (
            "2026 年接收硕士推免生拟录取名单公示\n"
            "姓名        拟录取专业代码     拟录取专业名称       复试成绩     备注\n"
            "冀星屹      081000          信息与通信工程       91.40      支教团\n"
            "通信工程（含宽带网络、\n"
            "边巴卓玛    085402                         85.64\n"
            "移动通信等）\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "中南民族大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://www.scuec.edu.cn/__local/sample.pdf",
                    "title": "2026年接收硕士推免生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "冀星屹")
        self.assertEqual(records[0]["major"], "081000")
        self.assertEqual(records[0]["admission_major"], "信息与通信工程")
        self.assertEqual(records[1]["person_name"], "边巴卓玛")
        self.assertEqual(records[1]["major"], "085402")
        self.assertEqual(records[1]["admission_major"], "通信工程（含宽带网络、移动通信等）")
        self.assertIn("复试成绩: 85.64", records[1]["remarks"])

    def test_parse_pdf_records_extracts_cqmu_postgraduate_score_rows(self):
        text = (
            "重庆医科大学第一临床学院2026年硕士研究生第一志愿复试成绩公示\n"
            "（按总成绩从高到低排序）\n"
            "专业代码：105123     专业名称（方向）：放射影像学     复试日期：2026年3月30日\n"
            "总成绩\n"
            "考生编号          姓名    初试成绩   复试成绩   是否拟录取\n"
            "（百分制）\n"
            "106316200004785   谢艾佳   394.00   93.10   84.52   是\n"
            "106316200004816   陈思思   397.00   90.50   83.84   是\n"
            "106316200004999   王某某   300.00   70.00   64.00   否\n"
            "注：一、总成绩=初试成绩÷5×0.6+复试成绩×0.4\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "重庆医科大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://www.hospital-cqmu.com/oss/sample.pdf",
                    "title": "重庆医科大学第一临床学院2026年硕士研究生第一志愿复试成绩公示PDF",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["student_id"], "106316200004785")
        self.assertEqual(records[0]["person_name"], "谢艾佳")
        self.assertEqual(records[0]["college"], "第一临床学院")
        self.assertEqual(records[0]["major"], "105123")
        self.assertEqual(records[0]["admission_major"], "105123 放射影像学")
        self.assertIn("初试成绩: 394.00", records[0]["remarks"])
        self.assertIn("是否拟录取: 是", records[0]["remarks"])
        self.assertFalse(records[0]["needs_review"])

    def test_parse_pdf_records_extracts_kmust_postgraduate_wrapped_rows(self):
        text = (
            "昆明理工大学2026年硕士研究生拟录取公示名单\n"
            "序号 考生编号 姓名 拟录取学院代码 拟录取学院名称 拟录取专业代码 拟录取专业（领域） 拟录取学习方式 拟录取类别\n"
            "1 106746000006901 刘福 001 国土资源工\n"
            "程学院 070901 矿物学、岩石学\n"
            "、矿床学 全日制 非定向\n"
            "就业 75 71 132 139 417 93.20 87.32\n"
            "2 106746000002313 徐福怡 001 国土资源工\n"
            "程学院 070903 古生物学与地层\n"
            "学 全日制 定向就\n"
            "业 57 33 131 81 302 82.60 69.28\n"
            "3 106136085700597 谌边一\n"
            "丁 001 国土资源工\n"
            "程学院 085703 地质工程 全日制 非定向\n"
            "就业 61 60 63 75 259 74.12 60.73\n"
            "4 106746000009263 王昭云 004 信息工程与\n"
            "自动化学院 085410 人工智能 非全日\n"
            "制 定向就\n"
            "业 69 72 113 92 346 81.10 73.96\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "昆明理工大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://www.kmust.edu.cn/system/_content/download.jsp?wbfileid=17827351",
                    "title": "昆明理工大学2026年硕士研究生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 4)
        self.assertEqual(records[0]["student_id"], "106746000006901")
        self.assertEqual(records[0]["person_name"], "刘福")
        self.assertEqual(records[0]["college"], "国土资源工程学院")
        self.assertEqual(records[0]["major"], "070901")
        self.assertEqual(records[0]["admission_major"], "070901 矿物学、岩石学、矿床学")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertIn("综合成绩: 87.32", records[0]["remarks"])
        self.assertEqual(records[1]["admission_major"], "070903 古生物学与地层学")
        self.assertIn("拟录取类别: 定向就业", records[1]["remarks"])
        self.assertEqual(records[2]["person_name"], "谌边一丁")
        self.assertEqual(records[2]["admission_major"], "085703 地质工程")
        self.assertEqual(records[3]["college"], "信息工程与自动化学院")
        self.assertIn("拟录取学习方式: 非全日制", records[3]["remarks"])

    def test_parse_pdf_records_extracts_ynnu_postgraduate_columns(self):
        text = (
            "云南师范大学2026年硕士研究生拟录取名单\n"
            "                         录取院               录取专                 录取学习\n"
            "    考生编号          考生姓名           录取院系所名称              录取专业名称        复试成绩      综合成绩   备注\n"
            "                         系所码               业代码                  方式\n"
            "106816000000001   鲁远驰    001   法学与社会学学院    030101   法学理论       全日制   85.20   79.5\n"
            "106816000000077         001\n"
            "                  美丽亚·赛衣丁     法学与社会学学院   030301   社会学       全日制  93.40   82.1\n"
            "106816000001071   韩宏钰   002   马克思主义学院   030506   研究\n"
            "106816000000287   石环    001   法学与社会学学院   035101   法律（非法学）   非全日制 89.80   74.3    双少生\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "云南师范大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://grs.ynnu.edu.cn/info/1035/1503.htm",
                    "title": "云南师范大学2026年硕士研究生招生考试拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 4)
        self.assertEqual(records[0]["student_id"], "106816000000001")
        self.assertEqual(records[0]["person_name"], "鲁远驰")
        self.assertEqual(records[0]["college"], "法学与社会学学院")
        self.assertEqual(records[0]["major"], "030101")
        self.assertEqual(records[0]["admission_major"], "030101 法学理论")
        self.assertIn("录取院系所码: 001", records[0]["remarks"])
        self.assertIn("复试成绩: 85.20", records[0]["remarks"])
        self.assertIn("综合成绩: 79.5", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "美丽亚·赛衣丁")
        self.assertEqual(records[1]["admission_major"], "030301 社会学")
        self.assertEqual(records[2]["person_name"], "韩宏钰")
        self.assertEqual(records[2]["admission_major"], "030506 研究")
        self.assertTrue(records[2]["needs_review"])
        self.assertEqual(records[3]["person_name"], "石环")
        self.assertIn("录取学习方式: 非全日制", records[3]["remarks"])
        self.assertIn("备注: 双少生", records[3]["remarks"])

    def test_clean_records_keeps_same_name_with_different_student_ids(self):
        with TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            records_path = tmp_path / "records.jsonl"
            clean_path = tmp_path / "records_clean.csv"
            summary_path = tmp_path / "summary.csv"
            base = {
                "school_name": "云南师范大学",
                "year": 2026,
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "person_name": "张丽",
                "college": "管理学院",
                "major": "125100",
                "admission_major": "125100 工商管理",
                "source_url": "https://grs.ynnu.edu.cn/info/1035/1503.htm",
                "title": "云南师范大学2026年硕士研究生招生考试拟录取名单公示",
                "needs_review": False,
            }
            rows = [
                {**base, "student_id": "106816000001801"},
                {**base, "student_id": "106156125100037"},
            ]
            records_path.write_text(
                "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows),
                encoding="utf-8",
            )

            summary = crawler.clean_records_to_outputs(records_path, clean_path, summary_path)
            with clean_path.open("r", encoding="utf-8-sig", newline="") as handle:
                cleaned = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 2)
        self.assertEqual({row["student_id"] for row in cleaned}, {"106816000001801", "106156125100037"})

    def test_parse_pdf_records_extracts_swjtu_recommendation_columns(self):
        text = (
            "招生              录取专业                     复试\n"
            "姓名     证件号码               录取院系所名册                   录取专业名称\n"
            "                    类型               代码                      成绩\n"
            "王鹏  532****031X 硕士 土木工程学院           081401   岩土工程            92\n"
            "闫鹏程   500****0317\n"
            "                硕士 信息科学与技术学院        085402   通信工程（含宽带网络、移动通信等） 85.2\n"
            "赵梓杰   510****1217   直博生 计算机与人工智能学院       081200   计算机科学与技术 94.25\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "西南交通大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://gsnews.swjtu.edu.cn/info/2127/32054.htm",
                    "title": "西南交通大学2026年推荐免试研究生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 3)
        self.assertEqual(records[0]["person_name"], "王鹏")
        self.assertEqual(records[0]["student_id"], "532****031X")
        self.assertEqual(records[0]["college"], "土木工程学院")
        self.assertEqual(records[0]["major"], "081401")
        self.assertEqual(records[0]["admission_major"], "081401 岩土工程")
        self.assertIn("招生类型: 硕士", records[0]["remarks"])
        self.assertIn("复试成绩: 92", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "闫鹏程")
        self.assertEqual(records[1]["admission_major"], "085402 通信工程（含宽带网络、移动通信等）")
        self.assertIn("复试成绩: 85.2", records[1]["remarks"])
        self.assertEqual(records[2]["person_name"], "赵梓杰")
        self.assertIn("招生类型: 直博生", records[2]["remarks"])

    def test_parse_pdf_records_extracts_swfu_recommendation_rows(self):
        text = (
            "西南林业大学2026年推荐免试研究生（含直博生）拟录取名单\n"
            "                                  报考院系          报考专业\n"
            "序号             考生编号    姓名    层次          报考院系               报考专业      学习形式   总成绩     录取类别\n"
            "                                   代码            代码\n"
            "1    106776106770001   黄梦江   直博    001   林学院    090702   森林培育         全日制    71.88 非定向就业\n"
            "2    106776106770002   付侯林   直博    001   林学院    090705   野生动植物保护与利用   全日制    73.25 非定向就业\n"
        )
        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "西南林业大学",
                    "document_type": "recommendation_exemption_list",
                    "year": "2026",
                    "source_url": "https://yjsy.swfu.edu.cn/info/1522/10996.htm",
                    "title": "西南林业大学2026年推荐免试研究生（含直博生）拟录取名单.pdf",
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "黄梦江")
        self.assertEqual(records[0]["student_id"], "106776106770001")
        self.assertEqual(records[0]["college"], "001 林学院")
        self.assertEqual(records[0]["major"], "090702")
        self.assertEqual(records[0]["admission_major"], "090702 森林培育")
        self.assertIn("层次: 直博", records[0]["remarks"])
        self.assertIn("总成绩: 71.88", records[0]["remarks"])
        self.assertFalse(records[0].get("needs_review"))

    def test_parse_pdf_records_extracts_tjus_recommendation_rows_with_masked_ids(self):
        text = (
            "天津体育学院 2026年接收推免生拟录取名单\n"
            "序号    姓名         身份证号            复试分数      报考专业\n"
            " 1    秦煜坤   140***********041X    92.6    运动人体科学\n"
            " 2    薛艺嘉   410***********3521    89.8    运动人体科学\n"
        )
        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "天津体育学院",
                    "document_type": "recommendation_exemption_list",
                    "year": "2026",
                    "source_url": "https://yjsb.tjus.edu.cn/info/1004/4527.htm",
                    "title": "天津体育学院2026年接收推免生拟录取名单.pdf",
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "秦煜坤")
        self.assertEqual(records[0]["student_id"], "140***********041X")
        self.assertEqual(records[0]["admission_major"], "运动人体科学")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertEqual(records[0]["remarks"], "复试分数: 92.6")
        self.assertFalse(records[0].get("needs_review"))

    def test_parse_pdf_records_extracts_hrbipe_recommendation_rows_without_header(self):
        text = (
            "哈尔滨体育学院2026年硕士研究生接收推免生拟录取名单\n"
            "                                           外国语 理论（专项）\n"
            "序号   姓名     拟录取专业         考生编号                        复试成绩         备注\n"
            "                                            成绩  考试成绩\n\n"
            "1    吴天庆    民族传统体育学   230904200208021112   71.00   94.00   87.10   拟录取\n"
            "2    李皓月    体育教育训练学   411424200409210027   90.00   82.67   84.87   拟录取\n"
        )
        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "哈尔滨体育学院",
                    "document_type": "recommendation_exemption_list",
                    "year": "2026",
                    "source_url": "https://www.hrbipe.edu.cn/yjsy/info/1011/1634.htm",
                    "title": "哈尔滨体育学院2026年硕士研究生接收推免生拟录取名单.pdf",
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "吴天庆")
        self.assertEqual(records[0]["student_id"], "230904200208021112")
        self.assertEqual(records[0]["admission_major"], "民族传统体育学")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertIn("复试成绩: 87.10", records[0]["remarks"])
        self.assertNotIn("考试成绩", {record["person_name"] for record in records})
        self.assertFalse(records[0].get("needs_review"))

    def test_parse_pdf_records_extracts_blcu_master_recommendation_rows(self):
        text = (
            "姓名          身份证号               拟录取专业代码及名称       复试成绩\n"
            "    赵同     4213***********625        020200-应用经济学    97.00\n"
            "阿丽米热·买买提   6529***********009       045300-国际中文教育    90.00\n"
        )
        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "北京语言大学",
                    "document_type": "recommendation_exemption_list",
                    "year": "2026",
                    "source_url": "https://yjsy.blcu.edu.cn/info/1071/6569.htm",
                    "title": "北京语言大学2026年硕士推免生拟录取名单.pdf",
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "赵同")
        self.assertEqual(records[0]["student_id"], "4213***********625")
        self.assertEqual(records[0]["major"], "020200")
        self.assertEqual(records[0]["admission_major"], "020200 应用经济学")
        self.assertEqual(records[0]["remarks"], "复试成绩: 97.00")
        self.assertEqual(records[1]["person_name"], "阿丽米热·买买提")
        self.assertFalse(records[0].get("needs_review"))

    def test_parse_pdf_records_extracts_blcu_direct_doctor_recommendation_rows(self):
        text = (
            "姓名         身份证号              拟录取专业代码及名称       录取导师   复试成绩\n"
            " 张帅   6201***********225   050102-语言学及应用语言学    熊仲儒    94.40\n"
            "张诗雨   5107***********649   050108-比较文学与世界文学    陈戎女    94.60\n"
            "                           第 1 页，共 1 页\n"
        )
        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "北京语言大学",
                    "document_type": "recommendation_exemption_list",
                    "year": "2026",
                    "source_url": "https://yjsy.blcu.edu.cn/info/1071/6569.htm",
                    "title": "北京语言大学2026年直博生拟录取名单.pdf",
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "张帅")
        self.assertEqual(records[0]["student_id"], "6201***********225")
        self.assertEqual(records[0]["major"], "050102")
        self.assertEqual(records[0]["admission_major"], "050102 语言学及应用语言学")
        self.assertEqual(records[0]["remarks"], "录取导师: 熊仲儒; 复试成绩: 94.40")
        self.assertNotIn("第 1 页，共 1 页", {record["person_name"] for record in records})
        self.assertFalse(records[0].get("needs_review"))

    def test_parse_pdf_records_extracts_smu_recommendation_rows(self):
        text = (
            "南方医科大学2026年招收推荐免试研究生拟录取名单公示\n"
            "序号   姓名    身份证号后6位     分委会       专业代码        专业名称           研究方向         学位类型   复试成绩    类别\n"
            "1   李杭栩    030012      法医学       101200      法医学           不分研究方向       学术型      89    硕士\n"
            "7   陈至臻    127058     公共管理学      120400     公共管理学         健康管理与促进       学术型     90.2   硕士\n"
            "102   任诗瞳    260425     基础医学、生物学      0710J3     生物信息学         不分研究方向       学术型     90.8   硕士\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "南方医科大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "year": "2026",
                    "source_url": "https://portal.smu.edu.cn/yzw/info/1002/11811.htm",
                    "title": "南方医科大学2026年招收推荐免试研究生拟录取名单公示.pdf",
                },
            )

        self.assertEqual(len(records), 3)
        self.assertEqual(records[0]["person_name"], "李杭栩")
        self.assertEqual(records[0]["student_id"], "030012")
        self.assertEqual(records[0]["college"], "法医学")
        self.assertEqual(records[0]["major"], "101200")
        self.assertEqual(records[0]["admission_major"], "法医学")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertIn("研究方向: 不分研究方向", records[0]["remarks"])
        self.assertIn("学位类型: 学术型", records[0]["remarks"])
        self.assertIn("复试成绩: 89", records[0]["remarks"])
        self.assertIn("类别: 硕士", records[0]["remarks"])
        self.assertEqual(records[2]["major"], "0710J3")
        self.assertEqual(records[2]["admission_major"], "生物信息学")
        self.assertFalse(records[0].get("needs_review"))

    def test_parse_pdf_records_extracts_cumt_recommendation_rows(self):
        text = (
            "中国矿业大学2026年接收推荐免试研究生拟录取名单公示\n"
            "           录取学院               录取专业                       复试综合\n"
            "序号   姓名            录取学院名称                     录取专业名称          招生类型      备注\n"
            "            代码                 代码                         成绩\n"
            "1    李牧     001   资源与地球科学学院   070500            地理学      96.40   硕士\n"
            "3    巨兰婷    001   资源与地球科学学院   070500            地理学      94.60   硕士    研究生支教团\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "中国矿业大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yz.cumt.edu.cn/2026jieshoutuimianyanjiushengniluqurenyuanmingdan.pdf",
                    "title": "中国矿业大学2026年接收推荐免试研究生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "李牧")
        self.assertEqual(records[0]["college"], "001 资源与地球科学学院")
        self.assertEqual(records[0]["admission_major"], "070500 地理学")
        self.assertIn("score 96.40", records[0]["remarks"])
        self.assertIn("硕士", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "巨兰婷")
        self.assertIn("研究生支教团", records[1]["remarks"])

    def test_parse_pdf_records_extracts_ahu_incoming_recommendation_rows(self):
        text = (
            "姓名 性别 毕业单位 拟录取学院 拟录取专业 考核成绩 备注\n"
            "张国豪 男 安徽大学 哲学学院 应用伦理 92.43\n"
            "沈玉竹 女 安徽大学 哲学学院 应用心理 92.2 研究生支教团\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "安徽大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://graschool.ahu.edu.cn/2026tuimian.pdf",
                    "title": "安徽大学2026年拟录取推免生公示名单",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "张国豪")
        self.assertEqual(records[0]["undergraduate_school"], "安徽大学")
        self.assertEqual(records[0]["college"], "哲学学院")
        self.assertEqual(records[0]["admission_major"], "应用伦理")
        self.assertIn("score 92.43", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "沈玉竹")
        self.assertNotIn("女", records[1]["person_name"])
        self.assertIn("研究生支教团", records[1]["remarks"])

    def test_parse_pdf_records_extracts_dlmu_recommendation_rows(self):
        text = (
            "                       大连海事大学2026年推免研究生（含直博生）拟录取考生名单\n"
            "     学位         拟录取学              拟录取专                 拟录取研   复式      就业    学习\n"
            "序号        姓名            拟录取学院名称            拟录取专业名称                                 备注\n"
            "     类别          院代码               业代码                 究方向码   成绩      方式    方式\n"
            "1    硕士   孙幼斌    001     航海学院     082302   交通信息工程及控制    04    86.99   非定向   全日制   研究生支教团\n"
            "\n"
            "2    硕士   闫梦园    001     航海学院     082302   交通信息工程及控制    01    90.69   非定向   全日制\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "大连海事大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://grs.dlmu.edu.cn/list.pdf",
                    "title": "大连海事大学2026年推免研究生（含直博生）拟录取考生名单.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "孙幼斌")
        self.assertEqual(records[0]["college"], "001 航海学院")
        self.assertEqual(records[0]["admission_major"], "082302 交通信息工程及控制")
        self.assertIn("direction 04", records[0]["remarks"])
        self.assertIn("score 86.99", records[0]["remarks"])
        self.assertIn("研究生支教团", records[0]["remarks"])
        self.assertNotIn("姓名", {record["person_name"] for record in records})

    def test_parse_pdf_records_extracts_dlmu_wrapped_and_letter_code_rows(self):
        text = (
            "135   硕士   李明     004    信息科学技术学院   085401                 01    83.49   非定向   全日制   工程硕博士专项\n"
            "197   硕士   王强     005    交通运输工程学院   0823Z3   交通运输工程      01    95.0   非定向   全日制\n"
            "289   硕士 张海明     007   环境与海洋工程学院   085501    机械工程      04    87.5    非定向   全日制\n"
            "415   硕士   李东阳    018                 0823Z4                 01    92.1    非定向   全日制\n"
            "428        赵明     001      航海学院       082302   交通信息工程及控制     01    88.8    非定向   全日制\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "大连海事大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://grs.dlmu.edu.cn/list.pdf",
                    "title": "大连海事大学2026年推免研究生（含直博生）拟录取考生名单.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 5)
        self.assertEqual(records[0]["admission_major"], "085401")
        self.assertEqual(records[1]["admission_major"], "0823Z3 交通运输工程")
        self.assertEqual(records[2]["person_name"], "张海明")
        self.assertEqual(records[3]["college"], "018")
        self.assertEqual(records[3]["admission_major"], "0823Z4")
        self.assertEqual(records[4]["person_name"], "赵明")
        self.assertIn("degree 直博", records[4]["remarks"])

    def test_parse_pdf_records_extracts_lnutcm_recommendation_rows(self):
        text = (
            "辽宁中医药大学2026年拟录取推免生（含直博生）名单\n"
            "                     拟录取专业               研究方             思想政治与 外语听说测试 专业素质测评\n"
            "序号   姓名    拟录取学院名称         拟录取专业名称             研究方向名称                           总分   备注\n"
            "                       代码                 向码             品德考查 （满分15分） （满分85分）\n"
            "1    李昕堉     药学院      100800     中药学      00   不区分研究方向    合格      13     76     89   直博生\n"
            "2    王淋     基础医学院     100601   中西医结合基础    01    中医方向      合格      10     75     85   硕士\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "辽宁中医药大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yjs.lnutcm.edu.cn/list.pdf",
                    "title": "辽宁中医药大学2026年拟录取推免生（含直博生）名单.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "李昕堉")
        self.assertEqual(records[0]["college"], "药学院")
        self.assertEqual(records[0]["admission_major"], "100800 中药学")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertIn("direction 00 不区分研究方向", records[0]["remarks"])
        self.assertIn("total 89", records[0]["remarks"])
        self.assertIn("直博生", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "王淋")

    def test_parse_pdf_records_extracts_dmu_recommendation_rows_with_continuations(self):
        text = (
            "大连医科大学2026年接收推荐优秀本科毕业生免试攻读硕士学位研究生拟录取名单\n"
            "院系所                                                            复试           是否\n"
            "         院系所名称        专业代码      专业名称     方向    研究方向     考生姓名           排名         备注\n"
            " 代码                                                            成绩           拟录取\n"
            "003   附属第一医院及所辖基地     105101    内科学      00   不区分研究方向   邢妍雨    79.2    1     是\n"
            "                      105101    内科学      00   不区分研究方向   高梓萌    89.6    1     是\n"
            "004   附属第二医院及所辖基地\n"
            "                      105123   放射影像学     00   不区分研究方向   张美玲     79     1     是\n"
            "008       护理学院                                          谢昕萍     83     2     是\n"
            "                      105400     护理      00   不区分研究方向\n"
            "                                                        任芷仪    76.5    3     是\n"
            "                      105701   中医内科学     00   不区分研究方向   吕泽阳    95.6    1     是\n"
            "010   中西医结合研究院及所辖基地\n"
            "                      105709   中西医结合临床   00   不区分研究方向   宋婉宁    92.6    1     是\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "大连医科大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yjs.dmu.edu.cn/master.pdf",
                    "title": "大连医科大学2026年接收推荐优秀本科毕业生免试攻读硕士学位研究生拟录取名单.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(
            [record["person_name"] for record in records],
            ["邢妍雨", "高梓萌", "张美玲", "谢昕萍", "任芷仪", "吕泽阳", "宋婉宁"],
        )
        self.assertEqual(records[0]["college"], "003 附属第一医院及所辖基地")
        self.assertEqual(records[1]["college"], "003 附属第一医院及所辖基地")
        self.assertEqual(records[2]["college"], "004 附属第二医院及所辖基地")
        self.assertEqual(records[3]["college"], "008 护理学院")
        self.assertEqual(records[4]["admission_major"], "105400 护理")
        self.assertIn("direction 00 不区分研究方向", records[4]["remarks"])
        self.assertIn("score 76.5", records[4]["remarks"])
        self.assertIn("admitted 是", records[4]["remarks"])
        self.assertEqual(records[5]["college"], "010 中西医结合研究院及所辖基地")
        self.assertEqual(records[6]["college"], "010 中西医结合研究院及所辖基地")

    def test_parse_pdf_records_extracts_dmu_direct_doctor_rows(self):
        text = (
            "大连医科大学2026年接收优秀应届本科毕业生直接攻读博士学位研究生拟录取名单\n"
            "院系所                                  研究方                         报考           考核   是否同意\n"
            "        院系所名称       专业代码     专业名称                研究方向                  考生姓名\n"
            " 代码                                  向代码                         导师           成绩    拟录取\n"
            "                                           1.乳腺癌及淋巴水肿疾病分子影像学研究\n"
            "004   附属第二医院及所辖基地   100210    外科学     01                         赵海东   马鸿壮    92    是\n"
            "                                            2.乳腺癌致病机制研究与临床转化研究\n"
            "011    肿瘤干细胞研究院     071009   细胞生物学    01       肿瘤RNA生物学方向        汪洋     冯靖    74    是\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "大连医科大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yjs.dmu.edu.cn/doctor.pdf",
                    "title": "大连医科大学2026年接收优秀应届本科毕业生直接攻读博士学位研究生拟录取名单.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual([record["person_name"] for record in records], ["马鸿壮", "冯靖"])
        self.assertEqual(records[0]["college"], "004 附属第二医院及所辖基地")
        self.assertEqual(records[0]["admission_major"], "100210 外科学")
        self.assertIn("advisor 赵海东", records[0]["remarks"])
        self.assertIn("research_direction 1.乳腺癌及淋巴水肿疾病分子影像学研究; 2.乳腺癌致病机制研究与临床转化研究", records[0]["remarks"])
        self.assertIn("degree 直博", records[0]["remarks"])
        self.assertEqual(records[1]["college"], "011 肿瘤干细胞研究院")
        self.assertIn("direction 01 肿瘤RNA生物学方向", records[1]["remarks"])

    def test_parse_pdf_records_splits_xzmu_recommendation_name_and_identity(self):
        text = (
            "西藏民族大学2026年接收推荐免试攻读硕士研究生拟录取名单（公示）\n"
            "考生姓名      身份证号          拟录取专业代码 拟录取专业名称       成绩\n"
            " 德吉 542329********0043    025300     税务      86.75\n"
            "旦增巴桑 540122********504X   030106    诉讼法学     82.30\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "西藏民族大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://www1.xzmu.edu.cn/yjsc/userfiles/file/list.pdf",
                    "title": "西藏民族大学2026年接收推荐免试攻读硕士研究生拟录取名单（公示）",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "德吉")
        self.assertEqual(records[0]["student_id"], "542329********0043")
        self.assertEqual(records[0]["admission_major"], "025300 税务")
        self.assertIn("score 86.75", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "旦增巴桑")
        self.assertEqual(records[1]["student_id"], "540122********504X")

    def test_parse_pdf_records_extracts_njtech_wrapped_adjustment_rows(self):
        text = (
            "南京工业大学 2026年调剂硕士研究生拟录取名单\n"
            "序号         考生编号           姓名       拟录取学院        拟录取专业       拟录取类别 拟录取学习形式 初试统考科目总分 复试成绩 综合成绩\n"
            "38      104976300307682   梁吉龙    材料科学与工程学院   080500材料科学与工程   非定向   全日制   244\n"
            "116.2   360.2\n"
            "机械与动力工程学院 080700动力工程及工程热物\n"
            "95      102876210211048   丁方磊    非定向   全日制   231   133.0   364.0\n"
            "（工业软件学院） 理\n"
            "柔性电子\n"
            "284     104036085402189   余杰    080300光学工程   非定向   全日制   246   246.0   370.2\n"
            "（未来技术）学院\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "南京工业大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://gra.njtech.edu.cn/list.pdf",
                    "title": "南京工业大学2026年调剂硕士研究生拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 3)
        self.assertEqual(records[0]["person_name"], "梁吉龙")
        self.assertEqual(records[0]["college"], "材料科学与工程学院")
        self.assertEqual(records[0]["admission_major"], "080500材料科学与工程")
        self.assertIn("total_score 360.2", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "丁方磊")
        self.assertEqual(records[1]["college"], "机械与动力工程学院（工业软件学院）")
        self.assertEqual(records[1]["admission_major"], "080700动力工程及工程热物理")
        self.assertEqual(records[2]["person_name"], "余杰")
        self.assertEqual(records[2]["college"], "柔性电子（未来技术）学院")
        self.assertEqual(records[2]["admission_major"], "080300光学工程")

    def test_parse_pdf_records_extracts_fudan_last_five_id_rows(self):
        text = (
            "\u8003\u751f\u7f16\u53f7\u540e\n"
            "         \u8003\u751f\u59d3\u540d   \u62df\u5f55\u53d6\u9662\u7cfb     \u521d\u8bd5\u603b\u6210\u7ee9 \u590d\u8bd5\u6210\u7ee9     \u603b\u6210\u7ee9       \u5907\u6ce8\n"
            " \u4e94\u4f4d\n\n"
            " 00465   \u674e*\u96c5    \u9a6c\u514b\u601d\u4e3b\u4e49\u5b66\u9662    406   90.75   85.98\n\n"
            " 09201   \u989c*\u71da    \u9a6c\u514b\u601d\u4e3b\u4e49\u5b66\u9662    366   85.00   79.10   \u5c11\u6570\u6c11\u65cf\u9aa8\u5e72\u8ba1\u5212\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u590d\u65e6\u5927\u5b66",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://gsao.fudan.edu.cn/list.pdf",
                    "title": "2026\u5e74\u8003\u8bd5\u62db\u751f\u7855\u58eb\u62df\u5f55\u53d6\u540d\u5355\u516c\u793a\uff08\u5b9a\uff09.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["student_id"], "00465")
        self.assertEqual(records[0]["person_name"], "\u674e*\u96c5")
        self.assertEqual(records[0]["college"], "\u9a6c\u514b\u601d\u4e3b\u4e49\u5b66\u9662")
        self.assertIn("initial_score 406", records[0]["remarks"])
        self.assertIn("\u5c11\u6570\u6c11\u65cf\u9aa8\u5e72\u8ba1\u5212", records[1]["remarks"])

    def test_parse_pdf_records_extracts_bucea_split_header_rows(self):
        text = (
            "\u5317\u4eac\u5efa\u7b51\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355\n"
            "\u5e8f                                         \u62df\u5f55\u53d6\u4e13                \u662f\u5426 \u521d\u8bd5\u6210\n"
            "       \u8003\u751f\u7f16\u53f7            \u8003\u751f\u59d3\u540d    \u5b66\u9662\u540d\u79f0                \u62df\u5f55\u53d6\u4e13\u4e1a\u540d\u79f0           \u590d\u8bd5\u6210\u7ee9 \u603b\u6210\u7ee9 \u5b66\u4e60\u65b9\u5f0f             \u5907\u6ce8\n"
            "\u53f7                                          \u4e1a\u4ee3\u7801               \u4e00\u5fd7\u613f  \u7ee9\n"
            "1    100166081300011    *\u94b6\u94a6   \u5efa\u7b51\u4e0e\u57ce\u5e02\u89c4\u5212\u5b66\u9662   081300     \u5efa\u7b51\u5b66     \u4e00\u5fd7\u613f   393   85.76   82.18   \u5168\u65e5\u5236\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u5317\u4eac\u5efa\u7b51\u5927\u5b66",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yjsy.bucea.edu.cn/list.pdf",
                    "title": "\u5317\u4eac\u5efa\u7b51\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["student_id"], "100166081300011")
        self.assertEqual(records[0]["person_name"], "*\u94b6\u94a6")
        self.assertEqual(records[0]["college"], "\u5efa\u7b51\u4e0e\u57ce\u5e02\u89c4\u5212\u5b66\u9662")
        self.assertEqual(records[0]["admission_major"], "081300 \u5efa\u7b51\u5b66")
        self.assertIn("\u4e00\u5fd7\u613f", records[0]["remarks"])

    def test_parse_pdf_records_extracts_college_major_exam_name_score_rows(self):
        text = (
            "\u5b89\u5fbd\u7406\u5de5\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355\u516c\u793a\uff08\u4e00\u5fd7\u613f\uff09\n"
            "\u7cfb\u6240\u7801    \u5b66\u9662\u540d\u79f0     \u4e13\u4e1a\u4ee3\u7801       \u4e13\u4e1a\u540d\u79f0           \u51c6\u8003\u8bc1\u53f7\u7801        \u8003\u751f\u59d3\u540d \u521d\u8bd5\u6210\u7ee9 \u590d\u8bd5\u6210\u7ee9 \u603b\u6210\u7ee9          \u62df\u5f55\u53d6\u7c7b\u578b\n"
            "001   \u5730\u7403\u4e0e\u73af\u5883\u5b66\u9662    081800 \u5730\u8d28\u8d44\u6e90\u4e0e\u5730\u8d28\u5de5\u7a0b 103616210005409 \u6613\u51e1\u68ee        361   79.68   75.19    \u5168\u65e5\u5236\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u5b89\u5fbd\u7406\u5de5\u5927\u5b66",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yjszs.aust.edu.cn/list.pdf",
                    "title": "\u5b89\u5fbd\u7406\u5de5\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["college"], "001 \u5730\u7403\u4e0e\u73af\u5883\u5b66\u9662")
        self.assertEqual(records[0]["admission_major"], "081800 \u5730\u8d28\u8d44\u6e90\u4e0e\u5730\u8d28\u5de5\u7a0b")
        self.assertEqual(records[0]["student_id"], "103616210005409")
        self.assertEqual(records[0]["person_name"], "\u6613\u51e1\u68ee")
        self.assertIn("initial_score 361", records[0]["remarks"])
        self.assertIn("total_score 75.19", records[0]["remarks"])

    def test_parse_pdf_records_extracts_dhu_postgraduate_rows(self):
        text = (
            "\u4e1c\u534e\u5927\u5b662026\u7ea7\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355\n"
            "\u5b66\u9662                                          \u62df\u5f55\u53d6\u4e13\n"
            "      \u5b66\u9662\u540d\u79f0      \u8003\u751f\u7f16\u53f7           \u59d3\u540d    \u5b66\u4f4d\u7c7b\u578b             \u62df\u5f55\u53d6\u4e13\u4e1a\u540d\u79f0     \u521d\u8bd5\u6210\u7ee9 \u590d\u8bd5\u6210\u7ee9       \u603b\u6210\u7ee9      \u5b66\u4e60\u65b9\u5f0f       \u5907\u6ce8\n"
            "\u4ee3\u7801                                           \u4e1a\u4ee3\u7801\n"
            "001   \u7eba\u7ec7\u5b66\u9662   102556****04639   \u8bb8*\u6db5   \u5b66\u672f\u5b66\u4f4d   0821Z5      \u975e\u7ec7\u9020\u6750\u6599\u4e0e\u5de5\u7a0b       385   195.26   580.26   \u5168\u65e5\u5236\n"
            "007   \u65ed\u65e5\u5de5\u5546\u7ba1\u7406\u5b66\u9662   102556****09913 \u5e93*\u70ed\u63d0\u00b7\u7c73\u5409\u63d0 \u4e13\u4e1a\u5b66\u4f4d   125100      \u5de5\u5546\u7ba1\u7406       94   170.4   264.4   \u5168\u65e5\u5236   \u9000\u5f79\u5927\u5b66\u751f\u58eb\u5175\u4e13\u9879\u8ba1\u5212\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u4e1c\u534e\u5927\u5b66",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yjszs.dhu.edu.cn/list.pdf",
                    "title": "\u4e1c\u534e\u5927\u5b662026\u7ea7\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["college"], "001 \u7eba\u7ec7\u5b66\u9662")
        self.assertEqual(records[0]["admission_major"], "0821Z5 \u975e\u7ec7\u9020\u6750\u6599\u4e0e\u5de5\u7a0b")
        self.assertIn("total_score 580.26", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "\u5e93*\u70ed\u63d0\u00b7\u7c73\u5409\u63d0")
        self.assertEqual(records[1]["student_id"], "102556****09913")
        self.assertIn("\u9000\u5f79\u5927\u5b66\u751f\u58eb\u5175\u4e13\u9879\u8ba1\u5212", records[1]["remarks"])

    def test_parse_pdf_records_preserves_shiep_score_only_rows(self):
        text = (
            "\u4e0a\u6d77\u7535\u529b\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355\u516c\u793a\uff08\u7b2c\u4e8c\u6279/\u5168\u65e5\u5236\uff09\n"
            "\u59d3\u540d        \u8003\u8bd5\u7f16\u53f7           \u521d\u8bd5\u6210\u7ee9     \u590d\u8bd5\u6210\u7ee9     \u603b\u6210\u7ee9        \u4e13\u9879\u8ba1\u5212\u7b49   \u62a5\u8003\u7c7b\u522b    \u62a5\u8003\u5b66\u4e60\u65b9\u5f0f\n"
            "\u536b\u5d07\u54f2    100046135204039   358.00   184.00   542.00     \u666e\u901a\u8ba1\u5212    \u975e\u5b9a\u5411\u5c31\u4e1a    \u5168\u65e5\u5236\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u4e0a\u6d77\u7535\u529b\u5927\u5b66",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yjsc.shiep.edu.cn/list.pdf",
                    "title": "\u4e0a\u6d77\u7535\u529b\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "\u536b\u5d07\u54f2")
        self.assertEqual(records[0]["student_id"], "100046135204039")
        self.assertIn("initial_score 358.00", records[0]["remarks"])
        self.assertIn("total_score 542.00", records[0]["remarks"])
        self.assertIn("\u975e\u5b9a\u5411\u5c31\u4e1a", records[0]["remarks"])

    def test_parse_pdf_records_extracts_ncut_admitted_retest_result_rows(self):
        text = (
            "\u5317\u65b9\u5de5\u4e1a\u5927\u5b66\u673a\u68b0\u4e0e\u6750\u6599\u5de5\u7a0b\u5b66\u96622026\u5e74\u7855\u58eb\u751f\u62db\u751f\u7b2c\u4e00\u5fd7\u613f\u8003\u751f\u590d\u8bd5\u7ed3\u679c(\u542b\u62df\u5f55\u53d6\u540d\u5355)\n"
            "\u5b66\u79d1\u4e13\u4e1a\u4ee3     \u5b66\u79d1\u4e13\u4e1a     \u5b66\u4e60\u65b9\u5f0f                               \u590d\u8bd5\u6210 \u603b\u6210\u7ee9                  \u662f\u5426\n"
            "\u5e8f\u53f7                                 \u8003\u751f\u7f16\u53f7            \u8003\u751f\u59d3\u540d\n"
            "\u7801        \u540d\u79f0       \u975e\u5168\u65e5                           \u603b\u6210\u7ee9 \u5230\u767e\u5206 \u5230\u767e\u5206            \u62df\u5f55\u53d6\n"
            "1    085507   \u5de5\u4e1a\u8bbe\u8ba1\u5de5\u7a0b    \u5168\u65e5\u5236    100096110900847   \u90b1\u6625\u840c   390   84.90   80.76    \u662f\n"
            "2    085507   \u5de5\u4e1a\u8bbe\u8ba1\u5de5\u7a0b    \u5168\u65e5\u5236    100096110900867   \u7530\u6d69\u7136   333   85.30   74.08    \u5426\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u5317\u65b9\u5de5\u4e1a\u5927\u5b66",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://cmm.ncut.edu.cn/result.pdf",
                    "title": "\u5317\u65b9\u5de5\u4e1a\u5927\u5b662026\u5e74\u7855\u58eb\u751f\u590d\u8bd5\u7ed3\u679c",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "\u90b1\u6625\u840c")
        self.assertEqual(records[0]["student_id"], "100096110900847")
        self.assertEqual(records[0]["admission_major"], "085507 \u5de5\u4e1a\u8bbe\u8ba1\u5de5\u7a0b")
        self.assertIn("initial_score 390", records[0]["remarks"])
        self.assertIn("admission_status \u662f", records[0]["remarks"])

    def test_parse_pdf_records_extracts_just_retest_ranking_rows(self):
        text = "\n".join(
            [
                "\u6c5f\u82cf\u79d1\u6280\u5927\u5b66\u7ecf\u6d4e\u7ba1\u7406\u5b66\u96622026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u590d\u8bd5\u6210\u7ee9\u6392\u5e8f\u8868",
                "\u751f\u6e90\u7c7b\u522b\uff08\u4e00\u5fd7\u613f\u6216\u8c03\u5242\uff09 \u590d\u8bd5\u4e13\u4e1a\u4ee3\u7801 \u590d\u8bd5\u4e13\u4e1a\u540d\u79f0 \u62a5\u8003\u5b66\u4e60\u65b9\u5f0f \u59d3\u540d \u8003\u751f\u7f16\u53f7 \u521d\u8bd5\u603b\u6210\u7ee9 \u590d\u8bd5\u603b\u6210\u7ee9 \u7efc\u5408\u6210\u7ee9 \u7efc\u5408\u6210\u7ee9\u6392\u540d \u662f\u5426\u62df\u5f55\u53d6",
                "\u4e00\u5fd7\u613f 120100 \u7ba1\u7406\u79d1\u5b66\u4e0e\u5de5\u7a0b \u5168\u65e5\u5236 \u9ec4\u4e9a\u5e0c 102896210602872 376 217.0 328.3 1 \u62df\u5f55\u53d6",
                "\u4e00\u5fd7\u613f 085500 \u673a\u68b0 \u5168\u65e5\u5236 \u8d75\u58eb\u8c6a 102896210303685 407 260.5 363.05 1 \u662f",
            ]
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u6c5f\u82cf\u79d1\u6280\u5927\u5b66",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://sem.just.edu.cn/list.pdf",
                    "title": "\u6c5f\u82cf\u79d1\u6280\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u590d\u8bd5\u62df\u5f55\u53d6\u540d\u5355\u516c\u793a.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "\u9ec4\u4e9a\u5e0c")
        self.assertEqual(records[0]["student_id"], "102896210602872")
        self.assertEqual(records[0]["major"], "120100")
        self.assertEqual(records[0]["admission_major"], "\u7ba1\u7406\u79d1\u5b66\u4e0e\u5de5\u7a0b")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertIn("\u62df\u5f55\u53d6", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "\u8d75\u58eb\u8c6a")
        self.assertEqual(records[1]["student_id"], "102896210303685")
        self.assertEqual(records[1]["admission_major"], "\u673a\u68b0")

    def test_parse_pdf_records_extracts_henu_recommendation_rows(self):
        text = (
            "\u6cb3\u5357\u5927\u5b662026\u5e74\u63a8\u514d\u751f\u62df\u5f55\u53d6\u540d\u5355\u516c\u793a\n"
            "\u62df\u5f55\u53d6\u9662\u7cfb  \u62df\u5f55\u53d6\u4e13\u4e1a  \u62df\u5f55\u53d6  \u62df\u5f55\u53d6  \u62df\u5f55\u53d6\n"
            "\u62df\u5f55\u53d6\u9662\u7cfb\u6240\u540d\u79f0  \u59d3\u540d  \u8bc1\u4ef6\u53f7\u7801  \u62df\u5f55\u53d6\u4e13\u4e1a\u540d\u79f0  \u5907\u6ce8\n"
            "\u6240\u4ee3\u7801  \u4ee3\u7801  \u5b66\u4e60\u65b9\u5f0f  \u7c7b\u578b  \u7c7b\u522b\n"
            "002  \u7ecf\u6d4e\u5b66\u9662  \u7a0b\u8bd7\u753b  371427********0028  020200  \u5e94\u7528\u7ecf\u6d4e\u5b66  \u5168\u65e5\u5236  \u7855\u58eb  \u975e\u5b9a\u5411  \u7814\u7a76\u751f\u652f\u6559\u56e2\n"
            "002  \u7ecf\u6d4e\u5b66\u9662  \u8c2d\u4e49\u542b  411627********3744  020200  \u5e94\u7528\u7ecf\u6d4e\u5b66  \u5168\u65e5\u5236  \u7855\u58eb  \u975e\u5b9a\u5411\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u6cb3\u5357\u5927\u5b66",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://grs.henu.edu.cn/list.pdf",
                    "title": "\u6cb3\u5357\u5927\u5b662026\u5e74\u63a8\u514d\u751f\u653b\u8bfb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355\u516c\u793a",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["college"], "002 \u7ecf\u6d4e\u5b66\u9662")
        self.assertEqual(records[0]["person_name"], "\u7a0b\u8bd7\u753b")
        self.assertEqual(records[0]["student_id"], "371427********0028")
        self.assertEqual(records[0]["admission_major"], "020200 \u5e94\u7528\u7ecf\u6d4e\u5b66")
        self.assertIn("\u7814\u7a76\u751f\u652f\u6559\u56e2", records[0]["remarks"])

    def test_parse_pdf_records_skips_repeated_header_continuation_rows(self):
        text = (
            "\u6cb3\u5317\u5e08\u8303\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355\uff08\u4e00\u5fd7\u613f\uff09\n"
            "\u521d\u8bd5  \u4e13\u4e1a  \u7efc\u5408  \u5916\u8bed\u53e3  \u590d\u8bd5  \u7efc\u5408\u6210\n"
            "\u5b66\u9662\u540d\u79f0  \u62df\u5f55\u53d6\u4e13\u4e1a\u540d\u79f0  \u5b66\u4e60\u65b9\u5f0f  \u8003\u751f\u7f16\u53f7  \u59d3\u540d  \u5907\u6ce8\n"
            "\u6210\u7ee9  \u57fa\u7840  \u80fd\u529b  \u8bed\u542c\u529b  \u603b\u5206  \u7ee9\n"
            "\u9a6c\u514b\u601d\u4e3b\u4e49\u5b66\u9662  \u54f2\u5b66  \u5168\u65e5\u5236  100946134408138  \u9ad8\u6dd1\u73ae  351  84  87.6  89.5  86.35  75.045\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u6cb3\u5317\u5e08\u8303\u5927\u5b66",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yjsy.hebtu.edu.cn/list.pdf",
                    "title": "\u6cb3\u5317\u5e08\u8303\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "\u9ad8\u6dd1\u73ae")
        self.assertEqual(records[0]["student_id"], "100946134408138")
        self.assertEqual(records[0]["admission_major"], "\u54f2\u5b66")

    def test_build_record_splits_student_id_prefix_from_college_field(self):
        record = crawler._build_record(
            {
                "school_name": "\u6e56\u5357\u4e2d\u533b\u836f\u5927\u5b66",
                "document_type": "postgraduate_admission_list",
                "source_url": "https://yjsy.hnucm.edu.cn/list.pdf",
                "title": "\u6e56\u5357\u4e2d\u533b\u836f\u5927\u5b662026\u5e74\u62df\u5f55\u53d6\u540d\u5355",
                "year": 2026,
            },
            {
                "person_name": "\u9093\u6587\u534e",
                "college": "105416431503384 \u4eba\u6587\u4e0e\u7ba1\u7406\u5b66\u9662",
                "admission_major": "125200",
            },
        )

        self.assertIsNotNone(record)
        self.assertEqual(record["student_id"], "105416431503384")
        self.assertEqual(record["college"], "\u4eba\u6587\u4e0e\u7ba1\u7406\u5b66\u9662")

    def test_build_record_rejects_pdf_header_fragments_and_score_names(self):
        document = {
            "school_name": "\u6e56\u5357\u4e2d\u533b\u836f\u5927\u5b66",
            "document_type": "postgraduate_admission_list",
            "source_url": "https://yjsy.hnucm.edu.cn/list.pdf",
            "title": "\u6e56\u5357\u4e2d\u533b\u836f\u5927\u5b662026\u5e74\u62df\u5f55\u53d6\u540d\u5355",
            "year": 2026,
        }

        self.assertIsNone(
            crawler._build_record(
                document,
                {"person_name": "\u4e1a\u4ee3\u7801", "college": "\u4e1a\u4ee3\u7801"},
            )
        )
        self.assertIsNone(
            crawler._build_record(
                document,
                {"person_name": "\u5206\u5236)", "college": "\u5206\u5236)", "admission_major": "\u5236)"},
            )
        )
        self.assertIsNone(
            crawler._build_record(
                document,
                {"person_name": "407.00", "college": "\u4e2d\u533b\u5b66\u9662", "admission_major": "105701"},
            )
        )

    def test_parse_pdf_records_extracts_yangtzeu_postgraduate_rows(self):
        text = (
            "\u957f\u6c5f\u5927\u5b662026\u5e74\u653b\u8bfb\u7855\u58eb\u5b66\u4f4d\u7814\u7a76\u751f\u62df\u5f55\u53d6\u8003\u751f\u540d\u5355\u6c47\u603b\u8868\n"
            "\u62df\u5f55\u53d6        \u521d\u8bd5\u603b   \u590d\u8bd5\u603b        \u62df\u5f55\u53d6   \u62df\u5f55\u53d6\n"
            "\u8003\u751f\u7f16\u53f7           \u59d3\u540d       \u62df\u5f55\u53d6\u5b66\u9662       \u62df\u5f55\u53d6\u4e13\u4e1a\u540d\u79f0       \u603b\u6210\u7ee9        \u5907\u6ce8\n"
            "\u4e13\u4e1a\u4ee3\u7801        \u5206     \u5206      \u5b66\u4e60\u65b9\u5f0f \u5b66\u4f4d\u7c7b\u522b\n"
            "\u5730\u7403\u79d1\u5b66\u5b66\u9662\uff08\u542b\u975e\u5e38\u89c4\u6cb9\n"
            "104896520222330   \u6bb5\u68a6    070900     \u5730\u8d28\u5b66     401   91.44   84.70   \u5168\u65e5\u5236   \u5b66\u672f\u5b66\u4f4d\n"
            "\u6c14\u6e56\u5317\u7701\u534f\u540c\u521b\u65b0\u4e2d\u5fc3\uff09\n"
            "104896520225701   \u6c6a\u94b0\u5f64   \u77f3\u6cb9\u5de5\u7a0b\u5b66\u9662   082001   \u6cb9\u6c14\u4e95\u5de5\u7a0b   401   95.02   86.13   \u5168\u65e5\u5236   \u5b66\u672f\u5b66\u4f4d\n"
            "\u9ec4\u535a\u5b87    \u57ce\u5e02\u5efa\u8bbe\u5b66\u9662   081400    \u571f\u6728\u5de5\u7a0b           82.12   82.12   \u5168\u65e5\u5236   \u5b66\u672f\u5b66\u4f4d   \u7acb\u529f\u8868\u5f70\u514d\u521d\u8bd5\u8003\u751f\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u957f\u6c5f\u5927\u5b66",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://gs.yangtzeu.edu.cn/list.pdf",
                    "title": "\u957f\u6c5f\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 3)
        self.assertEqual(records[0]["person_name"], "\u6bb5\u68a6")
        self.assertEqual(records[0]["college"], "\u5730\u7403\u79d1\u5b66\u5b66\u9662\uff08\u542b\u975e\u5e38\u89c4\u6cb9 \u6c14\u6e56\u5317\u7701\u534f\u540c\u521b\u65b0\u4e2d\u5fc3\uff09")
        self.assertEqual(records[0]["admission_major"], "070900 \u5730\u8d28\u5b66")
        self.assertIn("total_score 84.70", records[0]["remarks"])
        self.assertEqual(records[1]["college"], "\u77f3\u6cb9\u5de5\u7a0b\u5b66\u9662")
        self.assertEqual(records[2]["person_name"], "\u9ec4\u535a\u5b87")
        self.assertIn("\u7acb\u529f\u8868\u5f70\u514d\u521d\u8bd5\u8003\u751f", records[2]["remarks"])

    def test_parse_pdf_records_extracts_yangtzeu_recommendation_rows(self):
        text = (
            "\u957f\u6c5f\u5927\u5b66 2026 \u5e74\u514d\u8bd5\u653b\u8bfb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355\u516c\u793a\n"
            "\u62df\u5f55\u53d6                    \u62df\u5f55\u53d6        \u62df\u5f55\u53d6\n"
            "\u8003\u751f\u7f16\u53f7           \u59d3\u540d            \u62df\u5f55\u53d6\u5355\u4f4d\u540d\u79f0              \u590d\u8bd5\u6210\u7ee9    \u5907\u6ce8\n"
            "\u5355\u4f4d\u4ee3\u7801                  \u4e13\u4e1a\u4ee3\u7801       \u4e13\u4e1a\u540d\u79f0\n"
            "\u5730\u7403\u79d1\u5b66\u5b66\u9662\uff08\u542b\u975e\u5e38\u89c4\u6cb9\n"
            "104896104890119 \u5468\u68a6\u6b23     113                  081800   \u5730\u8d28\u8d44\u6e90\u4e0e\u5730\u8d28\u5de5\u7a0b   95.17\n"
            "\u6c14\u6e56\u5317\u7701\u534f\u540c\u521b\u65b0\u4e2d\u5fc3\uff09\n"
            "104896104890077 \u82cf\u5955\u94ed     111      \u77f3\u6cb9\u5de5\u7a0b\u5b66\u9662      082001    \u6cb9\u6c14\u4e95\u5de5\u7a0b      88.78\n"
            "104896104890080   \u8521\u743c    111      \u77f3\u6cb9\u5de5\u7a0b\u5b66\u9662      082002   \u6cb9\u6c14\u7530\u5f00\u53d1\u5de5\u7a0b     92.68   \u76f4\u535a\u751f\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u957f\u6c5f\u5927\u5b66",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://gs.yangtzeu.edu.cn/recommendation.pdf",
                    "title": "\u957f\u6c5f\u5927\u5b662026\u5e74\u514d\u8bd5\u653b\u8bfb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 3)
        self.assertEqual(records[0]["student_id"], "104896104890119")
        self.assertEqual(records[0]["person_name"], "\u5468\u68a6\u6b23")
        self.assertEqual(records[0]["college"], "113 \u5730\u7403\u79d1\u5b66\u5b66\u9662\uff08\u542b\u975e\u5e38\u89c4\u6cb9 \u6c14\u6e56\u5317\u7701\u534f\u540c\u521b\u65b0\u4e2d\u5fc3\uff09")
        self.assertEqual(records[0]["admission_major"], "081800 \u5730\u8d28\u8d44\u6e90\u4e0e\u5730\u8d28\u5de5\u7a0b")
        self.assertIn("reexam_score 95.17", records[0]["remarks"])
        self.assertEqual(records[1]["college"], "111 \u77f3\u6cb9\u5de5\u7a0b\u5b66\u9662")
        self.assertIn("\u76f4\u535a\u751f", records[2]["remarks"])

    def test_parse_pdf_records_extracts_swupl_two_line_score_rows(self):
        text = (
            "\u897f\u5357\u653f\u6cd5\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u62db\u751f\u8003\u8bd5\u62df\u5f55\u53d6\u540d\u5355-\u7b2c2\u6279\u590d\u8bd5\n"
            "035101-\u4e0d\u533a\u5206\u7814\u7a76\u65b9\u5411 \u7b2c01\u7ec4        \u975e\u5168\u65e5\u5236-\u6cd5\u5f8b\uff08\u975e\u6cd5\u5b66\uff09-\u975e\u5168\u65e5\u5236 \u7b2c01\u7ec4\n"
            "\u6392\u5e8f        \u8003\u751f\u7f16\u53f7          \u59d3\u540d    \u521d\u8bd5     \u9762\u8bd5       \u603b\u6210\u7ee9      \u5907\u6ce8   \u6392\u5e8f       \u8003\u751f\u7f16\u53f7          \u59d3\u540d    \u521d\u8bd5     \u9762\u8bd5       \u603b\u6210\u7ee9      \u5907\u6ce8\n"
            "1                            395   129.800   81.260        19                           363   123.600   75.540\n"
            "     106526235101343   \u6731*\u70e8                                      106526235103148   \u8096*\u60a6\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "\u897f\u5357\u653f\u6cd5\u5927\u5b66",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yjsy.swupl.edu.cn/list.pdf",
                    "title": "\u897f\u5357\u653f\u6cd5\u5927\u5b662026\u5e74\u7855\u58eb\u7814\u7a76\u751f\u62db\u751f\u8003\u8bd5\u62df\u5f55\u53d6\u540d\u5355",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["ranking"], "1")
        self.assertEqual(records[0]["student_id"], "106526235101343")
        self.assertEqual(records[0]["person_name"], "\u6731*\u70e8")
        self.assertEqual(records[0]["admission_major"], "035101-\u4e0d\u533a\u5206\u7814\u7a76\u65b9\u5411 \u7b2c01\u7ec4")
        self.assertIn("interview_score 129.800", records[0]["remarks"])
        self.assertEqual(records[1]["ranking"], "19")
        self.assertEqual(records[1]["student_id"], "106526235103148")
        self.assertEqual(records[1]["admission_major"], "\u975e\u5168\u65e5\u5236-\u6cd5\u5f8b\uff08\u975e\u6cd5\u5b66\uff09-\u975e\u5168\u65e5\u5236 \u7b2c01\u7ec4")

    def test_parse_pdf_records_extracts_zjut_postgraduate_rows_when_header_is_split(self):
        text = (
            "拟录取学院  拟录取学院名称  拟录取专业  拟录取专业名称  拟录取学习  专项计划  姓名  准考证号  身份证号\n"
            "001  化学工程学院  070300  化学  全日制  无  贺家瑞  103376210000375  430102********6017\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "浙江工业大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://www.yz.zjut.edu.cn/list.pdf",
                    "title": "浙江工业大学2026年硕士研究生拟录取名单",
                    "year": 2026,
                },
            )

        matched = [
            record
            for record in records
            if record["person_name"] == "贺家瑞" and record["student_id"] == "103376210000375"
        ]

        self.assertEqual(len(matched), 1)
        self.assertEqual(matched[0]["college"], "001 化学工程学院")
        self.assertEqual(matched[0]["admission_major"], "070300 化学")

    def test_parse_pdf_records_extracts_zjut_incoming_recommendation_rows(self):
        text = (
            "序号  姓名  身份证号（保密）  性别  拟录取类型  拟录取学院代码  拟录取学院名称  拟录取专业代码  拟录取专业名称\n"
            "1  蒋贤宇  331082********2612  男  硕士  001  化学工程学院  070300  化学\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "浙江工业大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://www.yz.zjut.edu.cn/tm.pdf",
                    "title": "浙江工业大学2026年拟接收推荐免试硕士研究生、直博生名单",
                    "year": 2026,
                },
            )

        matched = [
            record
            for record in records
            if record["person_name"] == "蒋贤宇" and record["student_id"] == "331082********2612"
        ]

        self.assertEqual(len(matched), 1)
        self.assertEqual(matched[0]["college"], "001 化学工程学院")
        self.assertEqual(matched[0]["admission_major"], "070300 化学")
        self.assertEqual(matched[0]["remarks"], "硕士")

    def test_parse_pdf_records_extracts_hebut_postgraduate_rows(self):
        text = (
            "录取\n"
            "序号  招生单元  考生编号  姓名  复试总分  初试总分  总成绩  学习方式  备注\n"
            "1  019电子信息工程学院080900电子科学与技术（01方向）  100806190109740  米阳阳  256.00  400  82.13  全日制\n"
            "23  014电气工程学院085801电气工程  102486122617708  董继泽  240.83  339  72.79  全日制  退役大学生士兵计划\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "河北工业大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yjs.hebut.edu.cn/docs/2026-04/list.pdf",
                    "title": "电子信息工程学院2026年硕士生一志愿考生拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["ranking"], "1")
        self.assertEqual(records[0]["college"], "019 电子信息工程学院")
        self.assertEqual(records[0]["admission_major"], "080900 电子科学与技术（01方向）")
        self.assertEqual(records[0]["student_id"], "100806190109740")
        self.assertEqual(records[0]["person_name"], "米阳阳")
        self.assertIn("retest_score 256.00", records[0]["remarks"])
        self.assertIn("total_score 82.13", records[0]["remarks"])
        self.assertEqual(records[1]["college"], "014 电气工程学院")
        self.assertEqual(records[1]["admission_major"], "085801 电气工程")
        self.assertIn("退役大学生士兵计划", records[1]["remarks"])

    def test_parse_pdf_records_extracts_hebut_recommendation_rows(self):
        text = (
            "河北工业大学2026年推免研究生拟录取名单\n"
            "序号  招生类型  学院代码  专业代码  专业名称  研究方向代码  研究方向名称  考生编号  姓名  性别  复试成绩  学制  专项计划  备注\n"
            "001  直博生  013  080700  动力工程及工程热物理  05  储能科学与工程  100806118450002  曾玥瑜  女  91.4  5年  普通计划\n"
            "003  硕士  011  025200  应用统计  02  大数据分析  100806100770004  封佳乐  男  91.5  3年  普通计划\n"
            "031  硕士  012  080400  仪器科学与技术  01  光视检测与诊断  100806100800032  吕雪松  男  90.6  3年  普通计划\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "河北工业大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yjs.hebut.edu.cn/docs/2025-10/recommend.pdf",
                    "title": "河北工业大学2026年推免研究生拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual([record["person_name"] for record in records], ["曾玥瑜", "封佳乐", "吕雪松"])
        self.assertEqual(records[0]["ranking"], "001")
        self.assertEqual(records[0]["college"], "013")
        self.assertEqual(records[0]["admission_major"], "080700 动力工程及工程热物理")
        self.assertIn("直博生", records[0]["remarks"])
        self.assertIn("direction 05 储能科学与工程", records[0]["remarks"])
        self.assertIn("score 91.4", records[0]["remarks"])
        self.assertEqual(records[1]["student_id"], "100806100770004")
        self.assertEqual(records[1]["admission_major"], "025200 应用统计")
        self.assertNotIn("别 成绩", [record["person_name"] for record in records])

    def test_parse_pdf_records_extracts_hgu_split_postgraduate_rows(self):
        text = (
            "河北地质大学2026年硕士研究生招生复试拟录取名单\n"
            "序号 姓名 考生编号 院系代码 院系名称 专业代码 专业名称 学习方式 初试总分 复试成绩 总成绩 录取结果 录取批次\n"
            "1 许延昌 100776001010035 001 地球科学学院 070900 地质学 全日制 400 93.2 83.96 拟录取 一志愿\n"
            "4 张净瑶\n"
            "100776001010101 001 地球科学学院 070900 地质学 全日制 408 86.8 83.16 拟录取 一志愿\n"
            "11 马嘉凯 100776001010029 001\n"
            "地球科学学院 070900 地质学 全日制 388 82.4 79.04 拟录取 一志愿\n"
            "18 滕若望 100776001010034 001 地球科学学院 070900 地质学\n"
            "用全日制 369 83 76.56 拟录取 一志愿\n"
            "140 马龙彪 100776003010006 003 085704 测绘工程 全日制 366 74.6 73.62 拟录取 一志愿\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "河北地质大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://www.hgu.edu.cn/list.pdf",
                    "title": "河北地质大学2026年硕士研究生招生复试拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual(
            [record["person_name"] for record in records],
            ["许延昌", "张净瑶", "马嘉凯", "滕若望", "马龙彪"],
        )
        self.assertEqual(records[1]["student_id"], "100776001010101")
        self.assertEqual(records[1]["college"], "001 地球科学学院")
        self.assertEqual(records[1]["admission_major"], "070900 地质学")
        self.assertIn("initial_score 408", records[1]["remarks"])
        self.assertIn("retest_score 86.8", records[1]["remarks"])
        self.assertIn("total_score 83.16", records[1]["remarks"])
        self.assertEqual(records[2]["student_id"], "100776001010029")
        self.assertEqual(records[3]["student_id"], "100776001010034")
        self.assertIn("total_score 76.56", records[3]["remarks"])
        self.assertEqual(records[4]["college"], "003")
        self.assertEqual(records[4]["admission_major"], "085704 测绘工程")
        self.assertFalse(any(record["needs_review"] for record in records))

    def test_parse_pdf_records_extracts_hgu_incoming_recommendation_rows(self):
        text = (
            "河北地质大学2026年推免生拟录取名单公示\n"
            "序号 姓名 录取计划类别 本科生学号 二级招生单位代码 二级招生单位名称 专业代码 专业名称 学习方式 复试总成绩\n"
            "1 邓浩文 普通计划 422145060203 006 经济学院 020200 应用经济学 全日制 92.24\n"
            "5 张霞 普通计划 422148010307 012 语言文化学院 055100 翻译 全日制 95.2\n"
            "8\n"
            "用\n"
            "司蒙蒙 普通计划 422146150524 007 管理学院 120200 工商管理学 全日制 89.04\n"
            "13 田雨欣 普通计划 422146130729 007 管理学院\n"
            "用 125300 会计 全日制 88.28\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "河北地质大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yjsxy.hgu.edu.cn/list.pdf",
                    "title": "河北地质大学2026年硕士研究生招生推免生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(
            [record["person_name"] for record in records],
            ["邓浩文", "张霞", "司蒙蒙", "田雨欣"],
        )
        self.assertEqual(records[0]["student_id"], "422145060203")
        self.assertEqual(records[0]["college"], "006 经济学院")
        self.assertEqual(records[0]["admission_major"], "020200 应用经济学")
        self.assertIn("plan_category 普通计划", records[0]["remarks"])
        self.assertIn("retest_total_score 92.24", records[0]["remarks"])
        self.assertEqual(records[2]["ranking"], "8")
        self.assertEqual(records[2]["admission_major"], "120200 工商管理学")
        self.assertEqual(records[3]["admission_major"], "125300 会计")
        self.assertFalse(any(record["needs_review"] for record in records))

    def test_parse_pdf_records_extracts_gxmzu_postgraduate_rows(self):
        text = (
            "民族学与社会学学院2026年硕士研究生招生拟录取考生汇总表（一志愿）\n"
            "序号\t复试批次\t考生编号\t姓名\t专业代码及名称\t研究方向\t拟录取情况\t总成绩\t排名\t复试\t外语听说\t思想政治\t初试总分\t备注\n"
            "1\t1\t106086202600186\t区盈盈\t030400民族学\t人类学与世界民族\t拟录取\t81.22\t1\t89.04\t81.00\t92.67\t367\n"
            "22\t1\t106086202603082\t莫雅微\t030400民族学\t中华民族学\t候补待录取\t63.80\t2\t71.80\t63.00\t77.00\t279\t少干计划；\n"
            "23\t1\t106086202604528\t夏诗尹\t035200社会工作\t不区分研究方向\t拟录取\t86.00\t1\t90.40\t90.00\t92.80\t408\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "广西民族大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://msy.gxmzu.edu.cn/list.pdf",
                    "title": "广西民族大学民族学与社会学学院2026年硕士研究生招生第一志愿拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual([record["person_name"] for record in records], ["区盈盈", "莫雅微", "夏诗尹"])
        self.assertEqual(records[0]["student_id"], "106086202600186")
        self.assertEqual(records[0]["admission_major"], "030400 民族学")
        self.assertIn("research_direction 人类学与世界民族", records[0]["remarks"])
        self.assertIn("admission_status 拟录取", records[0]["remarks"])
        self.assertIn("total_score 81.22", records[0]["remarks"])
        self.assertIn("initial_score 367", records[0]["remarks"])
        self.assertEqual(records[1]["ranking"], "22")
        self.assertIn("候补待录取", records[1]["remarks"])
        self.assertIn("少干计划", records[1]["remarks"])
        self.assertEqual(records[2]["admission_major"], "035200 社会工作")
        self.assertFalse(any(record["needs_review"] for record in records))

    def test_parse_pdf_records_extracts_gxu_incoming_recommendation_rows(self):
        text = (
            "广西大学2026年拟录取攻读硕士学位推免生名单\n"
            "录取学院代          姓名     招生类型 计划       录取专业代码   录取专业名称   学习方式   研究方向码 研究方向   录取类别 复试成绩\n"
            "001   公共管理学院   韦星宇    硕士   普通计划     120400   公共管理学    全日制    00   不区分研究方向 非定向    81.71\n"
            "003   法学院      程少倩    硕士   普通计划     030100   法学       全日制    00   不区分研究方向 非定向    84.20\n"
            "004   文学院      王琼金    硕士   研究生支教团   050100   中国语言文学   全日制    04   中国古代文学与文非定向    77.32\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "广西大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yjsc.gxu.edu.cn/list.pdf",
                    "title": "广西大学2026年拟录取攻读硕士学位推免生名单",
                    "year": 2026,
                },
            )

        self.assertEqual([record["person_name"] for record in records], ["韦星宇", "程少倩", "王琼金"])
        self.assertEqual(records[0]["college"], "001 公共管理学院")
        self.assertEqual(records[0]["admission_major"], "120400 公共管理学")
        self.assertIn("admission_type 硕士", records[0]["remarks"])
        self.assertIn("plan 普通计划", records[0]["remarks"])
        self.assertIn("study_mode 全日制", records[0]["remarks"])
        self.assertIn("retest_score 81.71", records[0]["remarks"])
        self.assertIn("研究生支教团", records[2]["remarks"])

    def test_parse_pdf_records_extracts_gxu_doctoral_admission_rows(self):
        text = (
            "广西大学2026年博士研究生第一批（一）拟录取名单\n"
            "序号 考生姓名 录取专业代码       录取专业名称      录取学院    总成绩\n"
            "1    赖密密    120400   公共管理学     公共管理学院    86.62\n"
            "14    王涛    120400   公共管理学     公共管理学院    84.44    中国-东盟研究院（东盟区域国别班）\n"
            "30   周婷玉    071000    生物学     生命科学与技术学院 85.00\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "广西大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yjsc.gxu.edu.cn/list.pdf",
                    "title": "广西大学2026年博士研究生第一批（一）拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual([record["person_name"] for record in records], ["赖密密", "王涛", "周婷玉"])
        self.assertEqual(records[0]["ranking"], "1")
        self.assertEqual(records[0]["college"], "公共管理学院")
        self.assertEqual(records[0]["admission_major"], "120400 公共管理学")
        self.assertIn("total_score 86.62", records[0]["remarks"])
        self.assertIn("中国-东盟研究院", records[1]["remarks"])
        self.assertEqual(records[2]["college"], "生命科学与技术学院")

    def test_parse_pdf_records_extracts_jlu_incoming_recommendation_rows(self):
        text = (
            "吉林大学2026年接收推荐免试攻读研究生拟录取名单\n"
            "院系                  录取专\n"
            "      院系名称     姓名            录取专业名称    学制 招生类型   备注\n"
            "所码                  业代码\n"
            "101   哲学社会学院   郭家禾 010101    马克思主义哲学   3   硕士\n"
            "101   哲学社会学院   崔鑫昊 010101    马克思主义哲学   5   直博生\n"
            "108   外国语学院   陈佳禹 050201    英语语言文学   3   硕士\n"
            "108   王云帆 050211   3   硕士\n"
            "208   沈煊雅 140200    国家安全学   3   硕士\n"
            "209   商学与管理学院 郝晓赟 020204    金融学   3   硕士\n"
            "301   数学学院   张腾 070100    数学   3   硕士   国优计划\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "吉林大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yjsy.jlu.edu.cn/list.pdf",
                    "title": "吉林大学2026年接收推荐免试攻读研究生拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual(
            [record["person_name"] for record in records],
            ["郭家禾", "崔鑫昊", "陈佳禹", "王云帆", "沈煊雅", "郝晓赟", "张腾"],
        )
        self.assertEqual(records[0]["college"], "101 哲学社会学院")
        self.assertEqual(records[0]["admission_major"], "010101 马克思主义哲学")
        self.assertEqual(records[0]["remarks"], "3 硕士")
        self.assertEqual(records[1]["remarks"], "5 直博生")
        self.assertEqual(records[3]["college"], "108 外国语学院")
        self.assertEqual(records[3]["admission_major"], "050211")
        self.assertEqual(records[4]["college"], "208")
        self.assertEqual(records[5]["college"], "209 商学与管理学院")
        self.assertEqual(records[6]["remarks"], "3 硕士 国优计划")

    def test_parse_pdf_records_extracts_college_section_recommendation_rows(self):
        text = (
            "黑龙江八一农垦大学 2026 年推免生拟录取名单\n"
            "学院        姓名     性别         拟录取专业   复试成绩\n"
            "          杨翼     女    生物学            93\n"
            "生命科学技术学院\n"
            "          王雨晴     女    生物学            77\n"
            "          谷镜人     男    农业工程           88\n"
            "工程学院\n"
            "          李旭东     男    农业工程           82\n"
            "农学院      高晟楠     女    作物学            74\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "黑龙江八一农垦大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yanjiu.byau.edu.cn/list.pdf",
                    "title": "黑龙江八一农垦大学2026年推免生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(
            [record["person_name"] for record in records],
            ["杨翼", "王雨晴", "谷镜人", "李旭东", "高晟楠"],
        )
        self.assertEqual(records[0]["college"], "生命科学技术学院")
        self.assertEqual(records[0]["admission_major"], "生物学")
        self.assertEqual(records[0]["remarks"], "女 复试成绩 93")
        self.assertEqual(records[3]["college"], "工程学院")
        self.assertEqual(records[4]["college"], "农学院")

    def test_parse_pdf_records_extracts_name_college_major_type_rows(self):
        text = (
            "哈尔滨工程大学2026年推荐免试硕士研究生、直博生拟录取名单\n"
            "姓名     拟录取学院    拟录取专业代码      拟录取专业名称      招生类型\n"
            "李涛     船舶工程学院     081500       水利工程       推免硕士\n"
            "王隽希     船舶工程学院     082401   船舶与海洋结构物设计制造   推免硕士\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "哈尔滨工程大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://yzb.hrbeu.edu.cn/list.pdf",
                    "title": "哈尔滨工程大学2026年推荐免试硕士研究生、直博生拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "李涛")
        self.assertEqual(records[0]["college"], "船舶工程学院")
        self.assertEqual(records[0]["admission_major"], "081500 水利工程")
        self.assertEqual(records[0]["remarks"], "推免硕士")

    def test_parse_pdf_records_extracts_sequence_name_type_school_major_rows(self):
        text = (
            "哈尔滨理工大学 2026年拟录取推免研究生名单\n"
            "序号    姓名          本科所在单位       拟录取学院                拟录取专业     方向码                拟录取研究方向                   复试成绩     备注\n"
            "1    范潇    直博生   哈尔滨理工大学   材料科学与化学工程学院    080500   材料科学与工程   00     不区分研究方向                                82.02   硕博专项\n"
            "14   刘智玄   硕士    哈尔滨理工大学   电气与电子工程学院      085801   电气工程      01                                            86.54   硕博专项\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "哈尔滨理工大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://graduate.hrbust.edu.cn/list.pdf",
                    "title": "哈尔滨理工大学2026年拟录取推免研究生名单",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "范潇")
        self.assertEqual(records[0]["undergraduate_school"], "哈尔滨理工大学")
        self.assertEqual(records[0]["college"], "材料科学与化学工程学院")
        self.assertEqual(records[0]["admission_major"], "080500 材料科学与工程")
        self.assertIn("直博生", records[0]["remarks"])
        self.assertIn("82.02", records[0]["remarks"])
        self.assertEqual(records[1]["admission_major"], "085801 电气工程")

    def test_parse_pdf_records_extracts_dlu_recommendation_major_name_rows(self):
        text = (
            "大连大学2026年接收推荐免试研究生拟录取名单公示\n"
            "序号   姓名       报考院系   拟录取专业代码    拟录取专业名称        研究方向       复试成绩 学习方式   层次\n"
            "1    葛思      美术学院     135600    美术与书法      文化遗产艺术创新与实践    86.6 全日制   硕士\n"
            "3   柳大可     中山临床学院    105124     超声医学        不区分研究方向      90.7 全日制   硕士\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "大连大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yjs.dlu.edu.cn/list.pdf",
                    "title": "大连大学2026年接收推荐免试研究生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "葛思")
        self.assertEqual(records[0]["college"], "美术学院")
        self.assertEqual(records[0]["major"], "135600")
        self.assertEqual(records[0]["admission_major"], "135600 美术与书法")
        self.assertIn("direction 文化遗产艺术创新与实践", records[0]["remarks"])
        self.assertIn("reexam_score 86.6", records[0]["remarks"])
        self.assertIn("全日制", records[0]["remarks"])
        self.assertIn("硕士", records[0]["remarks"])
        self.assertEqual(records[1]["admission_major"], "105124 超声医学")

    def test_parse_pdf_records_extracts_gxnu_postgraduate_rows(self):
        text = (
            "广西师范大学2026年硕士研究生招生考生拟录取名单公示（二）\n"
            "                         报考                                      研究方向代\n"
            "考生编号           姓名        学习方式 专业代码             专业名称                 初试总分   复试成绩 总成绩        专项计划   备注\n"
            "                         单位                                        码\n"
            "                  安克尔·\n"
            "107606440104402          001   全日制    120402   社会医学与卫生事业管理         00   334   83.30   75.05\n"
            "                  买合木提\n"
            "102006211309638   白骥川    009   全日制    070300        化学             04   268   78.20   65.90\n"
            "104766002120211   包晨莹    005   非全日制   125400      旅游管理             02   149   83.20   66.43\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "广西师范大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "http://www.yz.gxnu.edu.cn/list.pdf",
                    "title": "广西师范大学2026年硕士研究生招生考生拟录取名单公示（二）",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 3)
        self.assertEqual(records[0]["person_name"], "安克尔·买合木提")
        self.assertEqual(records[0]["student_id"], "107606440104402")
        self.assertEqual(records[0]["college"], "报考单位 001")
        self.assertEqual(records[0]["major"], "120402")
        self.assertEqual(records[0]["admission_major"], "120402 社会医学与卫生事业管理")
        self.assertIn("study_mode 全日制", records[0]["remarks"])
        self.assertIn("direction_code 00", records[0]["remarks"])
        self.assertIn("total_score 75.05", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "白骥川")
        self.assertEqual(records[1]["admission_major"], "070300 化学")
        self.assertIn("reexam_score 78.20", records[1]["remarks"])
        self.assertEqual(records[2]["admission_major"], "125400 旅游管理")
        self.assertIn("study_mode 非全日制", records[2]["remarks"])

    def test_parse_pdf_records_extracts_gxau_postgraduate_rows(self):
        text = (
            "广西艺术学院2026年硕士研究生招生调剂复试成绩及拟录取名单\n"
            "序号 学院 姓名 考生编号 复试准考证号 专业 研究方向 学习方式 初试总分 英语听说能力测试成绩 综合素质面试成绩 业务课测试 业务课测试成绩 复试总成绩 总成绩 拟录取结果 备注\n"
            "学术\n"
            "1 001人文学院 金帅笛 102846211905084 007 130100艺术学 01艺术理论 全日制 383 85.00 92.00 301艺术美学 92.00 91.65 84.13 拟录取\n"
            "学位\n"
            "学术\n"
            "2 001人文学院 蔡文慧 104756130100153 001 130100艺术学 01艺术理论 全日制 361 79.00 74.33 301艺术美学 74.00 74.35 73.27\n"
            "学位\n"
            "广西艺术学院2026年硕士研究生招生一志愿复试成绩及拟录取名单\n"
            "学术 302艺术管理\n"
            "5 001人文学院 吴梦莲 10 130100艺术学 02艺术管理 全日制 420 88.00 87.67 90.67 89.63 86.82 拟录取 正常上线\n"
            "学位 学概论\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "广西艺术学院",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://zsb.gxau.edu.cn/list.pdf",
                    "title": "广西艺术学院2026年硕士研究生招生复试成绩及拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "金帅笛")
        self.assertEqual(records[0]["student_id"], "102846211905084")
        self.assertEqual(records[0]["college"], "001人文学院")
        self.assertEqual(records[0]["major"], "130100")
        self.assertEqual(records[0]["admission_major"], "130100 艺术学")
        self.assertIn("direction 01艺术理论", records[0]["remarks"])
        self.assertIn("exam_number 007", records[0]["remarks"])
        self.assertIn("total_score 84.13", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "吴梦莲")
        self.assertEqual(records[1]["student_id"], "10")
        self.assertIn("course 302艺术管理学概论", records[1]["remarks"])

    def test_parse_pdf_records_extracts_lzjtu_postgraduate_rows(self):
        text = (
            "自动化与电气工程学院 2026 年硕士研究生拟录取名单（一志愿）公示\n"
            "考生编号           姓名    拟录取专业代码    拟录取专业名称\n"
            "107326001681834   王全喜    0802Z4   交通装备检测及控制工程\n"
            "107326107320010   王悦     080800      电气工程\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "兰州交通大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://dqxy.lzjtu.edu.cn/list.pdf",
                    "title": "自动化与电气工程学院2026年硕士研究生拟录取名单（一志愿）公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "王全喜")
        self.assertEqual(records[0]["student_id"], "107326001681834")
        self.assertEqual(records[0]["major"], "0802Z4")
        self.assertEqual(records[0]["admission_major"], "0802Z4 交通装备检测及控制工程")
        self.assertEqual(records[0]["college"], "自动化与电气工程学院")
        self.assertEqual(records[1]["admission_major"], "080800 电气工程")

    def test_parse_pdf_records_extracts_gmc_postgraduate_rows(self):
        text = (
            "贵州医科大学 2026年硕士研究生一志愿考生拟录取名单\n"
            "考生编号 姓名 学院代 码 拟录取学院名称 专业代 码 拟录取专业名称 拟录取研究方向码 拟录取研究方向名称 初试总分 复试成绩 总成绩\n"
            "106606000000001 李若含 005 护理学院 101100 护理学 00 不区分研究方向 327 67.04 66.22\n"
            "106606000004423 万佳模 010 生物医学工程学院 083100 生物医学工程 00 不区分研究方向 336 80.2 73.7\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "贵州医科大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yjsxy.gmc.edu.cn/list.pdf",
                    "title": "贵州医科大学2026年硕士研究生一志愿考生拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "李若含")
        self.assertEqual(records[0]["student_id"], "106606000000001")
        self.assertEqual(records[0]["college"], "005 护理学院")
        self.assertEqual(records[0]["major"], "101100")
        self.assertEqual(records[0]["admission_major"], "101100 护理学")
        self.assertIn("direction_code 00", records[0]["remarks"])
        self.assertIn("direction 不区分研究方向", records[0]["remarks"])
        self.assertIn("total_score 66.22", records[0]["remarks"])
        self.assertEqual(records[1]["admission_major"], "083100 生物医学工程")

    def test_parse_pdf_records_extracts_glut_split_college_major_name_row(self):
        text = (
            "桂林理工大学 计算机科学与工程学院/人工智能学院软件工程（专硕）专业2026年硕士研究生推免拟录取名单公示\n"
            "序号   考生性质   考生姓名   拟录取系院所代码   拟录取系院所名称    拟录取专业代码   拟录取专业名称   复试成绩    录取成绩    学位性质 学习方式          备注\n"
            "                              计算机科学与工程学             软件工程（专\n"
            "1    推免生    黄天池       007                  085405             71.75   71.75   专业学位   全日制   推免生--研究生支教团\n"
            "                               院/人工智能学院               硕）\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "桂林理工大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://cise.glut.edu.cn/list.pdf",
                    "title": "桂林理工大学计算机科学与工程学院软件工程专业2026年硕士研究生推免拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "黄天池")
        self.assertEqual(records[0]["college"], "007 计算机科学与工程学院/人工智能学院")
        self.assertEqual(records[0]["major"], "085405")
        self.assertEqual(records[0]["admission_major"], "085405 软件工程（专硕）")
        self.assertIn("reexam_score 71.75", records[0]["remarks"])
        self.assertIn("admission_score 71.75", records[0]["remarks"])
        self.assertIn("全日制", records[0]["remarks"])

    def test_parse_pdf_records_extracts_nxmu_postgraduate_rows(self):
        text = (
            "宁夏医科大学2026年硕士研究生招生调剂考生拟录取名单公示\n"
            "考生编号            姓名    专项计划   录取院系所码         录取院系所名称          录取专业代码      录取专业名称       录取研究方向码   录取研究方向名称   录取学习方式   初试总分   复试总成绩    录取成绩    备注\n"
            "101326990510116   杨镇仿    普通计划     001    第一临床医学院 （宁夏医科大学总医院 ）    100201       内科学           04      血液内科学       1       343    71.58   70.09\n"
            "102856210018834    李航    普通计划     002           基础医学院            0710Z1      化学生物学          00     不区分研究方向      1       287    65.44   61.42\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "宁夏医科大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://www.nxmu.edu.cn/list.pdf",
                    "title": "宁夏医科大学2026年硕士研究生招生调剂考生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "杨镇仿")
        self.assertEqual(records[0]["student_id"], "101326990510116")
        self.assertEqual(records[0]["college"], "001 第一临床医学院 （宁夏医科大学总医院 ）")
        self.assertEqual(records[0]["major"], "100201")
        self.assertEqual(records[0]["admission_major"], "100201 内科学")
        self.assertIn("plan 普通计划", records[0]["remarks"])
        self.assertIn("direction_code 04", records[0]["remarks"])
        self.assertIn("direction 血液内科学", records[0]["remarks"])
        self.assertIn("study_mode 1", records[0]["remarks"])
        self.assertIn("initial_score 343", records[0]["remarks"])
        self.assertIn("admission_score 70.09", records[0]["remarks"])
        self.assertEqual(records[1]["college"], "002 基础医学院")
        self.assertEqual(records[1]["admission_major"], "0710Z1 化学生物学")

    def test_parse_pdf_records_extracts_ujn_postgraduate_rows(self):
        text = (
            "济南大学2026年博士研究生拟录取名单公示\n"
            "序号      考生编号           姓名    报考院系代码        报考院系          报考专业代码     报考专业     学习形式 考试方式\n"
            "1    104276123301001   王秀强     301    建筑材料制备与测试技术重点实验室    080500   材料科学与工程   全日制   硕博连读\n"
            "9    104276123307002   朱龙飞     307       前沿交叉科学研究院        070300     化学      全日制   硕博连读\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "济南大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yz.ujn.edu.cn/list.pdf",
                    "title": "济南大学2026年博士研究生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "王秀强")
        self.assertEqual(records[0]["student_id"], "104276123301001")
        self.assertEqual(records[0]["college"], "301 建筑材料制备与测试技术重点实验室")
        self.assertEqual(records[0]["major"], "080500")
        self.assertEqual(records[0]["admission_major"], "080500 材料科学与工程")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertIn("study_mode 全日制", records[0]["remarks"])
        self.assertIn("exam_method 硕博连读", records[0]["remarks"])
        self.assertEqual(records[1]["admission_major"], "070300 化学")

    def test_parse_pdf_records_extracts_xzhmu_score_only_rows(self):
        text = (
            "徐州医科大学2026年硕士研究生一志愿拟录取名单\n"
            "考生编号          姓名   初试成绩   复试成绩     总成绩     备注\n"
            "10313*01557   白崇余    341   335.40   74.46\n"
            "10313*01264   白佳颐    381   356.96   81.42   士兵计划\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "徐州医科大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://yjs.xzhmu.edu.cn/list.pdf",
                    "title": "徐州医科大学2026年硕士研究生一志愿拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "白崇余")
        self.assertEqual(records[0]["student_id"], "10313*01557")
        self.assertIn("initial_score 341", records[0]["remarks"])
        self.assertIn("reexam_score 335.40", records[0]["remarks"])
        self.assertIn("total_score 74.46", records[0]["remarks"])
        self.assertIn("士兵计划", records[1]["remarks"])

    def test_parse_pdf_records_extracts_zstu_postgraduate_rows(self):
        text = (
            "浙江理工大学2026年硕士研究生拟录取名单公示\n"
            "报名号                  综合考核\n"
            "序号                   考生姓名           拟录取专业          录取类别 学位类型    备注\n"
            "     （准考证号码）                总成绩\n"
            "1    2024210401019   王启     89.24   纺织科学与工程        非定向   学术学位\n"
            "21   2024211001026   李佳霖    71.90   材料科学与工程        非定向   学术学位   科研项目博士\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "浙江理工大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://gradadmission.zstu.edu.cn/list.pdf",
                    "title": "浙江理工大学2026年硕士研究生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "王启")
        self.assertEqual(records[0]["student_id"], "2024210401019")
        self.assertEqual(records[0]["admission_major"], "纺织科学与工程")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertIn("total_score 89.24", records[0]["remarks"])
        self.assertIn("非定向", records[0]["remarks"])
        self.assertIn("学术学位", records[0]["remarks"])
        self.assertIn("科研项目博士", records[1]["remarks"])

    def test_parse_pdf_records_extracts_guet_split_postgraduate_rows(self):
        text = (
            "桂林电子科技大学建筑与交通工程学院2026年硕士研究生拟录取名单公示\n"
            "序号 复试学院代码及名称        考生编号        姓名 录取专业代码 录取专业名称 初试总分 复试总成绩 总成绩 学习形式 录取类别 备注\n"
            "1 015建筑与交通工程学院 105906543200735    陈梓航 082300 交通运输工程    256    75.12 63.16 全日制 非定向\n"
            "2    015建筑与交通工程学院    116466211300855 邓忠洋    086100    交通运输    302    61.28    60.84    全日制    非定向    计划调整补录取\n"
            "3    007数学与计算科学学院 105956701001783    曾勇平    070100 数学    345    79.88    74.44    全日制    非定向\n"
            "4\n"
            "105南宁研究院05    105956253004290    高歌    125300    会计    231    67.86    72.43    全日制    非定向\n"
            "5    010材料科学与工程学院    107016432510817    曹竞杰    080500 材料科学与工程 287    71.56    64.48    全日制    非定向\n"
            "6    015建筑与交通工程学院102946214011935 刘学    085900    土木水利    329    75.76    70.78    全日制    非定向\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "桂林电子科技大学",
                    "document_type": "postgraduate_admission_list",
                    "source_url": "https://www.guet.edu.cn/list.pdf",
                    "title": "建筑与交通工程学院2026年硕士研究生拟录取名单公示.pdf",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 6)
        self.assertEqual(records[0]["person_name"], "陈梓航")
        self.assertEqual(records[0]["student_id"], "105906543200735")
        self.assertEqual(records[0]["major"], "082300")
        self.assertEqual(records[0]["admission_major"], "082300 交通运输工程")
        self.assertEqual(records[1]["person_name"], "邓忠洋")
        self.assertEqual(records[1]["major"], "086100")
        self.assertEqual(records[1]["admission_major"], "086100 交通运输")
        self.assertIn("计划调整补录取", records[1]["remarks"])
        self.assertEqual(records[2]["person_name"], "曾勇平")
        self.assertEqual(records[2]["student_id"], "105956701001783")
        self.assertEqual(records[2]["college"], "007数学与计算科学学院")
        self.assertEqual(records[2]["admission_major"], "070100 数学")
        self.assertEqual(records[3]["person_name"], "高歌")
        self.assertEqual(records[3]["ranking"], "4")
        self.assertEqual(records[3]["college"], "105南宁研究院05")
        self.assertEqual(records[3]["admission_major"], "125300 会计")
        self.assertEqual(records[4]["person_name"], "曹竞杰")
        self.assertEqual(records[4]["admission_major"], "080500 材料科学与工程")
        self.assertIn("initial_score 287", records[4]["remarks"])
        self.assertEqual(records[5]["person_name"], "刘学")
        self.assertEqual(records[5]["student_id"], "102946214011935")
        self.assertEqual(records[5]["college"], "015建筑与交通工程学院")
        self.assertEqual(records[5]["admission_major"], "085900 土木水利")

    def test_parse_pdf_records_extracts_sdufe_wrapped_doctoral_rows(self):
        text = """
                       山东财经大学2026年“硕博连读”“申请考核”博士研究生拟录取名单（第二批）
序号        考生编号              姓名      录取学院       录取专业      报考导师   录取类别    学习方式   招生方式   拟录取总分

1    104566100300151   刘晓源       金融学院         金融学       彭红枫     非定向就业   全日制    硕博连读    92.33
2    104566100300152   丁紫茵       金融学院         金融学       彭红枫     非定向就业   全日制    硕博连读     91

                                              087100
4    104566100600130   高子恒       管理科学与工程学院              刘政敏     非定向就业   全日制    申请考核    89.7
                                              管理科学与工程
                                 中国经济研究院
6    104566102600114   樊云        （中国经济发展与安全   金融学       郭建峰     非定向就业   全日制    申请考核     90
                                 实验室）
        """
        document = {
            "school_name": "山东财经大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "source_url": "https://yjszs.sdufe.edu.cn/virtual_attach_file.vsb?e=.pdf",
            "title": "山东财经大学2026年博士研究生拟录取名单",
        }

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(Path("dummy.pdf"), document)

        self.assertEqual(len(records), 4)
        by_name = {record["person_name"]: record for record in records}
        self.assertEqual(by_name["刘晓源"]["admission_major"], "金融学")
        self.assertEqual(by_name["高子恒"]["college"], "管理科学与工程学院")
        self.assertEqual(by_name["高子恒"]["admission_major"], "087100 管理科学与工程")
        self.assertIn("报考导师: 刘政敏", by_name["高子恒"]["remarks"])
        self.assertEqual(by_name["樊云"]["college"], "中国经济研究院（中国经济发展与安全实验室）")
        self.assertEqual(by_name["樊云"]["admission_major"], "金融学")

    def test_parse_pdf_records_splits_shnu_adjustment_id_and_major_column(self):
        text = """
姓名    考生编号            调剂专业名称     录取类别
张婧琳   106116500600392 应用统计       非定向就业
杨焌翔   105746000001268 学科教学（物理）   非定向就业
        """
        document = {
            "school_name": "上海师范大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "source_url": "https://mathsc.shnu.edu.cn/list.pdf",
            "title": "调剂拟录取名单.pdf",
        }

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(Path("dummy.pdf"), document)

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "张婧琳")
        self.assertEqual(records[0]["student_id"], "106116500600392")
        self.assertEqual(records[0]["admission_major"], "应用统计")
        self.assertEqual(records[1]["student_id"], "105746000001268")
        self.assertEqual(records[1]["admission_major"], "学科教学（物理）")

    def test_parse_pdf_records_extracts_hainnu_master_admission_rows(self):
        text = """
海南师范大学2026年全国硕士研究生招生考试拟录取考生名单
序号      考生编号           姓名      拟录取学院代码    拟录取学院名称     拟录取专业代码       拟录取专业名称     学习方式  初试总成绩  复试总成绩  入学总成绩  加试1  加试2  备注
1    116586137012324   张嘉怡    001   马克思主义学院   030500   马克思主义理论        全日制    418    84.60   84.00   无     无     无
        """
        document = {
            "school_name": "海南师范大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "source_url": "https://yjsc.hainnu.edu.cn/list.pdf",
            "title": "海南师范大学2026年全国硕士研究生招生考试拟录取考生名单.pdf",
        }

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(Path("dummy.pdf"), document)

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "张嘉怡")
        self.assertEqual(records[0]["student_id"], "116586137012324")
        self.assertEqual(records[0]["college"], "001 马克思主义学院")
        self.assertEqual(records[0]["major"], "030500")
        self.assertEqual(records[0]["admission_major"], "030500 马克思主义理论")
        self.assertIn("study_mode 全日制", records[0]["remarks"])

    def test_parse_pdf_records_extracts_fjmu_recommendation_rows(self):
        text = """
福建医科大学2026年推免生拟录取名单
考生编号           姓名    复试成绩   拟录取类型   拟录取学院代码  拟录取学院名称   拟录取专业代码   拟录取专业名称  三级学科/研究方向   拟录取导师
103926103921001   林娇欣   94.40   直博生    001    基础医学院     100101   人体解剖与组织胚胎学     人体解剖学     陶武成
103926103920001   梁月珍   93.80   硕士     002   公共卫生学院     100401   流行病与卫生统计学                陈法
        """
        document = {
            "school_name": "福建医科大学",
            "year": 2026,
            "document_type": "incoming_recommendation_admission_list",
            "source_url": "https://yjsy.fjmu.edu.cn/list.pdf",
            "title": "福建医科大学2026年推免生拟录取名单.pdf",
        }

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(Path("dummy.pdf"), document)

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "林娇欣")
        self.assertEqual(records[0]["college"], "001 基础医学院")
        self.assertEqual(records[0]["major"], "100101")
        self.assertEqual(records[0]["admission_major"], "100101 人体解剖与组织胚胎学")
        self.assertIn("direction 人体解剖学", records[0]["remarks"])
        self.assertIn("advisor 陶武成", records[0]["remarks"])

    def test_parse_pdf_records_extracts_fjtcm_postgraduate_rows(self):
        text = """
福建中医药大学2026年拟录取统考硕士研究生名单公示（868人）
序号 招生学院名称 招生类型 招生专业代码 招生专业名称 姓名 考生编号 导师姓名 导师工作单位 初试成绩 复试成绩 成绩合计 是否拟录取
1    中医学院   学术型   100501   中医基础理论   郑锦虹     10***6210501004   夏淑洁    中医学院    302    84.49 72.45     是
5    中医学院   学术型   100502 中医临床基础     徐珂      10***6210502005   张喜奎    附属第二人民医 357    81.58 76.49     是
        """
        document = {
            "school_name": "福建中医药大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "source_url": "https://yjsy.fjtcm.edu.cn/list.pdf",
            "title": "公示版-2026年统考硕士复试录取公示名单.pdf",
        }

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(Path("dummy.pdf"), document)

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "郑锦虹")
        self.assertEqual(records[0]["college"], "中医学院")
        self.assertEqual(records[0]["major"], "100501")
        self.assertEqual(records[0]["admission_major"], "100501 中医基础理论")
        self.assertIn("advisor 夏淑洁", records[0]["remarks"])
        self.assertEqual(records[1]["admission_major"], "100502 中医临床基础")

    def test_parse_pdf_records_prioritizes_fjtcm_parser_over_broad_score_parser(self):
        text = """
福建中医药大学2026年拟录取统考硕士研究生名单公示（868人）
序号 招生学院名称 招生类型 招生专业代码 招生专业名称 姓名 考生编号 导师姓名 导师工作单位 初试成绩 复试成绩 成绩合计 是否拟录取
1    中医学院   学术型   100501   中医基础理论   郑锦虹     10***6210501004   夏淑洁    中医学院    302    84.49 72.45     是
5    中医学院   学术型   100502 中医临床基础     徐珂      10***6210502005   张喜奎    附属第二人民医 357    81.58 76.49     是
        """
        document = {
            "school_name": "福建中医药大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "source_url": "https://yjsy.fjtcm.edu.cn/list.pdf",
            "title": "公示版-2026年统考硕士复试录取公示名单.pdf",
        }

        with (
            patch.object(crawler, "extract_pdf_text", return_value=text),
            patch.object(
                crawler,
                "_records_from_gender_school_college_major_score_pdf_rows",
                return_value=[{"person_name": "通用解析误命中"}],
            ),
        ):
            records = crawler.parse_pdf_records(Path("dummy.pdf"), document)

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "郑锦虹")
        self.assertEqual(records[1]["person_name"], "徐珂")

    def test_parse_pdf_records_extracts_fjtcm_wrapped_real_world_rows(self):
        text = """
福建中医药大学2026年拟录取统考硕士研究生名单公示（868人）
序号 招生学院名称 招生类型 招生专业代码 招生专业名称 姓名 考生编号 导师姓名 导师工作单位 初试成绩 复试成绩 成绩合计 是否拟录取
17    中医学院    学术型    100505 中医诊断学    杨龙靖子 10***6432801561    中医学院    331    81.33 73.77    是
55    骨伤学院    专业学位型 105703 中医骨伤科学    王志宇    10***6215703144    林桦楠    厦门市中医院 378    85.23 80.42    是
104    针灸推拿学院 学术型    100512 针灸推拿学    严海欣    10***6210512010    王志福    福建中医药大学 346    71.56 70.38    是
130    针灸推拿学院 专业学位型 105900 针灸    孙梦娇    10***6215900064    张文兵    387    80.87 79.14    是
462    专业学位型 105701 中医内科学    王一诺    10***6215701237    375    90.78 82.89    是
580    专业学位型 105710 医，不授博士 魏建欣    10***6215710025    黄武松    附属第三人民医 385    81.53 79.27    是
851    学术型    1007Z1    陈筱莹    10***6210721007    高雪娟    福建中医药大学 350    85.41 77.71    是
        """
        document = {
            "school_name": "福建中医药大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "source_url": "https://yjsy.fjtcm.edu.cn/list.pdf",
            "title": "公示版-2026年统考硕士复试录取公示名单.pdf",
        }

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(Path("dummy.pdf"), document)

        self.assertEqual(len(records), 7)
        self.assertEqual(records[0]["person_name"], "杨龙靖子")
        self.assertEqual(records[0]["student_id"], "10***6432801561")
        self.assertEqual(records[1]["admission_major"], "105703 中医骨伤科学")
        self.assertEqual(records[2]["college"], "针灸推拿学院")
        self.assertEqual(records[3]["admission_major"], "105900 针灸")
        self.assertEqual(records[4]["person_name"], "王一诺")
        self.assertEqual(records[5]["person_name"], "魏建欣")
        self.assertEqual(records[6]["major"], "1007Z1")

    def test_parse_pdf_records_extracts_cueb_wrapped_doctoral_rows(self):
        text = """
首都经济贸易大学2026年博士研究生拟录取名单（第一批）
序号     学院         考生编号           姓名    拟录取专业    录取类别   拟录取导师   总成绩    拟录取情况
1    财政税务学院    100386120999029   石*颜    财政学     非定向   包健    92.2    拟录取
     城市经济与公共                           城市经济与战
16             100386120111033   田*鹏            非定向   李强    79.77   拟录取
       管理学院                              略管理
        """
        document = {
            "school_name": "首都经济贸易大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "source_url": "https://yjs.cueb.edu.cn/list.pdf",
            "title": "首都经济贸易大学2026年博士研究生拟录取名单（第一批）",
        }

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(Path("dummy.pdf"), document)

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["admission_major"], "财政学")
        self.assertIn("非定向", records[0]["remarks"])
        self.assertEqual(records[1]["college"], "城市经济与公共管理学院")
        self.assertEqual(records[1]["admission_major"], "城市经济与战略管理")

    def test_records_from_table_maps_admission_major_code_and_name_columns(self):
        rows = [
            [
                "姓名",
                "录取类型",
                "专项计划",
                "录取院系所",
                "录取专业代码",
                "录取专业名称",
                "录取研究方向",
                "面试成绩",
                "毕业单位",
                "毕业专业",
            ],
            [
                "查金诚",
                "直博生",
                "普通计划",
                "轻工与化学工程学院",
                "0822Z1",
                "生物质能源与材料",
                "不区分研究方向",
                "93.0",
                "大连工业大学",
                "生物质能源与材料",
            ],
        ]

        records = crawler._records_from_table(
            rows,
            {
                "school_name": "大连工业大学",
                "document_type": "postgraduate_admission_list",
                "source_url": "https://yjs.dep.dlpu.edu.cn/list.xlsx",
                "title": "拟录取名单.xlsx",
                "year": 2026,
            },
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "查金诚")
        self.assertEqual(records[0]["college"], "轻工与化学工程学院")
        self.assertEqual(records[0]["major"], "0822Z1")
        self.assertEqual(records[0]["admission_major"], "生物质能源与材料")
        self.assertEqual(records[0]["undergraduate_school"], "大连工业大学")
        self.assertEqual(records[0]["undergraduate_major"], "生物质能源与材料")

    def test_parse_pdf_records_extracts_tyut_recommendation_rows(self):
        text = (
            "太原理工大学接收2026届优秀本科毕业生免试攻读研究生拟录取名单\n"
            "录取学院                      录取专业                研究                      专项   招生    复试\n"
            "姓名     毕业院校                 录取学院名称                 录取专业名称             研究方向名称   学习方式\n"
            "                 代码                        代码                方向码                      计划   类型    成绩\n"
            "成瑞琦 太原理工大学       001 机械工程学院               080200 机械工程         05   矿山机械及其自动化   全日制         硕士生    92\n"
            "叶卓然 太原理工大学       001 机械工程学院               080200 机械工程         01   机械电子工程      全日制         直博生    95\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "太原理工大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://www.gs.tyut.edu.cn/list.pdf",
                    "title": "太原理工大学接收2026届优秀本科毕业生免试攻读研究生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "成瑞琦")
        self.assertEqual(records[0]["undergraduate_school"], "太原理工大学")
        self.assertEqual(records[0]["college"], "001 机械工程学院")
        self.assertEqual(records[0]["major"], "080200")
        self.assertEqual(records[0]["admission_major"], "080200 机械工程")
        self.assertIn("direction 05 矿山机械及其自动化", records[0]["remarks"])
        self.assertIn("全日制", records[0]["remarks"])
        self.assertIn("硕士生", records[0]["remarks"])
        self.assertIn("reexam_score 92", records[0]["remarks"])
        self.assertEqual(records[1]["person_name"], "叶卓然")
        self.assertIn("直博生", records[1]["remarks"])

    def test_parse_pdf_records_extracts_sxu_recommendation_rows(self):
        text = (
            "山西大学2026年推免生（含直博生）拟录取名单\n"
            "序号   姓名    录取层次     录取学院(中心、所)           录取专业代码         录取专业名称     复试成绩\n"
            "1    刘宝祯   直博生      科学技术哲学研究中心                010108    科学技术哲学     91.25\n"
            "10   雷鸣     硕士         哲学学院                   010101    马克思主义哲学    94.67\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "山西大学",
                    "document_type": "incoming_recommendation_admission_list",
                    "source_url": "https://yjszsw.sxu.edu.cn/docs/2025-10/list.pdf",
                    "title": "山西大学2026年推免生（含直博生）拟录取名单",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["person_name"], "刘宝祯")
        self.assertEqual(records[0]["ranking"], "1")
        self.assertEqual(records[0]["college"], "科学技术哲学研究中心")
        self.assertEqual(records[0]["major"], "010108")
        self.assertEqual(records[0]["admission_major"], "010108 科学技术哲学")
        self.assertEqual(records[0]["remarks"], "直博生 retest_score 91.25")
        self.assertEqual(records[1]["person_name"], "雷鸣")
        self.assertEqual(records[1]["college"], "哲学学院")
        self.assertEqual(records[1]["admission_major"], "010101 马克思主义哲学")
        self.assertEqual(records[1]["remarks"], "硕士 retest_score 94.67")

    def test_parse_pdf_records_extracts_lzu_law_recommendation_score_rows(self):
        text = (
            "兰州大学法学院2026年接收推荐免试攻读硕士学位研究生拟录取名单公示（法学）\n"
            "序号   报考专业     报考研究方向       姓名    专业面试成绩   外语口语及听力测试成绩   复试总成绩    备注\n"
            "1     法学       法学理论        舒保轩    93.33       92.00      93.07   拟录取\n"
            "2     法学    民商法学（含知识产权法）   苏紫轩     0.00       0.00       0.00    放弃\n"
            "3     法律（法学）     不区分研究方向       李璐0048    91.33       83.00      89.67   拟录取\n"
            "4     法律（法学）     不区分研究方向       江歆怡    87.33       74.00      84.67\n"
            "5     法律（非法学）     不区分研究方向       卡米拉·阿不都西库尔    84.00       90.50      85.30   拟录取\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "兰州大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://laws.lzu.edu.cn/laws/upload/files/list.pdf",
                    "title": "兰州大学法学院2026年接收推荐免试攻读硕士学位研究生拟录取名单公示",
                    "year": 2026,
                },
            )

        self.assertEqual(len(records), 4)
        self.assertEqual(records[0]["person_name"], "舒保轩")
        self.assertEqual(records[0]["admission_major"], "法学")
        self.assertEqual(records[0]["major"], "法学理论")
        self.assertIn("interview_score 93.33", records[0]["remarks"])
        self.assertIn("oral_listening_score 92.00", records[0]["remarks"])
        self.assertIn("total_score 93.07", records[0]["remarks"])
        self.assertIn("拟录取", records[0]["remarks"])
        self.assertNotIn("苏紫轩", {record["person_name"] for record in records})
        self.assertEqual(records[1]["person_name"], "李璐")
        self.assertEqual(records[2]["person_name"], "江歆怡")
        self.assertEqual(records[2]["admission_major"], "法律（法学）")
        self.assertEqual(records[3]["person_name"], "卡米拉·阿不都西库尔")

    def test_parse_pdf_records_extracts_plain_recommendation_name_list_after_intro(self):
        text = (
            "扬州大学法学院推荐 2025 年优秀应届本科毕业生免试攻读研究生通过答辩公示\n"
            "专家考核小组对申请者的资格、条件、科研成果、竞赛获奖等进行审核鉴定，"
            "现将通过答辩的学生名单公示\n"
            "如下：（以姓氏笔画排序）\n"
            "于乘苏、马天懋、邓雨晨、朱欣然、刘雨、刘弈彤、\n"
            "杜可凡、杜格格、杨刘晔、张如光、林群峰、周璐瑶、赵\n"
            "梦姣、贾鑫榕、顾乐融、康冉、董昊宇\n"
            "公示时间：2024 年 9 月 14 日-2024 年 9 月 17 日\n"
        )

        with patch.object(crawler, "extract_pdf_text", return_value=text):
            records = crawler.parse_pdf_records(
                Path("dummy.pdf"),
                {
                    "school_name": "扬州大学",
                    "document_type": "recommendation_exemption_list",
                    "source_url": "https://fxy.yzu.edu.cn/list.pdf",
                    "title": "扬州大学法学院推荐2025年优秀应届本科毕业生免试攻读研究生通过答辩公示",
                    "year": 2025,
                },
            )

        self.assertEqual(
            [record["person_name"] for record in records],
            [
                "于乘苏",
                "马天懋",
                "邓雨晨",
                "朱欣然",
                "刘雨",
                "刘弈彤",
                "杜可凡",
                "杜格格",
                "杨刘晔",
                "张如光",
                "林群峰",
                "周璐瑶",
                "赵梦姣",
                "贾鑫榕",
                "顾乐融",
                "康冉",
                "董昊宇",
            ],
        )
        self.assertEqual(records[0]["remarks"], "通过答辩")

    def test_records_from_table_skips_rows_where_person_name_is_identifier_only(self):
        records = crawler._records_from_table(
            [
                ["序号", "姓名", "身份证号", "拟录取学院", "拟录取专业", "复试成绩"],
                ["931", "652************52X", "网络空间安全学院", "083900网络空间安全", "86.6"],
            ],
            {
                "school_name": "北京邮电大学",
                "document_type": "recommendation_exemption_list",
                "source_url": "https://yzb.bupt.edu.cn/2026bstm.pdf",
                "title": "2026年北京邮电大学推荐免试攻读博士研究生拟录取名单",
                "year": 2026,
            },
        )

        self.assertEqual(records, [])

    def test_records_from_table_clears_major_code_name_misread_as_person_name(self):
        records = crawler._records_from_table(
            [
                ["考生编号", "姓名", "录取院系", "录取专业", "考试方式", "专项计划", "初试成绩", "复试成绩", "综合成绩", "录取类别", "学习方式"],
                ["105646000001365", "095132|资源利用与植物保护", "全国统考", "无", "291", "84.54", "71.37", "非定向", "全日制"],
            ],
            {
                "school_name": "华南农业大学",
                "document_type": "postgraduate_admission_list",
                "source_url": "https://yzb.scau.edu.cn/list.pdf",
                "title": "华南农业大学2026年硕士研究生拟录取名单",
                "year": 2026,
            },
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["student_id"], "105646000001365")
        self.assertEqual(records[0]["person_name"], "")
        self.assertEqual(records[0]["admission_major"], "095132|资源利用与植物保护")
        self.assertTrue(records[0]["needs_review"])

    def test_records_from_table_aligns_gdufe_wrapped_college_rows(self):
        records = crawler._records_from_table(
            [
                ["综合能", "专业基", "复试"],
                ["序号", "姓名", "考生编号", "学院名称", "专业代码", "专业名称", "初试总分", "总成绩", "考生类别", "学习方式", "拟录取", "备注"],
                ["力考核", "础考核", "总分"],
                ["财政税务学院（税务师学"],
                ["1", "谢雨轩", "105926360700744", "020203", "财政学", "401", "172.4", "84.0", "256.4", "657.4", "一志愿", "全日制", "拟录取1"],
                ["院）"],
            ],
            {
                "school_name": "广东财经大学",
                "document_type": "postgraduate_admission_list",
                "source_url": "https://yzb.gdufe.edu.cn/list.pdf",
                "title": "2026年度硕士拟录取状态信息表-3月21日财政税务学院.pdf",
                "year": 2026,
            },
        )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["person_name"], "谢雨轩")
        self.assertEqual(records[0]["college"], "财政税务学院（税务师学院）")
        self.assertEqual(records[0]["major"], "020203")
        self.assertEqual(records[0]["admission_major"], "财政学")
        self.assertEqual(records[0]["remarks"], "657.4")

    def test_records_from_table_skips_gdufe_not_admitted_rows(self):
        records = crawler._records_from_table(
            [
                ["序号", "姓名", "考生编号", "学院名称", "专业代码", "专业名称", "初试总分", "总成绩", "考生类别", "学习方式", "拟录取", "备注"],
                ["1", "陈彦谊", "105926441904327", "国家安全与发展研究院", "140200", "国家安全学", "355", "177.8", "76.0", "253.8", "608.8", "一志愿", "全日制", "拟录取1"],
                ["2", "叶梓瑶", "105926441904331", "国家安全与发展研究院", "140200", "国家安全学", "333", "119.4", "71.0", "190.4", "523.4", "一志愿", "全日制", "不予录取"],
            ],
            {
                "school_name": "广东财经大学",
                "document_type": "postgraduate_admission_list",
                "source_url": "https://yzb.gdufe.edu.cn/list.pdf",
                "title": "2026年度硕士拟录取状态信息表-3月19日国家安全学.pdf",
                "year": 2026,
            },
        )

        self.assertEqual([record["person_name"] for record in records], ["陈彦谊"])

    def test_records_from_table_skips_short_program_continuation_rows_without_identity(self):
        records = crawler._records_from_table(
            [
                ["考生编号", "姓名", "录取院系", "录取专业", "考试方式", "专项计划", "初试成绩", "复试成绩", "综合成绩", "录取类别", "学习方式"],
                ["人工智能与低空技术", "085401|新一代电子信息技术"],
            ],
            {
                "school_name": "华南农业大学",
                "document_type": "postgraduate_admission_list",
                "source_url": "https://yzb.scau.edu.cn/list.pdf",
                "title": "华南农业大学2026年硕士研究生拟录取名单",
                "year": 2026,
            },
        )

        self.assertEqual(records, [])

    def test_record_dedupe_key_ignores_source_url_for_same_person_program(self):
        base = {
            "school_name": "北京外国语大学",
            "year": 2026,
            "document_type": "incoming_recommendation_admission_list",
            "route": "recommendation_exemption",
            "person_name": "丁上茵",
            "student_id": "",
            "college": "英语学院",
            "admission_major": "英语语言文学",
        }
        first = {**base, "source_url": "https://graduate.bfsu.edu.cn/info/1048/4006.htm"}
        second = {**base, "source_url": "https://graduate.bfsu.edu.cn/info/1074/4016.htm"}

        self.assertEqual(crawler._record_dedupe_key(first), crawler._record_dedupe_key(second))

    def test_clean_records_drops_program_quota_rows_without_person_context(self):
        records = [
            {
                "school_name": "江西师范大学",
                "year": 2026,
                "document_type": "recommendation_exemption_list",
                "route": "recommendation_exemption",
                "person_name": "产品设计",
                "student_id": "",
                "college": "014",
                "major": "美术学院",
                "admission_major": "130504",
                "source_url": "https://yz.jxnu.edu.cn/quota.pdf",
                "title": "江西师范大学2026年推荐免试名额分配表",
            },
            {
                "school_name": "江西师范大学",
                "year": 2026,
                "document_type": "recommendation_exemption_list",
                "route": "recommendation_exemption",
                "person_name": "张三",
                "student_id": "104142026000001",
                "college": "014 美术学院",
                "major": "130504",
                "admission_major": "130504 产品设计",
                "source_url": "https://yz.jxnu.edu.cn/people.pdf",
                "title": "江西师范大学2026年推荐免试拟录取名单",
            },
        ]

        with TemporaryDirectory() as tmpdir:
            clean_path = Path(tmpdir) / "clean.csv"
            summary_path = Path(tmpdir) / "summary.csv"
            summary = crawler._clean_record_rows_to_outputs(
                records,
                clean_path,
                summary_path,
            )
            with clean_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(rows[0]["person_name"], "张三")

    def test_clean_records_preserves_masked_names_without_identifier(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "对外经济贸易大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "王**",
                    "student_id": "",
                    "college": "国际经济贸易学院",
                    "admission_major": "金融",
                    "source_url": "https://yjsy.uibe.edu.cn/list.pdf",
                },
                {
                    "school_name": "对外经济贸易大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "王**",
                    "student_id": "",
                    "college": "国际经济贸易学院",
                    "admission_major": "金融",
                    "source_url": "https://yjsy.uibe.edu.cn/list.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 2)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(len({row["record_id"] for row in output_rows}), 2)

    def test_clean_records_preserves_same_name_program_when_rankings_differ(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            base = {
                "school_name": "大连海事大学",
                "year": 2026,
                "document_type": "recommendation_exemption_list",
                "route": "recommendation_exemption",
                "person_name": "赵悦名",
                "student_id": "",
                "college": "006 航运经济与管理学院",
                "admission_major": "120200 工商管理学",
                "source_url": "https://grs.dlmu.edu.cn/list.pdf",
            }
            records = [
                base | {"ranking": "267", "remarks": "score 90.00"},
                base | {"ranking": "268", "remarks": "score 89.33"},
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 2)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual({row["ranking"] for row in output_rows}, {"267", "268"})
            self.assertEqual(len({row["record_id"] for row in output_rows}), 2)

    def test_clean_records_preserves_masked_name_ending_with_name_character(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "中央戏剧学院",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "*芷名",
                    "student_id": "",
                    "college": "戏剧学系",
                    "admission_major": "戏剧与影视",
                    "source_url": "https://chntheatre.edu.cn/list.pdf",
                }
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "*芷名")

    def test_clean_records_drops_rows_without_person_identity(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "东华大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "",
                    "student_id": "",
                    "college": "001",
                    "admission_major": "080500",
                    "source_url": "https://yjszs.dhu.edu.cn/list.pdf",
                },
                {
                    "school_name": "东华大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "张三",
                    "student_id": "",
                    "college": "服装与艺术设计学院",
                    "admission_major": "设计学",
                    "source_url": "https://yjszs.dhu.edu.cn/list.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_clean_records_keeps_valid_chinese_name_ending_with_ming(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "河北地质大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "王浩名",
                    "student_id": "100776006020004",
                    "college": "006 经济学院",
                    "admission_major": "025100 金融",
                    "ranking": "332",
                    "source_url": "https://www.hgu.edu.cn/list.pdf",
                }
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "王浩名")
            self.assertEqual(output_rows[0]["student_id"], "100776006020004")

    def test_clean_records_drops_navigation_fragments_without_program_context(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "广西民族大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "快速通道",
                    "student_id": "",
                    "source_url": "https://gxmzu.edu.cn/page.htm",
                },
                {
                    "school_name": "广西民族大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "合格",
                    "student_id": "瑚",
                    "source_url": "https://gxmzu.edu.cn/page.htm",
                },
                {
                    "school_name": "广西民族大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "合修",
                    "student_id": "林！",
                    "source_url": "https://gxmzu.edu.cn/page.htm",
                },
                {
                    "school_name": "广西民族大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "【 关闭窗口 】",
                    "student_id": "",
                    "college": "广西民族大学管理学院2026年硕士研究生招生调剂拟录取名单公示",
                    "source_url": "https://gxmzu.edu.cn/page.htm",
                },
                {
                    "school_name": "中国海洋大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "后台管理",
                    "student_id": "",
                    "major": "韩宁夫",
                    "source_url": "https://yz.ouc.edu.cn/?s=107",
                },
                {
                    "school_name": "广西民族大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "张三",
                    "student_id": "",
                    "admission_major": "030400 民族学",
                    "source_url": "https://gxmzu.edu.cn/page.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_clean_records_drops_graduate_school_navigation_labels(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "\u5185\u8499\u53e4\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u4e0b\u8f7d\u4e13\u533a",
                    "student_id": "",
                    "source_url": "https://yjsc.imnu.edu.cn/info/1004/5118.htm",
                },
                {
                    "school_name": "\u5185\u8499\u53e4\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u7855\u58eb\u62db\u751f",
                    "student_id": "",
                    "source_url": "https://yjsc.imnu.edu.cn/info/1004/5118.htm",
                },
                {
                    "school_name": "\u5185\u8499\u53e4\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u79d1\u7814\u5de5\u4f5c",
                    "student_id": "",
                    "source_url": "https://yjsc.imnu.edu.cn/info/1004/5118.htm",
                },
                {
                    "school_name": "\u5185\u8499\u53e4\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u5b66\u4f4d\u6388\u4e88",
                    "student_id": "",
                    "source_url": "https://yjsc.imnu.edu.cn/info/1004/5118.htm",
                },
                {
                    "school_name": "\u5185\u8499\u53e4\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u6559\u52a1\u7ba1\u7406",
                    "student_id": "",
                    "source_url": "https://yjsc.imnu.edu.cn/info/1004/5118.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u5916\u56fd\u8bed\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u6ee8\u6d77\u6821\u533a",
                    "major": "\u90ae\u7f16",
                    "student_id": "",
                    "source_url": "https://grad.tjfsu.edu.cn/info/1075/3519.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u5916\u56fd\u8bed\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u7814\u7a76\u751f\u9662",
                    "major": "\u90ae\u7f16",
                    "student_id": "",
                    "source_url": "https://grad.tjfsu.edu.cn/info/1075/3519.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5b66\u4e1a\u6307\u5bfc",
                    "major": "\u6570\u5b66\u5b66\u9662",
                    "student_id": "",
                    "source_url": "https://oaa.tju.edu.cn/info/1056/8290.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u6559\u5e08\u53d1\u5c55",
                    "major": "\u6570\u5b66\u5b66\u9662",
                    "student_id": "",
                    "source_url": "https://oaa.tju.edu.cn/info/1056/8290.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u529e\u516c\u7f51",
                    "major": "2021\u7ea7\uff082026\u5c4a\uff09\u5929\u6d25\u5927\u5b66\u533b\u5b66\u9662\u4e34\u5e8a\u533b\u5b66",
                    "student_id": "",
                    "source_url": "https://mstu.tju.edu.cn/info/1061/6201.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u56fe\u4e66\u9986",
                    "major": "2021\u7ea7\uff082026\u5c4a\uff09\u5929\u6d25\u5927\u5b66\u533b\u5b66\u9662\u4e34\u5e8a\u533b\u5b66",
                    "student_id": "",
                    "source_url": "https://mstu.tju.edu.cn/info/1061/6201.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5b66\u53f7",
                    "major": "2021\u7ea7\uff082026\u5c4a\uff09\u5929\u6d25\u5927\u5b66\u533b\u5b66\u9662\u4e34\u5e8a\u533b\u5b66",
                    "student_id": "",
                    "source_url": "https://mstu.tju.edu.cn/info/1061/6201.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u6559\u52a1\u5904",
                    "major": "2021\u7ea7\uff082026\u5c4a\uff09\u5929\u6d25\u5927\u5b66\u533b\u5b66\u9662\u4e34\u5e8a\u533b\u5b66",
                    "student_id": "",
                    "source_url": "https://mstu.tju.edu.cn/info/1061/6201.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u8d22\u52a1\u7cfb\u7edf",
                    "major": "2021\u7ea7\uff082026\u5c4a\uff09\u5929\u6d25\u5927\u5b66\u533b\u5b66\u9662\u4e34\u5e8a\u533b\u5b66",
                    "student_id": "",
                    "source_url": "https://mstu.tju.edu.cn/info/1061/6201.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u4f1a\u8ba1",
                    "admission_major": "2",
                    "remarks": "\u652f\u6559\u56e2",
                    "student_id": "",
                    "source_url": "https://ms.tjut.edu.cn/info/1059/5855.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5408\u4f5c\u9662\u6821",
                    "major": "\u5907\u6848\u53f7",
                    "student_id": "",
                    "source_url": "https://ms.tjut.edu.cn/info/1059/5855.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5de5\u5546\u7ba1\u7406\u5b66",
                    "admission_major": "3",
                    "remarks": "\u652f\u6559\u56e2",
                    "student_id": "",
                    "source_url": "https://ms.tjut.edu.cn/info/1059/5855.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u7ba1\u7406\u79d1\u5b66\u4e0e\u5de5\u7a0b",
                    "admission_major": "1",
                    "remarks": "\u652f\u6559\u56e2",
                    "student_id": "",
                    "source_url": "https://ms.tjut.edu.cn/info/1059/5855.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u652f\u6559\u56e2",
                    "major": "\u5929\u6d25\u7406\u5de5\u5927\u5b66\u7ba1\u7406\u5b66\u9662",
                    "student_id": "",
                    "source_url": "https://ms.tjut.edu.cn/info/1059/5855.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u4e3b\u8981",
                    "major": "\u53e6\u8bbe\u590d\u8bd5\u8bb0\u5f55\u5458\u548c\u534f\u8c03\u5458\u5404",
                    "student_id": "",
                    "source_url": "https://cs.tjut.edu.cn/info/1062/2753.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5929\u5f00\u676f",
                    "major": "\u53e6\u8bbe\u590d\u8bd5\u8bb0\u5f55\u5458\u548c\u534f\u8c03\u5458\u5404",
                    "student_id": "",
                    "source_url": "https://cs.tjut.edu.cn/info/1062/2753.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u521b\u65b0\u80fd\u529b",
                    "major": "\u53e6\u8bbe\u590d\u8bd5\u8bb0\u5f55\u5458\u548c\u534f\u8c03\u5458\u5404",
                    "student_id": "",
                    "source_url": "https://cs.tjut.edu.cn/info/1062/2753.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5408\u4f5c\u4ea4\u6d41",
                    "major": "\u79d1\u7814",
                    "student_id": "",
                    "source_url": "http://ylr.tjut.edu.cn/info/1026/2927.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5b66\u751f\u5de5\u4f5c",
                    "major": "\u79d1\u7814",
                    "student_id": "",
                    "source_url": "http://ylr.tjut.edu.cn/info/1026/2927.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u79d1\u7814\u9879\u76ee",
                    "major": "\u79d1\u7814",
                    "student_id": "",
                    "source_url": "http://ylr.tjut.edu.cn/info/1026/2927.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u4e2d\u533b\u836f\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u7855\u58eb",
                    "college": "105701",
                    "admission_major": "\u738b\u67d0\u67d0",
                    "student_id": "",
                    "source_url": "https://yjsy.tjutcm.edu.cn/info/1976/9429.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u4e2d\u533b\u836f\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u76f4\u535a\u751f",
                    "college": "100800",
                    "admission_major": "\u5e03\u67d0\u67d0",
                    "student_id": "",
                    "source_url": "https://yjsy.tjutcm.edu.cn/info/1976/9483.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u533b\u79d1\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u65b9\u5411",
                    "student_id": "\u9662\u7cfb\u6240",
                    "college": "\u521d\u8bd5\u603b \u590d\u8bd5\u6210",
                    "source_url": "https://gs.tmu.edu.cn/_upload/article/files/37/c0/a5cbb16440048fbdd5fdfdd8ef64/fd97bfc7-7f09-4d16-9486-871a9419a00f.pdf",
                },
                {
                    "school_name": "\u5185\u8499\u53e4\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u738b\u4e94",
                    "student_id": "",
                    "admission_major": "\u6559\u80b2\u5b66",
                    "source_url": "https://yjsc.imnu.edu.cn/info/1004/5118.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(output_rows[0]["person_name"], "\u738b\u4e94")

    def test_clean_records_dedupes_org_only_major_against_richer_record(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5468\u96e8\u6615",
                    "admission_major": "120100",
                    "student_id": "",
                    "source_url": "https://ms.tjut.edu.cn/info/1059/5855.htm",
                },
                {
                    "school_name": "\u5929\u6d25\u7406\u5de5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5468\u96e8\u6615",
                    "major": "\u5929\u6d25\u7406\u5de5\u5927\u5b66\u7ba1\u7406\u5b66\u9662",
                    "student_id": "",
                    "source_url": "https://ms.tjut.edu.cn/info/1059/5855.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(output_rows[0]["person_name"], "\u5468\u96e8\u6615")
        self.assertEqual(output_rows[0]["admission_major"], "120100")

    def test_clean_records_drops_lzufe_page_fragments_and_prefers_richer_rows(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "兰州财经大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "王雅萱",
                    "college": "国际经济与贸易学院",
                    "admission_major": "国际贸易学",
                    "source_url": "https://yjsy.lzufe.edu.cn/info/1093/3192.htm",
                },
                {
                    "school_name": "兰州财经大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "王雅萱",
                    "admission_major": "拟录取",
                    "source_url": "https://yjsy.lzufe.edu.cn/info/1093/3192.htm",
                },
                {
                    "school_name": "兰州财经大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "丁瑶",
                    "college": "经济学院",
                    "admission_major": "数字经济",
                    "source_url": "https://yjsy.lzufe.edu.cn/info/1093/3192.htm",
                },
                {
                    "school_name": "兰州财经大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "数字经济",
                    "admission_major": "拟录取",
                    "source_url": "https://yjsy.lzufe.edu.cn/info/1093/3192.htm",
                },
                {
                    "school_name": "兰州财经大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "金融工程",
                    "admission_major": "拟录取",
                    "source_url": "https://yjsy.lzufe.edu.cn/info/1093/3192.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 2)
        self.assertEqual(
            {(row["person_name"], row["admission_major"]) for row in output_rows},
            {("丁瑶", "数字经济"), ("王雅萱", "国际贸易学")},
        )

    def test_clean_records_drops_cdu_article_review_fragments(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "\u6210\u90fd\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u4e01\u742a",
                    "student_id": "110796000009999",
                    "major": "80",
                    "admission_major": "\u620f\u5267\u4e0e\u5f71\u89c6",
                    "remarks": "\u4e8c\u7b49\u6218\u529f",
                    "source_url": "https://yjsc.cdu.edu.cn/info/1028/8189.htm",
                },
                {
                    "school_name": "\u6210\u90fd\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u6768\u6c49\u56fd",
                    "admission_major": "\u4e00\u5ba1",
                    "source_url": "https://yjsc.cdu.edu.cn/info/1028/8189.htm",
                },
                {
                    "school_name": "\u6210\u90fd\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u9ad8\u5c71\u5c71",
                    "admission_major": "\u4e8c\u5ba1",
                    "source_url": "https://yjsc.cdu.edu.cn/info/1028/8189.htm",
                },
                {
                    "school_name": "\u6210\u90fd\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u5434\u542f\u7ea2",
                    "admission_major": "\u4e09\u5ba1",
                    "source_url": "https://yjsc.cdu.edu.cn/info/1028/8189.htm",
                },
                {
                    "school_name": "\u6210\u90fd\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u4e8c\u7b49\u6218\u529f",
                    "needs_review": True,
                    "source_url": "https://yjsc.cdu.edu.cn/info/1028/8189.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(output_rows[0]["person_name"], "\u4e01\u742a")
        self.assertEqual(output_rows[0]["admission_major"], "\u620f\u5267\u4e0e\u5f71\u89c6")

    def test_clean_records_prefers_qhnu_table_rows_over_page_fragments(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "\u9752\u6d77\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5fb7\u62c9\u63aa",
                    "college": "\u8ba1\u7b97\u673a\u5b66\u9662",
                    "remarks": "\u5168\u65e5\u5236",
                    "needs_review": True,
                    "source_url": "https://yjsb.qhnu.edu.cn/info/1059/2842.htm",
                },
                {
                    "school_name": "\u9752\u6d77\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u53b6\u4eae",
                    "college": "\u4f53\u80b2\u5b66\u9662",
                    "remarks": "\u5168\u65e5\u5236",
                    "needs_review": True,
                    "source_url": "https://yjsb.qhnu.edu.cn/info/1059/2842.htm",
                },
                {
                    "school_name": "\u9752\u6d77\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5fb7\u62c9\u63aa",
                    "major": "\u5404\u4f4d \u63a8\u514d\u751f",
                    "source_url": "https://yjsb.qhnu.edu.cn/info/1059/2842.htm",
                },
                {
                    "school_name": "\u9752\u6d77\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u4f53\u80b2\u5b66",
                    "major": "\u5404\u4f4d \u63a8\u514d\u751f",
                    "source_url": "https://yjsb.qhnu.edu.cn/info/1059/2842.htm",
                },
                {
                    "school_name": "\u9752\u6d77\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5b66\u4e60",
                    "major": "\u5404\u4f4d \u63a8\u514d\u751f",
                    "source_url": "https://yjsb.qhnu.edu.cn/info/1059/2842.htm",
                },
                {
                    "school_name": "\u9752\u6d77\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u6570\u5b66",
                    "major": "\u5404\u4f4d \u63a8\u514d\u751f",
                    "source_url": "https://yjsb.qhnu.edu.cn/info/1059/2842.htm",
                },
                {
                    "school_name": "\u9752\u6d77\u5e08\u8303\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u65b9\u5f0f",
                    "major": "\u5404\u4f4d \u63a8\u514d\u751f",
                    "source_url": "https://yjsb.qhnu.edu.cn/info/1059/2842.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 2)
        self.assertEqual(
            {(row["person_name"], row["college"], row["remarks"]) for row in output_rows},
            {("\u5fb7\u62c9\u63aa", "\u8ba1\u7b97\u673a\u5b66\u9662", "\u5168\u65e5\u5236"), ("\u53b6\u4eae", "\u4f53\u80b2\u5b66\u9662", "\u5168\u65e5\u5236")},
        )
        self.assertNotIn("\u5404\u4f4d \u63a8\u514d\u751f", {row["major"] for row in output_rows})

    def test_clean_records_drops_single_character_title_fragments_without_context(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "\u5e7f\u897f\u533b\u79d1\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u4f55\u4f73",
                    "college": "\u516c\u5171\u536b\u751f\u5b66\u9662",
                    "admission_major": "\u6d41\u884c\u75c5\u4e0e\u536b\u751f\u7edf\u8ba1\u5b66",
                    "source_url": "https://yjs.gxmu.edu.cn/zsgz/zsgg/P020250930676216023923.pdf",
                },
                {
                    "school_name": "\u5e7f\u897f\u533b\u79d1\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5e7f",
                    "source_url": "https://yjs.gxmu.edu.cn/zsgz/zsgg/P020250930676216023923.pdf",
                },
                {
                    "school_name": "\u5e7f\u897f\u533b\u79d1\u5927\u5b66",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u897f",
                    "source_url": "https://yjs.gxmu.edu.cn/zsgz/zsgg/P020250930676216023923.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(output_rows[0]["person_name"], "\u4f55\u4f73")

    def test_clean_records_drops_hebust_pdf_score_fragments_without_context(self):
        source_url = "http://yjsxy.web.hebust.edu.cn/docs/2026-04/1cc1c8ffd8134e019d18c7896d752bfb.pdf"
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "\u6cb3\u5317\u79d1\u6280\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u4e01\u51ef",
                    "student_id": "100826100703924",
                    "college": "007",
                    "major": "085500",
                    "admission_major": "\u673a\u68b0\u5de5\u7a0b\u5b66\u9662",
                    "source_url": source_url,
                },
                {
                    "school_name": "\u6cb3\u5317\u79d1\u6280\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "10\u5206",
                    "needs_review": True,
                    "source_url": source_url,
                },
                {
                    "school_name": "\u6cb3\u5317\u79d1\u6280\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u52a0\u8bd51\uff1a80",
                    "needs_review": True,
                    "source_url": source_url,
                },
                {
                    "school_name": "\u6cb3\u5317\u79d1\u6280\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u7ed3\u679c",
                    "needs_review": True,
                    "source_url": source_url,
                },
                {
                    "school_name": "\u6cb3\u5317\u79d1\u6280\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\uff08\u542b\u91cf\u5b50\u6280\u672f\u7b49\uff09",
                    "needs_review": True,
                    "source_url": source_url,
                },
                {
                    "school_name": "\u6cb3\u5317\u79d1\u6280\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u827e\u529b",
                    "needs_review": True,
                    "source_url": source_url,
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(output_rows[0]["person_name"], "\u4e01\u51ef")

    def test_clean_records_drops_qdu_pdf_fragments_without_context(self):
        source_url = "https://grad.qdu.edu.cn/__local/2/98/3D/B248B48C86D3D7841621B0C6706_4D28F42B_A8AE1.pdf"
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "\u9752\u5c9b\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u4e01\u4e16\u5353",
                    "student_id": "110656852003643",
                    "college": "023",
                    "major": "\u673a\u7535\u5de5\u7a0b\u5b66\u9662",
                    "admission_major": "085509",
                    "source_url": source_url,
                },
                {
                    "school_name": "\u9752\u5c9b\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u4fe1\u73ed\uff09",
                    "needs_review": True,
                    "source_url": source_url,
                },
                {
                    "school_name": "\u9752\u5c9b\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u7814\u7a76\u9662\uff09",
                    "needs_review": True,
                    "source_url": source_url,
                },
                {
                    "school_name": "\u9752\u5c9b\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u7f8e\u8fea\u5a1c\u00b7\u4e70",
                    "needs_review": True,
                    "source_url": source_url,
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(output_rows[0]["person_name"], "\u4e01\u4e16\u5353")

    def test_clean_records_drops_teacher_contacts_and_date_index_fragments(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "\u897f\u5b89\u7535\u5b50\u79d1\u6280\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5f20\u8001\u5e08",
                    "major": "\u957f\u4e09\u89d2\u57fa\u5730",
                    "source_url": "https://gr.xidian.edu.cn/info/1074/17124.htm",
                },
                {
                    "school_name": "\u897f\u5b89\u7535\u5b50\u79d1\u6280\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5218\u8001\u5e08",
                    "college": "\u6570\u5b66\u4e0e\u7edf\u8ba1\u5b66\u9662",
                    "needs_review": True,
                    "source_url": "https://gr.xidian.edu.cn/info/1074/17124.htm",
                },
                {
                    "school_name": "\u897f\u5b89\u90ae\u7535\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u62db\u751f\u52a8\u6001",
                    "major": "DATE",
                    "source_url": "https://gr.xiyou.edu.cn/xbwz.htm",
                },
                {
                    "school_name": "\u897f\u5b89\u90ae\u7535\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u674e\u6c38\u98de",
                    "major": "DATE",
                    "source_url": "https://gr.xiyou.edu.cn/xbwz.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

        self.assertEqual(summary["clean_rows"], 0)

    def test_clean_records_drops_notice_closing_text_without_identity(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "安徽财经大学",
                    "year": 2023,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "特此通知",
                    "source_url": "https://yz.aufe.edu.cn/2022/0909/c13923a185857/page.htm",
                }
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

        self.assertEqual(summary["clean_rows"], 0)

    def test_clean_records_drops_footer_link_labels(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "北京语言大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "友情链接",
                    "student_id": "",
                    "major": "版权所有",
                    "source_url": "https://yjsy.blcu.edu.cn/list.htm",
                },
                {
                    "school_name": "北京语言大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "常用链接",
                    "student_id": "",
                    "major": "版权所有",
                    "source_url": "https://yjsy.blcu.edu.cn/list.htm",
                },
                {
                    "school_name": "北京语言大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "王五",
                    "student_id": "",
                    "admission_major": "语言学及应用语言学",
                    "source_url": "https://yjsy.blcu.edu.cn/list.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "王五")

    def test_clean_records_drops_swu_navigation_labels_with_minor_major(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "西南大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "下载中心",
                    "student_id": "",
                    "major": "辅修",
                    "source_url": "https://xwcm.swu.edu.cn/info/1072/3706.htm",
                },
                {
                    "school_name": "西南大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "创制中心",
                    "student_id": "",
                    "major": "辅修",
                    "source_url": "https://xwcm.swu.edu.cn/info/1072/3706.htm",
                },
                {
                    "school_name": "西南大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "语合中心",
                    "student_id": "",
                    "major": "友情链接",
                    "source_url": "https://gjxy.swu.edu.cn/info/1003/5176.htm",
                },
                {
                    "school_name": "西南大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "媒体报道",
                    "student_id": "",
                    "major": "位置导航",
                    "source_url": "https://gjxy.swu.edu.cn/info/1003/5176.htm",
                },
                {
                    "school_name": "西南大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "微信",
                    "student_id": "",
                    "major": "邮编",
                    "source_url": "https://gcjsxy.swu.edu.cn/info/1147/4403.htm",
                },
                {
                    "school_name": "西南大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "党务公开",
                    "student_id": "",
                    "major": "辅修",
                    "source_url": "https://xwcm.swu.edu.cn/info/1072/3706.htm",
                },
                {
                    "school_name": "西南大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "张三",
                    "student_id": "106356310020001",
                    "admission_major": "新闻传播学",
                    "source_url": "https://xwcm.swu.edu.cn/info/1072/3706.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_clean_records_drops_scnu_pagination_label_fragments(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "华南师范大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "最后一页",
                    "student_id": "",
                    "major": "标签",
                    "source_url": "https://yz.scnu.edu.cn/a/20260407/672.html",
                },
                {
                    "school_name": "华南师范大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "张三",
                    "student_id": "105746000000001",
                    "admission_major": "教育学",
                    "source_url": "https://statics.scnu.edu.cn/pics/yz/2026/0407/sample.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_clean_records_drops_theme_toggle_labels(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "郑州轻工业大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "夜晚模式",
                    "student_id": "",
                    "college": "时间",
                    "major": "",
                    "source_url": "https://soft.zzuli.edu.cn/page.htm",
                },
                {
                    "school_name": "郑州轻工业大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "张三",
                    "student_id": "",
                    "college": "软件学院",
                    "major": "软件工程",
                    "source_url": "https://soft.zzuli.edu.cn/page.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_clean_records_drops_contact_person_fragments(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "大连理工大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "姜老师",
                    "student_id": "",
                    "major": "联 系 人",
                    "source_url": "https://info.dlut.edu.cn/page.htm",
                },
                {
                    "school_name": "大连理工大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "联系方式",
                    "student_id": "",
                    "major": "各位考生",
                    "source_url": "https://info.dlut.edu.cn/page.htm",
                },
                {
                    "school_name": "大连理工大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "刘可欣",
                    "student_id": "101416100050005",
                    "college": "硕士",
                    "admission_major": "（主校区）数学科学学院",
                    "source_url": "https://info.dlut.edu.cn/list.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "刘可欣")

    def test_clean_records_drops_header_like_person_names_without_identifier(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "东华大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "申请学院",
                    "student_id": "",
                    "college": "",
                    "admission_major": "接收",
                    "source_url": "https://yjszs.dhu.edu.cn/list.pdf",
                },
                {
                    "school_name": "东华大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "其他",
                    "student_id": "",
                    "college": "",
                    "admission_major": "接收",
                    "source_url": "https://yjszs.dhu.edu.cn/list.pdf",
                },
                {
                    "school_name": "东华大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "在机械",
                    "student_id": "",
                    "college": "",
                    "admission_major": "接收",
                    "source_url": "https://yjszs.dhu.edu.cn/list.pdf",
                },
                {
                    "school_name": "东华大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "李四",
                    "student_id": "",
                    "college": "机械工程学院",
                    "admission_major": "机械工程",
                    "source_url": "https://yjszs.dhu.edu.cn/list.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "李四")

    def test_clean_records_drops_parenthesized_category_labels_without_identifier(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "天津师范大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "（定向、非定向）",
                    "student_id": "",
                    "source_url": "https://zzyxz.tjnu.edu.cn/list.pdf",
                },
                {
                    "school_name": "天津师范大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "胡书妍",
                    "student_id": "",
                    "admission_major": "030300",
                    "source_url": "https://zzyxz.tjnu.edu.cn/list.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "胡书妍")

    def test_clean_record_repairs_identifier_name_shifted_into_major(self):
        record = {
            "school_name": "宁夏医科大学",
            "year": 2009,
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": "100899025000191",
            "student_id": "",
            "college": "",
            "admission_major": "陈仙梅",
            "source_url": "https://yzst.chsi.com.cn/example",
        }

        clean = crawler._clean_record(record)

        self.assertEqual(clean["person_name"], "陈仙梅")
        self.assertEqual(clean["student_id"], "100899025000191")
        self.assertEqual(clean["admission_major"], "")

    def test_clean_record_repairs_college_name_with_identifier_and_name(self):
        record = {
            "school_name": "北京大学",
            "year": 2026,
            "document_type": "recommendation_exemption_list",
            "route": "recommendation_exemption",
            "person_name": "公共卫生学院",
            "student_id": "101020268930288 宋雨擎",
            "college": "直博生",
            "admission_major": "公共卫生",
            "source_url": "https://yjsy.bjmu.edu.cn/list.pdf",
        }

        clean = crawler._clean_record(record)

        self.assertEqual(clean["person_name"], "宋雨擎")
        self.assertEqual(clean["student_id"], "101020268930288")
        self.assertEqual(clean["college"], "公共卫生学院")

    def test_clean_record_repairs_masked_identifier_gender_shifted_before_name(self):
        record = {
            "school_name": "重庆大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": "130******2442 女",
            "student_id": "赵馨雅",
            "college": "035200",
            "major": "社会工作",
            "admission_major": "专硕",
            "remarks": "公共管理学院",
            "source_url": "https://yz.cqu.edu.cn/upload/202605/4b1f5c4d.pdf",
        }

        clean = crawler._clean_record(record)

        self.assertEqual(clean["person_name"], "赵馨雅")
        self.assertEqual(clean["student_id"], "130******2442")
        self.assertIn("女", clean["remarks"])

    def test_clean_record_repairs_score_columns_before_identifier_and_name(self):
        record = {
            "school_name": "四川外国语大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": "88",
            "student_id": "412",
            "college": "",
            "major": "",
            "admission_major": "106506202000616 丁琳",
            "remarks": "",
            "source_url": "https://graduate.sisu.edu.cn/list.pdf",
        }

        clean = crawler._clean_record(record)

        self.assertEqual(clean["person_name"], "丁琳")
        self.assertEqual(clean["student_id"], "106506202000616")
        self.assertIn("score_columns_shifted", clean["quality_flags"])

    def test_clean_record_moves_plan_text_out_of_student_id(self):
        record = {
            "school_name": "北京体育大学",
            "year": 2025,
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": "乃菲莎·库尔班",
            "student_id": "100435202510456 少干计划",
            "admission_major": "体育经济与产业研究",
            "remarks": "286",
            "source_url": "https://zs.bsu.edu.cn/list.pdf",
        }

        clean = crawler._clean_record(record)

        self.assertEqual(clean["student_id"], "100435202510456")
        self.assertEqual(clean["remarks"], "286 少干计划")

    def test_clean_record_clears_header_text_left_in_ranking_field(self):
        record = {
            "school_name": "北京航空航天大学",
            "year": 2026,
            "document_type": "recommendation_exemption_list",
            "route": "recommendation_exemption",
            "person_name": "严子一",
            "student_id": "22377113",
            "college": "经济管理学院",
            "admission_major": "工程管理",
            "ranking": "专业排名",
            "source_url": "https://sem.buaa.edu.cn/list.pdf",
        }

        clean = crawler._clean_record(record)

        self.assertEqual(clean["person_name"], "严子一")
        self.assertEqual(clean["student_id"], "22377113")
        self.assertEqual(clean["ranking"], "")

    def test_clean_records_drops_header_like_names_even_with_non_identifier_text(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "清华大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "建筑学院",
                    "student_id": "清华大学",
                    "college": "",
                    "admission_major": "",
                    "source_url": "https://www.arch.tsinghua.edu.cn/list.pdf",
                },
                {
                    "school_name": "清华大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "姓名",
                    "student_id": "考生编号",
                    "college": "拟录取学院",
                    "admission_major": "拟录取专业",
                    "source_url": "https://www.arch.tsinghua.edu.cn/list.pdf",
                },
                {
                    "school_name": "清华大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "张三",
                    "student_id": "20260001",
                    "college": "建筑学院",
                    "admission_major": "建筑学",
                    "source_url": "https://www.arch.tsinghua.edu.cn/list.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_clean_records_drops_non_person_labels_with_score_like_identifier(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "江西财经大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "金融学(FRM方向)",
                    "student_id": "93.40099",
                    "college": "",
                    "admission_major": "",
                    "source_url": "https://finance.jxufe.edu.cn/news-show-5495.html",
                },
                {
                    "school_name": "中国社会科学院大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "公示期间",
                    "student_id": "",
                    "college": "",
                    "admission_major": "",
                    "source_url": "https://sg.ucass.edu.cn/info/1149/4984.htm",
                },
                {
                    "school_name": "北京中医药大学",
                    "year": 2026,
                    "document_type": "incoming_recommendation_admission_list",
                    "route": "recommendation_exemption",
                    "person_name": "院代码",
                    "student_id": "",
                    "college": "",
                    "admission_major": "类型",
                    "source_url": "https://yanjiusheng.bucm.edu.cn/list.pdf",
                },
                {
                    "school_name": "北京化工大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "类别",
                    "student_id": "学习方式",
                    "college": "",
                    "major": "（500分/300分）",
                    "admission_major": "",
                    "source_url": "https://graduate.buct.edu.cn/list.pdf",
                },
                {
                    "school_name": "首都医科大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "01",
                    "student_id": "",
                    "college": "001",
                    "major": "基础医学院",
                    "admission_major": "071003",
                    "source_url": "https://yjsh.ccmu.edu.cn/list.pdf",
                },
                {
                    "school_name": "中国舰船研究院(上海船舶设备研究所)",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "男",
                    "student_id": "",
                    "college": "",
                    "major": "",
                    "admission_major": "船舶与海洋工程",
                    "source_url": "https://yzst.chsi.com.cn/list.dhtml",
                },
                {
                    "school_name": "安徽工业大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "女",
                    "student_id": "",
                    "college": "能源与环境学院",
                    "major": "",
                    "admission_major": "环境工程",
                    "source_url": "https://jwc.ahut.edu.cn/list.htm",
                },
                {
                    "school_name": "海南大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "修复",
                    "student_id": "",
                    "college": "",
                    "major": "",
                    "admission_major": "",
                    "source_url": "https://gs.hainanu.edu.cn/list.pdf",
                },
                {
                    "school_name": "江西财经大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "张三",
                    "student_id": "20260001",
                    "college": "金融学院",
                    "admission_major": "金融学",
                    "source_url": "https://finance.jxufe.edu.cn/news-show-5495.html",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_clean_records_drops_gender_header_as_person_name(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "安徽工业大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "性别",
                    "student_id": "",
                    "college": "院系名称",
                    "major": "",
                    "admission_major": "",
                    "source_url": "https://nyhjxy.ahut.edu.cn/info/2598/6029.htm",
                },
                {
                    "school_name": "安徽工业大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "张三",
                    "student_id": "20260001",
                    "college": "能源与环境学院",
                    "admission_major": "环境工程",
                    "source_url": "https://nyhjxy.ahut.edu.cn/info/2598/6029.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_clean_records_drops_fragmented_pdf_header_labels(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "北京建筑大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "业代码",
                    "student_id": "号",
                    "college": "一志愿",
                    "admission_major": "绩",
                    "source_url": "https://yjsy.bucea.edu.cn/list.pdf",
                },
                {
                    "school_name": "北京建筑大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "张三",
                    "student_id": "100166085100001",
                    "college": "建筑学院",
                    "admission_major": "建筑",
                    "source_url": "https://yjsy.bucea.edu.cn/list.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_clean_records_drops_score_header_as_person_name(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "云南财经大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "总分",
                    "student_id": "别",
                    "college": "成绩",
                    "admission_major": "绩",
                    "source_url": "https://www.ynufe.edu.cn/list.pdf",
                },
                {
                    "school_name": "云南财经大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "张三",
                    "student_id": "106896123456789",
                    "college": "会计学院",
                    "admission_major": "会计",
                    "source_url": "https://www.ynufe.edu.cn/list.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_clean_records_drops_study_mode_fragment_as_person_name(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "南昌航空大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "非全日",
                    "student_id": "",
                    "college": "立功免初试退役人员",
                    "admission_major": "",
                    "source_url": "https://yjs.nchu.edu.cn/list.pdf",
                },
                {
                    "school_name": "南昌航空大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "张三",
                    "student_id": "104066123456789",
                    "college": "材料科学与工程学院",
                    "admission_major": "材料工程",
                    "source_url": "https://yjs.nchu.edu.cn/list.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_clean_records_drops_exam_subject_header_fragments(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            base = {
                "school_name": "北京联合大学",
                "year": 2026,
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "college": "",
                "admission_major": "",
                "source_url": "https://graduate.buu.edu.cn/list.pdf",
            }
            records = [
                base | {"person_name": "\u4e1a\u52a1\u8bfe\u4e00", "student_id": "\u5916\u8bed"},
                base | {"person_name": "业务一", "student_id": "外语"},
                base | {"person_name": "志愿 思想政", "student_id": "顺序号"},
                base | {"person_name": "管理类综", "student_id": "英语"},
                base | {"person_name": "绩×50％）", "student_id": "合能力"},
                base
                | {
                    "person_name": "李四",
                    "student_id": "100166085100002",
                    "college": "管理学院",
                    "admission_major": "工商管理",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "李四")

    def test_clean_records_drops_score_and_recommendation_form_text_as_person_names(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "\u5317\u4eac\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "78.7",
                    "student_id": "TM202513759",
                    "college": "\u56fd\u9645\u5173\u7cfb\u5b66\u9662",
                    "admission_major": "\u56fd\u9645\u5173\u7cfb",
                    "source_url": "https://admission.pku.edu.cn/list.pdf",
                },
                {
                    "school_name": "\u6cb3\u5357\u5de5\u4e1a\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u6750\u6599\u5ba1\u6838",
                    "student_id": "",
                    "college": "",
                    "major": "\u5b66\u79d1",
                    "admission_major": "",
                    "source_url": "https://cee.haut.edu.cn/list.htm",
                },
                {
                    "school_name": "\u9996\u90fd\u533b\u79d1\u5927\u5b66",
                    "year": 2025,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u4ee5\u4e0b\u8bf7\u63a8\u8350\u4eba\u586b\u5199\uff1a",
                    "student_id": "",
                    "college": "",
                    "major": "",
                    "admission_major": "",
                    "source_url": "https://yjsh.ccmu.edu.cn/recommendation-form.pdf",
                },
                {
                    "school_name": "\u6d4b\u8bd5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u5f20\u4e09",
                    "student_id": "20260001",
                    "college": "\u4fe1\u606f\u5b66\u9662",
                    "admission_major": "\u8ba1\u7b97\u673a\u79d1\u5b66",
                    "source_url": "https://example.edu.cn/list.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "\u5f20\u4e09")

    def test_clean_records_drops_gender_sequence_names_without_identifier(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            base = {
                "school_name": "重庆大学",
                "year": 2026,
                "document_type": "recommendation_exemption_list",
                "route": "recommendation_exemption",
                "source_url": "https://yz.cqu.edu.cn/list.pdf",
            }
            records = [
                base
                | {
                    "person_name": "女 001",
                    "student_id": "",
                    "college": "公共管理学院",
                    "admission_major": "硕士",
                },
                base | {"person_name": "李四", "student_id": "20260001"},
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "李四")

    def test_clean_records_drops_long_exam_subject_text_but_keeps_dotted_names(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "\u6d4b\u8bd5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u601d\u653f\u653f\u6cbb\u7d20\u8d28\u4e0e\u9053\u5fb7\u54c1\u8d28\u7efc\u5408\u9762\u8bd5",
                    "student_id": "\u82f1\u8bed",
                    "college": "",
                    "admission_major": "\u8ba1\u7b97\u673a\u79d1\u5b66",
                    "source_url": "https://example.edu.cn/list.pdf",
                },
                {
                    "school_name": "\u6d4b\u8bd5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "\u9762\u8bd5 (60%)",
                    "student_id": "\u603b\u5206",
                    "college": "",
                    "admission_major": "\u755c\u7267\u5b66",
                    "source_url": "https://example.edu.cn/list.pdf",
                },
                {
                    "school_name": "\u6d4b\u8bd5\u5927\u5b66",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": "\u52aa\u5c14\u6bd4\u4e9a \u00b7\u4e70\u4e70\u63d0",
                    "student_id": "",
                    "college": "\u519c\u5b66\u9662",
                    "admission_major": "\u519c\u5b66",
                    "source_url": "https://example.edu.cn/list.pdf",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "\u52aa\u5c14\u6bd4\u4e9a \u00b7\u4e70\u4e70\u63d0")

    def test_clean_records_drops_rows_missing_core_identity_fields(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            records = [
                {
                    "school_name": "样例大学",
                    "year": "",
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "李四",
                    "student_id": "",
                    "admission_major": "计算机科学与技术",
                    "source_url": "https://example.edu.cn/list.htm",
                },
                {
                    "school_name": "样例大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "王五",
                    "student_id": "",
                    "admission_major": "软件工程",
                    "source_url": "",
                },
                {
                    "school_name": "样例大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": "张三",
                    "student_id": "",
                    "admission_major": "数据科学",
                    "source_url": "https://example.edu.cn/valid.htm",
                },
            ]

            summary = crawler._clean_record_rows_to_outputs(
                records,
                output_dir / "records_clean.csv",
                output_dir / "summary.csv",
            )

            self.assertEqual(summary["clean_rows"], 1)
            with (output_dir / "records_clean.csv").open(
                "r", encoding="utf-8-sig", newline=""
            ) as handle:
                output_rows = list(csv.DictReader(handle))
            self.assertEqual(output_rows[0]["person_name"], "张三")

    def test_crawl_seed_documents_indexes_evidence_and_writes_records(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            seed = {
                "school_name": "东北农业大学",
                "source_type": "incoming_recommendation",
                "start_url": "https://yz.neau.edu.cn/index.htm",
            }
            pages = {
                "https://yz.neau.edu.cn/index.htm": FetchResponse(
                    url="https://yz.neau.edu.cn/index.htm",
                    status_code=200,
                    content_type="text/html; charset=utf-8",
                    content=(
                        '<a href="/result.htm">2026年接收推免生拟录取名单公示</a>'
                    ).encode("utf-8"),
                ),
                "https://yz.neau.edu.cn/result.htm": FetchResponse(
                    url="https://yz.neau.edu.cn/result.htm",
                    status_code=200,
                    content_type="text/html; charset=utf-8",
                    content=HTML_TABLE.encode("utf-8"),
                ),
            }

            def fake_fetch(url, timeout_seconds=20):
                return pages[url]

            summary = crawl_seed_documents(
                [seed],
                raw_dir=output_dir / "raw",
                processed_dir=output_dir / "processed",
                logs_dir=output_dir / "logs",
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                max_pages=5,
            )

            documents = [
                json.loads(line)
                for line in (output_dir / "processed" / "documents.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
            ]
            records = [
                json.loads(line)
                for line in (output_dir / "processed" / "records.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
            ]

        self.assertEqual(summary["documents_written"], 2)
        self.assertEqual(summary["records_written"], 1)
        self.assertEqual(documents[1]["document_type"], "incoming_recommendation_admission_list")
        self.assertEqual(records[0]["person_name"], "张三")

    def test_crawl_seed_documents_does_not_parse_homepage_redirect_with_seed_type(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            seed = {
                "school_name": "首都师范大学",
                "source_type": "postgraduate_admission",
                "start_url": "https://grad.cnu.edu.cn/zs1/sszs/missing.htm",
                "year": "2026",
                "document_type": "postgraduate_admission_list",
            }
            homepage = """
            <html><head><title>首都师范大学</title></head>
              <body>
                <p>地理・教学・AI・育人：</p>
                <p>信息公开</p>
                <p>招生</p>
              </body>
            </html>
            """

            def fake_fetch(url, timeout_seconds=20):
                return FetchResponse(
                    url="https://www.cnu.edu.cn/",
                    status_code=200,
                    content_type="text/html; charset=utf-8",
                    content=homepage.encode("utf-8"),
                )

            summary = crawl_seed_documents(
                [seed],
                raw_dir=output_dir / "raw",
                processed_dir=output_dir / "processed",
                logs_dir=output_dir / "logs",
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
            )

            documents = [
                json.loads(line)
                for line in (output_dir / "processed" / "documents.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
            ]
            records_path = output_dir / "processed" / "records.jsonl"

        self.assertEqual(summary["records_written"], 0)
        self.assertEqual(documents[0]["document_type"], "unknown")
        self.assertEqual(documents[0]["parse_status"], "parsed_no_records")
        self.assertFalse(records_path.exists())

    def test_crawl_seed_documents_carries_parent_year_to_generic_attachment(self):
        workbook_buffer = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["姓名", "录取专业"])
        sheet.append(["王五", "应用经济学"])
        workbook.save(workbook_buffer)

        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            seed = {
                "school_name": "样例大学",
                "source_type": "postgraduate_admission",
                "start_url": "https://yz.example.edu.cn/notice.htm",
            }
            pages = {
                "https://yz.example.edu.cn/notice.htm": FetchResponse(
                    url="https://yz.example.edu.cn/notice.htm",
                    status_code=200,
                    content_type="text/html; charset=utf-8",
                    content=(
                        "<html><head><title>2024年硕士研究生拟录取名单公示</title></head>"
                        '<body><a href="/upload/list.xlsx">附件1</a></body></html>'
                    ).encode("utf-8"),
                ),
                "https://yz.example.edu.cn/upload/list.xlsx": FetchResponse(
                    url="https://yz.example.edu.cn/upload/list.xlsx",
                    status_code=200,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    content=workbook_buffer.getvalue(),
                ),
            }

            def fake_fetch(url, timeout_seconds=20):
                return pages[url]

            crawl_seed_documents(
                [seed],
                raw_dir=output_dir / "raw",
                processed_dir=output_dir / "processed",
                logs_dir=output_dir / "logs",
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                max_pages=3,
            )

            records = [
                json.loads(line)
                for line in (output_dir / "processed" / "records.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
            ]

        self.assertEqual(records[0]["year"], 2024)

    def test_crawl_seed_documents_prefers_parent_year_for_attachment_publish_path(self):
        workbook_buffer = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["姓名", "录取专业"])
        sheet.append(["王五", "应用统计"])
        workbook.save(workbook_buffer)

        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            seed = {
                "school_name": "河北工业大学",
                "source_type": "incoming_recommendation",
                "start_url": "https://yjs.hebut.edu.cn/notice.htm",
                "year": "2026",
                "document_type": "incoming_recommendation_admission_list",
            }
            pages = {
                "https://yjs.hebut.edu.cn/notice.htm": FetchResponse(
                    url="https://yjs.hebut.edu.cn/notice.htm",
                    status_code=200,
                    content_type="text/html; charset=utf-8",
                    content=(
                        "<html><head><title>河北工业大学2026年推免研究生拟录取名单公示</title></head>"
                        '<body><a href="/docs/2025-10/list.xlsx">点击此处查看</a></body></html>'
                    ).encode("utf-8"),
                ),
                "https://yjs.hebut.edu.cn/docs/2025-10/list.xlsx": FetchResponse(
                    url="https://yjs.hebut.edu.cn/docs/2025-10/list.xlsx",
                    status_code=200,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    content=workbook_buffer.getvalue(),
                ),
            }

            def fake_fetch(url, timeout_seconds=20):
                return pages[url]

            crawl_seed_documents(
                [seed],
                raw_dir=output_dir / "raw",
                processed_dir=output_dir / "processed",
                logs_dir=output_dir / "logs",
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                max_pages=3,
            )

            records = [
                json.loads(line)
                for line in (output_dir / "processed" / "records.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
            ]
            documents = [
                json.loads(line)
                for line in (output_dir / "processed" / "documents.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
            ]

        self.assertEqual(records[0]["year"], 2026)
        self.assertEqual(documents[1]["document_type"], "incoming_recommendation_admission_list")

    def test_crawl_seed_documents_prefers_body_year_over_seed_year(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            seed = {
                "school_name": "样例大学",
                "source_type": "recommendation_exemption",
                "start_url": "https://yz.example.edu.cn/list.htm",
                "year": "2026",
                "document_type": "incoming_recommendation_admission_list",
            }
            pages = {
                "https://yz.example.edu.cn/list.htm": FetchResponse(
                    url="https://yz.example.edu.cn/list.htm",
                    status_code=200,
                    content_type="text/html; charset=utf-8",
                    content=(
                        '<html><head><title>样例大学研究生院</title></head><body>'
                        '<a href="/notice.htm">接收推免研究生拟录取名单公示</a>'
                        "</body></html>"
                    ).encode("utf-8"),
                ),
                "https://yz.example.edu.cn/notice.htm": FetchResponse(
                    url="https://yz.example.edu.cn/notice.htm",
                    status_code=200,
                    content_type="text/html; charset=utf-8",
                    content=(
                        "<html><head><title>样例大学研究生院</title></head><body>"
                        "<h2>样例大学2025年接收推免研究生拟录取名单公示</h2>"
                        f"{HTML_TABLE}</body></html>"
                    ).encode("utf-8"),
                ),
            }

            def fake_fetch(url, timeout_seconds=20):
                return pages[url]

            crawl_seed_documents(
                [seed],
                raw_dir=output_dir / "raw",
                processed_dir=output_dir / "processed",
                logs_dir=output_dir / "logs",
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                max_pages=2,
            )

            records = [
                json.loads(line)
                for line in (output_dir / "processed" / "records.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
            ]

        self.assertEqual(records[0]["year"], 2025)

    def test_crawl_seed_documents_parses_excel_download_without_url_suffix(self):
        workbook_buffer = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["姓名", "专业"])
        sheet.append(["Alice", "English"])
        workbook.save(workbook_buffer)

        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            seed = {
                "school_name": "Example University",
                "source_type": "recommendation_exemption",
                "start_url": "https://foreign.example.edu.cn/info/1.htm",
                "year": "2026",
                "document_type": "recommendation_exemption_list",
            }
            pages = {
                "https://foreign.example.edu.cn/info/1.htm": FetchResponse(
                    url="https://foreign.example.edu.cn/info/1.htm",
                    status_code=200,
                    content_type="text/html; charset=utf-8",
                    content=(
                        '<a href="/system/_content/download.jsp?wbfileid=1">'
                        "2026年免试推荐读研拟推荐名单.xlsx</a>"
                    ).encode("utf-8"),
                ),
                "https://foreign.example.edu.cn/system/_content/download.jsp?wbfileid=1": FetchResponse(
                    url="https://foreign.example.edu.cn/system/_content/download.jsp?wbfileid=1",
                    status_code=200,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    content=workbook_buffer.getvalue(),
                ),
            }

            def fake_fetch(url, timeout_seconds=20):
                return pages[url]

            crawl_seed_documents(
                [seed],
                raw_dir=output_dir / "raw",
                processed_dir=output_dir / "processed",
                logs_dir=output_dir / "logs",
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                max_pages=3,
            )

            records = [
                json.loads(line)
                for line in (output_dir / "processed" / "records.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
            ]

        self.assertEqual(records[0]["person_name"], "Alice")
        self.assertEqual(records[0]["document_type"], "recommendation_exemption_list")

    def test_crawl_seed_documents_uses_parent_page_as_attachment_referer(self):
        workbook_buffer = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["姓名", "专业"])
        sheet.append(["Alice", "English"])
        workbook.save(workbook_buffer)

        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            seed = {
                "school_name": "Example University",
                "source_type": "recommendation_exemption",
                "start_url": "https://foreign.example.edu.cn/info/1.htm",
                "year": "2026",
                "document_type": "recommendation_exemption_list",
            }
            attachment_url = "https://foreign.example.edu.cn/system/_content/download.jsp?wbfileid=1"
            pages = {
                "https://foreign.example.edu.cn/info/1.htm": FetchResponse(
                    url="https://foreign.example.edu.cn/info/1.htm",
                    status_code=200,
                    content_type="text/html; charset=utf-8",
                    content=(
                        '<a href="/system/_content/download.jsp?wbfileid=1">'
                        "2026年免试推荐读研拟推荐名单.xlsx</a>"
                    ).encode("utf-8"),
                ),
                attachment_url: FetchResponse(
                    url=attachment_url,
                    status_code=200,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    content=workbook_buffer.getvalue(),
                ),
            }
            calls = []

            def fake_fetch_url(url, timeout_seconds=20, referer=None):
                calls.append((url, referer))
                return pages[url]

            with patch.object(crawler, "fetch_url", side_effect=fake_fetch_url):
                crawl_seed_documents(
                    [seed],
                    raw_dir=output_dir / "raw",
                    processed_dir=output_dir / "processed",
                    logs_dir=output_dir / "logs",
                    sleeper=lambda seconds: None,
                    delay_seconds=0,
                    max_pages=3,
                )

        self.assertEqual(calls[1], (attachment_url, "https://foreign.example.edu.cn/info/1.htm"))

    def test_extract_candidate_links_includes_embedded_pdf_player_sources(self):
        html = """
        <html><body>
          <div
            pdfsrc="/_upload/article/files/list.pdf"
            sudyfile-attr="{'title':'航空学院2026届本科毕业生申请推免资格公示名单.pdf'}">
          </div>
        </body></html>
        """

        links = extract_candidate_links(
            html,
            "https://aero.example.edu.cn/2025/0912/page.htm",
            include_all_attachments=True,
        )

        self.assertEqual(len(links), 1)
        self.assertEqual(links[0].url, "https://aero.example.edu.cn/_upload/article/files/list.pdf")
        self.assertEqual(links[0].link_kind, "attachment")
        self.assertEqual(links[0].document_type, "recommendation_exemption_list")

    def test_extract_candidate_links_skips_next_previous_article_navigation(self):
        html = """
        <html><body>
          <a href="/system/_content/download.jsp?wbfileid=1">拟录取名单.xlsx</a>
          <a href="/info/1091/10414.htm">下一篇：关于公布2026年接收推荐免试攻读硕士研究生复试名单的通知&gt;</a>
        </body></html>
        """

        links = extract_candidate_links(
            html,
            "https://www.muhn.edu.cn/zsw/info/1091/10434.htm",
            include_all_attachments=True,
        )

        self.assertEqual([link.url for link in links], ["https://www.muhn.edu.cn/system/_content/download.jsp?wbfileid=1"])

    def test_extract_candidate_links_includes_vsb_pdf_iframe_sources(self):
        html = """
        <html><body>
          <script>
            showVsbpdfIframe("/__local/1/70/result.pdf","100%","600","0","border",[]);
          </script>
        </body></html>
        """

        links = extract_candidate_links(
            html,
            "https://cmm.ncut.edu.cn/info/1208/7082.htm",
            include_all_attachments=True,
        )

        self.assertEqual(len(links), 1)
        self.assertEqual(links[0].url, "https://cmm.ncut.edu.cn/__local/1/70/result.pdf")
        self.assertEqual(links[0].link_kind, "attachment")

    def test_extract_candidate_links_treats_virtual_attach_query_pdf_as_attachment(self):
        html = """
        <html><body>
          <script>
            showVsbpdfIframe("/virtual_attach_file.vsb?afc=abc&oid=1&tid=2&nid=3&e=.pdf","900","700","0","border",[]);
          </script>
        </body></html>
        """

        links = extract_candidate_links(
            html,
            "https://yjszs.sdufe.edu.cn/info/1034/2889.htm",
            include_all_attachments=True,
        )

        self.assertEqual(len(links), 1)
        self.assertEqual(
            links[0].url,
            "https://yjszs.sdufe.edu.cn/virtual_attach_file.vsb?afc=abc&oid=1&tid=2&nid=3&e=.pdf",
        )
        self.assertEqual(links[0].link_kind, "attachment")

    def test_request_url_percent_encodes_non_ascii_paths(self):
        safe_url = crawler._request_url(
            "https://mse.example.edu/upload/附件：2026届推免名单.pdf?t=1757678886198"
        )

        self.assertNotIn("附件", safe_url)
        self.assertIn("%E9%99%84%E4%BB%B6", safe_url)
        self.assertTrue(safe_url.endswith("?t=1757678886198"))

    def test_decode_response_text_skips_unknown_declared_charset(self):
        response = FetchResponse(
            url="https://yzb.example.edu.cn/page.htm",
            status_code=200,
            content_type="text/html; charset=yaml.null",
            content=b"<html><title>ok</title></html>",
        )

        self.assertEqual(crawler._decode_response_text(response), "<html><title>ok</title></html>")

    def test_response_suffix_prefers_xlsx_magic_over_legacy_excel_mime(self):
        workbook = Workbook()
        buffer = BytesIO()
        workbook.save(buffer)
        response = FetchResponse(
            url="https://yzb.example.edu.cn/list.xlsx",
            status_code=200,
            content_type="application/vnd.ms-excel;charset=UTF-8",
            content=buffer.getvalue(),
        )

        self.assertEqual(crawler._response_suffix(response), ".xlsx")

    def test_classify_document_treats_not_including_recommendation_as_postgraduate_admission(self):
        classification = classify_document(
            "\u5357\u4eac\u5927\u5b66\u73b0\u4ee3\u5de5\u5b66\u96622026\u5e74"
            "\u7855\u58eb\u7814\u7a76\u751f\u62db\u751f\u590d\u8bd5\u6210\u7ee9\u53ca"
            "\u62df\u5f55\u53d6\u540d\u5355\uff08\u4e0d\u542b\u63a8\u8350\u514d\u8bd5"
            "\u7855\u58eb\u751f\uff09.pdf",
            "https://eng.example.edu.cn/list.pdf",
            "",
        )

        self.assertEqual(classification["document_type"], "postgraduate_admission_list")

    def test_classify_document_treats_push_admission_as_recommendation(self):
        classification = classify_document(
            "\u6cb3\u5357\u5927\u5b662026\u5e74\u63a8\u514d\u751f\u653b\u8bfb\u7814\u7a76\u751f"
            "\u62df\u5f55\u53d6\u540d\u5355\u516c\u793a",
            "https://grs.henu.edu.cn/list.pdf",
            "",
        )

        self.assertEqual(classification["document_type"], "incoming_recommendation_admission_list")

    def test_records_from_table_aligns_missing_direction_group_before_exam_number(self):
        document = {
            "school_name": "\u5357\u4eac\u5927\u5b66",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "source_url": "https://eng.example.edu.cn/list.pdf",
            "title": "\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355",
        }
        rows = [
            [
                "\u5e8f\u53f7",
                "\u4e13\u4e1a",
                "\u7814\u7a76\u65b9\u5411\u7ec4",
                "\u8003\u751f\u7f16\u53f7",
                "\u59d3\u540d",
                "\u7b14\u8bd5\u6210\u7ee9",
                "\u9762\u8bd5\u6210\u7ee9",
                "\u590d\u8bd5\u6210\u7ee9",
                "\u603b\u6210\u7ee9",
                "\u62df\u5f55\u53d6\u7ed3\u679c",
                "\u5907\u6ce8",
            ],
            [
                "1",
                "\u5149\u5b66\u5de5\u7a0b",
                "102846213408975",
                "\u738b\u660e\u677e",
                "136",
                "121.00",
                "257.00",
                "662.00",
                "\u62df\u5f55\u53d6",
            ],
        ]

        records = crawler._records_from_table(rows, document)

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["document_type"], "postgraduate_admission_list")
        self.assertEqual(records[0]["person_name"], "\u738b\u660e\u677e")
        self.assertEqual(records[0]["student_id"], "102846213408975")
        self.assertEqual(records[0]["major"], "\u5149\u5b66\u5de5\u7a0b")

    def test_fetch_url_retries_certificate_failures_with_unverified_ssl_context(self):
        class FakeResponse:
            headers = {"Content-Type": "text/html; charset=utf-8"}
            status = 200

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, traceback):
                return False

            def read(self):
                return b"ok"

            def geturl(self):
                return "https://jwc.example.edu.cn/"

        calls = []

        def fake_urlopen(request, timeout=20, context=None):
            calls.append(context)
            if context is None:
                raise URLError(ssl.SSLCertVerificationError("certificate verify failed"))
            return FakeResponse()

        with patch.object(crawler, "urlopen", side_effect=fake_urlopen):
            response = crawler.fetch_url("https://jwc.example.edu.cn/")

        self.assertEqual(response.content, b"ok")
        self.assertEqual(len(calls), 2)
        self.assertIsNone(calls[0])
        self.assertIsInstance(calls[1], ssl.SSLContext)

    def test_fetch_url_falls_back_to_http_when_https_ssl_retry_still_fails(self):
        class FakeResponse:
            headers = {"Content-Type": "text/html; charset=utf-8"}
            status = 200

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, traceback):
                return False

            def read(self):
                return b"http ok"

            def geturl(self):
                return "http://jwc.example.edu.cn/"

        requested_urls = []

        def fake_urlopen(request, timeout=20, context=None):
            requested_urls.append(request.full_url)
            if request.full_url.startswith("https://"):
                raise URLError(ssl.SSLEOFError("EOF occurred in violation of protocol"))
            return FakeResponse()

        with patch.object(crawler, "urlopen", side_effect=fake_urlopen):
            response = crawler.fetch_url("https://jwc.example.edu.cn/")

        self.assertEqual(response.url, "http://jwc.example.edu.cn/")
        self.assertEqual(response.content, b"http ok")
        self.assertEqual(
            requested_urls,
            [
                "https://jwc.example.edu.cn/",
                "https://jwc.example.edu.cn/",
                "http://jwc.example.edu.cn/",
            ],
        )

    def test_fetch_url_uses_curl_when_ssl_and_plain_http_fallbacks_fail(self):
        def fake_urlopen(request, timeout=20, context=None):
            raise URLError(ssl.SSLError("sslv3 alert handshake failure"))

        curl_calls = []

        def fake_run(command, capture_output, timeout, check):
            curl_calls.append(command)
            return types.SimpleNamespace(
                returncode=0,
                stdout=b"<html><title>ok</title></html>",
                stderr=b"",
            )

        with patch.object(crawler, "urlopen", side_effect=fake_urlopen):
            with patch.object(crawler.subprocess, "run", side_effect=fake_run):
                response = crawler.fetch_url("https://yzb.example.edu.cn/page.htm")

        self.assertEqual(response.url, "https://yzb.example.edu.cn/page.htm")
        self.assertEqual(response.content, b"<html><title>ok</title></html>")
        self.assertEqual(response.content_type, "")
        self.assertEqual(len(curl_calls), 1)
        self.assertIn("-L", curl_calls[0])
        self.assertIn("--fail", curl_calls[0])
        self.assertIn("--noproxy", curl_calls[0])
        self.assertIn("*", curl_calls[0])
        if os.name == "nt":
            self.assertIn("--ssl-no-revoke", curl_calls[0])

    def test_fetch_url_uses_curl_when_urlopen_reports_gone_but_curl_succeeds(self):
        def fake_urlopen(request, timeout=20, context=None):
            raise crawler.HTTPError(request.full_url, 410, "Gone", {}, None)

        curl_calls = []

        def fake_run(command, capture_output, timeout, check):
            curl_calls.append(command)
            return types.SimpleNamespace(
                returncode=0,
                stdout=b"<html><title>ok</title></html>",
                stderr=b"",
            )

        with patch.object(crawler, "urlopen", side_effect=fake_urlopen):
            with patch.object(crawler.subprocess, "run", side_effect=fake_run):
                response = crawler.fetch_url("https://gs.example.edu.cn/page.psp")

        self.assertEqual(response.content, b"<html><title>ok</title></html>")
        self.assertEqual(len(curl_calls), 1)
        self.assertIn("--fail", curl_calls[0])
        self.assertIn("--noproxy", curl_calls[0])
        self.assertIn("*", curl_calls[0])
        if os.name == "nt":
            self.assertIn("--ssl-no-revoke", curl_calls[0])

    def test_fetch_url_retries_blocking_status_with_minimal_curl_headers(self):
        def fake_urlopen(request, timeout=20, context=None):
            raise crawler.HTTPError(request.full_url, 403, "Forbidden", {}, None)

        curl_calls = []

        def fake_run(command, capture_output, timeout, check):
            curl_calls.append(command)
            if len(curl_calls) == 1:
                return types.SimpleNamespace(
                    returncode=22,
                    stdout=b"",
                    stderr=b"curl: (22) The requested URL returned error: 403",
                )
            return types.SimpleNamespace(
                returncode=0,
                stdout=b"<html><title>ok</title></html>",
                stderr=b"",
            )

        with patch.object(crawler, "urlopen", side_effect=fake_urlopen):
            with patch.object(crawler.subprocess, "run", side_effect=fake_run):
                response = crawler.fetch_url("https://yjszs.example.edu.cn/info/1013/1734.htm")

        self.assertEqual(response.content, b"<html><title>ok</title></html>")
        self.assertEqual(len(curl_calls), 2)
        self.assertFalse(any(str(part).startswith("Referer:") for part in curl_calls[0]))
        self.assertFalse(any(str(part).startswith("Accept:") for part in curl_calls[1]))
        self.assertFalse(any(str(part).startswith("Accept-Language:") for part in curl_calls[1]))

    def test_write_raw_document_uses_pdf_magic_for_octet_stream_downloads(self):
        response = FetchResponse(
            url="https://jwc.example.edu/system/_content/download.jsp?wbfileid=1",
            status_code=200,
            content_type="application/octet-stream",
            content=b"%PDF-1.7\nbody",
        )

        with TemporaryDirectory() as temp_dir:
            path = crawler._write_raw_document(Path(temp_dir), response)

        self.assertEqual(path.suffix, ".pdf")

    def test_crawl_seed_documents_uses_seed_metadata_for_chsi_like_pages(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            seed = {
                "school_name": "样例大学",
                "source_type": "postgraduate_admission",
                "start_url": "https://yzst.chsi.com.cn/sch/viewBulletin--infoId-10.dhtml",
                "year": "2025",
                "document_type": "postgraduate_admission_list",
                "discovery_title": "样例大学2025年硕士研究生拟录取名单公示",
            }
            html = """
            <html><head><title>样例大学_院校信息_中国研究生招生信息网</title></head>
            <body>
              <table>
                <tr><th>姓名</th><th>拟录取专业</th></tr>
                <tr><td>赵六</td><td>软件工程</td></tr>
              </table>
            </body></html>
            """

            def fake_fetch(url, timeout_seconds=20):
                return FetchResponse(
                    url=url,
                    status_code=200,
                    content_type="text/html; charset=utf-8",
                    content=html.encode("utf-8"),
                )

            crawl_seed_documents(
                [seed],
                raw_dir=output_dir / "raw",
                processed_dir=output_dir / "processed",
                logs_dir=output_dir / "logs",
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
            )

            records = [
                json.loads(line)
                for line in (output_dir / "processed" / "records.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
            ]

        self.assertEqual(records[0]["year"], 2025)
        self.assertEqual(records[0]["document_type"], "postgraduate_admission_list")
        self.assertEqual(records[0]["route"], "postgraduate_exam_or_admission")

    def test_crawl_seed_documents_infers_year_from_chsi_body_text(self):
        with TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            seed = {
                "school_name": "Example Institute",
                "source_type": "postgraduate_admission",
                "start_url": "https://yzst.chsi.com.cn/sch/viewBulletin--infoId-10.dhtml",
                "document_type": "postgraduate_admission_list",
                "discovery_title": "Example Institute admission list",
            }
            html = """
            <html><head><title>Example Institute_院校信息_中国研究生招生信息网</title></head>
            <body>
              <h2>硕士研究生拟录取名单公示</h2>
              <p>Example Institute 2020年硕士研究生拟录取名单予以公示。</p>
              <table>
                <tr><th>姓名</th><th>录取专业</th></tr>
                <tr><td>Alice</td><td>Computer Science</td></tr>
              </table>
            </body></html>
            """

            def fake_fetch(url, timeout_seconds=20):
                return FetchResponse(
                    url=url,
                    status_code=200,
                    content_type="text/html; charset=utf-8",
                    content=html.encode("utf-8"),
                )

            crawl_seed_documents(
                [seed],
                raw_dir=output_dir / "raw",
                processed_dir=output_dir / "processed",
                logs_dir=output_dir / "logs",
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
            )

            records = [
                json.loads(line)
                for line in (output_dir / "processed" / "records.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
            ]

        self.assertEqual(records[0]["year"], 2020)

    def test_write_records_csv_flattens_jsonl_records(self):
        with TemporaryDirectory() as temp_dir:
            jsonl_path = Path(temp_dir) / "records.jsonl"
            csv_path = Path(temp_dir) / "records.csv"
            jsonl_path.write_text(
                json.dumps(
                    {
                        "school_name": "东北农业大学",
                        "year": 2026,
                        "document_type": "incoming_recommendation_admission_list",
                        "person_name": "张三",
                        "undergraduate_school": "东北农业大学",
                        "college": "计算机学院",
                        "admission_major": "计算机科学与技术",
                        "source_url": "https://yz.neau.edu.cn/result.htm",
                        "needs_review": False,
                    },
                    ensure_ascii=False,
                )
                + "\n",
                encoding="utf-8",
            )

            count = write_records_csv(jsonl_path, csv_path)

            with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(count, 1)
        self.assertEqual(rows[0]["school_name"], "东北农业大学")
        self.assertEqual(rows[0]["needs_review"], "false")

    def test_read_official_site_csv_accepts_official_url_rows(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "sites.csv"
            path.write_text(
                "school_name,official_url\n北京大学,https://www.pku.edu.cn/\n",
                encoding="utf-8-sig",
            )

            rows = read_official_site_csv(path)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["school_name"], "北京大学")
        self.assertEqual(rows[0]["official_url"], "https://www.pku.edu.cn/")

    def test_build_discovery_tasks_creates_all_school_query_plan(self):
        schools = [
            {
                "id": "1",
                "name": "北京大学",
                "province": "北京",
                "level": "本科",
                "tags": '["985","保研","研究生院"]',
            },
            {
                "id": "999",
                "name": "某职业技术学院",
                "province": "浙江",
                "level": "专科",
                "tags": "[]",
            },
        ]

        tasks = build_discovery_tasks(schools, years=[2026, 2025])

        self.assertEqual(len(tasks), 12)
        self.assertEqual(tasks[0]["school_name"], "北京大学")
        self.assertEqual(tasks[0]["source_type"], "recommendation_exemption")
        self.assertIn("北京大学", tasks[0]["search_query"])
        self.assertIn("2026", tasks[0]["search_query"])
        self.assertEqual(tasks[0]["eligibility_hint"], "recommended")
        self.assertEqual(tasks[-1]["eligibility_hint"], "low_priority")

    def test_write_discovery_tasks_csv_outputs_usable_queries(self):
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "discovery.csv"
            tasks = build_discovery_tasks(
                [{"id": "1", "name": "北京大学", "province": "北京", "level": "本科", "tags": "[]"}],
                years=[2026],
            )

            count = write_discovery_tasks_csv(tasks, output_path)

            with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(count, 3)
        self.assertEqual(rows[0]["school_name"], "北京大学")
        self.assertIn("site:edu.cn", rows[0]["search_query"])

    def test_parse_bing_search_results_extracts_result_rows(self):
        html = """
        <html><body>
          <li class="b_algo">
            <h2><a href="https://yz.example.edu.cn/info/1.htm">样例大学拟录取名单</a></h2>
            <p>硕士研究生拟录取名单公示</p>
          </li>
          <li class="b_algo">
            <h2><a href="https://news.example.com/2.htm">无关新闻</a></h2>
            <div class="b_caption"><p>校园新闻</p></div>
          </li>
        </body></html>
        """

        rows = parse_bing_search_results(html, "query text", limit=5)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["search_query"], "query text")
        self.assertEqual(rows[0]["result_rank"], 1)
        self.assertEqual(rows[0]["result_url"], "https://yz.example.edu.cn/info/1.htm")
        self.assertIn("拟录取", rows[0]["result_snippet"])

    def test_parse_bing_rss_search_results_extracts_items(self):
        rss = """<?xml version="1.0" encoding="utf-8"?>
        <rss version="2.0">
          <channel>
            <item>
              <title>样例大学拟录取名单</title>
              <link>https://yz.example.edu.cn/info/1.htm</link>
              <description>硕士研究生&lt;b&gt;拟录取&lt;/b&gt;名单公示</description>
            </item>
          </channel>
        </rss>
        """

        rows = parse_bing_rss_search_results(rss, "query text", limit=5)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["provider"], "bing-rss")
        self.assertEqual(rows[0]["result_title"], "样例大学拟录取名单")
        self.assertEqual(rows[0]["result_url"], "https://yz.example.edu.cn/info/1.htm")
        self.assertEqual(rows[0]["result_snippet"], "硕士研究生 拟录取 名单公示")

    def test_parse_duckduckgo_search_results_extracts_items_and_unwraps_links(self):
        html = """
        <html><body>
          <div class="result">
            <a class="result__a" href="/l/?uddg=https%3A%2F%2Fyz.example.edu.cn%2Finfo%2F1.htm">样例大学拟录取名单</a>
            <a class="result__snippet">硕士研究生拟录取名单公示</a>
          </div>
        </body></html>
        """

        rows = parse_duckduckgo_search_results(html, "query text", limit=5)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["provider"], "duckduckgo-html")
        self.assertEqual(rows[0]["result_url"], "https://yz.example.edu.cn/info/1.htm")
        self.assertIn("拟录取", rows[0]["result_snippet"])

    def test_collect_search_results_writes_chunked_result_csv(self):
        tasks = [
            {"search_query": "query one", "school_name": "A"},
            {"search_query": "query two", "school_name": "B"},
        ]
        html = """
        <html><body>
          <li class="b_algo">
            <h2><a href="https://yz.example.edu.cn/info/1.htm">样例大学拟录取名单</a></h2>
            <p>硕士研究生拟录取名单公示</p>
          </li>
        </body></html>
        """
        calls = []

        def fake_fetch(url, timeout_seconds=20):
            calls.append(url)
            return FetchResponse(
                url=url,
                status_code=200,
                content_type="text/html; charset=utf-8",
                content=html.encode("utf-8"),
            )

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "search_results.csv"
            summary = collect_search_results(
                tasks,
                output_path=output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
                start_index=1,
                max_tasks=1,
            )
            with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(summary["selected_tasks"], 1)
        self.assertEqual(summary["results_written"], 1)
        self.assertEqual(len(calls), 1)
        self.assertEqual(rows[0]["search_query"], "query two")

    def test_bing_rss_search_url_uses_chinese_market_without_english_filter(self):
        url = crawler._build_bing_rss_search_url("北京大学 2026 推免名单", count=5)
        params = parse_qs(urlparse(url).query)

        self.assertEqual(params["format"], ["rss"])
        self.assertEqual(params["mkt"], ["zh-CN"])
        self.assertEqual(params["setlang"], ["zh-CN"])
        self.assertNotIn("ensearch", params)

    def test_select_seed_rows_from_search_results_keeps_official_relevant_urls(self):
        tasks = build_discovery_tasks(
            [{"id": "1", "name": "北京大学", "province": "北京", "level": "本科", "tags": "[]"}],
            years=[2026],
        )
        search_results = [
            {
                "search_query": tasks[0]["search_query"],
                "result_title": "无关新闻",
                "result_url": "https://example.com/news",
                "result_snippet": "校园新闻",
            },
            {
                "search_query": tasks[0]["search_query"],
                "result_title": "北京大学2026届推荐免试攻读研究生名单公示",
                "result_url": "https://jwc.pku.edu.cn/info/1001/1.htm",
                "result_snippet": "推荐免试 名单 公示",
            },
            {
                "search_query": tasks[0]["search_query"],
                "result_title": "北京大学2026届推荐免试攻读研究生名单公示",
                "result_url": "https://jwc.pku.edu.cn/info/1001/1.htm",
                "result_snippet": "重复结果",
            },
        ]

        seeds = select_seed_rows_from_search_results(tasks, search_results)

        self.assertEqual(len(seeds), 1)
        self.assertEqual(seeds[0]["school_name"], "北京大学")
        self.assertEqual(seeds[0]["source_type"], "recommendation_exemption")
        self.assertEqual(seeds[0]["start_url"], "https://jwc.pku.edu.cn/info/1001/1.htm")
        self.assertEqual(seeds[0]["discovery_rank"], 2)

    def test_parse_official_site_seed_rows_keeps_recommendation_links(self):
        parse_rows = getattr(crawler, "parse_official_site_seed_rows", None)
        self.assertIsNotNone(parse_rows)
        html = """
        <html><body>
          <a href="/info/1001/1.htm">2026届推荐免试攻读研究生名单公示</a>
          <a href="https://news.example.com/irrelevant.htm">校园新闻</a>
          <a href="https://other.example.edu.cn/info/2.htm">2026届推荐免试名单公示</a>
        </body></html>
        """

        seeds = parse_rows(
            html,
            "https://jwc.example.edu.cn/tzgg/index.htm",
            school_name="Example University",
        )

        self.assertEqual(len(seeds), 1)
        self.assertEqual(seeds[0]["school_name"], "Example University")
        self.assertEqual(seeds[0]["source_type"], "recommendation_exemption")
        self.assertEqual(seeds[0]["document_type"], "recommendation_exemption_list")
        self.assertEqual(seeds[0]["start_url"], "https://jwc.example.edu.cn/info/1001/1.htm")
        self.assertEqual(seeds[0]["year"], 2026)

    def test_collect_official_site_seeds_writes_seed_csv(self):
        collect_official = getattr(crawler, "collect_official_site_seeds", None)
        self.assertIsNotNone(collect_official)
        html = """
        <html><body>
          <a href="/info/1001/1.htm">2026届推荐免试攻读研究生名单公示</a>
        </body></html>
        """

        def fake_fetch(url, timeout_seconds=20):
            return FetchResponse(
                url=url,
                status_code=200,
                content_type="text/html; charset=utf-8",
                content=html.encode("utf-8"),
            )

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "official_seeds.csv"
            summary = collect_official(
                [
                    {
                        "school_name": "Example University",
                        "start_url": "https://jwc.example.edu.cn/tzgg/index.htm",
                    }
                ],
                output_path,
                fetcher=fake_fetch,
                sleeper=lambda seconds: None,
                delay_seconds=0,
            )
            with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(summary["selected_sites"], 1)
        self.assertEqual(summary["seeds_written"], 1)
        self.assertEqual(summary["failed"], 0)
        self.assertEqual(rows[0]["school_name"], "Example University")
        self.assertEqual(rows[0]["start_url"], "https://jwc.example.edu.cn/info/1001/1.htm")

    def test_write_seed_rows_csv_outputs_seed_file_for_crawler(self):
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "seeds.csv"
            seeds = [
                {
                    "school_name": "北京大学",
                    "source_type": "recommendation_exemption",
                    "start_url": "https://jwc.pku.edu.cn/info/1001/1.htm",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "discovery_query": "query",
                    "discovery_title": "title",
                    "discovery_rank": 1,
                }
            ]

            count = write_seed_rows_csv(seeds, output_path)

            with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(count, 1)
        self.assertEqual(rows[0]["school_name"], "北京大学")
        self.assertEqual(rows[0]["start_url"], "https://jwc.pku.edu.cn/info/1001/1.htm")

    def test_read_seed_csv_preserves_optional_metadata_columns(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "seeds.csv"
            path.write_text(
                "school_name,source_type,start_url,year,document_type,discovery_title\n"
                "样例大学,postgraduate_admission,https://yz.example.edu.cn/a.htm,2025,postgraduate_admission_list,样例大学拟录取名单\n",
                encoding="utf-8-sig",
            )

            rows = read_seed_csv(path)

        self.assertEqual(rows[0]["year"], "2025")
        self.assertEqual(rows[0]["document_type"], "postgraduate_admission_list")
        self.assertEqual(rows[0]["discovery_title"], "样例大学拟录取名单")

    def test_slice_seed_rows_supports_chunked_crawls(self):
        seeds = [
            {"school_name": "A", "start_url": "https://a.edu.cn/1"},
            {"school_name": "B", "start_url": "https://b.edu.cn/1"},
            {"school_name": "C", "start_url": "https://c.edu.cn/1"},
        ]

        chunk = slice_seed_rows(seeds, start_index=1, max_seeds=1)

        self.assertEqual(chunk, [{"school_name": "B", "start_url": "https://b.edu.cn/1"}])

    def test_clean_records_to_outputs_deduplicates_and_summarizes(self):
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "records.jsonl"
            clean_csv = Path(temp_dir) / "clean.csv"
            summary_csv = Path(temp_dir) / "summary.csv"
            duplicate = {
                "school_name": "河南师范大学",
                "year": 2024,
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "person_name": "张三",
                "student_id": "1047612345",
                "undergraduate_school": "河南师范大学",
                "college": "001",
                "admission_major": "045104",
                "source_url": "https://www.htu.edu.cn/result.xlsx",
                "title": "拟录取名单",
                "needs_review": False,
            }
            review = duplicate | {
                "person_name": "",
                "student_id": "",
                "undergraduate_school": "",
                "admission_major": "",
                "source_url": "https://www.htu.edu.cn/review.xlsx",
                "needs_review": True,
            }
            input_path.write_text(
                "\n".join(
                    [
                        json.dumps(duplicate, ensure_ascii=False),
                        json.dumps(duplicate, ensure_ascii=False),
                        json.dumps(review, ensure_ascii=False),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            summary = clean_records_to_outputs(input_path, clean_csv, summary_csv)

            with clean_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                clean_rows = list(csv.DictReader(handle))
            with summary_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                summary_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["input_rows"], 3)
        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(clean_rows[0]["person_name_masked"], "张*")
        self.assertEqual(clean_rows[0]["student_id_masked"], "1047******")
        self.assertEqual(summary_rows[0]["record_count"], "1")
        self.assertEqual(summary_rows[0]["needs_review_count"], "0")

    def test_clean_records_repairs_numeric_name_shifted_from_sequence_column(self):
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "records.jsonl"
            clean_csv = Path(temp_dir) / "clean.csv"
            summary_csv = Path(temp_dir) / "summary.csv"
            input_path.write_text(
                json.dumps(
                    {
                        "school_name": "上海大学",
                        "year": 2026,
                        "document_type": "recommendation_exemption_list",
                        "route": "recommendation_exemption",
                        "person_name": "1",
                        "student_id": "吕烨佳",
                        "source_url": "https://ece.shu.edu.cn/info/1303/82114.htm",
                        "title": "推免名单公示",
                    },
                    ensure_ascii=False,
                )
                + "\n",
                encoding="utf-8",
            )

            clean_records_to_outputs(input_path, clean_csv, summary_csv)
            with clean_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                clean_rows = list(csv.DictReader(handle))

        self.assertEqual(clean_rows[0]["person_name"], "吕烨佳")
        self.assertEqual(clean_rows[0]["student_id"], "")
        self.assertEqual(clean_rows[0]["ranking"], "1")

    def test_build_summary_rows_treats_false_strings_as_not_needing_review(self):
        rows = [
            {
                "school_name": "测试大学",
                "year": "2025",
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "person_name": "张三",
                "student_id": "123",
                "needs_review": "False",
                "source_url": "https://example.edu.cn/list.pdf",
            },
            {
                "school_name": "测试大学",
                "year": "2025",
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "person_name": "李四",
                "student_id": "456",
                "needs_review": "true",
                "source_url": "https://example.edu.cn/list.pdf",
            },
        ]

        summary_rows = crawler._build_summary_rows(rows)

        self.assertEqual(summary_rows[0]["record_count"], 2)
        self.assertEqual(summary_rows[0]["needs_review_count"], 1)

    def test_clean_records_clears_numeric_person_name_when_identifier_is_present(self):
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "records.jsonl"
            clean_csv = Path(temp_dir) / "clean.csv"
            summary_csv = Path(temp_dir) / "summary.csv"
            input_path.write_text(
                json.dumps(
                    {
                        "school_name": "东北大学",
                        "year": 2026,
                        "document_type": "postgraduate_admission_list",
                        "route": "postgraduate_exam_or_admission",
                        "person_name": "003",
                        "student_id": "101456000002723",
                        "college": "085802",
                        "admission_major": "动力工程",
                        "source_url": "https://yz.neu.edu.cn/list.pdf",
                        "title": "东北大学2026年硕士研究生拟录取名单",
                    },
                    ensure_ascii=False,
                )
                + "\n",
                encoding="utf-8",
            )

            clean_records_to_outputs(input_path, clean_csv, summary_csv)
            with clean_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                clean_rows = list(csv.DictReader(handle))

        self.assertEqual(clean_rows[0]["person_name"], "")
        self.assertEqual(clean_rows[0]["student_id"], "101456000002723")
        self.assertIn("missing_person_name", clean_rows[0]["quality_flags"])

    def test_clean_records_drops_admission_notice_section_labels(self):
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "records.jsonl"
            clean_csv = Path(temp_dir) / "clean.csv"
            summary_csv = Path(temp_dir) / "summary.csv"
            base = {
                "school_name": "中国石油大学（华东）",
                "year": 2026,
                "document_type": "incoming_recommendation_admission_list",
                "route": "recommendation_exemption",
                "source_url": "https://yz.upc.edu.cn/notice.htm",
                "title": "接收推荐免试研究生通知",
            }
            input_path.write_text(
                "\n".join(
                    [
                        json.dumps(
                            base | {"person_name": "其他事项", "admission_major": "招收"},
                            ensure_ascii=False,
                        ),
                        json.dumps(
                            base | {"person_name": "奖学金", "admission_major": "招收"},
                            ensure_ascii=False,
                        ),
                        json.dumps(
                            base | {"person_name": "张三", "student_id": "20260001"},
                            ensure_ascii=False,
                        ),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            summary = crawler._clean_record_rows_to_outputs(
                list(crawler._iter_jsonl(input_path)),
                clean_csv,
                summary_csv,
            )
            with clean_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                clean_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(clean_rows[0]["person_name"], "张三")

    def test_clean_records_drops_numeric_name_with_score_like_identifier(self):
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "records.jsonl"
            clean_csv = Path(temp_dir) / "clean.csv"
            summary_csv = Path(temp_dir) / "summary.csv"
            base = {
                "school_name": "山东建筑大学",
                "year": 2026,
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "source_url": "https://www.sdjzu.edu.cn/list.pdf",
                "title": "硕士研究生拟录取名单",
            }
            input_path.write_text(
                "\n".join(
                    [
                        json.dumps(
                            base | {"person_name": "249", "student_id": "50.00"},
                            ensure_ascii=False,
                        ),
                        json.dumps(
                            base | {"person_name": "张三", "student_id": "100796000001373"},
                            ensure_ascii=False,
                        ),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            summary = crawler._clean_record_rows_to_outputs(
                list(crawler._iter_jsonl(input_path)),
                clean_csv,
                summary_csv,
            )
            with clean_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                clean_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(clean_rows[0]["person_name"], "张三")

    def test_clean_records_drops_numeric_name_with_college_shifted_identifier(self):
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "records.jsonl"
            clean_csv = Path(temp_dir) / "clean.csv"
            summary_csv = Path(temp_dir) / "summary.csv"
            base = {
                "school_name": "\u5b89\u5fbd\u5efa\u7b51\u5927\u5b66",
                "year": 2026,
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "source_url": "https://www.ahjzu.edu.cn/list.pdf",
                "title": "\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355",
            }
            input_path.write_text(
                "\n".join(
                    [
                        json.dumps(
                            base
                            | {
                                "person_name": "080500",
                                "student_id": "\u6750\u6599\u4e0e\u5316\u5b66\u5de5\u7a0b\u5b66\u9662",
                                "college": "\u6750\u6599\u4e0e\u5316\u5b66\u5de5\u7a0b\u5b66\u9662",
                                "major": "\u6750\u6599\u79d1\u5b66\u4e0e\u5de5\u7a0b",
                            },
                            ensure_ascii=False,
                        ),
                        json.dumps(
                            base
                            | {
                                "person_name": "\u5f20\u4e09",
                                "student_id": "108786340100596",
                            },
                            ensure_ascii=False,
                        ),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            summary = crawler._clean_record_rows_to_outputs(
                list(crawler._iter_jsonl(input_path)),
                clean_csv,
                summary_csv,
            )
            with clean_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                clean_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(clean_rows[0]["person_name"], "\u5f20\u4e09")

    def test_clean_records_drops_fragmented_college_suffix_rows(self):
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "records.jsonl"
            clean_csv = Path(temp_dir) / "clean.csv"
            summary_csv = Path(temp_dir) / "summary.csv"
            base = {
                "school_name": "\u5b89\u5fbd\u5efa\u7b51\u5927\u5b66",
                "year": 2026,
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "source_url": "https://www.ahjzu.edu.cn/list.pdf",
                "title": "\u7855\u58eb\u7814\u7a76\u751f\u62df\u5f55\u53d6\u540d\u5355",
            }
            input_path.write_text(
                "\n".join(
                    [
                        json.dumps(
                            base | {"person_name": "\u5de5\u7a0b", "student_id": "\u9662"},
                            ensure_ascii=False,
                        ),
                        json.dumps(
                            base
                            | {
                                "person_name": "\u5de5\u7a0b",
                                "student_id": "\u606f\u5de5\u7a0b\u5927\u5b66\uff09",
                            },
                            ensure_ascii=False,
                        ),
                        json.dumps(
                            base
                            | {
                                "person_name": "\u674e\u56db",
                                "student_id": "108786340100596",
                            },
                            ensure_ascii=False,
                        ),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            summary = crawler._clean_record_rows_to_outputs(
                list(crawler._iter_jsonl(input_path)),
                clean_csv,
                summary_csv,
            )
            with clean_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                clean_rows = list(csv.DictReader(handle))

        self.assertEqual(summary["clean_rows"], 1)
        self.assertEqual(clean_rows[0]["person_name"], "\u674e\u56db")

    def test_clean_record_repairs_direction_prefix_attached_to_person_name(self):
        record = {
            "school_name": "山东建筑大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": "01【机械工程】 刘如洋",
            "student_id": "104306371402773",
            "source_url": "https://www.sdjzu.edu.cn/list.pdf",
        }

        clean = crawler._clean_record(record)

        self.assertEqual(clean["person_name"], "刘如洋")
        self.assertEqual(clean["student_id"], "104306371402773")
        self.assertEqual(clean["admission_major"], "01【机械工程】")

    def test_clean_record_repairs_unclosed_direction_prefix_attached_to_person_name(self):
        record = {
            "school_name": "山东建筑大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": "00【不区分研究方张迪",
            "student_id": "106116001500061",
            "source_url": "https://www.sdjzu.edu.cn/list.pdf",
        }

        clean = crawler._clean_record(record)

        self.assertEqual(clean["person_name"], "张迪")
        self.assertEqual(clean["student_id"], "106116001500061")
        self.assertEqual(clean["admission_major"], "00【不区分研究方")

    def test_clean_record_repairs_direction_label_with_name_in_identifier_field(self):
        record = {
            "school_name": "山东建筑大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": "01【中国画】",
            "student_id": "刘润琪",
            "source_url": "https://www.sdjzu.edu.cn/list.pdf",
        }

        clean = crawler._clean_record(record)

        self.assertEqual(clean["person_name"], "刘润琪")
        self.assertEqual(clean["student_id"], "")
        self.assertEqual(clean["admission_major"], "01【中国画】")

    def test_clean_records_deduplicates_same_person_with_better_identifier(self):
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "records.jsonl"
            clean_csv = Path(temp_dir) / "clean.csv"
            summary_csv = Path(temp_dir) / "summary.csv"
            base = {
                "school_name": "上海大学",
                "year": 2026,
                "document_type": "recommendation_exemption_list",
                "route": "recommendation_exemption",
                "person_name": "吕烨佳",
                "source_url": "https://ece.shu.edu.cn/info/1303/82114.htm",
                "title": "推免名单公示",
            }
            input_path.write_text(
                "\n".join(
                    [
                        json.dumps(base | {"student_id": ""}, ensure_ascii=False),
                        json.dumps(base | {"student_id": "22122594"}, ensure_ascii=False),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            clean_records_to_outputs(input_path, clean_csv, summary_csv)
            with clean_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                clean_rows = list(csv.DictReader(handle))

        self.assertEqual(len(clean_rows), 1)
        self.assertEqual(clean_rows[0]["student_id"], "22122594")

    def test_clean_records_keeps_masked_same_name_when_partial_identifier_differs(self):
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "records.jsonl"
            clean_csv = Path(temp_dir) / "clean.csv"
            summary_csv = Path(temp_dir) / "summary.csv"
            base = {
                "school_name": "\u590d\u65e6\u5927\u5b66",
                "year": 2026,
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "person_name": "\u738b*",
                "college": "\u7ba1\u7406\u5b66\u9662",
                "source_url": "https://gsao.fudan.edu.cn/list.pdf",
                "title": "2026 admission list",
            }
            input_path.write_text(
                "\n".join(
                    [
                        json.dumps(base | {"student_id": "00001"}, ensure_ascii=False),
                        json.dumps(base | {"student_id": "00002"}, ensure_ascii=False),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            clean_records_to_outputs(input_path, clean_csv, summary_csv)
            with clean_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                clean_rows = list(csv.DictReader(handle))

        self.assertEqual(len(clean_rows), 2)
        self.assertEqual({row["student_id"] for row in clean_rows}, {"00001", "00002"})

    def test_merge_records_jsonl_to_outputs_combines_batches_and_keeps_best_duplicate(self):
        merge_records = getattr(crawler, "merge_records_jsonl_to_outputs", None)
        self.assertIsNotNone(merge_records)

        with TemporaryDirectory() as temp_dir:
            batch_one = Path(temp_dir) / "batch-one.jsonl"
            batch_two = Path(temp_dir) / "batch-two.jsonl"
            clean_csv = Path(temp_dir) / "master_clean.csv"
            summary_csv = Path(temp_dir) / "master_summary.csv"
            low_quality_duplicate = {
                "school_name": "Example University",
                "year": 2026,
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "person_name": "Alice",
                "student_id": "20260001",
                "undergraduate_school": "",
                "admission_major": "",
                "source_url": "https://example.edu/admission-2026.xlsx",
                "title": "2026 admission list",
                "needs_review": True,
            }
            high_quality_duplicate = low_quality_duplicate | {
                "undergraduate_school": "Example College",
                "admission_major": "Computer Science",
                "needs_review": False,
            }
            unique_record = {
                "school_name": "Example University",
                "year": 2026,
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "person_name": "Bob",
                "student_id": "20260002",
                "undergraduate_school": "Another College",
                "admission_major": "Mathematics",
                "source_url": "https://example.edu/admission-2026-extra.xlsx",
                "title": "2026 admission list extra",
                "needs_review": False,
            }
            batch_one.write_text(
                "\n".join([json.dumps(low_quality_duplicate), json.dumps(unique_record)]) + "\n",
                encoding="utf-8",
            )
            batch_two.write_text(json.dumps(high_quality_duplicate) + "\n", encoding="utf-8")

            summary = merge_records([batch_one, batch_two], clean_csv, summary_csv)

            with clean_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                clean_rows = list(csv.DictReader(handle))
            with summary_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                summary_rows = list(csv.DictReader(handle))

        alice = next(row for row in clean_rows if row["person_name"] == "Alice")
        self.assertEqual(summary["input_files"], 2)
        self.assertEqual(summary["input_rows"], 3)
        self.assertEqual(summary["clean_rows"], 2)
        self.assertEqual(alice["undergraduate_school"], "Example College")
        self.assertEqual(alice["needs_review"], "false")
        self.assertEqual(summary_rows[0]["record_count"], "2")
        self.assertEqual(summary_rows[0]["source_document_count"], "2")

    def test_export_public_records_csv_removes_direct_identifiers(self):
        export_public = getattr(crawler, "export_public_records_csv", None)
        self.assertIsNotNone(export_public)

        with TemporaryDirectory() as temp_dir:
            clean_csv = Path(temp_dir) / "records_clean.csv"
            public_csv = Path(temp_dir) / "records_public.csv"
            with clean_csv.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=crawler.CLEAN_RECORD_CSV_FIELDS)
                writer.writeheader()
                writer.writerow(
                    {
                        "record_id": "private-record-id",
                        "school_name": "Example University",
                        "year": "2026",
                        "document_type": "postgraduate_admission_list",
                        "route": "postgraduate_exam_or_admission",
                        "person_name": "Alice",
                        "person_name_masked": "A****",
                        "student_id": "20260001",
                        "student_id_masked": "2026****",
                        "undergraduate_school": "Example College",
                        "undergraduate_major": "",
                        "college": "",
                        "major": "",
                        "admission_major": "Computer Science",
                        "ranking": "",
                        "remarks": "",
                        "source_url": "https://example.edu/list.xlsx",
                        "title": "2026 admission list",
                        "needs_review": "false",
                        "quality_score": "93",
                        "quality_flags": "missing_student_id",
                    }
                )

            summary = export_public(clean_csv, public_csv)

            with public_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                public_rows = list(reader)
                public_fields = reader.fieldnames

        self.assertEqual(summary["input_rows"], 1)
        self.assertEqual(summary["public_rows"], 1)
        self.assertNotIn("person_name", public_fields)
        self.assertNotIn("student_id", public_fields)
        self.assertIn("person_name_masked", public_fields)
        self.assertIn("student_id_masked", public_fields)
        self.assertNotEqual(public_rows[0]["public_record_id"], "private-record-id")
        self.assertEqual(public_rows[0]["person_name_masked"], "A****")
        self.assertEqual(public_rows[0]["source_url"], "https://example.edu/list.xlsx")

    def test_request_headers_use_browser_like_identity_for_chsi_sources(self):
        headers = crawler._request_headers()

        self.assertIn("Mozilla/5.0", headers["User-Agent"])
        self.assertNotIn("compatible", headers["User-Agent"].lower())
        self.assertIn("Accept-Language", headers)


if __name__ == "__main__":
    unittest.main()
