from __future__ import annotations

import csv
import importlib.util
from pathlib import Path
import unittest

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
MODULE_PATH = ROOT / "scripts/one_off/official_sources/curate_bfa_2026_recommendation_candidate_numbers_auxiliary.py"
RAW_PDF = (
    ROOT
    / "data/raw/official_non_final_row_level_bfa_2026_recommendation_candidate_numbers/"
    "bfa_2026_recommendation_candidate_numbers.pdf"
)
CLEAN_DIR = ROOT / "data/cleaned/graduate_outcomes"
WORKBOOK = ROOT / "outputs/graduate_outcomes/graduate_outcomes_clean_data_package.xlsx"


def load_curate_module():
    if not MODULE_PATH.exists():
        raise AssertionError(f"missing parser module: {MODULE_PATH}")
    spec = importlib.util.spec_from_file_location("curate_bfa_auxiliary", MODULE_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"could not load {MODULE_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class BfaRecommendationCandidateNumbersAuxiliaryTest(unittest.TestCase):
    def test_parses_official_pdf_ocr_rows_with_candidate_numbers(self) -> None:
        curate = load_curate_module()

        rows = curate.parse_official_pdf(RAW_PDF)

        self.assertEqual(len(rows), 105)
        first = rows[0]
        self.assertEqual(first["school_name"], "\u5317\u4eac\u7535\u5f71\u5b66\u9662")
        self.assertEqual(first["source_scope"], "non_final_recommendation_candidate_number_list")
        self.assertEqual(first["coverage_counted"], "false")
        self.assertEqual(first["person_name"], "\u5218\u529b\u6e90")
        self.assertEqual(first["student_id"], "100506100500001")
        self.assertIn("not_final_admitted_list", first["exclusion_reason"])

        by_id = {row["student_id"]: row for row in rows}
        self.assertEqual(by_id["100506100500105"]["person_name"], "\u51af\u84dd\u6708")
        self.assertEqual(by_id["100506100500053"]["person_name"], "\u738b\u4f73\u79be")

    def test_auxiliary_records_are_available_in_cleaned_package_and_workbook(self) -> None:
        auxiliary_csv = CLEAN_DIR / "official_non_final_row_level_records.csv"
        self.assertTrue(auxiliary_csv.exists(), "auxiliary non-final row-level CSV should be in cleaned package")

        with auxiliary_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))

        bfa_rows = [row for row in rows if row["school_name"] == "\u5317\u4eac\u7535\u5f71\u5b66\u9662"]
        self.assertEqual(len(bfa_rows), 105)
        self.assertTrue(all(row["coverage_counted"] == "false" for row in bfa_rows))

        workbook = load_workbook(WORKBOOK, read_only=True)
        self.assertIn("Non_Final_Row_Level_Records", workbook.sheetnames)
        sheet = workbook["Non_Final_Row_Level_Records"]
        sheet_rows = max(sum(1 for _ in sheet.iter_rows(values_only=True)) - 1, 0)
        self.assertEqual(sheet_rows, len(rows))


if __name__ == "__main__":
    unittest.main()
