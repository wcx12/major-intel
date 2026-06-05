from __future__ import annotations

import csv
from pathlib import Path
import unittest

from openpyxl import load_workbook

from scripts.one_off.official_sources import curate_xza_2025_adjustment_score_auxiliary as curate


ROOT = Path(__file__).resolve().parents[3]
RAW_HTML = (
    ROOT
    / "data/raw/official_site_recommendation_websearch_web_20260602_batch491_xza_2025_adjustment_score_probe/"
    "www.xza.edu.cn/5e8c5de0a6671c67.htm"
)
CLEAN_DIR = ROOT / "data/cleaned/graduate_outcomes"
WORKBOOK = ROOT / "outputs/graduate_outcomes/graduate_outcomes_clean_data_package.xlsx"


class XzaAdjustmentScoreAuxiliaryTest(unittest.TestCase):
    def test_parses_official_score_table_with_scores(self) -> None:
        rows = curate.parse_official_score_rows(RAW_HTML.read_text(encoding="utf-8"))

        self.assertEqual(len(rows), 513)
        first = rows[0]
        self.assertEqual(first["school_name"], "西藏农牧大学")
        self.assertEqual(first["source_scope"], "non_final_adjustment_score_table")
        self.assertEqual(first["coverage_counted"], "false")
        self.assertEqual(first["person_name"], "李兆政")
        self.assertEqual(first["student_id"], "113475120250163")
        self.assertEqual(first["college"], "资源与环境学院")
        self.assertEqual(first["major"], "风景园林")
        self.assertEqual(first["retest_score"], "87.20")
        self.assertEqual(first["admission_score"], "74.10")
        self.assertIn("not_final_admitted_list", first["exclusion_reason"])

    def test_auxiliary_records_are_available_in_cleaned_package_and_workbook(self) -> None:
        auxiliary_csv = CLEAN_DIR / "official_non_final_row_level_records.csv"
        self.assertTrue(auxiliary_csv.exists(), "auxiliary non-final row-level CSV should be in cleaned package")

        with auxiliary_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))

        xza_rows = [row for row in rows if row["school_name"] == "\u897f\u85cf\u519c\u7267\u5927\u5b66"]
        self.assertEqual(len(xza_rows), 513)
        self.assertIn("source_scope", rows[0])
        self.assertIn("coverage_counted", rows[0])
        self.assertEqual(xza_rows[0]["source_scope"], "non_final_adjustment_score_table")
        self.assertEqual(xza_rows[0]["coverage_counted"], "false")

        workbook = load_workbook(WORKBOOK, read_only=True)
        self.assertIn("Non_Final_Row_Level_Records", workbook.sheetnames)
        sheet_rows = max(
            sum(1 for _ in workbook["Non_Final_Row_Level_Records"].iter_rows(values_only=True)) - 1,
            0,
        )
        self.assertEqual(sheet_rows, len(rows))


if __name__ == "__main__":
    unittest.main()
