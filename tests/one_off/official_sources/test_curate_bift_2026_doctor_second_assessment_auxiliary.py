from __future__ import annotations

import csv
import importlib.util
from pathlib import Path
import unittest

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
MODULE_PATH = ROOT / "scripts/one_off/official_sources/curate_bift_2026_doctor_second_assessment_auxiliary.py"
RAW_PDF = (
    ROOT
    / "data/raw/official_non_final_row_level_bift_2026_doctor_second_assessment/"
    "bift_2026_doctor_second_assessment_list.pdf"
)
CLEAN_DIR = ROOT / "data/cleaned/graduate_outcomes"
WORKBOOK = ROOT / "outputs/graduate_outcomes/graduate_outcomes_clean_data_package.xlsx"


def load_curate_module():
    if not MODULE_PATH.exists():
        raise AssertionError(f"missing parser module: {MODULE_PATH}")
    spec = importlib.util.spec_from_file_location("curate_bift_auxiliary", MODULE_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"could not load {MODULE_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class BiftDoctorSecondAssessmentAuxiliaryTest(unittest.TestCase):
    def test_parses_official_pdf_ocr_rows_with_review_scores(self) -> None:
        curate = load_curate_module()

        rows = curate.parse_official_pdf(RAW_PDF)

        self.assertEqual(len(rows), 23)
        first = rows[0]
        self.assertEqual(first["school_name"], "\u5317\u4eac\u670d\u88c5\u5b66\u9662")
        self.assertEqual(first["source_scope"], "non_final_doctor_second_batch_comprehensive_assessment_list")
        self.assertEqual(first["coverage_counted"], "false")
        self.assertEqual(first["research_direction"], "04\u6750\u6599\u667a\u80fd\u8bbe\u8ba1")
        self.assertEqual(first["college"], "003\u6750\u6599\u8bbe\u8ba1\u4e0e\u5de5\u7a0b\u5b66\u9662")
        self.assertEqual(first["major"], "140300\u8bbe\u8ba1\u5b66")
        self.assertEqual(first["application_no"], "1001299726")
        self.assertEqual(first["person_name"], "\u56de\u83b9\u83b9")
        self.assertEqual(first["material_review_score"], "53.40")
        self.assertEqual(first["entered_comprehensive_assessment"], "\u5426")
        self.assertIn("not_final_admitted_list", first["exclusion_reason"])

        entered = [row for row in rows if row["entered_comprehensive_assessment"] == "\u662f"]
        self.assertEqual(len(entered), 8)
        self.assertEqual(entered[0]["application_no"], "1001299739")
        self.assertEqual(entered[-1]["application_no"], "1001299759")

    def test_auxiliary_records_are_available_in_cleaned_package_and_workbook(self) -> None:
        auxiliary_csv = CLEAN_DIR / "official_non_final_row_level_records.csv"
        self.assertTrue(auxiliary_csv.exists(), "auxiliary non-final row-level CSV should be in cleaned package")

        with auxiliary_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))

        bift_rows = [row for row in rows if row["school_name"] == "\u5317\u4eac\u670d\u88c5\u5b66\u9662"]
        self.assertEqual(len(bift_rows), 23)
        self.assertIn("application_no", rows[0])
        self.assertIn("material_review_score", rows[0])
        self.assertTrue(all(row["coverage_counted"] == "false" for row in bift_rows))

        workbook = load_workbook(WORKBOOK, read_only=True)
        self.assertIn("Non_Final_Row_Level_Records", workbook.sheetnames)
        sheet = workbook["Non_Final_Row_Level_Records"]
        header = next(sheet.iter_rows(values_only=True))
        self.assertIn("application_no", header)
        self.assertIn("material_review_score", header)
        sheet_rows = max(sum(1 for _ in sheet.iter_rows(values_only=True)) - 1, 0)
        self.assertEqual(sheet_rows, len(rows))


if __name__ == "__main__":
    unittest.main()
