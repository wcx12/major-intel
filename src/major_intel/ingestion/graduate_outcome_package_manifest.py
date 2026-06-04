from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
CLEAN_DIR = ROOT / "data/cleaned/graduate_outcomes"
OUT_DIR = ROOT / "outputs/graduate_outcomes"
WORKBOOK = OUT_DIR / "graduate_outcomes_clean_data_package.xlsx"
MANIFEST = OUT_DIR / "package_manifest.json"

CLEAN_CSVS = [
    "master_records_public.csv",
    "master_records_clean.csv",
    "official_recommendation_school_coverage.csv",
    "official_recommendation_source_attempts.csv",
    "official_non_final_row_level_records.csv",
    "school_year_source_summary.csv",
    "undergraduate_school_outcome_summary.csv",
    "official_employment_report_sources.csv",
    "official_employment_report_metrics.csv",
]
PACKAGE_CSVS = [
    "remaining_uncovered_schools.csv",
    "remaining_uncovered_recheck_2026-06-03.csv",
    "remaining_uncovered_recheck_2026-06-04.csv",
]
CHECKSUM_PATHS = [
    *(CLEAN_DIR / name for name in CLEAN_CSVS),
    WORKBOOK,
    OUT_DIR / "README.md",
    *(OUT_DIR / name for name in PACKAGE_CSVS),
]


def csv_row_count(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return max(sum(1 for _ in csv.reader(handle)) - 1, 0)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def file_checksum(path: Path) -> dict[str, object]:
    data = path.read_bytes()
    return {"bytes": len(data), "sha256": hashlib.sha256(data).hexdigest()}


def workbook_row_counts() -> tuple[list[str], dict[str, int]]:
    workbook = load_workbook(WORKBOOK, read_only=True)
    counts = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        counts[sheet_name] = max(sum(1 for _ in sheet.iter_rows(values_only=True)) - 1, 0)
    return workbook.sheetnames, counts


def main() -> None:
    coverage_rows = read_csv(CLEAN_DIR / "official_recommendation_school_coverage.csv")
    uncovered = [
        row["school_name"]
        for row in coverage_rows
        if row["has_official_recommendation_records"].lower() != "true"
    ]
    covered = len(coverage_rows) - len(uncovered)
    workbook_sheets, workbook_counts = workbook_row_counts()

    manifest = {
        "generated_at": "2026-06-04",
        "package": "graduate_outcomes_clean_data_package",
        "status": f"blocked_at_{covered}_of_{len(coverage_rows)}_due_to_public_official_source_access",
        "coverage": {
            "target_schools": len(coverage_rows),
            "official_final_row_level_covered": covered,
            "official_final_row_level_uncovered": len(uncovered),
            "uncovered_schools": uncovered,
        },
        "csv_row_counts": {name: csv_row_count(CLEAN_DIR / name) for name in CLEAN_CSVS},
        "package_csv_row_counts": {name: csv_row_count(OUT_DIR / name) for name in PACKAGE_CSVS},
        "workbook_sheets": workbook_sheets,
        "workbook_data_row_counts": workbook_counts,
        "checksums": {
            path.relative_to(ROOT).as_posix(): file_checksum(path) for path in CHECKSUM_PATHS
        },
        "blocker_documents": [
            "docs/research/graduate_outcome_remaining_blockers_2026-06-03.md",
            "docs/research/graduate_outcome_unblock_checklist_2026-06-03.md",
        ],
    }
    MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(
        {
            "covered": covered,
            "uncovered": len(uncovered),
            "public_records": manifest["csv_row_counts"]["master_records_public.csv"],
            "workbook_public_records": workbook_counts["Public_Records"],
            "checksum_entries": len(manifest["checksums"]),
        }
    )


if __name__ == "__main__":
    main()
