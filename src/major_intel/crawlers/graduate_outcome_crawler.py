"""Crawl public postgraduate-outcome announcement lists.

The crawler is seed-driven: provide official university entry URLs, then it
follows same-site links that look like recommendation-exemption or postgraduate
admission public lists. It stores evidence documents and best-effort extracted
rows separately so downstream review can trace every fact to a source URL.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import ssl
import subprocess
import time
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, quote, urlencode, unquote, urljoin, urlparse
from urllib.request import Request, urlopen

from bs4 import BeautifulSoup
from openpyxl import load_workbook


DOCUMENT_SCHEMA_VERSION = "graduate_outcome_document/v1"
RECORD_SCHEMA_VERSION = "graduate_outcome_record/v1"
ATTACHMENT_SUFFIXES = {
    ".pdf",
    ".xls",
    ".xlsx",
    ".xlsm",
    ".doc",
    ".docx",
    ".wps",
}
BLOCKING_STATUS_CODES = {401, 403}
RETRYABLE_STATUS_CODES = {429, 500, 502, 503, 504}

DOCUMENT_TYPE_KEYWORDS = [
    (
        "incoming_recommendation_admission_list",
        [
            "接收推免",
            "推免生拟录取",
            "推免拟录取",
            "推免拟录取名单",
            "推免硕士研究生拟录取",
            "推免硕士研究生拟录取名单",
            "推免研究生拟录取名单",
            "推免研究生拟录取",
            "拟录取推免生",
            "推荐免试研究生拟录取",
            "推免生攻读研究生拟录取",
        ],
    ),
    (
        "recommendation_exemption_list",
        ["推荐免试", "推免名单", "免试攻读研究生拟录取", "免试攻读研究生名单", "免试攻读", "推免资格", "推免生名单"],
    ),
    (
        "postgraduate_admission_list",
        ["硕士研究生拟录取名单", "研究生拟录取名单", "拟录取名单"],
    ),
    (
        "postgraduate_admission_list",
        [
            "硕士研究生拟录取考生名单",
            "研究生拟录取考生名单",
            "拟录取考生名单",
            "待录取考生名单",
        ],
    ),
]

OFFICIAL_SITE_PORTAL_KEYWORDS = [
    "教务",
    "本科生院",
    "本科教学",
    "本科教育",
    "教学管理",
    "培养管理",
    "推免",
    "推荐免试",
]

HEADER_ALIASES = {
    "person_name": ["姓名", "考生姓名", "学生姓名", "推免生姓名"],
    "student_id": ["学号", "考生编号", "报名号", "申请编号", "编号"],
    "undergraduate_school": [
        "本科毕业院校",
        "本科毕业学校",
        "本科高校",
        "毕业院校",
        "推荐学校",
        "所在学校",
        "生源学校",
    ],
    "undergraduate_major": ["本科专业", "毕业专业", "所学专业"],
    "college": [
        "学院",
        "拟录取学院",
        "录取学院",
        "接收学院",
        "院系",
        "培养单位",
        "报考学院",
    ],
    "major": ["专业", "所在专业", "推荐专业"],
    "admission_major": ["拟录取专业", "录取专业", "接收专业", "报考专业", "申请专业", "专业名称"],
    "ranking": ["排名", "综合排名", "专业排名"],
    "remarks": ["备注", "说明", "录取类别", "学习方式"],
}

RECORD_CSV_FIELDS = [
    "school_name",
    "year",
    "document_type",
    "route",
    "person_name",
    "student_id",
    "undergraduate_school",
    "undergraduate_major",
    "college",
    "major",
    "admission_major",
    "ranking",
    "remarks",
    "source_url",
    "title",
    "needs_review",
]

DISCOVERY_SOURCE_TYPES = [
    {
        "source_type": "recommendation_exemption",
        "document_type": "recommendation_exemption_list",
        "query_template": '"{school_name}" "{year}届" "推荐免试" "名单" "公示" site:edu.cn',
    },
    {
        "source_type": "incoming_recommendation",
        "document_type": "incoming_recommendation_admission_list",
        "query_template": '"{school_name}" "{year}年" "接收推免生拟录取名单" site:edu.cn',
    },
    {
        "source_type": "postgraduate_admission",
        "document_type": "postgraduate_admission_list",
        "query_template": '"{school_name}" "{year}年硕士研究生拟录取名单" site:edu.cn',
    },
]

DISCOVERY_TASK_CSV_FIELDS = [
    "school_id",
    "school_name",
    "province",
    "level",
    "year",
    "source_type",
    "document_type",
    "eligibility_hint",
    "search_query",
    "preferred_domains",
    "status",
    "found_url",
    "notes",
]

SEED_CSV_FIELDS = [
    "school_name",
    "source_type",
    "start_url",
    "year",
    "document_type",
    "discovery_query",
    "discovery_title",
    "discovery_rank",
]

SEARCH_RESULT_CSV_FIELDS = [
    "search_query",
    "result_rank",
    "result_title",
    "result_url",
    "result_snippet",
    "provider",
    "captured_at",
]

CLEAN_RECORD_CSV_FIELDS = [
    "record_id",
    "school_name",
    "year",
    "document_type",
    "route",
    "person_name",
    "person_name_masked",
    "student_id",
    "student_id_masked",
    "undergraduate_school",
    "undergraduate_major",
    "college",
    "major",
    "admission_major",
    "ranking",
    "remarks",
    "source_url",
    "title",
    "needs_review",
    "quality_score",
    "quality_flags",
]

SUMMARY_CSV_FIELDS = [
    "school_name",
    "year",
    "document_type",
    "route",
    "record_count",
    "unique_person_count",
    "needs_review_count",
    "with_undergraduate_school_count",
    "with_admission_major_count",
    "source_document_count",
]

PUBLIC_RECORD_CSV_FIELDS = [
    "public_record_id",
    "school_name",
    "year",
    "document_type",
    "route",
    "person_name_masked",
    "student_id_masked",
    "undergraduate_school",
    "undergraduate_major",
    "college",
    "major",
    "admission_major",
    "ranking",
    "remarks",
    "source_url",
    "title",
    "needs_review",
    "quality_score",
    "quality_flags",
]

CHSI_SCHOOL_CSV_FIELDS = [
    "chsi_school_name",
    "chsi_sch_id",
    "chsi_school_url",
]

SCHOOL_SITE_CSV_FIELDS = [
    "school_id",
    "school_name",
    "province",
    "level",
    "eligibility_hint",
    "official_url",
    "matched_link_text",
    "source_url",
    "source_rank",
]


class FetchError(RuntimeError):
    def __init__(self, message: str, status_code: int | None = None):
        super().__init__(message)
        self.status_code = status_code


@dataclass(frozen=True)
class FetchResponse:
    url: str
    status_code: int
    content_type: str
    content: bytes


@dataclass(frozen=True)
class LinkCandidate:
    url: str
    text: str
    link_kind: str
    document_type: str
    matched_keywords: list[str]
    year_hint: int | None = None


Fetcher = Callable[[str, float], FetchResponse]
Sleeper = Callable[[float], None]


def classify_document(title: str, url: str, text: str = "") -> dict[str, Any]:
    haystack = f"{title} {unquote(url)} {text}"
    matched: list[str] = []
    document_type = "unknown"
    has_negative_recommendation_context = _has_negative_recommendation_context(haystack)
    compact_haystack = re.sub(r"\s+", "", haystack)
    if (
        not has_negative_recommendation_context
        and re.search(r"推免生.*拟录取", compact_haystack)
    ):
        document_type = "incoming_recommendation_admission_list"
        matched.append("推免生拟录取")

    for candidate_type, keywords in DOCUMENT_TYPE_KEYWORDS:
        if has_negative_recommendation_context and candidate_type in {
            "incoming_recommendation_admission_list",
            "recommendation_exemption_list",
        }:
            continue
        candidate_matches = [keyword for keyword in keywords if keyword in haystack]
        if candidate_matches and document_type == "unknown":
            document_type = candidate_type
        matched.extend(candidate_matches)

    return {
        "document_type": document_type,
        "matched_keywords": _unique_texts(matched),
    }


def _has_negative_recommendation_context(text: str) -> bool:
    compact = re.sub(r"\s+", "", text or "")
    return any(marker in compact for marker in ["不含推荐免试", "不含推免", "非推免"])


def extract_candidate_links(
    html: str,
    base_url: str,
    allowed_domains: set[str] | None = None,
    include_all_attachments: bool = False,
) -> list[LinkCandidate]:
    soup = BeautifulSoup(html, "html.parser")
    base_host = urlparse(base_url).netloc
    allowed = allowed_domains or {base_host}
    seen: set[str] = set()
    links: list[LinkCandidate] = []

    for anchor in soup.find_all("a"):
        href = (anchor.get("href") or "").strip()
        if not href or href.startswith("#") or href.lower().startswith("javascript:"):
            continue

        text = _clean_text(anchor.get_text(" ", strip=True))
        if _is_article_navigation_link(text):
            continue
        _add_candidate_link(
            href=href,
            text=text,
            base_url=base_url,
            allowed=allowed,
            include_all_attachments=include_all_attachments,
            seen=seen,
            links=links,
        )

    for tag in soup.find_all(True):
        for attr_name in ("pdfsrc", "filesrc"):
            href = (tag.get(attr_name) or "").strip()
            if not href:
                continue
            text = _clean_text(
                " ".join(
                    part
                    for part in [
                        _embedded_file_title(tag),
                        tag.get("title") or "",
                        tag.get_text(" ", strip=True),
                    ]
                    if part
                )
            )
            _add_candidate_link(
                href=href,
                text=text,
                base_url=base_url,
                allowed=allowed,
                include_all_attachments=include_all_attachments,
                seen=seen,
                links=links,
            )

    for script in soup.find_all("script"):
        script_text = script.string or script.get_text(" ", strip=False)
        for match in re.finditer(r"showVsbpdfIframe\(\s*['\"]([^'\"]+)['\"]", script_text or ""):
            _add_candidate_link(
                href=match.group(1),
                text="embedded pdf",
                base_url=base_url,
                allowed=allowed,
                include_all_attachments=include_all_attachments,
                seen=seen,
                links=links,
            )

    return links


def _is_article_navigation_link(text: str) -> bool:
    clean = _clean_text(text).strip("<>‹›«»")
    return bool(
        re.match(
            r"^(上一篇|下一篇|上一条|下一条|上一个|下一个|返回|关闭|首页|当前位置)\s*[：:]?",
            clean,
        )
    )


def _add_candidate_link(
    href: str,
    text: str,
    base_url: str,
    allowed: set[str],
    include_all_attachments: bool,
    seen: set[str],
    links: list[LinkCandidate],
) -> None:
    try:
        absolute_url = urljoin(base_url, href)
        parsed = urlparse(absolute_url)
    except ValueError:
        return
    if re.search(r"https?:", parsed.path, flags=re.I):
        return
    if parsed.scheme not in {"http", "https"} or parsed.netloc not in allowed:
        return

    link_kind = "attachment" if _looks_like_attachment_link(absolute_url, text) else "page"
    classification = classify_document(text, absolute_url, "")
    if classification["document_type"] == "unknown" and not (
        include_all_attachments and link_kind == "attachment"
    ):
        return
    if absolute_url in seen:
        return

    seen.add(absolute_url)
    links.append(
        LinkCandidate(
            url=absolute_url,
            text=text,
            link_kind=link_kind,
            document_type=classification["document_type"],
            matched_keywords=classification["matched_keywords"],
            year_hint=None,
        )
    )


def _embedded_file_title(tag: Any) -> str:
    raw = tag.get("sudyfile-attr") or tag.get("fileattr") or ""
    match = re.search(r"['\"]?title['\"]?\s*:\s*['\"]([^'\"]+)['\"]", raw)
    if match:
        return _clean_text(match.group(1))
    return ""


def _looks_like_attachment_link(url: str, text: str = "") -> bool:
    return _is_attachment_url(url) or _attachment_suffix_from_text(text) is not None


def _attachment_suffix_from_text(text: str) -> str | None:
    match = re.search(r"\.(pdf|xlsx|xlsm|xls|docx|doc|wps)\b", text or "", flags=re.I)
    if not match:
        return None
    return f".{match.group(1).lower()}"


def parse_html_records(html: str, document: dict[str, Any]) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    records: list[dict[str, Any]] = []

    shutcm_meta_records = _records_from_shutcm_meta_description(soup, document)
    if shutcm_meta_records:
        records.extend(shutcm_meta_records)

    for table in soup.find_all("table"):
        if table.find("table"):
            continue
        rows = [
            [_clean_text(cell.get_text(" ", strip=True)) for cell in row.find_all(["th", "td"])]
            for row in table.find_all("tr")
        ]
        gzucm_records = _records_from_school_college_candidate_admission_table(rows, document)
        if gzucm_records:
            records.extend(gzucm_records)
            continue
        cdutcm_records = _records_from_cdutcm_recommendation_table(rows, document)
        if cdutcm_records:
            records.extend(cdutcm_records)
            continue
        sut_records = _records_from_sut_recommendation_table(rows, document)
        if sut_records:
            records.extend(sut_records)
            continue
        candidate_major_records = _records_from_candidate_major_code_name_score_table(rows, document)
        if candidate_major_records:
            records.extend(candidate_major_records)
            continue
        subject_category_records = _records_from_subject_category_score_table(rows, document)
        if subject_category_records:
            records.extend(subject_category_records)
            continue
        if _table_is_aggregate_admission_count(rows):
            continue
        records.extend(_records_from_table(rows, document))

    text_soup = soup
    if records:
        text_soup = BeautifulSoup(str(soup), "html.parser")
        for table in text_soup.find_all("table"):
            table.decompose()

    replenishment_records = _records_from_aufe_replenishment_notice(text_soup, document)
    if replenishment_records:
        records.extend(replenishment_records)
        return records
    records.extend(_records_from_major_name_paragraphs(text_soup, document))
    records.extend(_records_from_structured_major_name_lines(text_soup, document))
    records.extend(_records_from_plain_recommendation_name_lines(text_soup, document))
    return records


def _records_from_shutcm_meta_description(
    soup: BeautifulSoup,
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "上海中医药大学" not in document_text and "shutcm.edu.cn" not in document_text:
        return []
    meta = soup.find("meta", attrs={"name": "description"})
    content = _clean_text(meta.get("content", "") if meta else "")
    if not all(term in content for term in ("序号", "姓名", "拟录取学院", "专业代码", "推荐学校")):
        return []
    payload = re.sub(r"^.*?推荐学校", "", content)
    pattern = re.compile(
        r"(?P<ranking>\d{1,3})"
        r"(?P<person_name>[\u4e00-\u9fff]{2,4})"
        r"(?P<college_code>A\d{2})"
        r"(?P<college>[\u4e00-\u9fff]+(?:学院|研究院))"
        r"(?P<major_code>\d{6}[A-Z]?)"
        r"(?P<major>.+?)"
        r"(?P<score>\d{2,3}(?:\.\d+)?)"
        r"(?P<undergraduate_school>.+?)"
        r"(?=\d{1,3}[\u4e00-\u9fff]{2,4}A\d{2}|$)"
    )
    records: list[dict[str, Any]] = []
    for match in pattern.finditer(payload):
        values = {key: _clean_text(value) for key, value in match.groupdict().items()}
        undergraduate_school = values["undergraduate_school"]
        if (
            not _looks_like_chinese_name(values["person_name"])
            or not _looks_like_major_code(values["major_code"])
            or not _looks_like_score_or_metric(values["score"])
            or not re.search(r"(大学|学院)$", undergraduate_school)
        ):
            continue
        record = _build_record(
            document,
            {
                "person_name": values["person_name"],
                "college": _clean_text(f"{values['college_code']} {values['college']}"),
                "major": values["major_code"],
                "admission_major": _clean_text(f"{values['major_code']} {values['major']}"),
                "ranking": values["ranking"],
                "remarks": _clean_text(
                    f"retest_score {values['score']}; undergraduate_school {undergraduate_school}; source_fragment meta_description"
                ),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_school_college_candidate_admission_table(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    header_index = _school_college_candidate_admission_header_index(rows)
    if header_index is None:
        return []

    records: list[dict[str, Any]] = []
    for values in rows[header_index + 1 :]:
        values = [_clean_text(value) for value in values]
        if len(values) < 8:
            continue
        (
            college_code,
            college_name,
            person_name,
            major_code,
            major_name,
            direction_code,
            direction_name,
            advisor,
            *extra_values,
        ) = values
        if not re.search(r"[\u4e00-\u9fff]{2,}", person_name):
            continue
        if not _looks_like_major_code(major_code):
            continue
        if not major_name or _looks_like_score_or_metric(major_name):
            continue

        remarks = _clean_text(
            " ".join(
                [
                    f"college_code {college_code}" if college_code else "",
                    f"direction {direction_code} {direction_name}".strip() if direction_code or direction_name else "",
                    f"advisor {advisor}" if advisor else "",
                    *extra_values,
                ]
            )
        )
        record = _build_record(
            document,
            {
                "person_name": person_name,
                "college": college_name,
                "major": major_code,
                "admission_major": _clean_text(f"{major_code} {major_name}"),
                "remarks": remarks,
            },
        )
        if record:
            records.append(record)
    return records


def _school_college_candidate_admission_header_index(rows: list[list[str]]) -> int | None:
    required = {
        "院所代码",
        "院所名称",
        "考生姓名",
        "录取专业代码",
        "录取专业名称",
        "研究方向代码",
        "研究方向名称",
        "接收导师",
    }
    for index, values in enumerate(rows[:20]):
        normalized = {_clean_text(value) for value in values}
        if required.issubset(normalized):
            return index
    return None


def _records_from_sut_recommendation_table(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "sut.edu.cn" not in document_text and "沈阳工业大学" not in document_text:
        return []

    header_index = _sut_recommendation_header_index(rows)
    if header_index is None:
        return []

    records: list[dict[str, Any]] = []
    for values in rows[header_index + 1 :]:
        record = _sut_recommendation_record(values, document)
        if record:
            records.append(record)
    return records


def _sut_recommendation_header_index(rows: list[list[str]]) -> int | None:
    required = {"序号", "姓名", "证件号码", "学院", "专业代码", "专业名称", "学习方式", "招生类型", "复试成绩"}
    for index, values in enumerate(rows[:20]):
        normalized = {_clean_text(value) for value in values}
        if required.issubset(normalized):
            return index
    return None


def _sut_recommendation_record(
    row: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in row]
    if len(values) < 10 or not values[0].isdigit():
        return None
    score = values[9]
    if not _looks_like_score_or_metric(score):
        return None
    person_name = re.sub(r"\s*\*\s*", "*", values[1])
    remarks = [
        f"gender {values[2]}",
        f"study_mode {values[7]}",
        f"admission_type {values[8]}",
        f"score {score}",
    ]
    return _build_record(
        document,
        {
            "person_name": person_name,
            "student_id": values[3],
            "college": values[4],
            "major": values[5],
            "admission_major": _clean_text(f"{values[5]} {values[6]}"),
            "ranking": values[0],
            "remarks": _clean_text("; ".join(part for part in remarks if part)),
        },
    )


def _records_from_candidate_major_code_name_score_table(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    header_index = _candidate_major_code_name_score_header_index(rows)
    if header_index is None:
        return []

    header = [_clean_text(value) for value in rows[header_index]]
    normalized_header = [_normalize_header(value) for value in header]
    name_index = _header_column_index(normalized_header, ["姓名", "考生姓名", "学生姓名"])
    major_code_index = _header_column_index(
        normalized_header,
        ["报考专业代码", "录取专业代码", "拟录取专业代码", "专业代码"],
    )
    major_name_index = _header_column_index(
        normalized_header,
        ["报考专业名称", "录取专业名称", "拟录取专业名称", "拟录取专业", "专业名称"],
    )
    if name_index is None or major_code_index is None or major_name_index is None:
        return []

    ranking_index = _header_column_index(normalized_header, ["序号", "排名"])
    college_index = _header_column_index(normalized_header, ["拟录取学院", "录取学院", "报考学院", "学院"])
    gender_index = _header_column_index(normalized_header, ["性别"])
    study_mode_index = _header_column_index(normalized_header, ["报考学习形式", "学习形式", "学习方式"])
    score_index = _header_column_index(normalized_header, ["复试成绩", "综合成绩", "总成绩"])
    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        values = [_clean_text(value) for value in row]
        if not any(values):
            continue
        values = _align_values_to_header(header, values)
        person_name = values[name_index] if name_index < len(values) else ""
        major_code = values[major_code_index] if major_code_index < len(values) else ""
        major_name = values[major_name_index] if major_name_index < len(values) else ""
        if not _looks_like_chinese_name(person_name):
            continue
        if not _looks_like_major_code(major_code):
            continue
        if not major_name or _looks_like_score_or_metric(major_name):
            continue

        remarks: list[str] = []
        if gender_index is not None and gender_index < len(values) and values[gender_index]:
            remarks.append(f"gender {values[gender_index]}")
        if study_mode_index is not None and study_mode_index < len(values) and values[study_mode_index]:
            remarks.append(f"study_mode {values[study_mode_index]}")
        if score_index is not None and score_index < len(values) and _looks_like_score_or_metric(values[score_index]):
            remarks.append(f"retest_score {values[score_index]}")

        record = _build_record(
            document,
            {
                "person_name": person_name,
                "college": values[college_index] if college_index is not None and college_index < len(values) else "",
                "major": major_code,
                "admission_major": _clean_text(f"{major_code} {major_name}"),
                "ranking": values[ranking_index] if ranking_index is not None and ranking_index < len(values) else "",
                "remarks": "; ".join(remarks),
            },
        )
        if record:
            records.append(record)
    return records


def _candidate_major_code_name_score_header_index(rows: list[list[str]]) -> int | None:
    for index, values in enumerate(rows[:20]):
        normalized = {_normalize_header(value) for value in values}
        if (
            any(value in normalized for value in {"姓名", "考生姓名", "学生姓名"})
            and any(value in normalized for value in {"报考专业代码", "录取专业代码", "拟录取专业代码", "专业代码"})
            and any(value in normalized for value in {"报考专业名称", "录取专业名称", "拟录取专业名称", "拟录取专业", "专业名称"})
        ):
            return index
    return None


def _records_from_subject_category_score_table(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    header_index = _subject_category_score_header_index(rows)
    if header_index is None:
        return []

    header = [_clean_text(value) for value in rows[header_index]]
    normalized_header = [_normalize_header(value) for value in header]
    student_id_index = _header_column_index(normalized_header, ["考生编号", "报名号"])
    name_index = _header_column_index(normalized_header, ["姓名", "考生姓名"])
    subject_index = _header_column_index(normalized_header, ["一级学科名称", "学科名称", "报考学科名称"])
    category_index = _header_column_index(normalized_header, ["报考类别", "录取类别"])
    score_index = _header_column_index(normalized_header, ["综合面试成绩", "面试成绩", "复试成绩", "综合成绩"])
    remarks_index = _header_column_index(normalized_header, ["备注"])
    if student_id_index is None or name_index is None or subject_index is None:
        return []

    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        values = [_clean_text(value) for value in row]
        if not any(values):
            continue
        values = _align_values_to_header(header, values)
        person_name = values[name_index] if name_index < len(values) else ""
        student_id = values[student_id_index] if student_id_index < len(values) else ""
        subject_name = values[subject_index] if subject_index < len(values) else ""
        if not _looks_like_chinese_name(person_name):
            continue
        if not _looks_like_identifier_only(student_id):
            continue
        if not subject_name or _looks_like_score_or_metric(subject_name):
            continue

        remarks: list[str] = []
        if category_index is not None and category_index < len(values) and values[category_index]:
            remarks.append(values[category_index])
        if score_index is not None and score_index < len(values) and values[score_index]:
            score = values[score_index]
            remarks.append(f"interview_score {score}" if _looks_like_score_or_metric(score) else f"interview_score {score}")
        if remarks_index is not None and remarks_index < len(values) and values[remarks_index]:
            remarks.append(values[remarks_index])

        record = _build_record(
            document,
            {
                "person_name": person_name,
                "student_id": student_id,
                "admission_major": subject_name,
                "remarks": _clean_text(" ".join(remarks)),
            },
        )
        if record:
            records.append(record)
    return records


def _subject_category_score_header_index(rows: list[list[str]]) -> int | None:
    for index, values in enumerate(rows[:20]):
        normalized = {_normalize_header(value) for value in values}
        if (
            any(value in normalized for value in {"姓名", "考生姓名"})
            and any(value in normalized for value in {"考生编号", "报名号"})
            and any(value in normalized for value in {"一级学科名称", "学科名称", "报考学科名称"})
        ):
            return index
    return None


def _header_column_index(normalized_header: list[str], aliases: list[str]) -> int | None:
    normalized_aliases = [_normalize_header(alias) for alias in aliases]
    for alias in normalized_aliases:
        for index, value in enumerate(normalized_header):
            if value == alias:
                return index
    for alias in normalized_aliases:
        for index, value in enumerate(normalized_header):
            if alias and alias in value:
                return index
    return None


def _records_from_cdutcm_recommendation_table(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "成都中医药大学" not in document_text and "cdutcm.edu.cn" not in document_text:
        return []

    header_index = _cdutcm_recommendation_header_index(rows)
    if header_index is None:
        return []

    college = _cdutcm_recommendation_college(document, rows[:header_index])
    records: list[dict[str, Any]] = []
    for values in rows[header_index + 1 :]:
        record = _cdutcm_recommendation_record(values, document, college)
        if record:
            records.append(record)
    return records


def _cdutcm_recommendation_header_index(rows: list[list[str]]) -> int | None:
    for index, values in enumerate(rows[:20]):
        normalized = [_clean_text(value) for value in values]
        if (
            "序号" in normalized
            and "考生姓名" in normalized
            and "报考专业代码" in normalized
            and "报考专业名称" in normalized
            and "综合面试成绩" in normalized
        ):
            return index
    return None


def _cdutcm_recommendation_college(
    document: dict[str, Any],
    heading_rows: list[list[str]],
) -> str:
    haystack = " ".join(
        [
            str(document.get("title") or ""),
            " ".join(" ".join(row) for row in heading_rows),
        ]
    )
    if "现代中药产业学院" in haystack:
        return "现代中药产业学院"

    match = re.search(r"([\u4e00-\u9fff]{2,30}学院)", haystack)
    if not match:
        return ""
    college = match.group(1)
    school_name = _clean_text(document.get("school_name"))
    if school_name and college.startswith(school_name):
        college = college[len(school_name) :]
    return _clean_text(college)


def _cdutcm_recommendation_record(
    values: list[str],
    document: dict[str, Any],
    college: str,
) -> dict[str, Any] | None:
    if len(values) < 9:
        return None
    (
        ranking,
        person_name,
        degree_type,
        major_code,
        major_name,
        direction_code,
        direction_name,
        interview_score,
        advisor,
        *extra_values,
    ) = values
    if not ranking.isdigit():
        return None
    if not _looks_like_chinese_name(person_name):
        return None
    if not _looks_like_major_code(major_code):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None
    if not _looks_like_score_or_metric(interview_score):
        return None

    return _build_record(
        document,
        {
            "ranking": ranking,
            "person_name": person_name,
            "college": college,
            "major": major_code,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": _clean_text(
                " ".join(
                    [
                        degree_type,
                        f"direction {direction_code}" if direction_code else "",
                        direction_name,
                        f"interview_score {interview_score}",
                        f"advisor {advisor}" if advisor else "",
                        *extra_values,
                    ]
                )
            ),
        },
    )


def _table_is_aggregate_admission_count(rows: list[list[str]]) -> bool:
    for row in rows[:3]:
        compact = "".join(row)
        if not compact:
            continue
        if "录取人数" in compact and not any(identity in compact for identity in ("姓名", "考生编号", "学号")):
            return True
    return False


def _records_from_major_name_paragraphs(
    soup: BeautifulSoup,
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    current_major = ""
    seen: set[tuple[str, str]] = set()

    for block in soup.find_all(["p", "li"]):
        if _has_navigation_ancestor(block):
            continue
        line = _clean_text(block.get_text(" ", strip=True))
        if not line:
            continue
        if _paragraph_breaks_major_name_list(line):
            current_major = ""
            continue
        major, inline_names = _major_and_candidate_names_from_paragraph(line)
        if major:
            current_major = major
            for name in inline_names:
                key = (current_major, name)
                if key in seen:
                    continue
                seen.add(key)
                record = _build_record(document, {"person_name": name, "major": current_major})
                if record:
                    records.append(record)
            continue
        if not current_major:
            continue
        for name in _split_candidate_names(line):
            key = (current_major, name)
            if key in seen:
                continue
            seen.add(key)
            record = _build_record(document, {"person_name": name, "major": current_major})
            if record:
                records.append(record)
    return records


def _has_navigation_ancestor(block: Any) -> bool:
    for parent in [block, *block.parents]:
        attrs = getattr(parent, "attrs", None)
        if not attrs:
            continue
        hints: list[str] = []
        for key in ("class", "id", "role", "aria-label"):
            value = attrs.get(key)
            if isinstance(value, (list, tuple)):
                hints.extend(str(item).lower() for item in value)
            elif value:
                hints.append(str(value).lower())
        joined = " ".join(hints)
        if re.search(r"\b(nav|navi|menu|aside|sidebar|breadcrumb|crumb|footer|header)\b", joined):
            return True
        if any(marker in joined for marker in ("wp-navi", "wp_nav", "subnav")):
            return True
    return False


def _records_from_structured_major_name_lines(
    soup: BeautifulSoup,
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()

    for container in _article_content_containers(soup):
        lines = [
            _clean_text(line)
            for line in container.get_text("\n", strip=True).splitlines()
            if _clean_text(line)
        ]
        current_major = ""
        for index, line in enumerate(lines):
            if _structured_line_breaks_major_name_list(line):
                current_major = ""
                continue
            next_line = lines[index + 1] if index + 1 < len(lines) else ""
            count_context = _line_is_count_marker(next_line) or _line_has_inline_count_marker(line)
            major = _structured_major_from_line(line, count_context=count_context)
            if major and count_context:
                current_major = major
                continue
            if _line_is_count_marker(line):
                continue
            if not current_major:
                continue
            for name in _split_candidate_names_with_parenthetical_notes(line):
                key = (current_major, name)
                if key in seen:
                    continue
                seen.add(key)
                record = _build_record(document, {"person_name": name, "major": current_major})
                if record:
                    records.append(record)
    return records


def _records_from_aufe_replenishment_notice(
    soup: BeautifulSoup,
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    school_name = str(document.get("school_name") or "")
    title = str(document.get("title") or "")
    source_url = str(document.get("source_url") or "")
    if "安徽财经大学" not in school_name and "aufe.edu.cn" not in source_url:
        return []
    if "递补" not in title:
        return []

    records: list[dict[str, Any]] = []
    for container in _article_content_containers(soup):
        lines = [
            _clean_text(line)
            for line in container.get_text("\n", strip=True).splitlines()
            if _clean_text(line)
        ]
        joined = " ".join(lines)
        if "递补" not in joined or "学生信息" not in joined:
            continue
        person_name = _value_after_label(lines, "姓名")
        major = _value_after_label(lines, "专业名称")
        if not _looks_like_chinese_name(person_name):
            continue
        remarks = _clean_text(
            " ".join(
                value
                for value in [
                    _labeled_remark(lines, "综合成绩"),
                    _labeled_remark(lines, "综合成绩排名"),
                    _labeled_remark(lines, "排名人数"),
                    _labeled_remark(lines, "GPA"),
                ]
                if value
            )
        )
        record = _build_record(
            document,
            {
                "person_name": person_name,
                "major": major,
                "remarks": remarks,
            },
        )
        if record:
            records.append(record)
    return records


def _labeled_remark(lines: list[str], label: str) -> str:
    value = _value_after_label(lines, label)
    return f"{label} {value}" if value else ""


def _value_after_label(lines: list[str], label: str) -> str:
    normalized_label = _normalize_header(label)
    for index, line in enumerate(lines):
        clean_line = _clean_text(line)
        normalized_line = _normalize_header(clean_line)
        value = ""
        if normalized_line == normalized_label:
            value = _next_label_value(lines, index)
        else:
            match = re.match(rf"^{re.escape(label)}\s*[：:]\s*(.*)$", clean_line)
            if match:
                value = _clean_text(match.group(1))
                if not value:
                    value = _next_label_value(lines, index)
        if value:
            return value
    return ""


def _next_label_value(lines: list[str], index: int) -> str:
    if index + 1 >= len(lines):
        return ""
    value = _clean_text(lines[index + 1]).strip("：:")
    if index + 2 < len(lines):
        next_value = _clean_text(lines[index + 2])
        if re.fullmatch(r"\d", value) and re.fullmatch(r"\d+(?:\.\d+)?", next_value):
            return f"{value}{next_value}"
    return value


def _records_from_plain_recommendation_name_lines(
    soup: BeautifulSoup,
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    seen: set[str] = set()

    for container in _article_content_containers(soup):
        lines = [
            _clean_text(line)
            for line in container.get_text("\n", strip=True).splitlines()
            if _clean_text(line)
        ]
        collecting = False
        for line in lines:
            if re.search(r"(候补|特此|如有异议|公示期|联系人|联系电话|上一篇|下一篇)", line):
                collecting = False
                continue
            if re.search(r"(学生名单|名单公示).*?如下", line):
                collecting = True
                continue
            if not collecting:
                continue
            for name in _split_candidate_names_with_parenthetical_notes(line):
                if name in seen:
                    continue
                seen.add(name)
                record = _build_record(document, {"person_name": name})
                if record:
                    records.append(record)
    return records


def _article_content_containers(soup: BeautifulSoup) -> list[Any]:
    containers: list[Any] = []
    class_markers = (
        "v_news_content",
        "wp_articlecontent",
        "article_content",
        "article-content",
        "news_content",
        "news-content",
        "nry_main",
        "ej_nry",
    )
    for tag in soup.find_all(True):
        classes = tag.get("class") or []
        class_text = " ".join(str(item) for item in classes).lower()
        tag_id = str(tag.get("id") or "").lower()
        if any(marker in class_text or marker in tag_id for marker in class_markers):
            containers.append(tag)
    return containers


def _structured_major_from_line(line: str, *, count_context: bool = False) -> str:
    clean = _clean_text(line).strip("：: ")
    if not clean:
        return ""
    clean = re.sub(r"[（(]\s*\d+\s*人\s*[)）]$", "", clean).strip()
    if len(clean) < 2 or len(clean) > 30:
        return ""
    if re.fullmatch(r"[\d.]+", clean):
        return ""
    if _line_is_count_marker(clean):
        return ""
    if not count_context and _split_candidate_names_with_parenthetical_notes(clean):
        return ""
    if re.search(r"(公告|公示|名单|专业名称|推免|推荐|共计|特此|学院|大学|本科生院|通知|工作|细则|年月日)", clean):
        return ""
    if re.search(r"(首页|导航|更多|下载|新闻|通知|招生|招聘|教师|教授|院士|讲坛|活动|中心|平台|制度|流程|校友)", clean):
        return ""
    return clean


def _line_is_count_marker(line: str) -> bool:
    clean = _clean_text(line).strip()
    return bool(re.fullmatch(r"[（(]?\s*\d+\s*人\s*[)）]?", clean))


def _line_has_inline_count_marker(line: str) -> bool:
    return bool(re.search(r"[（(]\s*\d+\s*人\s*[)）]$", _clean_text(line)))


def _structured_line_breaks_major_name_list(line: str) -> bool:
    clean = _clean_text(line)
    return bool(re.search(r"(特此|公告日期|发布日期|联系人|联系电话|邮箱|如有异议|公示期)", clean))


def _split_candidate_names_with_parenthetical_notes(line: str) -> list[str]:
    clean = _clean_text(line).strip("。；;，,、 ")
    if not clean:
        return []
    names: list[str] = []
    for token in re.split(r"[、，,；;\s]+", clean):
        token = _clean_text(token).strip("。；;，,、 ")
        token = re.sub(r"[（(][^）)]{1,30}[）)]$", "", token).strip()
        if not token:
            continue
        if not _looks_like_chinese_name(token):
            return []
        names.append(token)
    return names


def _major_and_candidate_names_from_paragraph(line: str) -> tuple[str, list[str]]:
    major, tail = _split_major_heading_and_tail(line)
    if major:
        return major, _split_candidate_names(tail)
    return _major_from_paragraph_heading(line), []


def _split_major_heading_and_tail(line: str) -> tuple[str, str]:
    if len(line) > 160:
        return "", ""
    match = re.match(
        r"^(?:[一二三四五六七八九十]+[、.．]\s*)?(.{2,40}?)(?:专业)?\s*[:：]\s*(.*)$",
        line,
    )
    if not match:
        return "", ""
    major = _normalize_major_heading(match.group(1))
    if not major:
        return "", ""
    return major, _clean_text(match.group(2))


def _major_from_paragraph_heading(line: str) -> str:
    if len(line) > 40:
        return ""
    match = re.search(
        r"^(?:[一二三四五六七八九十]+[、.．]\s*)?([\u4e00-\u9fffA-Za-z0-9（）()·]{2,30}?)(?:专业方向|专业)$",
        line,
    )
    if not match:
        colon_major, tail = _split_major_heading_and_tail(line)
        return colon_major if colon_major and not tail else ""
    return _normalize_major_heading(re.sub(r"(专业方向|专业)$", "", match.group(1)))


def _normalize_major_heading(value: str) -> str:
    major = _clean_text(value)
    major = re.sub(r"(专业方向|专业)$", "", major)
    major = major.strip(" :：")
    if not major:
        return ""
    if major in {"学院", "学校", "专业", "方向", "研究生", "组长", "副组长", "成员", "工作职责"}:
        return ""
    if re.search(r"(联系|地址|邮箱|电话|附件|上一篇|下一篇|当前位置|发布|浏览|公示|名单|情况|专项|异议|其中|作者|编辑|发布时间|点击次数|体检|常规检查|化验检查|必检项目)", major):
        return ""
    if len(major) > 30:
        return ""
    return major


def _paragraph_breaks_major_name_list(line: str) -> bool:
    if not line:
        return False
    if line.startswith(("其中", "如对", "若对", "以上", "附件", "联系人", "联系地址")):
        return True
    return bool(re.search(r"(专项|入选情况|异议|反馈|邮箱|电话|教务办|上一篇|下一篇)", line))


def _split_candidate_names(line: str) -> list[str]:
    line = _clean_text(line).strip("。；;，,、 ")
    if not line:
        return []
    names: list[str] = []
    tokens = [token for token in re.split(r"[、，,；;\s]+", line) if token]
    if not tokens:
        return []

    index = 0
    while index < len(tokens):
        token = _clean_text(tokens[index]).strip("。.")
        if re.fullmatch(r"[\u4e00-\u9fff]", token) and index + 1 < len(tokens):
            next_token = _clean_text(tokens[index + 1]).strip("。.")
            if re.fullmatch(r"[\u4e00-\u9fff]", next_token):
                names.append(f"{token}{next_token}")
                index += 2
                continue
        if _looks_like_chinese_name(token):
            names.append(token)
            index += 1
            continue
        return []
    return names


def parse_excel_records(path: Path, document: dict[str, Any]) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []

    try:
        for sheet in workbook.worksheets:
            rows = [
                [_clean_text(cell) for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            admission_status_records = _records_from_recommendation_admission_status_rows(
                rows,
                document,
            )
            if admission_status_records:
                records.extend(admission_status_records)
                continue
            records.extend(_records_from_recommendation_status_name_columns(rows, document))
            records.extend(_records_from_table(rows, document))
    finally:
        workbook.close()

    return records


def parse_legacy_xls_records(path: Path, document: dict[str, Any]) -> list[dict[str, Any]]:
    converted_path = _convert_legacy_xls_to_xlsx(path)
    if not converted_path or not converted_path.exists():
        return []
    return parse_excel_records(converted_path, document)


def _convert_legacy_xls_to_xlsx(path: Path) -> Path | None:
    if os.name != "nt":
        return None
    input_path = path.resolve()
    output_path = input_path.with_suffix(".converted.xlsx")
    if output_path.exists() and output_path.stat().st_mtime >= input_path.stat().st_mtime:
        return output_path

    def ps_quote(value: str) -> str:
        return "'" + value.replace("'", "''") + "'"

    script = "\n".join(
        [
            f"$in = {ps_quote(str(input_path))}",
            f"$out = {ps_quote(str(output_path))}",
            "$excel = $null",
            "$wb = $null",
            "try {",
            "  $excel = New-Object -ComObject Excel.Application",
            "  $excel.Visible = $false",
            "  $excel.DisplayAlerts = $false",
            "  $wb = $excel.Workbooks.Open($in)",
            "  $wb.SaveAs($out, 51)",
            "} finally {",
            "  if ($wb -ne $null) { $wb.Close($false) | Out-Null }",
            "  if ($excel -ne $null) { $excel.Quit() | Out-Null }",
            "}",
            "Write-Output $out",
        ]
    )
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", script],
            check=False,
            capture_output=True,
            text=True,
            timeout=90,
        )
    except (OSError, subprocess.TimeoutExpired):
        return None
    if result.returncode != 0:
        return None
    return output_path if output_path.exists() else None


def _records_from_recommendation_status_name_columns(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if document.get("document_type") != "recommendation_exemption_list":
        return []

    for header_index, row in enumerate(rows[:5]):
        formal_columns = [
            (column_index, value)
            for column_index, value in enumerate(row)
            if _is_formal_recommendation_status_header(value)
        ]
        if not formal_columns:
            continue

        records: list[dict[str, Any]] = []
        for column_index, status_label in formal_columns:
            for values in rows[header_index + 1 :]:
                cell = values[column_index] if column_index < len(values) else ""
                for name in _split_candidate_names_with_parenthetical_notes(cell):
                    record = _build_record(
                        document,
                        {"person_name": name, "remarks": status_label},
                    )
                    if record:
                        records.append(record)
        if records:
            return records

    return []


def _is_formal_recommendation_status_header(value: str) -> bool:
    clean = _clean_text(value)
    if not clean or re.search(r"(候补|递补|备选|待定)", clean):
        return False
    return bool(re.search(r"(正式|拟获.*推免资格|获.*推免资格|取得.*推免资格|推免资格.*名单)", clean))


def parse_docx_records(path: Path, document: dict[str, Any]) -> list[dict[str, Any]]:
    try:
        from docx import Document
    except ModuleNotFoundError:
        return []

    doc = Document(path)
    records: list[dict[str, Any]] = []
    for table in doc.tables:
        rows = [
            [_clean_text(cell.text) for cell in row.cells]
            for row in table.rows
        ]
        records.extend(_records_from_table(rows, document))
    return records


def parse_pdf_records(path: Path, document: dict[str, Any]) -> list[dict[str, Any]]:
    text = extract_pdf_text(path)
    if not text:
        return []
    kmust_records = _records_from_kmust_postgraduate_pdf_text(text, document)
    if not kmust_records and _is_kmust_postgraduate_document(document) and path.exists():
        pypdf_text = _extract_pdf_text_with_pypdf(path)
        if pypdf_text and pypdf_text != text:
            kmust_records = _records_from_kmust_postgraduate_pdf_text(pypdf_text, document)
    if kmust_records:
        return kmust_records
    ynnu_records = _records_from_ynnu_postgraduate_pdf_text(text, document)
    if ynnu_records:
        return ynnu_records
    swjtu_records = _records_from_swjtu_recommendation_pdf_text(text, document)
    if swjtu_records:
        return swjtu_records
    rows = _rows_from_text_lines(text)
    swfu_records = _records_from_swfu_recommendation_pdf_rows(rows, document)
    if swfu_records:
        return swfu_records
    tjus_records = _records_from_tjus_recommendation_pdf_rows(rows, document)
    if tjus_records:
        return tjus_records
    hrbipe_records = _records_from_hrbipe_recommendation_pdf_rows(rows, document)
    if hrbipe_records:
        return hrbipe_records
    blcu_records = _records_from_blcu_recommendation_pdf_rows(rows, document)
    if blcu_records:
        return blcu_records
    smu_records = _records_from_smu_recommendation_pdf_rows(rows, document)
    if smu_records:
        return smu_records
    plain_recommendation_records = _records_from_plain_recommendation_pdf_name_lines(
        text,
        document,
    )
    if plain_recommendation_records:
        return plain_recommendation_records
    nbu_records = _records_from_nbu_recommendation_pdf_rows(rows, document)
    if nbu_records:
        return nbu_records
    jxnu_records = _records_from_jxnu_recommendation_pdf_rows(rows, document)
    if jxnu_records:
        return jxnu_records
    hainnu_records = _records_from_hainnu_postgraduate_pdf_rows(rows, document)
    if hainnu_records:
        return hainnu_records
    fjmu_records = _records_from_fjmu_recommendation_pdf_rows(rows, document)
    if fjmu_records:
        return fjmu_records
    fjtcm_records = _records_from_fjtcm_postgraduate_pdf_rows(rows, document)
    if fjtcm_records:
        return fjtcm_records
    cueb_records = _records_from_cueb_doctoral_pdf_rows(rows, document)
    if cueb_records:
        return cueb_records
    scnu_records = _records_from_scnu_postgraduate_pdf_rows(rows, document)
    if scnu_records:
        return scnu_records
    xupt_records = _records_from_xupt_recommendation_pdf_rows(rows, document)
    if xupt_records:
        return xupt_records
    syphu_records = _records_from_syphu_postgraduate_pdf_rows(rows, document)
    if syphu_records:
        return syphu_records
    scuec_records = _records_from_scuec_recommendation_pdf_rows(rows, document)
    if scuec_records:
        return scuec_records
    cqmu_records = _records_from_cqmu_postgraduate_pdf_rows(rows, document)
    if cqmu_records:
        return cqmu_records
    gender_school_college_score_records = (
        _records_from_gender_school_college_major_score_pdf_rows(rows, document)
    )
    if gender_school_college_score_records:
        return gender_school_college_score_records
    njtech_records = _records_from_njtech_adjustment_pdf_rows(rows, document)
    if njtech_records:
        return njtech_records
    hgu_records = _records_from_hgu_pdf_rows(rows, document)
    if hgu_records:
        return hgu_records
    gxmzu_records = _records_from_gxmzu_pdf_rows(rows, document)
    if gxmzu_records:
        return gxmzu_records
    gxu_records = _records_from_gxu_pdf_rows(rows, document)
    if gxu_records:
        return gxu_records
    if _is_gxu_document(document):
        return []
    hebut_records = _records_from_hebut_pdf_rows(rows, document)
    if hebut_records:
        return hebut_records
    jlu_records = _records_from_jlu_pdf_text_rows(rows, document)
    if jlu_records:
        return jlu_records
    dlu_records = _records_from_dlu_recommendation_pdf_rows(rows, document)
    if dlu_records:
        return dlu_records
    nxmu_records = _records_from_nxmu_postgraduate_pdf_rows(rows, document)
    if nxmu_records:
        return nxmu_records
    ujn_records = _records_from_ujn_postgraduate_pdf_rows(rows, document)
    if ujn_records:
        return ujn_records
    xzhmu_records = _records_from_xzhmu_score_only_pdf_rows(rows, document)
    if xzhmu_records:
        return xzhmu_records
    zstu_records = _records_from_zstu_postgraduate_pdf_rows(rows, document)
    if zstu_records:
        return zstu_records
    gmc_records = _records_from_gmc_postgraduate_pdf_rows(rows, document)
    if gmc_records:
        return gmc_records
    gxau_records = _records_from_gxau_postgraduate_pdf_rows(rows, document)
    if gxau_records:
        return gxau_records
    lzjtu_records = _records_from_lzjtu_postgraduate_pdf_rows(rows, document)
    if lzjtu_records:
        return lzjtu_records
    guet_records = _records_from_guet_postgraduate_pdf_rows(rows, document)
    if guet_records:
        return guet_records
    gxnu_records = _records_from_gxnu_postgraduate_pdf_rows(rows, document)
    if gxnu_records:
        return gxnu_records
    glut_records = _records_from_glut_split_major_pdf_rows(rows, document)
    if glut_records:
        return glut_records
    college_section_records = _records_from_college_section_pdf_text_rows(rows, document)
    if college_section_records:
        return college_section_records
    yangtzeu_records = _records_from_yangtzeu_pdf_rows(rows, document)
    if yangtzeu_records:
        return yangtzeu_records
    sxu_records = _records_from_sxu_pdf_rows(rows, document)
    if sxu_records:
        return sxu_records
    tyut_records = _records_from_tyut_recommendation_pdf_rows(rows, document)
    if tyut_records:
        return tyut_records
    lzu_law_records = _records_from_lzu_law_recommendation_pdf_rows(rows, document)
    if lzu_law_records:
        return lzu_law_records
    lnutcm_records = _records_from_lnutcm_recommendation_pdf_rows(rows, document)
    if lnutcm_records:
        return lnutcm_records
    dmu_records = _records_from_dmu_recommendation_pdf_rows(rows, document)
    if dmu_records:
        return dmu_records
    major_code_name_records = _records_from_major_code_name_pdf_rows(rows, document)
    if major_code_name_records:
        return major_code_name_records
    fudan_records = _records_from_fudan_postgraduate_pdf_rows(rows, document)
    if fudan_records:
        return fudan_records
    swupl_records = _records_from_swupl_postgraduate_pdf_rows(rows, document)
    if swupl_records:
        return swupl_records
    bucea_records = _records_from_bucea_postgraduate_pdf_rows(rows, document)
    if bucea_records:
        return bucea_records
    dlmu_records = _records_from_dlmu_recommendation_pdf_rows(rows, document)
    if dlmu_records:
        return dlmu_records
    cumt_records = _records_from_cumt_recommendation_pdf_rows(rows, document)
    if cumt_records:
        return cumt_records
    xzmu_records = _records_from_xzmu_recommendation_pdf_rows(rows, document)
    if xzmu_records:
        return xzmu_records
    college_major_score_records = _records_from_college_major_exam_name_score_rows(rows, document)
    if college_major_score_records:
        return college_major_score_records
    dhu_records = _records_from_dhu_postgraduate_pdf_rows(rows, document)
    if dhu_records:
        return dhu_records
    shiep_records = _records_from_shiep_score_only_pdf_rows(rows, document)
    if shiep_records:
        return shiep_records
    ncut_records = _records_from_ncut_retest_result_pdf_rows(rows, document)
    if ncut_records:
        return ncut_records
    henu_records = _records_from_henu_recommendation_pdf_rows(rows, document)
    if henu_records:
        return henu_records
    just_records = _records_from_just_retest_ranking_pdf_rows(rows, document)
    if just_records:
        return just_records
    gdufe_records = _records_from_gdufe_postgraduate_pdf_rows(rows, document)
    if gdufe_records:
        return gdufe_records
    sdufe_records = _records_from_sdufe_doctoral_pdf_rows(rows, document)
    if sdufe_records:
        return sdufe_records
    xmu_records = _records_from_xmu_sectioned_postgraduate_pdf_rows(rows, document)
    if xmu_records:
        return xmu_records
    records = _records_from_table(rows, document)
    records.extend(_records_from_college_code_first_text_rows(rows, document))
    records.extend(_records_from_zjut_pdf_text_rows(rows, document))
    return records


def _records_from_xmu_sectioned_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    school_name = str(document.get("school_name") or "")
    source_url = str(document.get("source_url") or "")
    title = str(document.get("title") or "")
    if (
        "厦门大学" not in school_name
        and "xmu.edu.cn" not in source_url
        and "厦门大学" not in title
    ):
        return []

    document_text = " ".join(" ".join(row) for row in rows[:12])
    if "专业代码和专业名称" not in document_text:
        return []

    college = _xmu_college_name_from_context(rows, document)
    current_major_code = ""
    current_major_name = ""
    records: list[dict[str, Any]] = []
    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if not values:
            continue
        row_text = _clean_text(" ".join(values))
        major_match = re.search(r"专业代码和专业名称[:：]\s*([0-9A-Z]{6})\s*(.+)", row_text)
        if major_match:
            current_major_code = major_match.group(1)
            current_major_name = _clean_text(major_match.group(2))
            continue
        if not current_major_code or len(values) < 7 or not re.fullmatch(r"\d+", values[0]):
            continue

        candidate_values = list(values)
        if len(candidate_values) >= 2:
            identity_match = re.fullmatch(r"([0-9Xx]{10,})\s+([\u4e00-\u9fff]{2,5}(?:·[\u4e00-\u9fff]{1,4})?)", candidate_values[1])
            if identity_match:
                candidate_values = [
                    candidate_values[0],
                    identity_match.group(1),
                    identity_match.group(2),
                    *candidate_values[2:],
                ]
        if len(candidate_values) < 8:
            continue

        ranking, student_id, person_name = candidate_values[:3]
        initial_score, reexam_score, total_score = candidate_values[3:6]
        study_mode = candidate_values[6]
        admission_category = candidate_values[7]
        extra_remarks = _clean_text(" ".join(candidate_values[8:]))
        if (
            not _looks_like_identifier_only(student_id)
            or not _looks_like_chinese_name(person_name)
            or not _looks_like_score_or_metric(initial_score)
            or not _looks_like_score_or_metric(reexam_score)
            or not _looks_like_score_or_metric(total_score)
        ):
            continue

        remarks = [
            f"initial_score {initial_score}",
            f"reexam_score {reexam_score}",
            f"total_score {total_score}",
            f"study_mode {study_mode}",
            f"admission_category {admission_category}",
        ]
        if extra_remarks:
            remarks.append(f"remark {extra_remarks}")

        record = _build_record(
            document,
            {
                "ranking": ranking,
                "student_id": student_id,
                "person_name": person_name,
                "college": college,
                "major": current_major_code,
                "admission_major": _clean_text(f"{current_major_code} {current_major_name}"),
                "remarks": "; ".join(remarks),
            },
        )
        if record:
            records.append(record)
    return records


def _xmu_college_name_from_context(
    rows: list[list[str]],
    document: dict[str, Any],
) -> str:
    contexts = [" ".join(" ".join(row) for row in rows[:5]), str(document.get("title") or "")]
    for context in contexts:
        text = _clean_text(context)
        match = re.search(r"([\u4e00-\u9fff、与]+(?:学院|研究院)(?:[\u4e00-\u9fff、与]*(?:学院|研究院))*)\s*2026年", text)
        if match:
            return match.group(1)
    return ""


def _records_from_jxnu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    school_name = str(document.get("school_name") or "")
    source_url = str(document.get("source_url") or "")
    title = str(document.get("title") or "")
    if (
        "江西师范大学" not in school_name
        and "jxnu.edu.cn" not in source_url
        and "江西师范大学" not in title
    ):
        return []

    header_text = " ".join(" ".join(row) for row in rows[:10])
    if "推免生接收名单" in header_text or all(
        term in header_text for term in ("录取院系所", "考生编号", "姓名", "录取专业名称")
    ):
        return _records_from_jxnu_incoming_recommendation_rows(rows, document)
    if "推荐免试攻读研究生资格名单" in header_text or all(
        term in header_text for term in ("学院名称", "专业代码", "专业名称", "姓名", "综合成绩")
    ):
        return _records_from_jxnu_recommendation_qualification_rows(rows, document)
    return []


def _records_from_jxnu_incoming_recommendation_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if len(values) < 7:
            continue
        college_code, college_name, student_id, person_name, gender, major_code, major_name = values[:7]
        remark = _clean_text(" ".join(values[7:]))
        if (
            not re.fullmatch(r"\d{3}", college_code)
            or not _looks_like_identifier_only(student_id)
            or not _looks_like_chinese_name(person_name)
            or gender not in {"男", "女"}
            or not re.fullmatch(r"\d{6}[A-Z]?", major_code)
        ):
            continue
        remarks = [f"gender {gender}"]
        if remark:
            if re.search(r"(直博生|硕士|博士)", remark):
                remarks.append(f"admission_type {remark}")
            else:
                remarks.append(f"remark {remark}")

        record = _build_record(
            document,
            {
                "student_id": student_id,
                "person_name": person_name,
                "college": _clean_text(f"{college_code} {college_name}"),
                "major": major_code,
                "admission_major": _clean_text(f"{major_code} {major_name}"),
                "remarks": "; ".join(remarks),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_jxnu_recommendation_qualification_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if len(values) < 7:
            continue
        college_code, college_name, major_code, major_name, person_name, gender, composite_score = values[:7]
        if (
            not re.fullmatch(r"\d{3}", college_code)
            or not re.fullmatch(r"\d{6}[A-Z]?", major_code)
            or not _looks_like_chinese_name(person_name)
            or gender not in {"男", "女"}
            or not _looks_like_score_or_metric(composite_score)
        ):
            continue
        record = _build_record(
            document,
            {
                "person_name": person_name,
                "college": _clean_text(f"{college_code} {college_name}"),
                "major": major_code,
                "undergraduate_major": _clean_text(f"{major_code} {major_name}"),
                "remarks": _clean_text(f"gender {gender}; composite_score {composite_score}"),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_nbu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    school_name = str(document.get("school_name") or "")
    source_url = str(document.get("source_url") or "")
    title = str(document.get("title") or "")
    if (
        "宁波大学" not in school_name
        and "nbu.edu.cn" not in source_url
        and "宁波大学" not in title
    ):
        return []

    header_text = " ".join(" ".join(row) for row in rows[:10])
    required_terms = ("姓名", "性别", "复试成绩", "接收学院", "接收专业代码", "接收专业名称", "录取类型")
    if not all(term in header_text for term in required_terms):
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if len(values) < 7:
            continue
        if len(values) > 7:
            values = values[:6] + [" ".join(values[6:])]
        person_name, gender, retest_score, college, major_code, major_name, admission_type = values[:7]
        if (
            not _looks_like_chinese_name(person_name)
            or gender not in {"男", "女"}
            or not _looks_like_score_or_metric(retest_score)
            or not re.fullmatch(r"\d{6}[A-Z]?", major_code)
        ):
            continue

        record = _build_record(
            document,
            {
                "person_name": person_name,
                "college": college,
                "major": major_code,
                "admission_major": _clean_text(f"{major_code} {major_name}"),
                "remarks": _clean_text(
                    f"gender {gender}; retest_score {retest_score}; admission_type {admission_type}"
                ),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_hainnu_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if "海南师范大学" not in str(document.get("school_name") or ""):
        return []
    header_text = " ".join(" ".join(row) for row in rows[:8])
    if not all(term in header_text for term in ("拟录取", "学院代码", "专业代码", "专业名称")):
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        values = [_clean_text(value) for value in values if _clean_text(value)]
        if len(values) < 11 or not re.fullmatch(r"\d+", values[0]) or not _looks_like_identifier_only(values[1]):
            continue
        if not _looks_like_chinese_name(values[2]):
            continue
        college_code, college_name = values[3], values[4]
        major_code, major_name = values[5], values[6]
        if not re.fullmatch(r"\d{3}", college_code) or not re.fullmatch(r"\d{6}", major_code):
            continue
        record = _build_record(
            document,
            {
                "ranking": values[0],
                "student_id": values[1],
                "person_name": values[2],
                "college": _clean_text(f"{college_code} {college_name}"),
                "major": major_code,
                "admission_major": _clean_text(f"{major_code} {major_name}"),
                "remarks": _clean_text(
                    " ".join(
                        [
                            f"study_mode {values[7]}",
                            f"initial_score {values[8]}",
                            f"reexam_score {values[9]}",
                            f"total_score {values[10]}",
                            *values[11:],
                        ]
                    )
                ),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_fjmu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if "福建医科大学" not in str(document.get("school_name") or ""):
        return []
    header_text = " ".join(" ".join(row) for row in rows[:8])
    if not all(term in header_text for term in ("推免生拟录取名单", "拟录取学院", "拟录取专业")):
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        values = [_clean_text(value) for value in values if _clean_text(value)]
        if len(values) < 8 or not _looks_like_identifier_only(values[0]):
            continue
        student_id, person_name, reexam_score, degree_type = values[:4]
        if not _looks_like_chinese_name(person_name) or not _looks_like_score_or_metric(reexam_score):
            continue
        if not re.fullmatch(r"\d{3}", values[4]) or not re.fullmatch(r"\d{6}[A-Z]?", values[6]):
            continue
        direction = ""
        advisor = ""
        if len(values) >= 10:
            direction = values[8]
            advisor = values[9]
        elif len(values) >= 9:
            advisor = values[8]
        record = _build_record(
            document,
            {
                "student_id": student_id,
                "person_name": person_name,
                "college": _clean_text(f"{values[4]} {values[5]}"),
                "major": values[6],
                "admission_major": _clean_text(f"{values[6]} {values[7]}"),
                "remarks": _clean_text(
                    " ".join(
                        [
                            degree_type,
                            f"reexam_score {reexam_score}",
                            f"direction {direction}" if direction else "",
                            f"advisor {advisor}" if advisor else "",
                        ]
                    )
                ),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_fjtcm_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if "福建中医药大学" not in str(document.get("school_name") or ""):
        return []
    header_text = " ".join(" ".join(row) for row in rows[:12])
    if not all(term in header_text for term in ("拟录取统考硕士", "招生学院", "招生专业")):
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        values = [_clean_text(value) for value in values if _clean_text(value)]
        record = _fjtcm_postgraduate_pdf_record(values, document)
        if record:
            records.append(record)
    return records


def _fjtcm_postgraduate_pdf_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 6 or not re.fullmatch(r"\d+", values[0]):
        return None
    if values[-1] != "是":
        return None
    sequence = values[0]
    fields = values[1:-1]
    parsed = _fjtcm_parse_program_fields(fields)
    if not parsed:
        return None
    college, degree_type, major_code, major_name, cursor = parsed
    trailing_person = _fjtcm_split_trailing_person_from_major_name(major_name)
    if trailing_person and cursor < len(fields) and _looks_like_identifier_only(fields[cursor]):
        major_name, person_name = trailing_person
        student_id = fields[cursor]
        cursor += 1
    else:
        person_student = _fjtcm_parse_person_student(fields, cursor)
        if not person_student:
            return None
        person_name, student_id, cursor = person_student
    if not _looks_like_chinese_name(person_name) or not _looks_like_identifier_only(student_id):
        return None

    rest = fields[cursor:]
    score_values = _score_values_from_fragments(rest)
    remarks = [degree_type]
    advisor = _fjtcm_advisor_text(rest)
    if advisor:
        remarks.append(f"advisor {advisor}")
    if len(score_values) >= 1:
        remarks.append(f"initial_score {score_values[0]}")
    if len(score_values) >= 2:
        remarks.append(f"reexam_score {score_values[1]}")
    if len(score_values) >= 3:
        remarks.append(f"total_score {score_values[2]}")
    record = _build_record(
        document,
        {
            "ranking": sequence,
                "student_id": student_id,
                "person_name": person_name,
                "college": college,
                "major": major_code,
                "admission_major": _clean_text(f"{major_code} {major_name}") if major_name else major_code,
                "remarks": _clean_text(" ".join(remarks)),
            },
        )
    return record


def _fjtcm_parse_program_fields(values: list[str]) -> tuple[str, str, str, str, int] | None:
    if not values:
        return None

    if len(values) >= 2 and values[0] in {"学术型", "专业学位型"}:
        major = _split_fjtcm_major_code_name(values[1])
        if major:
            major_code, major_name = major
            cursor = 2
            if not major_name and cursor < len(values):
                next_value = values[cursor]
                next_next = values[cursor + 1] if cursor + 1 < len(values) else ""
                if not (_looks_like_chinese_name(next_value) and _looks_like_identifier_only(next_next)):
                    major_name = next_value
                    cursor += 1
            return "", values[0], major_code, major_name, cursor

    parsed_program = _fjtcm_split_degree_program(values[0])
    if parsed_program:
        college, degree_type, major_code, major_name = parsed_program
        return college, degree_type, major_code, major_name, 1

    if len(values) >= 2:
        parsed_program = _fjtcm_split_degree_program(values[1])
        if parsed_program and not parsed_program[0]:
            _, degree_type, major_code, major_name = parsed_program
            return values[0], degree_type, major_code, major_name, 2

        college_degree = _fjtcm_split_college_degree(values[0])
        major = _split_fjtcm_major_code_name(values[1])
        if college_degree and major:
            college, degree_type = college_degree
            major_code, major_name = major
            return college, degree_type, major_code, major_name, 2

    if len(values) >= 3 and values[1] in {"学术型", "专业学位型"}:
        college = values[0]
        degree_type = values[1]
        major = _split_fjtcm_major_code_name(values[2])
        if not major:
            return None
        major_code, major_name = major
        cursor = 3
        if not major_name and cursor < len(values):
            next_value = values[cursor]
            next_next = values[cursor + 1] if cursor + 1 < len(values) else ""
            if not (_looks_like_chinese_name(next_value) and _looks_like_identifier_only(next_next)):
                major_name = next_value
                cursor += 1
        return college, degree_type, major_code, major_name, cursor

    return None


def _fjtcm_split_degree_program(value: str) -> tuple[str, str, str, str] | None:
    match = re.fullmatch(
        rf"(?:(.+?)\s+)?(学术型|专业学位型)\s+({_FJTCM_MAJOR_CODE_PATTERN})(?:\s+(.+))?",
        _clean_text(value),
    )
    if not match:
        return None
    return (
        _clean_text(match.group(1) or ""),
        match.group(2),
        match.group(3),
        _clean_text(match.group(4) or ""),
    )


def _fjtcm_split_college_degree(value: str) -> tuple[str, str] | None:
    match = re.fullmatch(r"(.+?)\s+(学术型|专业学位型)", _clean_text(value))
    if not match:
        return None
    return _clean_text(match.group(1)), match.group(2)


def _fjtcm_parse_person_student(values: list[str], cursor: int) -> tuple[str, str, int] | None:
    if cursor >= len(values):
        return None
    combined = re.fullmatch(r"(.+?)\s+([0-9Xx*]{6,})", values[cursor])
    if combined:
        return _clean_text(combined.group(1)), combined.group(2), cursor + 1
    if cursor + 1 >= len(values):
        return None
    return values[cursor], values[cursor + 1], cursor + 2


def _fjtcm_split_trailing_person_from_major_name(major_name: str) -> tuple[str, str] | None:
    if " " not in major_name:
        return None
    major_part, person_part = major_name.rsplit(" ", 1)
    if not major_part or not _looks_like_chinese_name(person_part):
        return None
    return _clean_text(major_part), person_part


def _fjtcm_advisor_text(values: list[str]) -> str:
    parts: list[str] = []
    for value in values:
        if re.search(r"\d{2,3}(?:\.\d+)?", value or ""):
            break
        if value in {"录取后选", "报导师", "院", "队", "学"}:
            continue
        parts.append(value)
    return _clean_text(" ".join(parts))


_FJTCM_MAJOR_CODE_PATTERN = r"(?:\d{6}[A-Z]?|\d{4}[A-Z]\d)"


def _split_fjtcm_major_code_name(value: str) -> tuple[str, str] | None:
    match = re.fullmatch(rf"({_FJTCM_MAJOR_CODE_PATTERN})(?:\s+(.+))?", _clean_text(value))
    if not match:
        return None
    return match.group(1), _clean_text(match.group(2) or "")


def _records_from_cueb_doctoral_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if "首都经济贸易大学" not in str(document.get("school_name") or ""):
        return []
    header_text = " ".join(" ".join(row) for row in rows[:8])
    if "博士研究生拟录取名单" not in header_text:
        return []

    records: list[dict[str, Any]] = []
    index = 0
    while index < len(rows):
        values = [_clean_text(value) for value in rows[index] if _clean_text(value)]
        if not values:
            index += 1
            continue
        direct_record = _cueb_doctoral_direct_record(values, document)
        if direct_record:
            records.append(direct_record)
            index += 1
            continue
        if (
            len(values) == 2
            and index + 2 < len(rows)
            and _cueb_doctoral_wrapped_data_row(rows[index + 1])
            and len([value for value in rows[index + 2] if _clean_text(value)]) == 2
        ):
            suffix_values = [_clean_text(value) for value in rows[index + 2] if _clean_text(value)]
            record = _cueb_doctoral_wrapped_record(values, rows[index + 1], suffix_values, document)
            if record:
                records.append(record)
                index += 3
                continue
        index += 1
    return records


def _cueb_doctoral_direct_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 9 or not re.fullmatch(r"\d+", values[0]) or not _looks_like_identifier_only(values[2]):
        return None
    if "拟录取" not in values[-1]:
        return None
    return _build_record(
        document,
        {
            "ranking": values[0],
            "college": values[1],
            "student_id": values[2],
            "person_name": values[3],
            "admission_major": values[4],
            "remarks": _clean_text(
                " ".join([values[5], f"advisor {values[6]}", f"total_score {values[7]}", values[8]])
            ),
        },
    )


def _cueb_doctoral_wrapped_data_row(row: list[str]) -> bool:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    return len(values) >= 7 and re.fullmatch(r"\d+", values[0] or "") and _looks_like_identifier_only(values[1] or "")


def _cueb_doctoral_wrapped_record(
    prefix_values: list[str],
    row: list[str],
    suffix_values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    if len(values) < 7 or "拟录取" not in values[-1]:
        return None
    return _build_record(
        document,
        {
            "ranking": values[0],
            "college": _clean_text(f"{prefix_values[0]}{suffix_values[0]}"),
            "student_id": values[1],
            "person_name": values[2],
            "admission_major": _clean_text(f"{prefix_values[1]}{suffix_values[1]}"),
            "remarks": _clean_text(
                " ".join([values[3], f"advisor {values[4]}", f"total_score {values[5]}", values[6]])
            ),
        },
    )


def _records_from_scnu_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if "华南师范大学" not in str(document.get("school_name") or ""):
        return []
    header_text = " ".join(" ".join(row) for row in rows[:8])
    if not all(term in header_text for term in ("复试及拟录取情况汇总表", "拟录取专业")):
        return []

    college = _scnu_pdf_college(rows, document)
    records: list[dict[str, Any]] = []
    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        record = _scnu_postgraduate_pdf_record(values, college, document)
        if record:
            records.append(record)
    return records


def _scnu_pdf_college(rows: list[list[str]], document: dict[str, Any]) -> str:
    for row in rows[:12]:
        text = _clean_text(" ".join(row))
        match = re.search(
            r"招生单位名称(?:（盖章）)?[:：]\s*(.+?)(?:[（(]\d{3}[）)]|\s+制表日期|$)",
            text,
        )
        if match:
            return _clean_text(match.group(1))
    title = str(document.get("title") or "")
    match = re.match(r"\d{3}\s*(.+?)\s*20\d{2}年", title)
    return _clean_text(match.group(1)) if match else ""


def _scnu_postgraduate_pdf_record(
    values: list[str],
    college: str,
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if (
        len(values) < 10
        or not re.fullmatch(r"\d+", values[0])
        or not _looks_like_identifier_only(values[1])
        or not re.fullmatch(r"[\u4e00-\u9fff·]{2,12}", values[2])
    ):
        return None

    major_index = next(
        (
            index
            for index, value in enumerate(values[3:], start=3)
            if _looks_like_scnu_major_code_name(value)
        ),
        None,
    )
    if major_index is None:
        return None

    admission_major = values[major_index]
    suffix = values[major_index + 1 :]
    embedded_study_mode = re.fullmatch(r"(.+?)\s+(全日制|非全日制)", admission_major)
    if embedded_study_mode:
        admission_major = _clean_text(embedded_study_mode.group(1))
        suffix = [embedded_study_mode.group(2), *suffix]

    prefix = values[3:major_index]
    exam_method = prefix[0] if prefix else ""
    adjustment_category = prefix[1] if len(prefix) > 1 else ""
    special_plan_parts = [
        value
        for value in prefix[2:]
        if not _looks_like_scnu_score_value(value) and value not in {"无"}
    ]
    score_values = [value for value in prefix[2:] if _looks_like_scnu_score_value(value)]
    score_labels = ("初试成绩", "复试成绩", "最终成绩")
    score_parts = list(zip(score_labels[-len(score_values) :], score_values))

    study_mode = suffix[0] if len(suffix) > 0 else ""
    admission_category = suffix[1] if len(suffix) > 1 else ""
    admission_status = suffix[2] if len(suffix) > 2 else ""
    extra_remarks = " ".join(suffix[3:]) if len(suffix) > 3 else ""
    remarks = _join_labeled_parts(
        [
            ("考试方式", exam_method),
            ("调剂类别", adjustment_category),
            ("专项计划", " ".join(special_plan_parts)),
            *score_parts,
            ("学习方式", study_mode),
            ("录取类别", admission_category),
            ("是否拟录取", admission_status),
            ("备注", extra_remarks),
        ]
    )
    return _build_record(
        document,
        {
            "ranking": values[0],
            "student_id": values[1],
            "person_name": values[2],
            "college": college,
            "admission_major": admission_major,
            "remarks": remarks,
        },
    )


def _looks_like_scnu_major_code_name(value: str) -> bool:
    return bool(re.fullmatch(r"[（(]\d{6}[）)].+", value or ""))


def _looks_like_scnu_score_value(value: str) -> bool:
    return bool(re.fullmatch(r"\d{1,3}(?:\.\d+)?", value or ""))


def _records_from_xupt_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if "西安邮电大学" not in str(document.get("school_name") or ""):
        return []
    header_text = " ".join(" ".join(row) for row in rows[:8])
    if not all(term in header_text for term in ("推荐免试研究生拟录取名单", "后四位", "拟录取学院")):
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if (
            len(values) < 5
            or not _looks_like_chinese_name(values[0])
            or not re.fullmatch(r"[0-9Xx]{4}", values[1])
            or not re.fullmatch(r"\d{6}", values[3])
        ):
            continue
        records.append(
            _build_record(
                document,
                {
                    "person_name": values[0],
                    "student_id": values[1],
                    "college": values[2],
                    "major": values[3],
                    "admission_major": _clean_text(f"{values[3]} {values[4]}"),
                    "remarks": "证件后四位",
                },
            )
        )
    return [record for record in records if record]


def _records_from_syphu_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if "沈阳药科大学" not in str(document.get("school_name") or ""):
        return []
    header_text = " ".join(" ".join(row) for row in rows[:8])
    if not all(term in header_text for term in ("考生编号", "复试", "拟录取")):
        return []

    records: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if not values:
            continue

        record = _syphu_wrapped_record_from_rows(rows, index, document)
        if record:
            records.append(record)
            continue

        record = _syphu_record_from_values(values, document)
        if record:
            records.append(record)
    return records


def _syphu_wrapped_record_from_rows(
    rows: list[list[str]],
    index: int,
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in rows[index] if _clean_text(value)]
    if (
        index == 0
        or index + 1 >= len(rows)
        or len(values) < 10
        or not _looks_like_identifier_only(values[0])
        or not _looks_like_chinese_name(values[1])
    ):
        return None

    previous = [_clean_text(value) for value in rows[index - 1] if _clean_text(value)]
    following = [_clean_text(value) for value in rows[index + 1] if _clean_text(value)]
    if len(previous) < 2 or len(following) < 3:
        return None

    batch = _clean_text(f"{previous[0].split()[0]}{following[0]}")
    college_prefix = _clean_text(" ".join(previous[0].split()[1:]))
    college = _clean_text(f"{college_prefix}{following[1]}")
    direction = _clean_text(f"{previous[1]}{following[2]}")
    if not (
        re.fullmatch(r"复试(?:一志愿|调剂(?:第?[一二三四五六七八九十]+))批次", batch)
        and re.fullmatch(r"\d{3}【.+】", college)
        and re.fullmatch(r"\d{6}【.+】", values[2])
    ):
        return None

    score_values = values[3:8]
    if not all(_looks_like_syphu_score_value(value) for value in score_values):
        return None
    status_text = _clean_text(" ".join(values[9:]))
    return _syphu_build_record(
        document,
        student_id=values[0],
        person_name=values[1],
        batch=batch,
        college=college,
        admission_major=_clean_text(f"{values[2]} {direction}"),
        score_values=score_values,
        ranking=values[8],
        status_text=status_text,
    )


def _syphu_record_from_values(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    text = _clean_text(" ".join(values))
    match = re.match(
        r"^(?P<student_id>\d{10,})\s+(?P<person_name>[\u4e00-\u9fff·]{2,20})\s+(?P<tail>.+)$",
        text,
    )
    if not match:
        return None

    tail = match.group("tail")
    batch_match = re.search(
        r"复试(?:一志愿|调剂(?:第?[一二三四五六七八九十]+))批次",
        tail,
    )
    if not batch_match:
        return None

    batch = batch_match.group(0)
    rest = _clean_text(tail[batch_match.end() :])
    college_match = re.match(r"^(?P<college>\d{3}【[^】]+】)\s*(?P<rest>.+)$", rest)
    if not college_match:
        return None

    college = college_match.group("college")
    rest = college_match.group("rest")
    score_match = re.match(
        r"^(?P<major>.+?)\s+(?P<scores>-?\d+(?:\.\d+)?(?:\s+-?\d+(?:\.\d+)?){4})(?P<tail>(?:\s+.+)?)$",
        rest,
    )
    if not score_match:
        return None

    admission_major = _clean_text(score_match.group("major"))
    score_values = score_match.group("scores").split()
    ranking, status_text = _syphu_split_ranking_and_status(score_match.group("tail"))
    if not (
        admission_major
        and re.search(r"\d{6}【.+】", admission_major)
        and all(_looks_like_syphu_score_value(value) for value in score_values)
    ):
        return None

    return _syphu_build_record(
        document,
        student_id=match.group("student_id"),
        person_name=match.group("person_name"),
        batch=batch,
        college=college,
        admission_major=admission_major,
        score_values=score_values,
        ranking=ranking,
        status_text=status_text,
    )


def _syphu_build_record(
    document: dict[str, Any],
    *,
    student_id: str,
    person_name: str,
    batch: str,
    college: str,
    admission_major: str,
    score_values: list[str],
    ranking: str,
    status_text: str,
) -> dict[str, Any] | None:
    retest_result = _syphu_retest_result(status_text)
    admission_opinion = _syphu_admission_opinion(status_text)
    admission_category = _syphu_admission_category(status_text)
    remarks = _join_labeled_parts(
        [
            ("复试批次", batch),
            ("初试总分", score_values[0] if len(score_values) > 0 else ""),
            ("专业考核", score_values[1] if len(score_values) > 1 else ""),
            ("综合素质考核", score_values[2] if len(score_values) > 2 else ""),
            ("复试总分", score_values[3] if len(score_values) > 3 else ""),
            ("总分", score_values[4] if len(score_values) > 4 else ""),
            ("复试结果", retest_result),
            ("录取意见", admission_opinion),
            ("录取类别", admission_category),
            ("备注", status_text),
        ]
    )
    major_code_match = re.search(r"(\d{6})【", admission_major)
    return _build_record(
        document,
        {
            "student_id": student_id,
            "person_name": person_name,
            "college": college,
            "major": major_code_match.group(1) if major_code_match else "",
            "admission_major": admission_major,
            "ranking": ranking,
            "remarks": remarks,
        },
    )


def _syphu_split_ranking_and_status(value: str) -> tuple[str, str]:
    text = _clean_text(value)
    if not text:
        return "", ""
    match = re.match(r"^(?P<ranking>\d+)\s*(?P<status>.*)$", text)
    if match:
        return match.group("ranking"), _clean_text(match.group("status") or "")
    return "", text


def _looks_like_syphu_score_value(value: str) -> bool:
    return bool(re.fullmatch(r"-?\d{1,3}(?:\.\d+)?", value or ""))


def _syphu_retest_result(status_text: str) -> str:
    if "缺考" in status_text:
        return "缺考"
    if "复试不合格" in status_text:
        return "复试不合格"
    if "复试合格" in status_text:
        return "复试合格"
    return ""


def _syphu_admission_opinion(status_text: str) -> str:
    if "放弃录取" in status_text:
        return "放弃录取"
    if "拟录取" in status_text:
        return "拟录取"
    return ""


def _syphu_admission_category(status_text: str) -> str:
    match = re.search(r"(全日制|非全日制)(?:非定向|定向)", status_text)
    return match.group(0) if match else ""


def _records_from_scuec_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if "中南民族大学" not in str(document.get("school_name") or ""):
        return []
    header_text = " ".join(" ".join(row) for row in rows[:16])
    if not all(term in header_text for term in ("姓名", "拟录取专业代码", "复试成绩")):
        return []

    records: list[dict[str, Any]] = []
    index = 0
    while index < len(rows):
        values = [_clean_text(value) for value in rows[index] if _clean_text(value)]
        if not values:
            index += 1
            continue

        wrapped_record = _scuec_wrapped_recommendation_record(rows, index, document)
        if wrapped_record:
            records.append(wrapped_record)
            index += 3
            continue

        record = _scuec_recommendation_record(values, document)
        if record:
            records.append(record)
        index += 1

    return records


def _scuec_wrapped_recommendation_record(
    rows: list[list[str]],
    index: int,
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if index + 2 >= len(rows):
        return None
    prefix_values = [_clean_text(value) for value in rows[index] if _clean_text(value)]
    current_values = [_clean_text(value) for value in rows[index + 1] if _clean_text(value)]
    suffix_values = [_clean_text(value) for value in rows[index + 2] if _clean_text(value)]
    if not (
        len(prefix_values) == 1
        and len(current_values) >= 3
        and len(suffix_values) == 1
        and _looks_like_chinese_name(current_values[0])
        and re.fullmatch(r"\d{6}", current_values[1])
        and _looks_like_syphu_score_value(current_values[2])
    ):
        return None
    major_name = _clean_text(f"{prefix_values[0]}{suffix_values[0]}")
    if not re.search(r"[\u4e00-\u9fff]", major_name):
        return None
    remarks = _join_labeled_parts(
        [
            ("复试成绩", current_values[2]),
            ("备注", " ".join(current_values[3:])),
        ]
    )
    return _build_record(
        document,
        {
            "person_name": current_values[0],
            "major": current_values[1],
            "admission_major": major_name,
            "remarks": remarks,
        },
    )


def _scuec_recommendation_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if (
        len(values) < 4
        or not _looks_like_chinese_name(values[0])
        or not re.fullmatch(r"\d{6}", values[1])
        or _looks_like_syphu_score_value(values[2])
        or not _looks_like_syphu_score_value(values[3])
    ):
        return None
    remarks = _join_labeled_parts(
        [
            ("复试成绩", values[3]),
            ("备注", " ".join(values[4:])),
        ]
    )
    return _build_record(
        document,
        {
            "person_name": values[0],
            "major": values[1],
            "admission_major": values[2],
            "remarks": remarks,
        },
    )


def _records_from_cqmu_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if document.get("document_type") != "postgraduate_admission_list":
        return []
    school_name = str(document.get("school_name") or "")
    source_url = str(document.get("source_url") or "")
    if "重庆医科大学" not in school_name and "hospital-cqmu.com" not in source_url:
        return []
    header_text = " ".join(" ".join(row) for row in rows[:12])
    if not all(term in header_text for term in ("考生编号", "姓名", "是否拟录取")):
        return []

    records: list[dict[str, Any]] = []
    current_major_code = ""
    current_major_name = ""
    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if not values:
            continue

        line = _clean_text(" ".join(values))
        major_code_match = re.search(r"专业代码[:：]\s*([0-9A-Z]{5,})", line)
        major_name_match = re.search(r"专业名称（方向）[:：]\s*(.+?)(?:\s+复试日期[:：]|$)", line)
        if major_code_match:
            current_major_code = major_code_match.group(1)
        if major_name_match:
            current_major_name = _clean_text(major_name_match.group(1))

        record = _cqmu_postgraduate_score_record(
            values,
            document,
            current_major_code=current_major_code,
            current_major_name=current_major_name,
        )
        if record:
            records.append(record)
    return records


def _cqmu_postgraduate_score_record(
    values: list[str],
    document: dict[str, Any],
    *,
    current_major_code: str,
    current_major_name: str,
) -> dict[str, Any] | None:
    if (
        len(values) < 6
        or not _looks_like_identifier_only(values[0])
        or not _looks_like_chinese_name(values[1])
        or not all(_looks_like_syphu_score_value(value) for value in values[2:5])
    ):
        return None
    admission_status = values[5]
    if admission_status not in {"是", "拟录取"}:
        return None

    remarks = _join_labeled_parts(
        [
            ("初试成绩", values[2]),
            ("复试成绩", values[3]),
            ("总成绩", values[4]),
            ("是否拟录取", admission_status),
        ]
    )
    return _build_record(
        document,
        {
            "student_id": values[0],
            "person_name": values[1],
            "college": "第一临床学院",
            "major": current_major_code,
            "admission_major": _clean_text(f"{current_major_code} {current_major_name}"),
            "remarks": remarks,
        },
    )


def _score_values_from_fragments(values: list[str]) -> list[str]:
    scores: list[str] = []
    for value in values:
        scores.extend(re.findall(r"\d{2,3}(?:\.\d+)?", value or ""))
    return scores


def _records_from_kmust_postgraduate_pdf_text(
    text: str, document: dict[str, Any]
) -> list[dict[str, Any]]:
    if not _is_kmust_postgraduate_document(document, text):
        return []
    if "硕士研究生拟录取公示名单" not in text and "拟录取学院名称" not in text:
        return []

    compact = re.sub(r"\s+", " ", text.replace("\u3000", " ")).strip()
    row_starts = list(
        re.finditer(r"(?<!\d)(\d{1,5})\s+(\d{15})\s+(.+?)\s+(\d{3})\s+", compact)
    )
    if not row_starts:
        return []

    records: list[dict[str, Any]] = []
    for index, match in enumerate(row_starts):
        ranking, student_id, person_name, college_code = match.groups()
        end = row_starts[index + 1].start() if index + 1 < len(row_starts) else len(compact)
        chunk = compact[match.end() : end].strip()
        parsed = re.match(r"(.+?)\s+([0-9A-Z]{5,6})\s+(.+)", chunk)
        if not parsed:
            continue
        college_name, major_code, tail = parsed.groups()
        tail_match = re.match(
            r"(.+?)\s+(全日\s*制|非全日\s*制)\s+((?:非定向|定向)\s*就\s*业|非定向就业|定向就业)\s+(.+)",
            tail,
        )
        if not tail_match:
            continue
        major_name, study_mode, admission_category, score_text = tail_match.groups()
        score_values = re.findall(r"\d+(?:\.\d+)?", score_text)
        remarks_parts = [
            f"拟录取学院代码: {college_code}",
            f"拟录取学习方式: {_kmust_join_wrapped_text(study_mode)}",
            f"拟录取类别: {_kmust_join_wrapped_text(admission_category)}",
        ]
        score_labels = [
            "政治理论",
            "外国语",
            "业务课一",
            "业务课二",
            "初试成绩总分",
            "复试成绩",
            "综合成绩",
        ]
        for label, value in zip(score_labels, score_values):
            remarks_parts.append(f"{label}: {value}")
        trailing_note = re.sub(r"^(?:\d+(?:\.\d+)?\s*){1,7}", "", score_text).strip()
        if trailing_note:
            remarks_parts.append(f"备注: {_kmust_join_wrapped_text(trailing_note)}")

        records.append(
            {
                "school_name": document.get("school_name") or "昆明理工大学",
                "year": document.get("year"),
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "person_name": _kmust_join_wrapped_text(person_name),
                "student_id": student_id,
                "college": _kmust_join_wrapped_text(college_name),
                "major": major_code,
                "admission_major": _clean_text(
                    f"{major_code} {_kmust_join_wrapped_text(major_name)}"
                ),
                "ranking": ranking,
                "remarks": "; ".join(remarks_parts),
                "source_url": document.get("source_url"),
                "title": document.get("title"),
                "needs_review": False,
            }
        )
    return records


def _kmust_join_wrapped_text(value: str) -> str:
    text = _clean_text(value)
    text = re.sub(r"(?<=[\u4e00-\u9fff、，；：（）()])\s+(?=[\u4e00-\u9fff、，；：（）()])", "", text)
    text = re.sub(r"\s+([、，。；：）)])", r"\1", text)
    text = re.sub(r"([（(])\s+", r"\1", text)
    return text


def _is_kmust_postgraduate_document(
    document: dict[str, Any], text: str = ""
) -> bool:
    school_name = str(document.get("school_name") or "")
    source_url = str(document.get("source_url") or "")
    title = str(document.get("title") or "")
    haystack = f"{school_name} {source_url} {title} {text[:500]}"
    return "昆明理工大学" in haystack or "kmust.edu.cn" in haystack


def _records_from_ynnu_postgraduate_pdf_text(
    text: str, document: dict[str, Any]
) -> list[dict[str, Any]]:
    if not _is_ynnu_postgraduate_document(document, text):
        return []
    if "云南师范大学" not in text or "录取院系所名称" not in text:
        return []

    compact = re.sub(r"\s+", " ", text.replace("\u3000", " ")).strip()
    row_starts = list(re.finditer(r"(?<!\d)(\d{15})\s+", compact))
    if not row_starts:
        return []

    records: list[dict[str, Any]] = []
    for index, match in enumerate(row_starts):
        student_id = match.group(1)
        end = row_starts[index + 1].start() if index + 1 < len(row_starts) else len(compact)
        chunk = compact[match.end() : end].strip()
        parsed = _parse_ynnu_postgraduate_chunk(chunk)
        if not parsed:
            continue
        (
            person_name,
            college_code,
            college,
            major_code,
            major_name,
            study_mode,
            retest_score,
            total_score,
            trailing_note,
            needs_review,
        ) = parsed
        remarks_parts = [f"录取院系所码: {college_code}"]
        if study_mode:
            remarks_parts.append(f"录取学习方式: {study_mode}")
        if retest_score:
            remarks_parts.append(f"复试成绩: {retest_score}")
        if total_score:
            remarks_parts.append(f"综合成绩: {total_score}")
        if trailing_note:
            remarks_parts.append(f"备注: {trailing_note}")
        records.append(
            {
                "school_name": document.get("school_name") or "云南师范大学",
                "year": document.get("year"),
                "document_type": "postgraduate_admission_list",
                "route": "postgraduate_exam_or_admission",
                "person_name": _clean_text(person_name),
                "student_id": student_id,
                "college": _clean_text(college),
                "major": major_code,
                "admission_major": _clean_text(f"{major_code} {major_name}"),
                "ranking": "",
                "remarks": "; ".join(remarks_parts),
                "source_url": document.get("source_url"),
                "title": document.get("title"),
                "needs_review": needs_review,
            }
        )
    return records


def _parse_ynnu_postgraduate_chunk(
    chunk: str,
) -> tuple[str, str, str, str, str, str, str, str, str, bool] | None:
    row_match: re.Match[str] | None
    if re.match(r"\d{3}\s+", chunk):
        row_match = re.match(
            r"(?P<college_code>\d{3})\s+"
            r"(?P<name>.+?)\s+"
            r"(?P<college>.+?)\s+"
            r"(?P<major_code>[0-9A-Z]{5,6})\s+"
            r"(?P<tail>.+)",
            chunk,
        )
    else:
        row_match = re.match(
            r"(?P<name>.+?)\s+"
            r"(?P<college_code>\d{3})\s+"
            r"(?P<college>.+?)\s+"
            r"(?P<major_code>[0-9A-Z]{5,6})\s+"
            r"(?P<tail>.+)",
            chunk,
        )
    if not row_match:
        return None

    tail_match = re.match(
        r"(?P<major_name>.+?)\s+"
        r"(?P<study_mode>全日制|非全日制)\s+"
        r"(?P<retest_score>\d+(?:\.\d+)?)\s+"
        r"(?P<total_score>\d+(?:\.\d+)?)"
        r"(?:\s+(?P<note>.+))?$",
        row_match.group("tail"),
    )
    if not tail_match:
        partial_tail = _clean_text(row_match.group("tail"))
        if not partial_tail:
            return None
        return (
            row_match.group("name"),
            row_match.group("college_code"),
            row_match.group("college"),
            row_match.group("major_code"),
            partial_tail,
            "",
            "",
            "",
            "PDF文本层缺少学习方式或成绩，建议复核原文",
            True,
        )

    return (
        row_match.group("name"),
        row_match.group("college_code"),
        row_match.group("college"),
        row_match.group("major_code"),
        tail_match.group("major_name"),
        tail_match.group("study_mode"),
        tail_match.group("retest_score"),
        tail_match.group("total_score"),
        tail_match.group("note") or "",
        False,
    )


def _is_ynnu_postgraduate_document(
    document: dict[str, Any], text: str = ""
) -> bool:
    school_name = str(document.get("school_name") or "")
    source_url = str(document.get("source_url") or "")
    title = str(document.get("title") or "")
    haystack = f"{school_name} {source_url} {title} {text[:500]}"
    return "云南师范大学" in haystack or "ynnu.edu.cn" in haystack


_SWJTU_MASKED_ID_RE = r"\d{3}\*{4}[0-9Xx]{4}"


def _records_from_swjtu_recommendation_pdf_text(
    text: str, document: dict[str, Any]
) -> list[dict[str, Any]]:
    if not _is_swjtu_recommendation_document(document, text):
        return []
    if "录取院系所" not in text or "录取专业名称" not in text:
        return []

    records: list[dict[str, Any]] = []
    pending_identity_line = ""
    for raw_line in text.splitlines():
        line = _clean_text(raw_line)
        if not line:
            continue

        parsed = _parse_swjtu_recommendation_line(line)
        if parsed:
            records.append(_build_swjtu_recommendation_record(document, parsed))
            pending_identity_line = ""
            continue

        if re.fullmatch(
            rf"[\u4e00-\u9fff·]{{2,12}}\s+{_SWJTU_MASKED_ID_RE}",
            line,
        ):
            pending_identity_line = line
            continue

        if pending_identity_line and re.match(r"^(?:硕士|直博生?|博士)\s+", line):
            parsed = _parse_swjtu_recommendation_line(f"{pending_identity_line} {line}")
            if parsed:
                records.append(_build_swjtu_recommendation_record(document, parsed))
            pending_identity_line = ""
            continue

        if pending_identity_line:
            pending_identity_line = ""

    return records


def _parse_swjtu_recommendation_line(line: str) -> dict[str, str] | None:
    match = re.match(
        rf"^\s*(?P<person_name>[\u4e00-\u9fff·]{{2,12}})\s+"
        rf"(?P<student_id>{_SWJTU_MASKED_ID_RE})\s+"
        r"(?P<degree_type>硕士|直博生?|博士)\s+"
        r"(?P<college>.+?)\s+"
        r"(?P<major_code>[0-9A-Z]{4,6})\s+"
        r"(?P<major_name>.+?)\s+"
        r"(?P<retest_score>\d+(?:\.\d+)?)\s*$",
        line,
    )
    return match.groupdict() if match else None


def _build_swjtu_recommendation_record(
    document: dict[str, Any],
    parsed: dict[str, str],
) -> dict[str, Any]:
    major_code = parsed["major_code"]
    remarks = _join_labeled_parts(
        [
            ("招生类型", parsed["degree_type"]),
            ("复试成绩", parsed["retest_score"]),
        ]
    )
    return {
        "school_name": document.get("school_name") or "西南交通大学",
        "year": document.get("year"),
        "document_type": document.get("document_type") or "recommendation_exemption_list",
        "route": "recommendation_exemption",
        "person_name": parsed["person_name"],
        "student_id": parsed["student_id"],
        "college": _clean_text(parsed["college"]),
        "major": major_code,
        "admission_major": _clean_text(f"{major_code} {parsed['major_name']}"),
        "ranking": "",
        "remarks": remarks,
        "source_url": document.get("source_url"),
        "title": document.get("title"),
        "needs_review": False,
    }


def _is_swjtu_recommendation_document(
    document: dict[str, Any], text: str = ""
) -> bool:
    school_name = str(document.get("school_name") or "")
    source_url = str(document.get("source_url") or "")
    title = str(document.get("title") or "")
    haystack = f"{school_name} {source_url} {title} {text[:500]}"
    return "西南交通大学" in haystack or "swjtu.edu.cn" in haystack


def _records_from_swfu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = _document_and_row_haystack(document, rows)
    if "西南林业大学" not in document_text and "swfu.edu.cn" not in document_text:
        return []
    if "考生编号" not in document_text or "录取类别" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        values = [_clean_text(value) for value in values if _clean_text(value)]
        if len(values) < 10:
            continue
        sequence, student_id, person_name, degree_type = values[:4]
        college_code, college_name, major_code, major_name, study_mode = values[4:9]
        if not (
            sequence.isdigit()
            and _looks_like_identifier_only(student_id)
            and _looks_like_chinese_name(person_name)
            and re.fullmatch(r"\d{3}", college_code or "")
            and _looks_like_major_code(major_code)
        ):
            continue

        tail = _clean_text(" ".join(values[9:]))
        total_score = ""
        admission_category = ""
        score_match = re.search(r"\d+(?:\.\d+)?", tail)
        if score_match:
            total_score = score_match.group(0)
            admission_category = _clean_text(
                f"{tail[: score_match.start()]} {tail[score_match.end() :]}"
            )
        else:
            admission_category = tail

        record = _build_record(
            document,
            {
                "ranking": sequence,
                "student_id": student_id,
                "person_name": person_name,
                "college": _clean_text(f"{college_code} {college_name}"),
                "major": major_code,
                "admission_major": _clean_text(f"{major_code} {major_name}"),
                "remarks": _join_labeled_parts(
                    [
                        ("层次", degree_type),
                        ("学习形式", study_mode),
                        ("总成绩", total_score),
                        ("录取类别", admission_category),
                    ]
                ),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_tjus_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = _document_and_row_haystack(document, rows)
    if "天津体育学院" not in document_text and "tjus.edu.cn" not in document_text:
        return []
    if "身份证号" not in document_text or "复试分数" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        values = [_clean_text(value) for value in values if _clean_text(value)]
        if len(values) < 5:
            continue
        sequence, person_name, student_id, retest_score = values[:4]
        admission_major = _clean_text(" ".join(values[4:]))
        if not (
            sequence.isdigit()
            and _looks_like_chinese_name(person_name)
            and _looks_like_identifier_only(student_id)
            and _looks_like_score_or_metric(retest_score)
            and admission_major
        ):
            continue

        record = _build_record(
            document,
            {
                "ranking": sequence,
                "person_name": person_name,
                "student_id": student_id,
                "admission_major": admission_major,
                "remarks": _join_labeled_parts([("复试分数", retest_score)]),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_hrbipe_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = _document_and_row_haystack(document, rows)
    if "哈尔滨体育学院" not in document_text and "hrbipe.edu.cn" not in document_text:
        return []
    if "拟录取专业" not in document_text or "考生编号" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        values = [_clean_text(value) for value in values if _clean_text(value)]
        if len(values) < 8:
            continue
        sequence, person_name, admission_major, student_id = values[:4]
        foreign_language_score, theory_score, retest_score = values[4:7]
        note = _clean_text(" ".join(values[7:]))
        if not (
            sequence.isdigit()
            and _looks_like_chinese_name(person_name)
            and _looks_like_identifier_only(student_id)
            and _looks_like_score_or_metric(retest_score)
        ):
            continue

        record = _build_record(
            document,
            {
                "ranking": sequence,
                "person_name": person_name,
                "student_id": student_id,
                "admission_major": admission_major,
                "remarks": _join_labeled_parts(
                    [
                        ("外国语成绩", foreign_language_score),
                        ("理论（专项）考试成绩", theory_score),
                        ("复试成绩", retest_score),
                        ("备注", note),
                    ]
                ),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_blcu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = _document_and_row_haystack(document, rows)
    if "北京语言大学" not in document_text and "blcu.edu.cn" not in document_text:
        return []
    if "身份证号" not in document_text or "拟录取专业代码及名称" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        values = [_clean_text(value) for value in values if _clean_text(value)]
        if len(values) < 4:
            continue
        person_name = values[0]
        student_id = values[1]
        major_code, major_name = _split_blcu_major_code_name(values[2])
        if len(values) >= 5:
            advisor = values[3]
            retest_score = values[4]
        else:
            advisor = ""
            retest_score = values[3]
        if not (
            _looks_like_chinese_name(person_name)
            and _looks_like_identifier_only(student_id)
            and major_code
            and major_name
            and _looks_like_score_or_metric(retest_score)
        ):
            continue

        record = _build_record(
            document,
            {
                "person_name": person_name,
                "student_id": student_id,
                "major": major_code,
                "admission_major": _clean_text(f"{major_code} {major_name}"),
                "remarks": _join_labeled_parts(
                    [
                        ("录取导师", advisor),
                        ("复试成绩", retest_score),
                    ]
                ),
            },
        )
        if record:
            records.append(record)
    return records


def _split_blcu_major_code_name(value: str) -> tuple[str, str]:
    text = _clean_text(value)
    match = re.fullmatch(r"([0-9A-Z]{4,6})[-－]\s*(.+)", text, flags=re.I)
    if not match:
        return "", ""
    return match.group(1), _clean_text(match.group(2))


def _records_from_smu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = _document_and_row_haystack(document, rows, row_limit=20)
    if "南方医科大学" not in document_text and "smu.edu.cn" not in document_text:
        return []
    if not all(term in document_text for term in ["身份证号后6位", "分委会", "复试成绩"]):
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if len(values) < 10:
            continue
        ranking, person_name, identity_suffix, college, major_code, major_name = values[:6]
        research_direction, degree_type, retest_score, category = values[6:10]
        if not (
            ranking.isdigit()
            and _looks_like_chinese_name(person_name)
            and re.fullmatch(r"[0-9Xx]{6}", identity_suffix)
            and re.fullmatch(r"[0-9A-Z]{6}", major_code, flags=re.I)
            and _looks_like_score_or_metric(retest_score)
        ):
            continue

        record = _build_record(
            document,
            {
                "ranking": ranking,
                "person_name": person_name,
                "student_id": identity_suffix,
                "college": college,
                "major": major_code,
                "admission_major": major_name,
                "remarks": _join_labeled_parts(
                    [
                        ("研究方向", research_direction),
                        ("学位类型", degree_type),
                        ("复试成绩", retest_score),
                        ("类别", category),
                    ]
                ),
            },
        )
        if record:
            records.append(record)
    return records


def _document_and_row_haystack(
    document: dict[str, Any],
    rows: list[list[str]],
    row_limit: int = 8,
) -> str:
    return " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("source_url") or ""),
            str(document.get("title") or ""),
            " ".join(" ".join(row) for row in rows[:row_limit]),
        ]
    )


def _records_from_sdufe_doctoral_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    header_text = " ".join(" ".join(row) for row in rows[:6])
    if not all(
        marker in header_text
        for marker in ["考生编号", "录取学院", "录取专业", "报考导师", "招生方式"]
    ):
        return []

    records: list[dict[str, Any]] = []
    pending_singletons: list[str] = []
    current_record: dict[str, Any] | None = None
    current_needs_major = False
    current_needs_college = False
    inside_table = False

    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if not values:
            continue
        if _is_sdufe_doctoral_header_row(values):
            inside_table = True
            pending_singletons = []
            continue
        if not inside_table:
            continue

        if _is_sdufe_doctoral_data_row(values):
            prefix_values = pending_singletons
            pending_singletons = []
            record = _sdufe_doctoral_record(values, prefix_values, document)
            if not record:
                current_record = None
                current_needs_major = False
                current_needs_college = False
                continue

            records.append(record)
            current_record = record
            current_needs_major = _sdufe_needs_major_continuation(
                record.get("admission_major") or ""
            )
            current_needs_college = _has_unclosed_chinese_bracket(record.get("college") or "")
            continue

        if len(values) == 1:
            value = values[0]
            if current_record and current_needs_major:
                current_record["admission_major"] = _clean_text(
                    f"{current_record.get('admission_major') or ''} {value}"
                )
                _refresh_record_raw_row(current_record, "admission_major")
                current_needs_major = _sdufe_needs_major_continuation(
                    current_record.get("admission_major") or ""
                )
                continue
            if current_record and current_needs_college:
                current_record["college"] = _clean_text(f"{current_record.get('college') or ''}{value}")
                _refresh_record_raw_row(current_record, "college")
                current_needs_college = _has_unclosed_chinese_bracket(
                    current_record.get("college") or ""
                )
                continue
            if not _sdufe_ignorable_singleton(value):
                pending_singletons.append(value)

    return records


def _is_sdufe_doctoral_data_row(values: list[str]) -> bool:
    return (
        len(values) >= 9
        and values[0].isdigit()
        and bool(re.fullmatch(r"\d{10,}", values[1]))
        and bool(re.search(r"[\u4e00-\u9fff]", values[2]))
    )


def _is_sdufe_doctoral_header_row(values: list[str]) -> bool:
    text = " ".join(values)
    return all(marker in text for marker in ["序号", "考生编号", "录取学院", "录取专业"])


def _sdufe_doctoral_record(
    values: list[str],
    prefix_values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    prefix_code_parts = [
        value for value in prefix_values if re.fullmatch(r"\d{4,6}[A-Za-z]?", value)
    ]
    prefix_college_parts = [value for value in prefix_values if value not in prefix_code_parts]

    if len(values) >= 10:
        sequence, student_id, person_name, college, admission_major = values[:5]
        tutor, admission_category, study_mode, admission_method, score = values[5:10]
    else:
        sequence, student_id, person_name, college = values[:4]
        admission_major = " ".join(prefix_code_parts)
        tutor, admission_category, study_mode, admission_method, score = values[4:9]

    if prefix_college_parts:
        college = _clean_text("".join(prefix_college_parts) + college)
    if prefix_code_parts and not admission_major:
        admission_major = " ".join(prefix_code_parts)

    remarks = _join_labeled_parts(
        [
            ("报考导师", tutor),
            ("录取类别", admission_category),
            ("学习方式", study_mode),
            ("招生方式", admission_method),
            ("拟录取总分", score),
        ]
    )
    return _build_record(
        document,
        {
            "ranking": sequence,
            "student_id": student_id,
            "person_name": person_name,
            "college": college,
            "admission_major": admission_major,
            "remarks": remarks,
        },
    )


def _join_labeled_parts(parts: list[tuple[str, str]]) -> str:
    return "; ".join(f"{label}: {value}" for label, value in parts if value)


def _sdufe_needs_major_continuation(value: str) -> bool:
    return not value or bool(re.fullmatch(r"\d{4,6}[A-Za-z]?", value))


def _has_unclosed_chinese_bracket(value: str) -> bool:
    return value.count("（") > value.count("）") or value.count("(") > value.count(")")


def _sdufe_ignorable_singleton(value: str) -> bool:
    return value in {"序号", "考生编号", "姓名"} or bool(re.fullmatch(r"\d{1,3}(?:\.\d+)?", value))


def _refresh_record_raw_row(record: dict[str, Any], field: str) -> None:
    try:
        raw_row = json.loads(record.get("raw_row_json") or "{}")
    except json.JSONDecodeError:
        raw_row = {}
    raw_row[field] = record.get(field) or ""
    record["raw_row_json"] = json.dumps(raw_row, ensure_ascii=False, separators=(",", ":"))
    record["needs_review"] = not bool(
        record.get("person_name")
        and (
            record.get("admission_major")
            or record.get("major")
            or record.get("undergraduate_school")
        )
    )


def _records_from_gender_school_college_major_score_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_type = document.get("document_type") or ""
    if document_type not in {
        "incoming_recommendation_admission_list",
        "recommendation_exemption_list",
    }:
        return []

    header_text = " ".join(" ".join(row) for row in rows[:8])
    if not re.search(r"姓名\s*性别\s*毕业单位\s*拟录取学院\s*拟录取专业\s*考核成绩", header_text):
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        line = _clean_text(" ".join(row))
        if not line:
            continue
        if re.search(r"姓名\s*性别\s*毕业单位", line):
            continue
        if re.search(r"第\s*\d+\s*页\s*[，,]\s*共\s*\d+\s*页", line):
            continue

        match = re.match(r"^(?P<name>[\u4e00-\u9fff·]{2,8})\s+(?P<gender>男|女)\s+(?P<rest>.+)$", line)
        if not match:
            continue
        tokens = match.group("rest").split()
        score_index = None
        for index in range(len(tokens) - 1, -1, -1):
            if re.fullmatch(r"\d+(?:\.\d+)?", tokens[index]):
                score_index = index
                break
        if score_index is None or score_index < 3:
            continue

        score = tokens[score_index]
        remarks = [f"gender {match.group('gender')}", f"score {score}"]
        extra_remark = " ".join(tokens[score_index + 1 :])
        if extra_remark:
            remarks.append(extra_remark)
        record = _build_record(
            document,
            {
                "person_name": match.group("name"),
                "undergraduate_school": tokens[0],
                "college": tokens[1],
                "admission_major": " ".join(tokens[2:score_index]),
                "remarks": "; ".join(remarks),
            },
        )
        if record:
            records.append(record)

    return records


def _records_from_plain_recommendation_pdf_name_lines(
    text: str,
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if document.get("document_type") != "recommendation_exemption_list":
        return []

    records: list[dict[str, Any]] = []
    seen: set[str] = set()
    collecting = False
    pending_intro = False
    pending_name_prefix = ""
    remark = "通过答辩" if "通过答辩" in f"{text} {document.get('title', '')}" else ""

    for raw_line in text.splitlines():
        line = _clean_text(raw_line)
        if not line:
            continue
        if re.search(r"(学生名单|拟推荐名单|推荐名单|推免名单).*?(公示)?如下", line):
            collecting = True
            pending_intro = False
            continue
        if re.search(r"(学生名单|拟推荐名单|推荐名单|推免名单).*?(公示)?$", line):
            pending_intro = True
            continue
        if pending_intro and re.search(r"^如下", line):
            collecting = True
            pending_intro = False
            continue
        pending_intro = False
        if not collecting:
            continue
        if re.search(r"(公示时间|公示期|如有异议|联系电话|联系邮箱|邮箱|电话|学院$|大学$)", line):
            collecting = False
            pending_name_prefix = ""
            continue
        if pending_name_prefix:
            line = f"{pending_name_prefix}{line}"
            pending_name_prefix = ""
        split_name_match = re.fullmatch(r"(.+)[、，,\s]([\u4e00-\u9fff])", line)
        if split_name_match:
            line = split_name_match.group(1)
            pending_name_prefix = split_name_match.group(2)
        for name in _split_candidate_names_with_parenthetical_notes(line):
            if name in seen:
                continue
            seen.add(name)
            record = _build_record(
                document,
                {"person_name": name, "remarks": remark},
            )
            if record:
                records.append(record)

    return records


def _records_from_dlmu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "dlmu.edu.cn" not in document_text and "大连海事大学" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        record = _dlmu_recommendation_pdf_record(row, document)
        if record:
            records.append(record)
    return records


def _dlmu_recommendation_pdf_record(
    row: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    if len(values) < 9:
        return None
    if not values[0].isdigit():
        return None

    degree_and_name = re.fullmatch(r"(硕士|博士|直博)\s+(.+)", values[1] or "")
    if degree_and_name:
        values = [values[0], degree_and_name.group(1), degree_and_name.group(2), *values[2:]]
    elif values[1] not in {"硕士", "博士", "直博"}:
        values = [values[0], "直博", *values[1:]]
    if len(values) < 9:
        return None

    if not _looks_like_code(values[3], 3):
        return None

    remaining = values[4:]
    college_name = ""
    if not _looks_like_dlmu_major_code(remaining[0]):
        if len(remaining) < 6:
            return None
        college_name = remaining.pop(0)

    if not remaining or not _looks_like_dlmu_major_code(remaining[0]):
        return None
    major_code = remaining.pop(0)

    major_name = ""
    if len(remaining) >= 5 and not (
        re.fullmatch(r"\d{2}", remaining[0] or "") and _looks_like_score_or_metric(remaining[1])
    ):
        major_name = remaining.pop(0)

    if len(remaining) < 4:
        return None
    direction, score, employment, study_mode, *extra_values = remaining
    if not _looks_like_score_or_metric(score):
        return None

    remarks = [
        f"degree {values[1]}",
        f"direction {direction}",
        f"score {score}",
        f"employment {employment}",
        f"study_mode {study_mode}",
        *extra_values,
    ]
    return _build_record(
        document,
        {
            "person_name": values[2],
            "college": _clean_text(f"{values[3]} {college_name}"),
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "ranking": values[0],
            "remarks": _clean_text("; ".join(part for part in remarks if part)),
        },
    )


def _looks_like_dlmu_major_code(value: str) -> bool:
    return bool(re.fullmatch(r"\d{4}[0-9A-Z]\d", value or ""))


def _records_from_lnutcm_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "lnutcm.edu.cn" not in document_text and "辽宁中医药大学" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if len(values) < 12 or not values[0].isdigit():
            continue
        if not _looks_like_code(values[3], 6) or not re.fullmatch(r"\d{2}", values[5] or ""):
            continue
        if not all(_looks_like_score_or_metric(value) for value in values[8:11]):
            continue

        remarks = [
            f"direction {values[5]} {values[6]}",
            f"political {values[7]}",
            f"foreign_score {values[8]}",
            f"professional_score {values[9]}",
            f"total {values[10]}",
            *values[11:],
        ]
        record = _build_record(
            document,
            {
                "person_name": values[1],
                "college": values[2],
                "admission_major": _clean_text(f"{values[3]} {values[4]}"),
                "ranking": values[0],
                "remarks": _clean_text("; ".join(part for part in remarks if part)),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_dmu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
            " ".join(" ".join(row) for row in rows[:8]),
        ]
    )
    if "dmu.edu.cn" not in document_text and "大连医科大学" not in document_text:
        return []
    if "直接攻读博士" in document_text:
        return _records_from_dmu_direct_doctor_pdf_rows(rows, document)
    return _records_from_dmu_master_recommendation_pdf_rows(rows, document)


def _records_from_dmu_master_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    context = {
        "college_code": "",
        "college_name": "",
        "major_code": "",
        "major_name": "",
        "direction_code": "",
        "direction_name": "",
    }

    index = 0
    while index < len(rows):
        values = [_clean_text(value) for value in rows[index] if _clean_text(value)]
        if _dmu_skip_row(values):
            index += 1
            continue

        consumed_next = False
        record: dict[str, Any] | None = None
        if _dmu_is_college_code(values[0]):
            record, consumed_next = _dmu_master_record_from_college_row(rows, index, values, context, document)
        elif _looks_like_code(values[0], 6):
            record, consumed_next = _dmu_master_record_from_major_row(rows, index, values, context, document)
        elif _dmu_is_direction_code(values[0]):
            record, consumed_next = _dmu_master_record_from_direction_row(rows, index, values, context, document)
        elif _dmu_has_score_rank_tail(values) and len(values) >= 4:
            record = _dmu_build_recommendation_record(
                document,
                context,
                person_name=values[-4],
                score=values[-3],
                rank=values[-2],
                admitted=values[-1],
            )

        if record:
            records.append(record)
        index += 2 if consumed_next else 1

    return records


def _dmu_master_record_from_college_row(
    rows: list[list[str]],
    index: int,
    values: list[str],
    context: dict[str, str],
    document: dict[str, Any],
) -> tuple[dict[str, Any] | None, bool]:
    if len(values) >= 4 and _looks_like_code(values[2], 6):
        _dmu_update_college(context, values[0], values[1])
        _dmu_update_major(context, values[2], values[3])
        if len(values) >= 10 and _dmu_is_direction_code(values[4]) and _dmu_has_score_rank_tail(values):
            _dmu_update_direction(context, values[4], values[5])
            return (
                _dmu_build_recommendation_record(
                    document,
                    context,
                    person_name=values[-4],
                    score=values[-3],
                    rank=values[-2],
                    admitted=values[-1],
                ),
                False,
            )
        return None, False

    if len(values) >= 6 and _dmu_has_score_rank_tail(values):
        next_values = _dmu_next_values(rows, index)
        if len(next_values) >= 4 and _looks_like_code(next_values[0], 6) and not _dmu_has_score_rank_tail(next_values):
            _dmu_update_college(context, values[0], values[1])
            _dmu_update_major(context, next_values[0], next_values[1])
            if _dmu_is_direction_code(next_values[2]):
                _dmu_update_direction(context, next_values[2], next_values[3])
            return (
                _dmu_build_recommendation_record(
                    document,
                    context,
                    person_name=values[-4],
                    score=values[-3],
                    rank=values[-2],
                    admitted=values[-1],
                ),
                True,
            )
        _dmu_update_college(context, values[0], values[1])
        return (
            _dmu_build_recommendation_record(
                document,
                context,
                person_name=values[-4],
                score=values[-3],
                rank=values[-2],
                admitted=values[-1],
            ),
            False,
        )

    if len(values) >= 2:
        _dmu_update_college(context, values[0], values[1])
    return None, False


def _dmu_master_record_from_major_row(
    rows: list[list[str]],
    index: int,
    values: list[str],
    context: dict[str, str],
    document: dict[str, Any],
) -> tuple[dict[str, Any] | None, bool]:
    if len(values) < 2:
        return None, False
    previous_major_code = context.get("major_code") or ""
    _dmu_update_major(context, values[0], values[1])
    if len(values) >= 4 and _dmu_is_direction_code(values[2]):
        _dmu_update_direction(context, values[2], values[3])
    if len(values) >= 8 and _dmu_has_score_rank_tail(values):
        next_values = _dmu_next_values(rows, index)
        consumed_next = False
        if (
            values[0] != previous_major_code
            and len(next_values) >= 2
            and _dmu_is_college_code(next_values[0])
            and not _dmu_has_score_rank_tail(next_values)
        ):
            _dmu_update_college(context, next_values[0], next_values[1])
            consumed_next = True
        return (
            _dmu_build_recommendation_record(
                document,
                context,
                person_name=values[-4],
                score=values[-3],
                rank=values[-2],
                admitted=values[-1],
            ),
            consumed_next,
        )
    return None, False


def _dmu_master_record_from_direction_row(
    rows: list[list[str]],
    index: int,
    values: list[str],
    context: dict[str, str],
    document: dict[str, Any],
) -> tuple[dict[str, Any] | None, bool]:
    if len(values) < 2:
        return None, False
    if len(values) >= 6 and _dmu_has_score_rank_tail(values):
        next_values = _dmu_next_values(rows, index)
        if len(next_values) >= 4 and _dmu_is_college_code(next_values[0]) and _looks_like_code(next_values[2], 6):
            _dmu_update_college(context, next_values[0], next_values[1])
            _dmu_update_major(context, next_values[2], next_values[3])
            _dmu_update_direction(context, values[0], values[1])
            return (
                _dmu_build_recommendation_record(
                    document,
                    context,
                    person_name=values[-4],
                    score=values[-3],
                    rank=values[-2],
                    admitted=values[-1],
                ),
                True,
            )
        _dmu_update_direction(context, values[0], values[1])
        return (
            _dmu_build_recommendation_record(
                document,
                context,
                person_name=values[-4],
                score=values[-3],
                rank=values[-2],
                admitted=values[-1],
            ),
            False,
        )

    _dmu_update_direction(context, values[0], values[1])
    return None, False


def _records_from_dmu_direct_doctor_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    pending_research_directions: list[str] = []
    index = 0
    while index < len(rows):
        values = [_clean_text(value) for value in rows[index] if _clean_text(value)]
        if _dmu_skip_row(values):
            index += 1
            continue
        if _dmu_is_research_direction_line(values):
            pending_research_directions.append(values[0])
            index += 1
            continue
        if not (_dmu_is_college_code(values[0]) and len(values) >= 9 and _looks_like_code(values[2], 6)):
            index += 1
            continue

        following_directions: list[str] = []
        lookahead = index + 1
        while lookahead < len(rows):
            next_values = [_clean_text(value) for value in rows[lookahead] if _clean_text(value)]
            if _dmu_is_research_direction_line(next_values):
                following_directions.append(next_values[0])
                lookahead += 1
                continue
            break

        direction_name = ""
        if len(values) >= 10:
            direction_name = values[5]
            advisor = values[6]
            person_name = values[7]
            score = values[8]
            admitted = values[9]
        else:
            advisor = values[5]
            person_name = values[6]
            score = values[7]
            admitted = values[8]

        research_directions = pending_research_directions + following_directions
        pending_research_directions = []
        remarks = [
            "degree 直博",
            _clean_text(f"direction {values[4]} {direction_name}"),
            f"advisor {advisor}",
            f"score {score}",
            f"admitted {admitted}",
        ]
        if research_directions:
            remarks.insert(2, "research_direction " + "; ".join(research_directions))
        record = _build_record(
            document,
            {
                "person_name": person_name,
                "college": _clean_text(f"{values[0]} {values[1]}"),
                "admission_major": _clean_text(f"{values[2]} {values[3]}"),
                "remarks": _clean_text("; ".join(part for part in remarks if part)),
            },
        )
        if record:
            records.append(record)
        index = lookahead
    return records


def _dmu_build_recommendation_record(
    document: dict[str, Any],
    context: dict[str, str],
    *,
    person_name: str,
    score: str,
    rank: str,
    admitted: str,
) -> dict[str, Any] | None:
    remarks = [
        _clean_text(f"direction {context.get('direction_code', '')} {context.get('direction_name', '')}"),
        f"score {score}",
        f"admitted {admitted}",
    ]
    return _build_record(
        document,
        {
            "person_name": person_name,
            "college": _clean_text(f"{context.get('college_code', '')} {context.get('college_name', '')}"),
            "admission_major": _clean_text(f"{context.get('major_code', '')} {context.get('major_name', '')}"),
            "ranking": rank,
            "remarks": _clean_text("; ".join(part for part in remarks if part)),
        },
    )


def _dmu_update_college(context: dict[str, str], code: str, name: str) -> None:
    context["college_code"] = code
    context["college_name"] = name


def _dmu_update_major(context: dict[str, str], code: str, name: str) -> None:
    context["major_code"] = code
    context["major_name"] = name


def _dmu_update_direction(context: dict[str, str], code: str, name: str) -> None:
    context["direction_code"] = code
    context["direction_name"] = name


def _dmu_next_values(rows: list[list[str]], index: int) -> list[str]:
    if index + 1 >= len(rows):
        return []
    return [_clean_text(value) for value in rows[index + 1] if _clean_text(value)]


def _dmu_skip_row(values: list[str]) -> bool:
    if not values:
        return True
    text = " ".join(values)
    return bool(
        "第 " in text
        or text.startswith("第")
        or "拟录取名单" in text
        or "院系所名称" in text
        or "专业代码" in text
        or "考生姓名" in text
        or text in {"研究", "院系所 复试 是否", "院系所 研究方 报考 考核 是否同意", "代码 成绩 拟录取", "代码 向代码 导师 成绩 拟录取"}
    )


def _dmu_has_score_rank_tail(values: list[str]) -> bool:
    return len(values) >= 4 and _looks_like_score_or_metric(values[-3]) and values[-2].isdigit() and values[-1] in {"是", "否"}


def _dmu_is_college_code(value: str) -> bool:
    return bool(re.fullmatch(r"\d{3}", value or ""))


def _dmu_is_direction_code(value: str) -> bool:
    return bool(re.fullmatch(r"\d{2}", value or ""))


def _dmu_is_research_direction_line(values: list[str]) -> bool:
    return len(values) == 1 and bool(re.match(r"\d+\.", values[0] or ""))


def _records_from_just_retest_ranking_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "just.edu.cn" not in document_text and "江苏科技大学" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        record = _just_retest_ranking_record(row, document)
        if record:
            records.append(record)
    return records


def _just_retest_ranking_record(
    row: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [
        _clean_text(value)
        for value in re.split(r"\s+", " ".join(_clean_text(value) for value in row))
        if _clean_text(value)
    ]
    if len(values) < 8 or values[0] not in {"一志愿", "调剂"}:
        return None

    major_index = next(
        (index for index, value in enumerate(values[1:], start=1) if _looks_like_major_code(value)),
        None,
    )
    if major_index is None:
        return None

    mode_index = next(
        (
            index
            for index in range(major_index + 1, len(values))
            if _looks_like_study_mode(values[index])
        ),
        None,
    )
    if mode_index is None or mode_index + 2 >= len(values):
        return None

    major_code = values[major_index]
    major_name = _clean_text(" ".join(values[major_index + 1 : mode_index]))
    study_mode = values[mode_index]
    person_name = values[mode_index + 1]
    student_id = values[mode_index + 2]
    score_values = values[mode_index + 3 :]

    if not (
        major_name
        and _looks_like_chinese_name(person_name)
        and _looks_like_identifier_only(student_id)
    ):
        return None

    status_index = next(
        (
            index
            for index, value in enumerate(score_values)
            if value in {"拟录取", "是", "否", "待录取"}
        ),
        None,
    )
    ranking = ""
    if status_index is not None and status_index > 0:
        ranking_candidate = score_values[status_index - 1]
        if re.fullmatch(r"\d+", ranking_candidate):
            ranking = ranking_candidate

    return _build_record(
        document,
        {
            "person_name": person_name,
            "student_id": student_id,
            "major": major_code,
            "admission_major": major_name,
            "ranking": ranking,
            "remarks": _clean_text(" ".join([values[0], study_mode, *score_values])),
        },
    )


def _records_from_hgu_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "hgu.edu.cn" not in document_text and "河北地质大学" not in document_text:
        return []

    recommendation_records = _records_from_hgu_recommendation_pdf_rows(rows, document)
    if recommendation_records:
        return recommendation_records

    records: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        record = _hgu_postgraduate_pdf_record(rows, index, document)
        if record:
            records.append(record)
    return records


def _hgu_postgraduate_pdf_record(
    rows: list[list[str]],
    index: int,
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = _hgu_candidate_values(rows, index)
    if not values or not values[0].isdigit():
        return None

    parsed = _hgu_parse_candidate_values(values)
    if not parsed:
        return None
    (
        sequence,
        person_name,
        student_id,
        college_code,
        college_name,
        major_code,
        major_name,
        study_mode,
        initial_score,
        retest_score,
        total_score,
        admission_status,
        admission_batch,
        extra_values,
    ) = parsed
    if not (
        sequence.isdigit()
        and _looks_like_chinese_name(person_name)
        and _looks_like_identifier_only(student_id)
        and re.fullmatch(r"\d{3}", college_code or "")
        and _looks_like_major_code(major_code)
        and _looks_like_score_or_metric(initial_score)
        and _looks_like_score_or_metric(retest_score)
        and _looks_like_score_or_metric(total_score)
    ):
        return None

    return _build_record(
        document,
        {
            "ranking": sequence,
            "person_name": person_name,
            "student_id": student_id,
            "college": _clean_text(f"{college_code} {college_name}"),
            "major": major_code,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": _clean_text(
                " ".join(
                    [
                        study_mode,
                        f"initial_score {initial_score}",
                        f"retest_score {retest_score}",
                        f"total_score {total_score}",
                        admission_status,
                        admission_batch,
                        *extra_values,
                    ]
                )
            ),
        },
    )


def _records_from_hgu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    header_text = " ".join(" ".join(row) for row in rows[:15])
    document_type = str(document.get("document_type") or "")
    if not (
        document_type == "incoming_recommendation_admission_list"
        or "推免" in document_text
        or "推免" in header_text
    ):
        return []

    records: list[dict[str, Any]] = []
    for index, _row in enumerate(rows):
        values = _hgu_recommendation_candidate_values(rows, index)
        parsed = _hgu_parse_recommendation_candidate_values(values)
        if not parsed:
            continue
        (
            sequence,
            person_name,
            plan_category,
            student_id,
            college_code,
            college_name,
            major_code,
            major_name,
            study_mode,
            retest_total_score,
        ) = parsed
        records.append(
            _build_record(
                document,
                {
                    "ranking": sequence,
                    "person_name": person_name,
                    "student_id": student_id,
                    "college": _clean_text(f"{college_code} {college_name}"),
                    "major": major_code,
                    "admission_major": _clean_text(f"{major_code} {major_name}"),
                    "remarks": _clean_text(
                        " ".join(
                            [
                                f"plan_category {plan_category}",
                                f"study_mode {study_mode}",
                                f"retest_total_score {retest_total_score}",
                            ]
                        )
                    ),
                },
            )
        )
    return records


def _hgu_recommendation_candidate_values(rows: list[list[str]], index: int) -> list[str]:
    values = _hgu_row_values(rows[index])
    if not values or not (values[0].isdigit() and len(values[0]) <= 4):
        return values

    candidate = list(values)
    for row in rows[index + 1 : index + 6]:
        if _hgu_parse_recommendation_candidate_values(candidate):
            break
        next_values = _hgu_row_values(row)
        if not next_values:
            continue
        if next_values[0].isdigit() and len(next_values[0]) <= 4:
            break
        candidate.extend(next_values)
    return candidate


def _hgu_parse_recommendation_candidate_values(
    values: list[str],
) -> tuple[str, str, str, str, str, str, str, str, str, str] | None:
    if len(values) < 7:
        return None
    sequence = values[0]
    if not (sequence.isdigit() and len(sequence) <= 4):
        return None

    identity = _hgu_recommendation_identity(values, 1)
    if not identity:
        return None
    person_name, plan_category, student_id, offset = identity
    if len(values) <= offset:
        return None

    college_code = values[offset]
    if not re.fullmatch(r"\d{3}", college_code or ""):
        return None
    rest = values[offset + 1 :]
    if len(rest) < 4:
        return None
    study_mode, retest_total_score = rest[-2:]
    if study_mode not in {"全日制", "非全日制"}:
        return None
    if not _looks_like_score_or_metric(retest_total_score):
        return None

    major_text = _clean_text(" ".join(rest[:-2]))
    major_match = re.fullmatch(
        r"(?P<college_name>.+?)\s+(?P<major_code>\d{6})\s+(?P<major_name>.+)",
        major_text,
    )
    if not major_match:
        return None
    college_name = major_match.group("college_name")
    major_code = major_match.group("major_code")
    major_name = major_match.group("major_name")
    if not (
        _looks_like_chinese_name(person_name)
        and _looks_like_identifier_only(student_id)
        and _looks_like_major_code(major_code)
    ):
        return None
    return (
        sequence,
        person_name,
        plan_category,
        student_id,
        college_code,
        college_name,
        major_code,
        major_name,
        study_mode,
        retest_total_score,
    )


def _hgu_recommendation_identity(
    values: list[str],
    offset: int,
) -> tuple[str, str, str, int] | None:
    if len(values) <= offset:
        return None
    joined_match = re.fullmatch(
        r"(?P<person_name>[\u4e00-\u9fff]{2,4})\s+"
        r"(?P<plan_category>.+?计划)\s+"
        r"(?P<student_id>\d{8,})",
        values[offset],
    )
    if joined_match:
        data = joined_match.groupdict()
        return (
            data["person_name"],
            data["plan_category"],
            data["student_id"],
            offset + 1,
        )

    if len(values) <= offset + 1:
        return None
    split_plan_match = re.fullmatch(r"(?P<plan_category>.+?计划)\s+(?P<student_id>\d{8,})", values[offset + 1])
    if split_plan_match and _looks_like_chinese_name(values[offset]):
        data = split_plan_match.groupdict()
        return (
            values[offset],
            data["plan_category"],
            data["student_id"],
            offset + 2,
        )

    if len(values) <= offset + 2:
        return None
    if (
        _looks_like_chinese_name(values[offset])
        and values[offset + 1].endswith("计划")
        and _looks_like_identifier_only(values[offset + 2])
    ):
        return values[offset], values[offset + 1], values[offset + 2], offset + 3
    return None


def _hgu_candidate_values(rows: list[list[str]], index: int) -> list[str]:
    values = _hgu_row_values(rows[index])
    if len(values) >= 13:
        return values
    if len(values) < 2 or not (values[0].isdigit() and _looks_like_chinese_name(values[1])):
        return values

    candidate = list(values)
    for row in rows[index + 1 : index + 7]:
        if _hgu_parse_candidate_values(candidate):
            break
        next_values = _hgu_row_values(row)
        if not next_values:
            continue
        if next_values[0].isdigit() and len(next_values[0]) <= 4:
            break
        candidate.extend(next_values)
    return candidate


def _hgu_parse_candidate_values(
    values: list[str],
) -> tuple[str, str, str, str, str, str, str, str, str, str, str, str, str, list[str]] | None:
    if len(values) < 12:
        return None
    sequence, person_name, student_id, college_code = values[:4]
    if not (
        sequence.isdigit()
        and _looks_like_chinese_name(person_name)
        and _looks_like_identifier_only(student_id)
        and re.fullmatch(r"\d{3}", college_code or "")
    ):
        return None

    rest = values[4:]
    if len(rest) >= 8 and _looks_like_major_code(rest[0]):
        college_name = ""
        major_code, major_name = rest[:2]
        offset = 2
    elif len(rest) >= 9 and _looks_like_major_code(rest[1]):
        college_name = rest[0]
        major_code, major_name = rest[1:3]
        offset = 3
    else:
        return None

    remaining = rest[offset:]
    if len(remaining) < 6:
        return None
    study_mode, initial_score, retest_score, total_score, admission_status, admission_batch = remaining[:6]
    return (
        sequence,
        person_name,
        student_id,
        college_code,
        college_name,
        major_code,
        major_name,
        study_mode,
        initial_score,
        retest_score,
        total_score,
        admission_status,
        admission_batch,
        remaining[6:],
    )


def _hgu_row_values(row: list[str]) -> list[str]:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    if len(values) == 1 and " " in values[0]:
        values = [_clean_text(value) for value in re.split(r"\s+", values[0]) if _clean_text(value)]
    normalized: list[str] = []
    for value in values:
        value = re.sub(r"^[仅转用禁示严载、]+(?=(?:非)?全日制)", "", value)
        if value in {"仅", "转", "用", "禁", "示", "严", "载", "、", "次"}:
            continue
        normalized.append(value)
    values = normalized
    return values


def _records_from_gxmzu_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    header_text = " ".join(" ".join(row) for row in rows[:12])
    if not (
        "广西民族大学" in document_text
        or "gxmzu.edu.cn" in document_text.lower()
        or ("拟录取考生汇总表" in header_text and "专业代码及名称" in header_text)
    ):
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        record = _gxmzu_postgraduate_pdf_record(values, document)
        if record:
            records.append(record)
    return records


def _gxmzu_postgraduate_pdf_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in values if _clean_text(value)]
    if len(values) < 12:
        return None
    sequence, retest_batch, student_id, person_name = values[:4]
    if not (
        sequence.isdigit()
        and retest_batch.isdigit()
        and _looks_like_identifier_only(student_id)
        and _looks_like_chinese_name(person_name)
    ):
        return None
    major_match = re.fullmatch(r"(?P<major_code>\d{6})\s*(?P<major_name>.+)", values[4])
    if not major_match:
        return None
    research_direction = values[5]
    admission_status = values[6]
    total_score = values[7]
    program_rank = values[8]
    retest_score = values[9]
    if "录取" not in admission_status:
        return None
    if not (
        _looks_like_score_or_metric(total_score)
        and program_rank.isdigit()
        and _looks_like_score_or_metric(retest_score)
    ):
        return None

    initial_score = ""
    score_parts: list[str] = []
    extra_values: list[str] = []
    for offset in range(len(values) - 1, 9, -1):
        candidate = values[offset]
        if re.fullmatch(r"\d{2,3}", candidate):
            initial_score = candidate
            score_parts = values[10:offset]
            extra_values = values[offset + 1 :]
            break

    major_code = major_match.group("major_code")
    major_name = major_match.group("major_name")
    remarks = _clean_text(
        " ".join(
            [
                f"retest_batch {retest_batch}",
                f"research_direction {research_direction}",
                f"admission_status {admission_status}",
                f"total_score {total_score}",
                f"program_rank {program_rank}",
                f"retest_score {retest_score}",
                f"score_fields {' '.join(score_parts)}" if score_parts else "",
                f"initial_score {initial_score}" if initial_score else "",
                *extra_values,
            ]
        )
    )
    return _build_record(
        document,
        {
            "ranking": sequence,
            "person_name": person_name,
            "student_id": student_id,
            "major": major_code,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": remarks,
        },
    )


def _records_from_gxu_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if not _is_gxu_document(document):
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        record = (
            _gxu_incoming_recommendation_record(values, document)
            or _gxu_doctoral_admission_record(values, document)
            or _gxu_master_admission_record(values, document)
            or _gxu_college_code_admission_record(values, document)
            or _gxu_special_plan_admission_record(values, document)
        )
        if record:
            records.append(record)
    return records


def _is_gxu_document(document: dict[str, Any]) -> bool:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    return "广西大学" in document_text or "gxu.edu.cn" in document_text


def _gxu_incoming_recommendation_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 11:
        return None
    college_code, college_name, person_name, admission_type, plan, major_code, major_name = values[:7]
    if not (
        re.fullmatch(r"\d{3}", college_code or "")
        and _looks_like_chinese_name(person_name)
        and _looks_like_major_code(major_code)
        and admission_type in {"硕士", "直博生", "博士"}
    ):
        return None

    study_mode = values[7]
    direction_code = values[8]
    direction, admission_category = _split_gxu_direction_and_category(values[9])
    retest_score = values[10]
    remarks = _clean_text(
        " ".join(
            part
            for part in [
                f"admission_type {admission_type}",
                f"plan {plan}",
                f"study_mode {study_mode}",
                f"direction_code {direction_code}",
                f"research_direction {direction}",
                f"admission_category {admission_category}" if admission_category else "",
                f"retest_score {retest_score}" if _looks_like_score_or_metric(retest_score) else "",
                *values[11:],
            ]
            if part
        )
    )
    return _build_record(
        document,
        {
            "person_name": person_name,
            "college": _clean_text(f"{college_code} {college_name}"),
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": remarks,
        },
    )


def _gxu_doctoral_admission_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 5:
        return None
    sequence, person_name, major_code, major_name = values[:4]
    college = values[4]
    total_score = values[5] if len(values) > 5 else ""
    extra_values = values[6:]
    if not total_score:
        college_match = re.fullmatch(r"(.+?学院)\s+(\d{1,3}(?:\.\d+)?)", college)
        if college_match:
            college = college_match.group(1)
            total_score = college_match.group(2)
    if not (
        sequence.isdigit()
        and _looks_like_chinese_name(person_name)
        and _looks_like_major_code(major_code)
        and college.endswith("学院")
        and _looks_like_score_or_metric(total_score)
    ):
        return None
    return _build_record(
        document,
        {
            "ranking": sequence,
            "person_name": person_name,
            "college": college,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": _clean_text(" ".join([f"total_score {total_score}", *extra_values])),
        },
    )


def _gxu_master_admission_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 7:
        return None
    if not (values[0].isdigit() and _looks_like_major_code(values[2])):
        return None

    sequence = values[0]
    college = values[1]
    major_code = values[2]
    major_name = values[3]
    direction_code = values[4] if len(values) > 4 else ""
    direction = ""
    study_mode = ""
    student_id = ""
    person_name = ""
    score_values: list[str] = []

    if len(values) >= 9 and _looks_like_study_mode(values[6]) and _looks_like_identifier_only(values[7]):
        direction = values[5]
        study_mode = values[6]
        student_id = values[7]
        person_name = values[8]
        score_values = values[9:]
    elif len(values) >= 8 and _looks_like_study_mode(values[5]) and _looks_like_identifier_only(values[6]):
        study_mode = values[5]
        student_id = values[6]
        person_name = values[7]
        score_values = values[8:]
    else:
        return None

    if not _looks_like_chinese_name(person_name):
        return None
    remarks = _gxu_score_remarks(direction_code, direction, study_mode, score_values)
    return _build_record(
        document,
        {
            "ranking": sequence,
            "person_name": person_name,
            "student_id": student_id,
            "college": college,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": remarks,
        },
    )


def _gxu_college_code_admission_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 7:
        return None
    college_code, college_name, person_name, student_id, major_code = values[:5]
    if not (
        re.fullmatch(r"\d{3}", college_code or "")
        and _looks_like_chinese_name(person_name)
        and _looks_like_identifier_only(student_id)
        and _looks_like_major_code(major_code)
    ):
        return None

    major_name = values[5]
    direction_code = values[6] if len(values) > 6 else ""
    direction = values[7] if len(values) > 7 else ""
    study_mode = values[8] if len(values) > 8 else ""
    extra_values = values[9:]

    split_major = re.fullmatch(r"(.+?)\s+(\d{2})", major_name)
    if split_major and len(values) >= 8:
        major_name = split_major.group(1)
        direction_code = split_major.group(2)
        direction = values[6]
        study_mode = values[7]
        extra_values = values[8:]

    direction, mode_from_direction = _split_gxu_direction_and_mode(direction)
    if not _looks_like_study_mode(study_mode) and mode_from_direction:
        study_mode = mode_from_direction
    elif not _looks_like_study_mode(study_mode) and _looks_like_study_mode(direction):
        study_mode = direction
        direction = ""

    return _build_record(
        document,
        {
            "person_name": person_name,
            "student_id": student_id,
            "college": _clean_text(f"{college_code} {college_name}"),
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": _clean_text(
                " ".join(
                    part
                    for part in [
                        f"direction_code {direction_code}" if direction_code else "",
                        f"research_direction {direction}" if direction else "",
                        f"study_mode {study_mode}" if study_mode else "",
                        *extra_values,
                    ]
                    if part
                )
            ),
        },
    )


def _gxu_special_plan_admission_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 4 or not values[0].isdigit():
        return None
    combined = values[1]
    match = re.match(
        r"^(?P<name>[\u4e00-\u9fff]{2,4})\s+(?P<student_id>\d{12,})\s*(?P<plan>退役士兵计划|少数民族骨干计划)\s+(?P<mode>全日制|非全日制)\s*(?P<rest>.*)$",
        combined,
    )
    if not match:
        return None

    rest = _clean_text(match.group("rest"))
    major_name = ""
    major_code = ""
    if rest:
        rest_match = re.match(r"(.+?)\s+(\d{6})$", rest)
        if rest_match:
            major_name = rest_match.group(1)
            major_code = rest_match.group(2)
    if not major_code:
        for index, value in enumerate(values[2:], start=2):
            if _looks_like_major_code(value):
                major_code = value
                major_name = " ".join(part for part in [rest, *values[2:index]] if part)
                break
    if not major_code or not major_name:
        return None

    score_values = values[2:]
    return _build_record(
        document,
        {
            "ranking": values[0],
            "person_name": match.group("name"),
            "student_id": match.group("student_id"),
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": _clean_text(
                " ".join(
                    [
                        f"plan {match.group('plan')}",
                        f"study_mode {match.group('mode')}",
                        *score_values,
                    ]
                )
            ),
        },
    )


def _split_gxu_direction_and_category(value: str) -> tuple[str, str]:
    text = _clean_text(value)
    match = re.fullmatch(r"(.+?)\s+(非定向|定向)$", text)
    if match:
        return _clean_text(match.group(1)), match.group(2)
    return text, ""


def _split_gxu_direction_and_mode(value: str) -> tuple[str, str]:
    text = _clean_text(value)
    match = re.fullmatch(r"(.+?)\s+(全日制|非全日制)$", text)
    if match:
        return _clean_text(match.group(1)), match.group(2)
    return text, ""


def _looks_like_study_mode(value: str) -> bool:
    return value in {"全日制", "非全日制"}


def _gxu_score_remarks(
    direction_code: str,
    direction: str,
    study_mode: str,
    score_values: list[str],
) -> str:
    labels = ["initial_score", "retest_score", "total_score"]
    score_parts = []
    extra_values = []
    for value in score_values:
        if len(score_parts) < len(labels) and _looks_like_score_or_metric(value):
            score_parts.append(f"{labels[len(score_parts)]} {value}")
        else:
            extra_values.append(value)
    return _clean_text(
        " ".join(
            part
            for part in [
                f"direction_code {direction_code}" if direction_code else "",
                f"research_direction {direction}" if direction else "",
                f"study_mode {study_mode}" if study_mode else "",
                *score_parts,
                *extra_values,
            ]
            if part
        )
    )


def _records_from_hebut_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "hebut.edu.cn" not in document_text and "河北工业大学" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        record = _hebut_postgraduate_pdf_record(values, document) or _hebut_recommendation_pdf_record(
            values,
            document,
        )
        if record:
            records.append(record)
    return records


def _hebut_postgraduate_pdf_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in values if _clean_text(value)]
    if len(values) < 8 or not values[0].isdigit():
        return None

    ranking, unit, student_id, person_name = values[:4]
    if not _looks_like_identifier_only(student_id) or not _looks_like_chinese_name(person_name):
        return None
    unit_parts = _hebut_split_unit(unit)
    if not unit_parts:
        return None
    college, admission_major = unit_parts

    reexam_score, initial_score, total_score, study_mode = values[4:8]
    if not (
        _looks_like_score_or_metric(reexam_score)
        and _looks_like_score_or_metric(initial_score)
        and _looks_like_score_or_metric(total_score)
    ):
        return None
    remarks = _clean_text(
        " ".join(
            [
                f"retest_score {reexam_score}",
                f"initial_score {initial_score}",
                f"total_score {total_score}",
                study_mode,
                *values[8:],
            ]
        )
    )
    return _build_record(
        document,
        {
            "ranking": ranking,
            "student_id": student_id,
            "person_name": person_name,
            "college": college,
            "admission_major": admission_major,
            "remarks": remarks,
        },
    )


def _hebut_split_unit(value: str) -> tuple[str, str] | None:
    match = re.fullmatch(r"(\d{3})(.+?)(\d{6})(.+)", _clean_text(value))
    if not match:
        return None
    college_code, college_name, major_code, major_name = match.groups()
    return (
        _clean_text(f"{college_code} {college_name}"),
        _clean_text(f"{major_code} {major_name}"),
    )


def _hebut_recommendation_pdf_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in values if _clean_text(value)]
    parsed = _hebut_recommendation_prefix(values)
    if not parsed:
        return None
    ranking, admission_type, college_code, major_code, major_name, index = parsed

    student_id_index = _hebut_find_student_id_index(values, index)
    if student_id_index is None:
        return None
    direction_values = values[index:student_id_index]
    direction = _hebut_direction_text(direction_values)
    student_id = values[student_id_index]
    identity = _hebut_recommendation_identity(values[student_id_index + 1 :])
    if not identity:
        return None
    person_name, gender, score, remaining = identity
    if not _looks_like_chinese_name(person_name):
        return None

    remarks = _clean_text(
        " ".join(
            part
            for part in [
                admission_type,
                f"direction {direction}" if direction else "",
                gender,
                f"score {score}" if score else "",
                *remaining,
            ]
            if part
        )
    )
    return _build_record(
        document,
        {
            "ranking": ranking,
            "student_id": student_id,
            "person_name": person_name,
            "college": college_code,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": remarks,
        },
    )


def _hebut_recommendation_prefix(
    values: list[str],
) -> tuple[str, str, str, str, str, int] | None:
    if not values:
        return None

    compact_match = re.fullmatch(
        r"(\d{3})\s+(\S+)\s+(\d{3})\s+(\d{6})(?:\s+(.+))?",
        values[0],
    )
    if compact_match:
        ranking, admission_type, college_code, major_code, major_tail = compact_match.groups()
        index = 1
        if major_tail:
            major_name = major_tail
        elif len(values) > index:
            major_name = values[index]
            index += 1
        else:
            return None
        if admission_type not in {"硕士", "直博生"}:
            return None
        return ranking, admission_type, college_code, major_code, major_name, index

    if len(values) < 5 or not re.fullmatch(r"\d{3}", values[0] or ""):
        return None
    ranking, admission_type = values[:2]
    if admission_type not in {"硕士", "直博生"}:
        return None

    code_match = re.fullmatch(r"(\d{3})\s+(\d{6})(?:\s+(.+))?", values[2])
    if code_match:
        college_code, major_code, major_tail = code_match.groups()
        index = 3
        if major_tail:
            major_name = major_tail
        elif len(values) > index:
            major_name = values[index]
            index += 1
        else:
            return None
        return ranking, admission_type, college_code, major_code, major_name, index

    if len(values) < 6 or not (re.fullmatch(r"\d{3}", values[2]) and _looks_like_major_code(values[3])):
        return None
    return ranking, admission_type, values[2], values[3], values[4], 5


def _hebut_find_student_id_index(values: list[str], start_index: int) -> int | None:
    for index in range(start_index, len(values)):
        value = values[index]
        if _looks_like_identifier_only(value) and len(re.sub(r"\D", "", value)) >= 10:
            return index
    return None


def _hebut_direction_text(values: list[str]) -> str:
    if not values:
        return ""
    if re.fullmatch(r"\d{2}", values[0] or ""):
        return _clean_text(" ".join(values))
    return _clean_text(" ".join(values))


def _hebut_recommendation_identity(values: list[str]) -> tuple[str, str, str, list[str]] | None:
    if not values:
        return None
    combined = re.fullmatch(
        r"([\u4e00-\u9fff·]{1,12})\s+([男女])\s+(\d+(?:\.\d+)?)",
        values[0],
    )
    if combined:
        person_name, gender, score = combined.groups()
        return person_name, gender, score, values[1:]

    person_name = values[0]
    if len(values) < 2:
        return person_name, "", "", []

    gender = ""
    score = ""
    remaining_index = 1
    gender_score_match = re.fullmatch(r"([男女])\s+(\d+(?:\.\d+)?)", values[1])
    if gender_score_match:
        gender, score = gender_score_match.groups()
        remaining_index = 2
    elif values[1] in {"男", "女"}:
        gender = values[1]
        remaining_index = 2
        if len(values) > 2 and _looks_like_score_or_metric(values[2]):
            score = values[2]
            remaining_index = 3
    return person_name, gender, score, values[remaining_index:]


def _records_from_yangtzeu_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "yangtzeu.edu.cn" not in document_text and "\u957f\u6c5f\u5927\u5b66" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for index, values in enumerate(rows):
        record = _yangtzeu_record_from_window(rows, index, document)
        if not record:
            continue
        key = (
            record.get("student_id") or "",
            record.get("person_name") or "",
            record.get("college") or "",
            record.get("admission_major") or "",
        )
        if key in seen:
            continue
        seen.add(key)
        records.append(record)
    return records


def _yangtzeu_record_from_window(
    rows: list[list[str]],
    index: int,
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = _yangtzeu_clean_values(rows[index])
    if not values:
        return None

    context_index = index
    if _looks_like_major_code(values[0]):
        identity = _yangtzeu_previous_identity(rows, index)
        if not identity:
            return None
        values = [identity[0], identity[1], *values]
        context_index = identity[2]
    elif _looks_like_code(_yangtzeu_normalize_unit_code(values[0]), 3):
        identity = _yangtzeu_previous_identity(rows, index)
        if not identity:
            return None
        values = [identity[0], identity[1], _yangtzeu_normalize_unit_code(values[0]), *values[1:]]
        context_index = identity[2]

    college_context = _yangtzeu_college_context(rows, context_index)
    return _yangtzeu_record_from_values(values, college_context, document)


def _yangtzeu_record_from_values(
    values: list[str],
    college_context: str,
    document: dict[str, Any],
) -> dict[str, Any] | None:
    identity = _yangtzeu_split_identity(values)
    if not identity:
        return None
    student_id, person_name, rest = identity
    if not _looks_like_chinese_name(person_name):
        return None

    rest = [_yangtzeu_normalize_unit_code(value) for value in rest]
    rest = [value for value in rest if value]
    if not rest:
        return None

    unit_code = ""
    if _looks_like_code(rest[0], 3):
        unit_code = rest.pop(0)

    college = ""
    if rest and not _looks_like_major_code(rest[0]):
        college = rest.pop(0)
    elif college_context:
        college = college_context
    if unit_code and college:
        college = _clean_text(f"{unit_code} {college}")

    if len(rest) < 3 or not _looks_like_major_code(rest[0]):
        return None
    major_code, major_name = rest[:2]
    if not major_name or _looks_like_score_or_metric(major_name):
        return None

    score_values: list[str] = []
    extra_values: list[str] = []
    for value in rest[2:]:
        if _looks_like_score_or_metric(value) and len(score_values) < 3:
            score_values.append(value)
        else:
            extra_values.append(value)
    if not score_values:
        return None

    if len(score_values) >= 3:
        score_text = [
            f"initial_score {score_values[0]}",
            f"reexam_score {score_values[1]}",
            f"total_score {score_values[2]}",
        ]
    elif len(score_values) == 2:
        score_text = [
            f"reexam_score {score_values[0]}",
            f"total_score {score_values[1]}",
        ]
    else:
        score_text = [f"reexam_score {score_values[0]}"]

    return _build_record(
        document,
        {
            "student_id": student_id,
            "person_name": person_name,
            "college": college,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": _clean_text(" ".join([*score_text, *extra_values])),
        },
    )


def _yangtzeu_split_identity(values: list[str]) -> tuple[str, str, list[str]] | None:
    if not values:
        return None

    first = values[0]
    combined = re.fullmatch(r"([0-9Xx*]{10,})\s+([\u4e00-\u9fff\u00b7\u8def]{2,6})", first)
    if combined:
        return combined.group(1), combined.group(2), values[1:]

    if _looks_like_identifier_only(first) and len(values) >= 2:
        return first, values[1], values[2:]

    if _looks_like_chinese_name(first):
        return "", first, values[1:]

    return None


def _yangtzeu_previous_identity(
    rows: list[list[str]],
    index: int,
) -> tuple[str, str, int] | None:
    for previous_index in range(index - 1, max(-1, index - 6), -1):
        values = _yangtzeu_clean_values(rows[previous_index])
        if not values:
            continue
        identity = _yangtzeu_split_identity(values)
        if identity and not identity[2]:
            return identity[0], identity[1], previous_index
        if _looks_like_major_code(values[0]):
            return None
    return None


def _yangtzeu_college_context(rows: list[list[str]], index: int) -> str:
    before_reversed: list[str] = []
    for previous_index in range(index - 1, max(-1, index - 6), -1):
        values = _yangtzeu_clean_values(rows[previous_index])
        if not values:
            continue
        if _yangtzeu_is_data_or_header_row(values):
            break
        if _yangtzeu_looks_like_college_fragment(values):
            fragment = " ".join(values)
            before_reversed.append(fragment)

    after: list[str] = []
    for next_index in range(index + 1, min(len(rows), index + 6)):
        values = _yangtzeu_clean_values(rows[next_index])
        if not values:
            continue
        if _yangtzeu_is_data_or_header_row(values):
            break
        if _yangtzeu_looks_like_college_fragment(values):
            fragment = " ".join(values)
            after.append(fragment)
            if "\uff09" in fragment:
                break

    before = list(reversed(before_reversed))
    return _yangtzeu_normalize_college_context(_clean_text(" ".join([*before, *after])))


def _yangtzeu_normalize_college_context(value: str) -> str:
    parts = value.split()
    for index, part in enumerate(parts):
        if "\uff08" not in part or "\uff09" in part:
            continue
        collected = [part]
        for next_part in parts[index + 1 :]:
            collected.append(next_part)
            if "\uff09" in next_part:
                return _clean_text(" ".join(collected))
    return value


def _yangtzeu_clean_values(values: list[str]) -> list[str]:
    cleaned: list[str] = []
    for value in values:
        text = _clean_text(value)
        if not text or _yangtzeu_is_watermark_fragment(text):
            continue
        cleaned.append(text)
    return cleaned


def _yangtzeu_normalize_unit_code(value: str) -> str:
    match = re.fullmatch(r"[\u590d\u8f6c\u5236\u8f7d\u7981\u4e25\uff01\uff0c\u7528\u793a\u516c\u4ec5]*(\d{3})", value or "")
    return match.group(1) if match else value


def _yangtzeu_is_watermark_fragment(value: str) -> bool:
    return value in {
        "\u590d",
        "\u8f6c",
        "\u5236",
        "\u8f7d",
        "\u7981",
        "\u4e25",
        "\uff01",
        "\uff0c",
        "\u7528",
        "\u793a",
        "\u516c",
        "\u4ec5",
    }


def _yangtzeu_is_data_or_header_row(values: list[str]) -> bool:
    joined = " ".join(values)
    if re.search(r"(\u8003\u751f\u7f16\u53f7|\u59d3\u540d|\u62df\u5f55\u53d6|\u4e13\u4e1a\u4ee3\u7801|\u590d\u8bd5\u6210\u7ee9)", joined):
        return True
    if _yangtzeu_split_identity(values):
        return True
    return bool(values and _looks_like_major_code(values[0]))


def _yangtzeu_looks_like_college_fragment(values: list[str]) -> bool:
    text = " ".join(values)
    if len(text) > 40:
        return False
    if re.search(r"(\u5b66\u9662|\u4e2d\u5fc3|\u533b\u9662|\u533b\u5b66\u90e8)", text):
        return True
    return "\uff09" in text and bool(re.search(r"[\u4e00-\u9fff]", text))


def _records_from_college_major_exam_name_score_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for values in rows:
        record = _college_major_exam_name_score_record(values, document)
        if record:
            records.append(record)
    return records


def _college_major_exam_name_score_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 7:
        return None
    college_code, college_name = values[:2]
    if not _looks_like_code(college_code, 3):
        return None
    rest = _clean_text(" ".join(values[2:]))
    match = re.fullmatch(
        r"(?P<major_code>\d{4}[0-9A-Z]{2})\s+"
        r"(?P<major_name>.+?)\s+"
        r"(?P<student_id>[0-9Xx*]{10,})\s+"
        r"(?P<person_name>[\u4e00-\u9fff路*]{1,12})\s+"
        r"(?P<initial_score>\d{2,3})\s+"
        r"(?P<reexam_score>\d{1,3}(?:\.\d+)?)\s+"
        r"(?P<total_score>\d{1,3}(?:\.\d+)?)(?:\s+(?P<extra>.*))?",
        rest,
    )
    if not match:
        return None
    groups = match.groupdict()
    person_name = groups["person_name"]
    if not _looks_like_masked_or_chinese_name(person_name):
        return None
    remarks = _clean_text(
        " ".join(
            [
                f"initial_score {groups['initial_score']}",
                f"reexam_score {groups['reexam_score']}",
                f"total_score {groups['total_score']}",
                groups.get("extra") or "",
            ]
        )
    )
    return _build_record(
        document,
        {
            "college": _clean_text(f"{college_code} {college_name}"),
            "admission_major": _clean_text(f"{groups['major_code']} {groups['major_name']}"),
            "student_id": groups["student_id"],
            "person_name": person_name,
            "remarks": remarks,
        },
    )


def _records_from_dhu_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    header_text = " ".join(" ".join(row) for row in rows[:10])
    if not (
        "\u4e1c\u534e\u5927\u5b66" in document_text
        or "dhu.edu.cn" in document_text.lower()
        or (
            "\u5b66\u4f4d\u7c7b\u578b" in header_text
            and "\u62df\u5f55\u53d6\u4e13\u4e1a\u540d\u79f0" in header_text
            and "\u521d\u8bd5\u6210\u7ee9" in header_text
        )
    ):
        return []

    pattern = re.compile(
        r"^(?P<college_code>\d{3})\s+"
        r"(?P<college>\S+?)\s+"
        r"(?P<student_id>\d{6}[0-9Xx*]{5,})\s+"
        r"(?P<person_name>.+?)\s+"
        r"(?P<degree_type>\u5b66\u672f\u5b66\u4f4d|\u4e13\u4e1a\u5b66\u4f4d)\s+"
        r"(?P<major_code>[0-9A-Za-z]{6})\s+"
        r"(?P<major_name>.+?)\s+"
        r"(?P<initial_score>\d{1,3})\s+"
        r"(?P<reexam_score>\d+(?:\.\d+)?)\s+"
        r"(?P<total_score>\d+(?:\.\d+)?)\s+"
        r"(?P<study_mode>\u5168\u65e5\u5236|\u975e\u5168\u65e5\u5236)"
        r"(?:\s+(?P<extra_remarks>.+))?$"
    )

    records: list[dict[str, Any]] = []
    for values in rows:
        line = _clean_text(" ".join(values))
        match = pattern.match(line)
        if not match:
            continue
        data = match.groupdict()
        remarks = [
            f"degree_type {data['degree_type']}",
            f"initial_score {data['initial_score']}",
            f"reexam_score {data['reexam_score']}",
            f"total_score {data['total_score']}",
            f"study_mode {data['study_mode']}",
        ]
        if data.get("extra_remarks"):
            remarks.append(data["extra_remarks"])

        record = _build_record(
            document,
            {
                "college": _clean_text(f"{data['college_code']} {data['college']}"),
                "student_id": data["student_id"],
                "person_name": data["person_name"],
                "admission_major": _clean_text(f"{data['major_code']} {data['major_name']}"),
                "remarks": "; ".join(remarks),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_shiep_score_only_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    header_text = " ".join(" ".join(row) for row in rows[:10])
    if not (
        "\u4e0a\u6d77\u7535\u529b\u5927\u5b66" in document_text
        or "shiep.edu.cn" in document_text.lower()
        or ("\u4e13\u9879\u8ba1\u5212\u7b49" in header_text and "\u62a5\u8003\u5b66\u4e60\u65b9\u5f0f" in header_text)
    ):
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        if len(values) < 8:
            continue
        person_name, student_id, initial_score, reexam_score, total_score = values[:5]
        plan = values[5]
        application_category = values[6]
        study_mode = values[7]
        extra = _clean_text(" ".join(values[8:]))
        if not _looks_like_masked_or_chinese_name(person_name):
            continue
        if not _looks_like_identifier_only(student_id):
            continue
        if not (
            _looks_like_score_or_metric(initial_score)
            and _looks_like_score_or_metric(reexam_score)
            and _looks_like_score_or_metric(total_score)
        ):
            continue

        remarks = _clean_text(
            " ".join(
                [
                    f"initial_score {initial_score}",
                    f"reexam_score {reexam_score}",
                    f"total_score {total_score}",
                    f"special_plan {plan}",
                    f"application_category {application_category}",
                    f"study_mode {study_mode}",
                    extra,
                ]
            )
        )
        record = _build_record(
            document,
            {
                "person_name": person_name,
                "student_id": student_id,
                "remarks": remarks,
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_ncut_retest_result_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    header_text = " ".join(" ".join(row) for row in rows[:12])
    if not (
        "\u5317\u65b9\u5de5\u4e1a\u5927\u5b66" in document_text
        or "ncut.edu.cn" in document_text.lower()
        or ("\u662f\u5426" in header_text and "\u62df\u5f55\u53d6" in header_text and "\u8003\u751f\u7f16\u53f7" in header_text)
    ):
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        if len(values) < 10:
            continue
        ranking, major_code, major_name, study_mode, student_id, person_name = values[:6]
        initial_score, reexam_score, total_score, admission_status = values[6:10]
        extra = _clean_text(" ".join(values[10:]))
        if not re.fullmatch(r"\d{1,4}", ranking or ""):
            continue
        if not re.fullmatch(r"\d{6}", major_code or ""):
            continue
        if not _looks_like_identifier_only(student_id):
            continue
        if not _looks_like_masked_or_chinese_name(person_name):
            continue
        if admission_status not in {"\u662f", "\u62df\u5f55\u53d6"}:
            continue
        if not (
            _looks_like_score_or_metric(initial_score)
            and _looks_like_score_or_metric(reexam_score)
            and _looks_like_score_or_metric(total_score)
        ):
            continue

        remarks = _clean_text(
            " ".join(
                [
                    f"study_mode {study_mode}",
                    f"initial_score {initial_score}",
                    f"reexam_score {reexam_score}",
                    f"total_score {total_score}",
                    f"admission_status {admission_status}",
                    extra,
                ]
            )
        )
        record = _build_record(
            document,
            {
                "ranking": ranking,
                "student_id": student_id,
                "person_name": person_name,
                "admission_major": _clean_text(f"{major_code} {major_name}"),
                "remarks": remarks,
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_henu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "\u6cb3\u5357\u5927\u5b66" not in document_text or "\u63a8\u514d" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        if len(values) < 9:
            continue
        college_code, college_name, name, identity, major_code, major_name = values[:6]
        if not re.fullmatch(r"\d{3}", college_code or ""):
            continue
        if not _looks_like_chinese_name(name):
            continue
        if not re.fullmatch(r"[\dXx*]{10,}", re.sub(r"\s+", "", identity or "")):
            continue
        if not re.fullmatch(r"\d{6}", major_code or ""):
            continue
        record = _build_record(
            document,
            {
                "college": _clean_text(f"{college_code} {college_name}"),
                "person_name": name,
                "student_id": identity,
                "admission_major": _clean_text(f"{major_code} {major_name}"),
                "remarks": _clean_text(" ".join(values[6:])),
            },
        )
        if record:
            records.append(record)
    return records


def _records_from_bucea_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "北京建筑大学" not in document_text and "bucea.edu.cn" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        record = _bucea_postgraduate_record(values, document)
        if record:
            records.append(record)
    return records


def _bucea_postgraduate_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 10:
        return None
    ranking, student_id, person_name, college, major_code, major_name = values[:6]
    if not re.fullmatch(r"\d{1,4}", ranking or ""):
        return None
    if not _looks_like_identifier_only(student_id):
        return None
    if not _looks_like_masked_or_chinese_name(person_name):
        return None
    if not _looks_like_major_code(major_code):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None
    return _build_record(
        document,
        {
            "ranking": ranking,
            "student_id": student_id,
            "person_name": person_name,
            "college": college,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": _clean_text(" ".join(values[6:])),
        },
    )


def _records_from_swupl_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "西南政法大学" not in document_text and "swupl.edu.cn" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    left_major = ""
    right_major = ""
    index = 0
    while index < len(rows):
        values = rows[index]
        group_headers = _swupl_group_headers(values)
        if group_headers:
            left_major = group_headers[0]
            right_major = group_headers[1] if len(group_headers) > 1 else group_headers[0]
            index += 1
            continue

        if index + 1 >= len(rows):
            index += 1
            continue
        next_values = rows[index + 1]
        pair_records = _swupl_records_from_score_identity_pair(
            values,
            next_values,
            document,
            left_major,
            right_major,
        )
        if pair_records:
            records.extend(pair_records)
            index += 2
            continue
        index += 1

    return records


def _swupl_group_headers(values: list[str]) -> list[str]:
    headers = [
        value
        for value in values
        if (
            "组" in value
            and not re.search(r"(排序|考生编号|姓名|页|共|第\s*\d+\s*页)", value)
            and not re.fullmatch(r"\d+", value)
        )
    ]
    return headers[:2]


def _swupl_records_from_score_identity_pair(
    score_values: list[str],
    identity_values: list[str],
    document: dict[str, Any],
    left_major: str,
    right_major: str,
) -> list[dict[str, Any]]:
    chunks: list[tuple[int, int, str]] = []
    if _swupl_score_chunk(score_values, 0):
        chunks.append((0, 0, left_major))
    if _swupl_score_chunk(score_values, 4):
        chunks.append((4, 2, right_major or left_major))

    records: list[dict[str, Any]] = []
    for score_offset, identity_offset, major in chunks:
        if identity_offset + 1 >= len(identity_values):
            continue
        student_id = identity_values[identity_offset]
        person_name = identity_values[identity_offset + 1]
        if not _looks_like_identifier_only(student_id):
            continue
        if not _looks_like_masked_or_chinese_name(person_name):
            continue
        ranking = score_values[score_offset]
        initial_score = score_values[score_offset + 1]
        interview_score = score_values[score_offset + 2]
        total_score = score_values[score_offset + 3]
        records.append(
            _build_record(
                document,
                {
                    "ranking": ranking,
                    "student_id": student_id,
                    "person_name": person_name,
                    "admission_major": major,
                    "remarks": _clean_text(
                        " ".join(
                            [
                                f"initial_score {initial_score}",
                                f"interview_score {interview_score}",
                                f"total_score {total_score}",
                            ]
                        )
                    ),
                },
            )
        )
    return [record for record in records if record]


def _swupl_score_chunk(values: list[str], offset: int) -> bool:
    return (
        offset + 3 < len(values)
        and bool(re.fullmatch(r"\d{1,4}", values[offset] or ""))
        and _looks_like_score_or_metric(values[offset + 1])
        and _looks_like_score_or_metric(values[offset + 2])
        and _looks_like_score_or_metric(values[offset + 3])
    )


def _records_from_fudan_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "\u590d\u65e6\u5927\u5b66" not in document_text and "gsao.fudan.edu.cn" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        if len(values) < 6:
            continue
        last_five, person_name, college, initial_score, reexam_score, total_score = values[:6]
        if not re.fullmatch(r"\d{5}", last_five or ""):
            continue
        if not _looks_like_masked_or_chinese_name(person_name):
            continue
        if not college or _looks_like_score_or_metric(college):
            continue
        if not re.fullmatch(r"\d{2,3}", initial_score or ""):
            continue
        if not (_looks_like_score_or_metric(reexam_score) and _looks_like_score_or_metric(total_score)):
            continue
        remarks = _clean_text(
            " ".join(
                [
                    f"initial_score {initial_score}",
                    f"reexam_score {reexam_score}",
                    f"total_score {total_score}",
                    *values[6:],
                ]
            )
        )
        records.append(
            _build_record(
                document,
                {
                    "student_id": last_five,
                    "person_name": person_name,
                    "college": college,
                    "remarks": remarks,
                },
            )
        )
    return records


def _records_from_jlu_pdf_text_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "吉林大学" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    last_college_by_code: dict[str, str] = {}
    for values in rows:
        record = _jlu_incoming_recommendation_record(values, document, last_college_by_code)
        if record:
            records.append(record)
    return records


def _jlu_incoming_recommendation_record(
    values: list[str],
    document: dict[str, Any],
    last_college_by_code: dict[str, str] | None = None,
) -> dict[str, Any] | None:
    if len(values) < 4:
        return None
    college_code = values[0]
    if not re.fullmatch(r"\d{3}", college_code or ""):
        return None

    remaining = values[1:]
    if not remaining:
        return None

    college_name = (last_college_by_code or {}).get(college_code, "")
    name_code_match = re.fullmatch(r"(.+?)\s+(\d{4}[0-9A-Z]{2})", remaining[0] or "")
    first_is_person_followed_by_major_code = bool(
        len(remaining) > 1
        and _looks_like_chinese_name(remaining[0])
        and not _jlu_looks_like_college_name(remaining[0])
        and re.fullmatch(r"\d{4}[0-9A-Z]{2}", remaining[1] or "")
    )
    if (
        not name_code_match
        and not re.fullmatch(r"\d{4}[0-9A-Z]{2}", remaining[0] or "")
        and not first_is_person_followed_by_major_code
    ):
        college_name = remaining.pop(0)
        if last_college_by_code is not None:
            last_college_by_code[college_code] = college_name

    if len(remaining) < 3:
        return None
    name_code_match = re.fullmatch(r"(.+?)\s+(\d{4}[0-9A-Z]{2})", remaining[0] or "")
    if name_code_match:
        name, major_code = name_code_match.groups()
        inline_college_name = _jlu_inline_college_name(name)
        if inline_college_name:
            college_name, name = inline_college_name
            if last_college_by_code is not None:
                last_college_by_code[college_code] = college_name
        remaining = remaining[1:]
    else:
        if len(remaining) < 4:
            return None
        name, major_code = remaining[:2]
        remaining = remaining[2:]

    if not remaining:
        return None
    if remaining[0].isdigit():
        major_name = ""
        study_years = remaining[0]
        admission_type = remaining[1] if len(remaining) > 1 else ""
        extra = remaining[2:]
    else:
        if len(remaining) < 3:
            return None
        major_name, study_years, admission_type = remaining[:3]
        extra = remaining[3:]

    if not _looks_like_chinese_name(name):
        return None
    if not re.fullmatch(r"\d{4}[0-9A-Z]{2}", major_code or ""):
        return None
    if not study_years.isdigit():
        return None
    if admission_type not in {"硕士", "直博生"}:
        return None
    college_value = _clean_text(f"{college_code} {college_name}") if college_name else college_code
    remarks = _clean_text(" ".join([study_years, admission_type, *extra]))
    return _build_record(
        document,
        {
            "college": college_value,
            "person_name": name,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": remarks,
        },
    )


def _jlu_looks_like_college_name(value: str) -> bool:
    return bool(re.search(r"(学院|中心|研究院|医院|学部|系|所|校区|实验室)", value or ""))


def _jlu_inline_college_name(value: str) -> tuple[str, str] | None:
    match = re.fullmatch(
        r"(.+(?:学院|中心|研究院|医院|学部|系|所|校区|实验室))\s+([\u4e00-\u9fff·]{1,8})",
        value or "",
    )
    if not match:
        return None
    return match.group(1), match.group(2)


def _records_from_college_section_pdf_text_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    header_index = _college_section_header_index(rows)
    if header_index is None:
        return []

    records: list[dict[str, Any]] = []
    current_college = ""
    pending_without_college: list[list[str]] = []
    for values in rows[header_index + 1 :]:
        if not values:
            continue
        if len(values) == 1:
            if _jlu_looks_like_college_name(values[0]):
                current_college = values[0]
                if pending_without_college:
                    for pending_values in pending_without_college:
                        record = _college_section_record(pending_values, current_college, document)
                        if record:
                            records.append(record)
                    pending_without_college = []
            continue
        record = _college_section_record(values, current_college, document)
        if record:
            if len(values) >= 5 and _jlu_looks_like_college_name(values[0]):
                current_college = values[0]
            records.append(record)
        elif len(values) >= 4 and not current_college and values[1] in {"男", "女"}:
            pending_without_college.append(values)
    return records


def _college_section_header_index(rows: list[list[str]]) -> int | None:
    for index, values in enumerate(rows[:30]):
        normalized = [_clean_text(value) for value in values]
        if (
            "学院" in normalized
            and "姓名" in normalized
            and "性别" in normalized
            and any("拟录取专业" in value for value in normalized)
            and any("复试成绩" in value for value in normalized)
        ):
            return index
    return None


def _college_section_record(
    values: list[str],
    current_college: str,
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) >= 5 and _jlu_looks_like_college_name(values[0]):
        college, name, gender, major, score = values[:5]
    elif len(values) >= 4 and current_college:
        college = current_college
        name, gender, major, score = values[:4]
    else:
        return None

    if gender not in {"男", "女"}:
        return None
    if not _looks_like_chinese_name(name):
        return None
    if _looks_like_non_person_label(name):
        return None
    if not major or _looks_like_non_person_label(major):
        return None
    remarks = _clean_text(f"{gender} 复试成绩 {score}") if score else gender
    return _build_record(
        document,
        {
            "person_name": name,
            "college": college,
            "admission_major": major,
            "remarks": remarks,
        },
    )


def _records_from_major_code_name_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for values in rows:
        record = _name_college_major_type_record(values, document) or _sequence_name_type_school_major_record(values, document)
        if record:
            records.append(record)
    return records


def _records_from_sxu_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    header_text = " ".join(" ".join(row) for row in rows[:20])
    if not (
        "山西大学" in document_text
        or "sxu.edu.cn" in document_text
        or "山西大学" in header_text
    ):
        return []
    if not any(_is_sxu_recommendation_header(row) for row in rows[:20]):
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        record = _sxu_recommendation_record(values, document)
        if record:
            records.append(record)
    return records


def _is_sxu_recommendation_header(values: list[str]) -> bool:
    normalized = [_clean_text(value) for value in values]
    return (
        "序号" in normalized
        and "姓名" in normalized
        and "录取层次" in normalized
        and any("录取学院" in value for value in normalized)
        and "录取专业代码" in normalized
        and "录取专业名称" in normalized
    )


def _sxu_recommendation_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 7:
        return None
    ranking, person_name, admission_level, college, major_code, major_name, retest_score = values[:7]
    extra_values = values[7:]
    if not ranking.isdigit():
        return None
    if not _looks_like_chinese_name(person_name):
        return None
    if admission_level not in {"直博生", "硕士"}:
        return None
    if not college or _looks_like_score_or_metric(college):
        return None
    if not _looks_like_major_code(major_code):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None
    if not _looks_like_score_or_metric(retest_score):
        return None

    return _build_record(
        document,
        {
            "ranking": ranking,
            "person_name": person_name,
            "college": college,
            "major": major_code,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": _clean_text(
                " ".join([admission_level, f"retest_score {retest_score}", *extra_values])
            ),
        },
    )


def _records_from_tyut_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    header_text = " ".join(" ".join(row) for row in rows[:20])
    if "太原理工大学" not in document_text and "tyut.edu.cn" not in document_text.lower():
        return []
    if not (
        "毕业院校" in header_text
        and "录取学院名称" in header_text
        and "录取专业名称" in header_text
        and "研究方向名称" in header_text
    ):
        return []

    records: list[dict[str, Any]] = []
    last_college_by_code: dict[str, str] = {}
    for values in rows:
        record = _tyut_recommendation_record(values, document, last_college_by_code)
        if record:
            records.append(record)
    return records


def _tyut_recommendation_record(
    values: list[str],
    document: dict[str, Any],
    last_college_by_code: dict[str, str],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in values if _clean_text(value)]
    if len(values) < 8:
        return None

    first_parts = values[0].split(maxsplit=1)
    if len(first_parts) == 2 and _looks_like_chinese_name(first_parts[0]):
        person_name, undergraduate_school = first_parts
        remaining = values[1:]
    elif len(values) >= 9 and _looks_like_chinese_name(values[0]):
        person_name, undergraduate_school = values[0], values[1]
        remaining = values[2:]
    else:
        return None

    if not _looks_like_chinese_name(person_name):
        return None
    if _looks_like_non_person_label(person_name):
        return None

    college_code = ""
    college_name = ""
    college_match = re.fullmatch(r"(?P<code>\d{3})\s+(?P<name>.+)", remaining[0])
    if college_match:
        college_code = college_match.group("code")
        college_name = college_match.group("name")
        last_college_by_code[college_code] = college_name
        remaining = remaining[1:]
    elif re.fullmatch(r"\d{3}", remaining[0]):
        college_code = remaining[0]
        college_name = last_college_by_code.get(college_code, "")
        remaining = remaining[1:]
    else:
        return None

    if len(remaining) < 5:
        return None
    major_code = ""
    major_name = ""
    major_match = re.fullmatch(r"(?P<code>\d{4,6}[A-Z]?\d*)\s+(?P<name>.+)", remaining[0])
    if major_match:
        major_code = major_match.group("code")
        major_name = major_match.group("name")
        remaining = remaining[1:]
    elif (
        len(remaining) >= 6
        and _looks_like_major_code(remaining[0])
        and not re.fullmatch(r"\d{2}", remaining[1])
    ):
        major_code, major_name = remaining[:2]
        remaining = remaining[2:]
    else:
        return None

    if not _looks_like_major_code(major_code):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None

    direction_code = ""
    direction_name = ""
    if remaining and re.fullmatch(r"\d{2}", remaining[0]):
        direction_code = remaining[0]
        remaining = remaining[1:]
    if len(remaining) >= 4:
        direction_name = remaining[0]
        remaining = remaining[1:]

    if len(remaining) < 3:
        return None
    reexam_score = remaining[-1]
    admission_type = remaining[-2]
    study_mode = remaining[-3]
    special_plan = _clean_text(" ".join(remaining[:-3]))
    if not _looks_like_score_or_metric(reexam_score):
        return None
    if "硕士" not in admission_type and "博士" not in admission_type and "直博" not in admission_type:
        return None
    if study_mode not in {"全日制", "非全日制"}:
        return None

    college = _clean_text(f"{college_code} {college_name}") if college_name else college_code
    remarks = _clean_text(
        " ".join(
            [
                f"direction {direction_code} {direction_name}".strip()
                if direction_code or direction_name
                else "",
                special_plan,
                study_mode,
                admission_type,
                f"reexam_score {reexam_score}",
            ]
        )
    )
    return _build_record(
        document,
        {
            "person_name": person_name,
            "undergraduate_school": undergraduate_school,
            "college": college,
            "major": major_code,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": remarks,
        },
    )


def _records_from_dlu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    header_text = " ".join(" ".join(row) for row in rows[:20])
    if not ("大连大学" in document_text or "dlu.edu.cn" in document_text.lower()):
        return []
    if not ("拟录取专业代码" in header_text and "拟录取专业名称" in header_text and "学习方式" in header_text):
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        record = _dlu_recommendation_record(values, document)
        if record:
            records.append(record)
    return records


def _dlu_recommendation_record(values: list[str], document: dict[str, Any]) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in values if _clean_text(value)]
    if len(values) < 8:
        return None
    ranking, person_name, college, major_code, major_name = values[:5]
    study_mode_index = next(
        (
            index
            for index in range(5, len(values))
            if values[index] in {"全日制", "非全日制"} or re.fullmatch(r"\d{1,3}(?:\.\d+)?\s+(全日制|非全日制)", values[index])
        ),
        None,
    )
    if study_mode_index is None or study_mode_index < 6 or study_mode_index + 1 >= len(values):
        return None
    score_mode_match = re.fullmatch(r"(?P<score>\d{1,3}(?:\.\d+)?)\s+(?P<mode>全日制|非全日制)", values[study_mode_index])
    if score_mode_match:
        direction = _clean_text(" ".join(values[5:study_mode_index]))
        reexam_score = score_mode_match.group("score")
        study_mode = score_mode_match.group("mode")
    else:
        if study_mode_index < 7:
            return None
        direction = _clean_text(" ".join(values[5 : study_mode_index - 1]))
        reexam_score = values[study_mode_index - 1]
        study_mode = values[study_mode_index]
    level = values[study_mode_index + 1]
    extra_values = values[study_mode_index + 2 :]

    if not ranking.isdigit():
        return None
    if not _looks_like_chinese_name(person_name):
        return None
    if not college or _looks_like_score_or_metric(college):
        return None
    if not _looks_like_major_code(major_code):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None
    if not _looks_like_score_or_metric(reexam_score):
        return None

    remarks = _clean_text(
        " ".join(
            [
                f"direction {direction}" if direction else "",
                f"reexam_score {reexam_score}",
                study_mode,
                level,
                *extra_values,
            ]
        )
    )
    return _build_record(
        document,
        {
            "ranking": ranking,
            "person_name": person_name,
            "college": college,
            "major": major_code,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": remarks,
        },
    )


def _records_from_nxmu_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if not ("宁夏医科大学" in document_text or "nxmu.edu.cn" in document_text.lower()):
        return []

    header_text = " ".join(" ".join(row) for row in rows[:12])
    if not all(
        term in header_text
        for term in ("考生编号", "录取院系所码", "录取专业代码", "录取专业名称", "录取成绩")
    ):
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        record = _nxmu_postgraduate_record(row, document)
        if record:
            records.append(record)
    return records


def _nxmu_postgraduate_record(
    row: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    if len(values) < 13:
        return None

    (
        student_id,
        person_name,
        plan,
        college_code,
        college_name,
        major_code,
        major_name,
        direction_code,
        direction_name,
        study_mode,
        initial_score,
        reexam_score,
        admission_score,
        *extra,
    ) = values

    if not _looks_like_identifier_only(student_id):
        return None
    if not _looks_like_chinese_name(person_name):
        return None
    if not re.fullmatch(r"\d{3}", college_code or ""):
        return None
    if not re.fullmatch(r"\d{4}[0-9A-Za-z]{2}", major_code or ""):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None
    if not all(_looks_like_score_or_metric(value) for value in (initial_score, reexam_score, admission_score)):
        return None

    remarks = [
        f"plan {plan}",
        f"direction_code {direction_code}",
        f"direction {direction_name}",
        f"study_mode {study_mode}",
        f"initial_score {initial_score}",
        f"reexam_score {reexam_score}",
        f"admission_score {admission_score}",
    ]
    extra_text = " ".join(value for value in extra if value)
    if extra_text:
        remarks.append(extra_text)

    return _build_record(
        document,
        {
            "person_name": person_name,
            "student_id": student_id,
            "college": f"{college_code} {college_name}",
            "major": major_code,
            "admission_major": f"{major_code} {major_name}",
            "remarks": "; ".join(remarks),
        },
    )


def _records_from_ujn_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if not ("济南大学" in document_text or "ujn.edu.cn" in document_text.lower()):
        return []

    header_text = " ".join(" ".join(row) for row in rows[:12])
    if not all(term in header_text for term in ("考生编号", "报考院系代码", "报考专业代码", "报考专业")):
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        record = _ujn_postgraduate_record(row, document)
        if record:
            records.append(record)
    return records


def _ujn_postgraduate_record(
    row: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    if len(values) < 9:
        return None

    sequence, student_id, person_name, college_code, college_name, major_code, major_name, study_mode, exam_method = values[:9]
    if not sequence.isdigit():
        return None
    if not _looks_like_identifier_only(student_id):
        return None
    if not _looks_like_chinese_name(person_name):
        return None
    if not re.fullmatch(r"\d{3}", college_code or ""):
        return None
    if not re.fullmatch(r"\d{4}[0-9A-Za-z]{2}", major_code or ""):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None

    return _build_record(
        document,
        {
            "person_name": person_name,
            "student_id": student_id,
            "college": f"{college_code} {college_name}",
            "major": major_code,
            "admission_major": f"{major_code} {major_name}",
            "ranking": sequence,
            "remarks": f"study_mode {study_mode}; exam_method {exam_method}",
        },
    )


def _records_from_xzhmu_score_only_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if not ("徐州医科大学" in document_text or "xzhmu.edu.cn" in document_text.lower()):
        return []

    header_text = " ".join(" ".join(row) for row in rows[:12])
    if not all(term in header_text for term in ("考生编号", "姓名", "初试成绩", "复试成绩")):
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        record = _xzhmu_score_only_record(row, document)
        if record:
            records.append(record)
    return records


def _xzhmu_score_only_record(
    row: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    if len(values) < 5:
        return None
    student_id, person_name, initial_score, reexam_score, total_score, *extra = values
    if not re.fullmatch(r"[0-9Xx*]{8,}", student_id or ""):
        return None
    if not _looks_like_chinese_name(person_name):
        return None
    if not all(_looks_like_score_or_metric(value) for value in (initial_score, reexam_score, total_score)):
        return None
    remarks = [
        f"initial_score {initial_score}",
        f"reexam_score {reexam_score}",
        f"total_score {total_score}",
    ]
    extra_text = " ".join(value for value in extra if value)
    if extra_text:
        remarks.append(extra_text)
    return _build_record(
        document,
        {
            "person_name": person_name,
            "student_id": student_id,
            "remarks": "; ".join(remarks),
        },
    )


def _records_from_zstu_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if not ("浙江理工大学" in document_text or "zstu.edu.cn" in document_text.lower()):
        return []

    header_text = " ".join(" ".join(row) for row in rows[:12])
    if not all(term in header_text for term in ("报名号", "考生姓名", "拟录取专业", "综合考核")):
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        record = _zstu_postgraduate_record(row, document)
        if record:
            records.append(record)
    return records


def _zstu_postgraduate_record(
    row: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    if len(values) < 7:
        return None
    ranking, student_id, person_name, total_score, admission_major, admission_category, degree_type, *extra = values
    if not ranking.isdigit():
        return None
    if not _looks_like_identifier_only(student_id):
        return None
    if not _looks_like_chinese_name(person_name):
        return None
    if not _looks_like_score_or_metric(total_score):
        return None
    if not admission_major or _looks_like_score_or_metric(admission_major):
        return None
    remarks = [
        f"total_score {total_score}",
        admission_category,
        degree_type,
    ]
    extra_text = " ".join(value for value in extra if value)
    if extra_text:
        remarks.append(extra_text)
    return _build_record(
        document,
        {
            "person_name": person_name,
            "student_id": student_id,
            "admission_major": admission_major,
            "ranking": ranking,
            "remarks": "; ".join(remarks),
        },
    )


def _records_from_gmc_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
            " ".join(" ".join(row) for row in rows[:8]),
        ]
    )
    if not ("贵州医科大学" in document_text or "gmc.edu.cn" in document_text.lower()):
        return []
    if not all(term in document_text for term in ("考生编号", "拟录取专业名称", "拟录取研究方向")):
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        values = _gmc_row_values(row)
        record = _gmc_postgraduate_record(values, document)
        if record:
            records.append(record)
    return records


def _gmc_postgraduate_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 10:
        return None
    student_id, person_name, college_code, college_name, major_code, major_name = values[:6]
    direction_code = values[6] if len(values) > 6 else ""
    direction_name = values[7] if len(values) > 7 else ""
    initial_score = values[8] if len(values) > 8 else ""
    score_tokens: list[str] = []
    for value in values[9:]:
        score_tokens.extend(_clean_text(part) for part in value.split() if _clean_text(part))
    reexam_score = score_tokens[0] if len(score_tokens) > 0 else ""
    total_score = score_tokens[1] if len(score_tokens) > 1 else ""

    if not _looks_like_identifier_only(student_id):
        return None
    if not _looks_like_chinese_name(person_name):
        return None
    if not re.fullmatch(r"\d{3}", college_code or ""):
        return None
    if not re.fullmatch(r"\d{4,6}[A-Z0-9]*", major_code or ""):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None
    if not direction_code or _looks_like_score_or_metric(direction_name):
        return None

    remarks = [
        f"direction_code {direction_code}",
        f"direction {direction_name}",
        f"initial_score {initial_score}",
        f"reexam_score {reexam_score}",
        f"total_score {total_score}",
    ]
    return _build_record(
        document,
        {
            "person_name": person_name,
            "student_id": student_id,
            "college": f"{college_code} {college_name}",
            "major": major_code,
            "admission_major": f"{major_code} {major_name}",
            "remarks": "; ".join(value for value in remarks if value and not value.endswith(" ")),
        },
    )


def _gmc_row_values(row: list[str]) -> list[str]:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    if len(values) == 1 and re.match(r"^\d{10,}\s+", values[0]):
        return [_clean_text(value) for value in values[0].split() if _clean_text(value)]
    return values


def _records_from_gxau_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if not ("广西艺术学院" in document_text or "gxau.edu.cn" in document_text.lower()):
        return []

    header_text = " ".join(" ".join(row) for row in rows[:12])
    if "复试成绩及拟录取名单" not in header_text and "拟录取" not in header_text:
        return []

    records: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        values = _gxau_row_values(row)
        if not values or not values[0].isdigit():
            continue
        if "拟录取" not in values:
            continue
        record = _gxau_postgraduate_record(values, rows, index, document)
        if record:
            records.append(record)
    return records


def _gxau_postgraduate_record(
    values: list[str],
    rows: list[list[str]],
    index: int,
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 13:
        return None

    sequence = values[0]
    if len(values) > 2 and _looks_like_chinese_name(values[2]):
        college = values[1]
        person_name = values[2]
        remaining = values[3:]
    elif len(values) > 1 and _looks_like_chinese_name(values[1]):
        college = _gxau_context_college(rows, index)
        person_name = values[1]
        remaining = values[2:]
    else:
        return None

    student_id = remaining[0]
    if len(remaining) >= 5 and _looks_like_identifier_only(remaining[0]) and re.fullmatch(r"\d{1,4}", remaining[1] or ""):
        exam_number = remaining[1]
        major_value = remaining[2]
        direction = remaining[3] if len(remaining) > 3 else ""
        study_mode = remaining[4] if len(remaining) > 4 else ""
        scores = remaining[5:]
    else:
        exam_number = remaining[0]
        major_value = remaining[1] if len(remaining) > 1 else ""
        direction = remaining[2] if len(remaining) > 2 else ""
        study_mode = remaining[3] if len(remaining) > 3 else ""
        scores = remaining[4:]

    major_code, major_name = _split_gxau_major_code_name(major_value)
    if not major_code or not major_name:
        return None
    if study_mode not in {"全日制", "非全日制"}:
        return None

    parsed_scores = _gxau_score_fields(scores, rows, index)
    if not parsed_scores:
        return None

    remarks = [
        f"direction {direction}",
        f"study_mode {study_mode}",
        f"degree_type {_gxau_degree_type(rows, index)}",
        f"initial_score {parsed_scores['initial_score']}",
        f"english_score {parsed_scores['english_score']}",
        f"interview_score {parsed_scores['interview_score']}",
        f"course {parsed_scores['course']}",
        f"course_score {parsed_scores['course_score']}",
        f"reexam_score {parsed_scores['reexam_score']}",
        f"total_score {parsed_scores['total_score']}",
        f"status {parsed_scores['status']}",
    ]
    if exam_number and exam_number != student_id:
        remarks.append(f"exam_number {exam_number}")
    if parsed_scores.get("extra"):
        remarks.append(parsed_scores["extra"])

    return _build_record(
        document,
        {
            "person_name": person_name,
            "student_id": student_id,
            "college": college,
            "major": major_code,
            "admission_major": f"{major_code} {major_name}",
            "ranking": sequence,
            "remarks": "; ".join(value for value in remarks if value and not value.endswith(" ")),
        },
    )


def _gxau_score_fields(
    values: list[str],
    rows: list[list[str]],
    index: int,
) -> dict[str, str] | None:
    if len(values) < 7:
        return None
    initial_score = values[0]
    english_score = values[1] if len(values) > 1 else ""
    interview_score = values[2] if len(values) > 2 else ""
    offset = 3
    if len(values) > offset and not _looks_like_score_or_metric(values[offset]):
        course = values[offset]
        offset += 1
    else:
        course = _gxau_context_course(rows, index)
    if len(values) <= offset + 3:
        return None
    course_score = values[offset]
    reexam_score = values[offset + 1]
    total_score = values[offset + 2]
    status = values[offset + 3]
    extra = " ".join(values[offset + 4:])
    if status != "拟录取":
        return None
    if not all(
        _gxau_score_or_absent(value)
        for value in [initial_score, english_score, interview_score, course_score, reexam_score, total_score]
    ):
        return None
    return {
        "initial_score": initial_score,
        "english_score": english_score,
        "interview_score": interview_score,
        "course": course,
        "course_score": course_score,
        "reexam_score": reexam_score,
        "total_score": total_score,
        "status": status,
        "extra": extra,
    }


def _split_gxau_major_code_name(value: str) -> tuple[str, str]:
    match = re.match(r"^(\d{6}[A-Z]?)(.+)$", _clean_text(value))
    if not match:
        return "", ""
    return match.group(1), _clean_text(match.group(2))


def _gxau_score_or_absent(value: str) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.\d+)?", value or "") or value == "缺考")


def _gxau_degree_type(rows: list[list[str]], index: int) -> str:
    nearby = _gxau_nearby_tokens(rows, index, before=1, after=1)
    if "学术" in nearby and "学位" in nearby:
        return "学术学位"
    if "专业" in nearby and "学位" in nearby:
        return "专业学位"
    return ""


def _gxau_context_college(rows: list[list[str]], index: int) -> str:
    for row in reversed(rows[max(0, index - 3):index]):
        values = _gxau_row_values(row)
        for value in values:
            if re.match(r"^\d{3}[\u4e00-\u9fff]", value or "") and not _gxau_course_fragment(value):
                suffix = ""
                if index + 1 < len(rows):
                    next_values = [_clean_text(item) for item in rows[index + 1] if _clean_text(item)]
                    if next_values and next_values[0] in {"学院", "研究院"}:
                        suffix = next_values[0]
                return f"{value}{suffix}"
    return ""


def _gxau_context_course(rows: list[list[str]], index: int) -> str:
    fragments: list[str] = []
    for value in _gxau_nearby_tokens(rows, index, before=2, after=2):
        if not fragments:
            if _gxau_course_fragment(value):
                fragments.append(value)
            continue
        if _gxau_course_continuation(value):
            fragments.append(value)
    return "".join(fragments)


def _gxau_nearby_tokens(
    rows: list[list[str]],
    index: int,
    *,
    before: int,
    after: int,
) -> list[str]:
    tokens: list[str] = []
    start = max(0, index - before)
    stop = min(len(rows), index + after + 1)
    for row_index in range(start, stop):
        if row_index == index:
            continue
        tokens.extend(_gxau_row_values(rows[row_index]))
    return tokens


def _gxau_row_values(row: list[str]) -> list[str]:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    if len(values) == 1:
        text = values[0]
        if re.match(r"^\d+\s+", text) or re.match(r"^(学术|专业)\s+", text) or text.startswith("学位 "):
            return [_clean_text(value) for value in text.split() if _clean_text(value)]
    return values


def _gxau_course_fragment(value: str) -> bool:
    text = _clean_text(value)
    return bool(re.match(r"^\d{3}[\u4e00-\u9fff]", text or "") and "学院" not in text and "研究院" not in text)


def _gxau_course_continuation(value: str) -> bool:
    text = _clean_text(value)
    if not text or text in {"学术", "专业", "学位", "学院", "研究院"}:
        return False
    if re.match(r"^\d{3}", text):
        return False
    if re.search(r"(上线|拟录取|少数民族|政策|第\s*\d+\s*页)", text):
        return False
    return bool(re.fullmatch(r"[\u4e00-\u9fff（）()与、]+", text))


def _records_from_lzjtu_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
            " ".join(" ".join(row) for row in rows[:8]),
        ]
    )
    if not ("兰州交通大学" in document_text or "lzjtu.edu.cn" in document_text.lower()):
        return []
    if not ("考生编号" in document_text and "拟录取专业" in document_text):
        return []

    college = _lzjtu_college_from_document(document_text)
    records: list[dict[str, Any]] = []
    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if len(values) == 1 and re.match(r"^\d{10,}\s+", values[0]):
            values = [_clean_text(value) for value in values[0].split() if _clean_text(value)]
        if len(values) < 4:
            continue
        student_id, person_name, major_code = values[:3]
        major_name = "".join(values[3:])
        if not _looks_like_identifier_only(student_id):
            continue
        if not _looks_like_chinese_name(person_name):
            continue
        if not re.fullmatch(r"\d{4,6}[A-Z0-9]*", major_code or ""):
            continue
        if not major_name or _looks_like_score_or_metric(major_name):
            continue
        records.append(
            _build_record(
                document,
                {
                    "person_name": person_name,
                    "student_id": student_id,
                    "college": college,
                    "major": major_code,
                    "admission_major": f"{major_code} {major_name}",
                },
            )
        )
    return records


def _records_from_guet_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if document.get("document_type") != "postgraduate_admission_list":
        return []
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "桂林电子科技大学" not in document_text and "guet.edu.cn" not in document_text.lower():
        return []

    records: list[dict[str, Any]] = []
    pending_ranking = ""
    for values in rows:
        cleaned_values = [_clean_text(value) for value in values if _clean_text(value)]
        if len(cleaned_values) == 1 and re.fullmatch(r"\d+", cleaned_values[0]):
            pending_ranking = cleaned_values[0]
            continue
        record = _guet_postgraduate_pdf_record(
            cleaned_values,
            document,
            pending_ranking=pending_ranking,
        )
        if record:
            records.append(record)
            pending_ranking = ""
    return records


def _guet_postgraduate_pdf_record(
    values: list[str],
    document: dict[str, Any],
    *,
    pending_ranking: str = "",
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in values if _clean_text(value)]
    if not values:
        return None

    ranking = ""
    college = ""
    student_id = ""
    person_name = ""
    major_code = ""
    major_name = ""
    tail: list[str] = []

    if pending_ranking and len(values) >= 8:
        ranking = pending_ranking
        college, student_id, person_name = values[:3]
        major_code, major_name, major_tail = _split_guet_major_code_name_score(values[3])
        if major_code and not major_name and len(values) > 4:
            major_name = values[4]
            tail = values[5:]
        else:
            tail = [*major_tail, *values[4:]]

    if not person_name and len(values) >= 4 and re.fullmatch(r"\d+", values[0]):
        ranking = values[0]
        college_identity_name = re.fullmatch(
            r"(?P<college>.+?(?:学院|研究院\d{0,2}|中心|系|部))"
            r"(?P<student_id>[0-9Xx*]{10,})\s+"
            r"(?P<person_name>[\u4e00-\u9fff·]{2,6})",
            values[1],
        )
        if college_identity_name and _looks_like_major_code(values[2]):
            college = college_identity_name.group("college")
            student_id = college_identity_name.group("student_id")
            person_name = college_identity_name.group("person_name")
            major_code = values[2]
            major_name = values[3] if len(values) > 3 else ""
            tail = values[4:]

    if not person_name and len(values) >= 4 and re.fullmatch(r"\d+", values[0]):
        ranking = values[0]
        college_identity = re.fullmatch(
            r"(?P<college>.+?)\s+(?P<student_id>[0-9Xx*]{10,})",
            values[1],
        )
        if college_identity and _looks_like_chinese_name(values[2]):
            major_code, major_name, major_tail = _split_guet_major_code_name_score(values[3])
            if major_code:
                college = college_identity.group("college")
                student_id = college_identity.group("student_id")
                person_name = values[2]
                if not major_name and len(values) > 4:
                    major_name = values[4]
                    tail = values[5:]
                else:
                    tail = [*major_tail, *values[4:]]

    if not person_name and len(values) >= 6 and re.fullmatch(r"\d+", values[0]):
        ranking = values[0]
        college = values[1]
        if (
            _looks_like_identifier_only(values[2])
            and _looks_like_chinese_name(values[3])
        ):
            student_id = values[2]
            person_name = values[3]
            major_code, major_name, major_tail = _split_guet_major_code_name_score(values[4])
            if major_code and not major_name and len(values) > 5:
                major_name = values[5]
                tail = values[6:]
            else:
                tail = [*major_tail, *values[5:]]
        else:
            identity_name = re.fullmatch(
                r"(?P<student_id>[0-9Xx*]{10,})\s+(?P<person_name>[\u4e00-\u9fff·]{2,6})",
                values[2],
            )
            if identity_name:
                student_id = identity_name.group("student_id")
                person_name = identity_name.group("person_name")
                major_code, major_name, major_tail = _split_guet_major_code_name_score(values[3])
                if major_code and not major_name and len(values) > 4:
                    major_name = values[4]
                    tail = values[5:]
                else:
                    tail = [*major_tail, *values[4:]]

    if not person_name:
        compact_head = re.fullmatch(
            r"(?P<ranking>\d+)\s+(?P<college>.+?)\s+(?P<student_id>[0-9Xx*]{10,})",
            values[0],
        )
        if compact_head and len(values) >= 2:
            compact_major = re.fullmatch(
                r"(?P<person_name>[\u4e00-\u9fff·]{2,6})\s+"
                r"(?P<major_value>.+)",
                values[1],
            )
            if compact_major:
                major_code, major_name, major_tail = _split_guet_major_code_name_score(
                    compact_major.group("major_value")
                )
                ranking = compact_head.group("ranking")
                college = compact_head.group("college")
                student_id = compact_head.group("student_id")
                person_name = compact_major.group("person_name")
                tail = [*major_tail, *values[2:]]
                if not major_name and tail and not _looks_like_score_or_metric(tail[0]):
                    major_name = tail[0]
                    tail = tail[1:]

    if not (
        ranking
        and college
        and _looks_like_identifier_only(student_id)
        and _looks_like_chinese_name(person_name)
        and _looks_like_major_code(major_code)
        and major_name
        and not _looks_like_score_or_metric(major_name)
    ):
        return None

    return _build_record(
        document,
        {
            "ranking": ranking,
            "person_name": person_name,
            "student_id": student_id,
            "college": college,
            "major": major_code,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": _clean_text(" ".join(tail)),
        },
    )


def _split_guet_major_code_name_score(value: str) -> tuple[str, str, list[str]]:
    parts = _clean_text(value).split()
    if not parts or not _looks_like_major_code(parts[0]):
        return "", "", []
    if len(parts) == 1:
        return parts[0], "", []

    extra: list[str] = []
    name_parts = parts[1:]
    if len(name_parts) > 1 and _looks_like_score_or_metric(name_parts[-1]):
        extra.append(f"initial_score {name_parts.pop()}")
    return parts[0], _clean_text(" ".join(name_parts)), extra


def _lzjtu_college_from_document(document_text: str) -> str:
    match = re.search(r"([\u4e00-\u9fff]+学院)", document_text)
    return match.group(1) if match else ""


def _records_from_gxnu_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if not ("广西师范大学" in document_text or "yz.gxnu.edu.cn" in document_text.lower()):
        return []

    header_text = " ".join(" ".join(row) for row in rows[:12])
    if not all(term in header_text for term in ("考生编号", "姓名", "学习方式", "专业代码", "复试成绩", "总成绩")):
        return []

    records: list[dict[str, Any]] = []
    pending_name_prefix = ""
    pending_record_values: list[str] | None = None

    for row in rows:
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if not values:
            continue

        if len(values) == 1:
            value = values[0]
            if _gxnu_ignorable_pdf_fragment(value):
                continue
            if pending_record_values and _gxnu_name_fragment(value):
                record = _gxnu_postgraduate_record(
                    pending_record_values,
                    document,
                    person_name=_clean_text(f"{pending_name_prefix}{value}"),
                )
                if record:
                    records.append(record)
                pending_record_values = None
                pending_name_prefix = ""
                continue
            if _gxnu_name_fragment(value):
                pending_name_prefix = value
            continue

        values = _gxnu_normalized_pdf_values(values)
        record = _gxnu_postgraduate_record(values, document)
        if record:
            records.append(record)
            pending_name_prefix = ""
            pending_record_values = None
            continue

        if pending_name_prefix and _gxnu_values_need_pending_name(values):
            pending_record_values = values

    if pending_record_values and pending_name_prefix:
        record = _gxnu_postgraduate_record(
            pending_record_values,
            document,
            person_name=pending_name_prefix,
        )
        if record:
            records.append(record)

    return records


def _gxnu_normalized_pdf_values(values: list[str]) -> list[str]:
    if not values:
        return values
    match = re.fullmatch(r"(\d{10,})\s+(.+)", values[0])
    if match:
        return [match.group(1), _clean_text(match.group(2)), *values[1:]]
    return values


def _gxnu_values_need_pending_name(values: list[str]) -> bool:
    return (
        len(values) >= 9
        and _looks_like_identifier_only(values[0])
        and re.fullmatch(r"\d{3}", values[1] or "") is not None
        and values[2] in {"全日制", "非全日制"}
        and re.fullmatch(r"\d{6}", values[3] or "") is not None
    )


def _gxnu_postgraduate_record(
    values: list[str],
    document: dict[str, Any],
    person_name: str | None = None,
) -> dict[str, Any] | None:
    if person_name is None:
        if len(values) < 10:
            return None
        student_id, name, unit_code, study_mode, major_code, major_name = values[:6]
        tail = values[6:]
    else:
        if len(values) < 9:
            return None
        student_id, unit_code, study_mode, major_code, major_name = values[:5]
        name = person_name
        tail = values[5:]

    if not _looks_like_identifier_only(student_id):
        return None
    if not _gxnu_name_fragment(name):
        return None
    if not re.fullmatch(r"\d{3}", unit_code or ""):
        return None
    if study_mode not in {"全日制", "非全日制"}:
        return None
    if not re.fullmatch(r"\d{6}", major_code or ""):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None

    direction_code = tail[0] if len(tail) > 0 else ""
    initial_score = tail[1] if len(tail) > 1 else ""
    reexam_score = tail[2] if len(tail) > 2 else ""
    total_score = tail[3] if len(tail) > 3 else ""
    extra = " ".join(tail[4:])
    remarks = [
        f"study_mode {study_mode}",
        f"direction_code {direction_code}",
        f"initial_score {initial_score}",
        f"reexam_score {reexam_score}",
        f"total_score {total_score}",
    ]
    if extra:
        remarks.append(extra)

    return _build_record(
        document,
        {
            "person_name": name,
            "student_id": student_id,
            "college": f"报考单位 {unit_code}",
            "major": major_code,
            "admission_major": f"{major_code} {major_name}",
            "remarks": "; ".join(remarks),
        },
    )


def _gxnu_name_fragment(value: str) -> bool:
    text = _clean_text(value)
    if not text or _gxnu_ignorable_pdf_fragment(text):
        return False
    return bool(re.fullmatch(r"[\u4e00-\u9fffA-Za-z·]{2,24}", text))


def _gxnu_ignorable_pdf_fragment(value: str) -> bool:
    text = _clean_text(value)
    if not text:
        return True
    if text.lower() in {"ww", "w.", "yz", ".g", "xn", "u.", "ed", "cn"}:
        return True
    if re.fullmatch(r"第\s*\d+\s*页[，,]\s*共\s*\d+\s*页", text):
        return True
    return bool(re.search(r"(考生编号|姓名|学习方式|专业代码|专业名称|初试|复试|总成绩|专项计划|备注|报考|研究方向代|单位|^码$)", text))


def _records_from_glut_split_major_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if not ("桂林理工大学" in document_text or "glut.edu.cn" in document_text.lower()):
        return []

    header_text = " ".join(" ".join(row) for row in rows[:8])
    if not all(term in header_text for term in ("考生性质", "考生姓名", "拟录取专业代码", "复试成绩")):
        return []

    records: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if len(values) < 10:
            continue
        sequence, candidate_type, person_name, college_code, major_code = values[:5]
        if not sequence.isdigit():
            continue
        if not _looks_like_chinese_name(person_name):
            continue
        if not re.fullmatch(r"\d{3}", college_code or ""):
            continue
        if not _looks_like_major_code(major_code):
            continue
        if not (_looks_like_score_or_metric(values[5]) and _looks_like_score_or_metric(values[6])):
            continue

        previous_fragments = _glut_split_name_fragments(rows[index - 1] if index > 0 else [])
        next_fragments = _glut_split_name_fragments(rows[index + 1] if index + 1 < len(rows) else [])
        college_name = _clean_text(
            f"{previous_fragments[0] if previous_fragments else ''}{next_fragments[0] if next_fragments else ''}"
        )
        major_name = _clean_text(
            f"{previous_fragments[1] if previous_fragments else ''}{next_fragments[1] if next_fragments else ''}"
        )
        remarks = _clean_text(
            " ".join(
                [
                    candidate_type,
                    f"reexam_score {values[5]}",
                    f"admission_score {values[6]}",
                    *values[7:],
                ]
            )
        )
        record = _build_record(
            document,
            {
                "ranking": sequence,
                "person_name": person_name,
                "college": _clean_text(f"{college_code} {college_name}"),
                "major": major_code,
                "admission_major": _clean_text(f"{major_code} {major_name}"),
                "remarks": remarks,
            },
        )
        if record:
            records.append(record)

    return records


def _glut_split_name_fragments(row: list[str]) -> tuple[str, str] | None:
    values = [_clean_text(value) for value in row if _clean_text(value)]
    if len(values) != 2:
        return None
    if any(_looks_like_score_or_metric(value) for value in values):
        return None
    return values[0], values[1]


def _records_from_lzu_law_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    header_text = " ".join(" ".join(row) for row in rows[:20])
    if not ("兰州大学" in document_text or "lzu.edu.cn" in document_text.lower()):
        return []
    if not (
        "报考专业" in header_text
        and "报考研究方向" in header_text
        and "专业面试成绩" in header_text
        and "复试总成绩" in header_text
    ):
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        record = _lzu_law_recommendation_record(values, document)
        if record:
            records.append(record)
    return records


def _lzu_law_recommendation_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = [_clean_text(value) for value in values if _clean_text(value)]
    if len(values) < 7:
        return None
    sequence, admission_major, research_direction, person_name = values[:4]
    interview_score, oral_listening_score, total_score = values[4:7]
    status = values[7] if len(values) >= 8 else ""
    extra_values = values[8:] if len(values) >= 8 else []
    person_name = _lzu_law_clean_person_name(person_name)

    if not sequence.isdigit():
        return None
    if status and "拟录取" not in status:
        return None
    if not (_looks_like_chinese_name(person_name) or _looks_like_dotted_person_name(person_name)):
        return None
    if not admission_major or _looks_like_score_or_metric(admission_major):
        return None
    if not research_direction or _looks_like_score_or_metric(research_direction):
        return None
    if not (
        _looks_like_score_or_metric(interview_score)
        and _looks_like_score_or_metric(oral_listening_score)
        and _looks_like_score_or_metric(total_score)
    ):
        return None

    return _build_record(
        document,
        {
            "ranking": sequence,
            "person_name": person_name,
            "major": research_direction,
            "admission_major": admission_major,
            "remarks": _clean_text(
                " ".join(
                    [
                        f"interview_score {interview_score}",
                        f"oral_listening_score {oral_listening_score}",
                        f"total_score {total_score}",
                        status,
                        *extra_values,
                    ]
                )
            ),
        },
    )


def _lzu_law_clean_person_name(value: str) -> str:
    text = _clean_text(value)
    match = re.fullmatch(r"(?P<name>[\u4e00-\u9fff·]{2,6})\d{2,4}", text)
    return match.group("name") if match else text


def _name_college_major_type_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 5:
        return None
    name, college, major_code, major_name, admission_type = values[:5]
    if not _looks_like_chinese_name(name):
        return None
    if not _jlu_looks_like_college_name(college):
        return None
    if not _looks_like_major_code(major_code):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None
    if not re.search(r"(推免|硕士|直博)", admission_type or ""):
        return None
    return _build_record(
        document,
        {
            "person_name": name,
            "college": college,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": admission_type,
        },
    )


def _sequence_name_type_school_major_record(
    values: list[str],
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if len(values) < 8:
        return None
    sequence, name, admission_type, undergraduate_school, college, major_code, major_name = values[:7]
    if not sequence.isdigit():
        return None
    if not _looks_like_chinese_name(name):
        return None
    if admission_type not in {"硕士", "直博生"}:
        return None
    if not _jlu_looks_like_college_name(college):
        return None
    if not _looks_like_major_code(major_code):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None
    return _build_record(
        document,
        {
            "ranking": sequence,
            "person_name": name,
            "undergraduate_school": undergraduate_school,
            "college": college,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": _clean_text(" ".join([admission_type, *values[7:]])),
        },
    )


def _looks_like_major_code(value: str) -> bool:
    return bool(re.fullmatch(r"\d{4}[0-9A-Z]{2}", value or ""))


def _records_from_college_code_first_text_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for values in rows:
        if len(values) < 6:
            continue
        college_code, college_name, student_id, person_name, major_code, major_name = values[:6]
        if not re.fullmatch(r"\d{2,4}", college_code or ""):
            continue
        if not _looks_like_identifier_only(student_id):
            continue
        if not _looks_like_masked_or_chinese_name(person_name):
            continue
        if not re.fullmatch(r"\d{4,6}", major_code or ""):
            continue
        normalized = {
            "college": _clean_text(f"{college_code} {college_name}"),
            "student_id": student_id,
            "person_name": person_name,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": _clean_text(" ".join(values[8:])) if len(values) > 8 else "",
        }
        record = _build_record(document, normalized)
        if record:
            records.append(record)
    return records


def _looks_like_masked_or_chinese_name(value: str) -> bool:
    text = _clean_text(value)
    return bool(re.fullmatch(r"[\u4e00-\u9fff·*]{1,12}", text)) and bool(
        re.search(r"[\u4e00-\u9fff]", text)
    )


def _records_from_zjut_pdf_text_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for values in rows:
        record = _zjut_postgraduate_record(values, document) or _zjut_recommendation_record(values, document)
        if record:
            records.append(record)
    return records


def _zjut_postgraduate_record(values: list[str], document: dict[str, Any]) -> dict[str, Any] | None:
    if len(values) < 9:
        return None
    college_code, college_name, major_code, major_name, study_mode, plan, name, exam_no, identity = values[:9]
    if not (_looks_like_code(college_code, 3) and _looks_like_code(major_code, 6)):
        return None
    if study_mode not in {"全日制", "非全日制"}:
        return None
    if not _looks_like_chinese_name(name):
        return None
    if not _looks_like_identifier_only(exam_no):
        return None
    if not _looks_like_masked_identifier(identity):
        return None
    return _build_record(
        document,
        {
            "college": _clean_text(f"{college_code} {college_name}"),
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "person_name": name,
            "student_id": exam_no,
            "remarks": _clean_text(f"{study_mode} {plan} {identity}"),
        },
    )


def _zjut_recommendation_record(values: list[str], document: dict[str, Any]) -> dict[str, Any] | None:
    if len(values) < 9:
        return None
    sequence, name, identity, gender, admission_type, college_code, college_name, major_code, major_name = values[:9]
    if not sequence.isdigit():
        return None
    if not _looks_like_chinese_name(name):
        return None
    if not _looks_like_masked_identifier(identity):
        return None
    if gender not in {"男", "女"}:
        return None
    if not (_looks_like_code(college_code, 3) and _looks_like_code(major_code, 6)):
        return None
    return _build_record(
        document,
        {
            "person_name": name,
            "student_id": identity,
            "college": _clean_text(f"{college_code} {college_name}"),
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "ranking": sequence,
            "remarks": admission_type,
        },
    )


def _records_from_cumt_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "中国矿业大学" not in document_text and "cumt.edu.cn" not in document_text:
        return []
    if "推荐免试" not in document_text and "tuimian" not in document_text.lower():
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        record = _cumt_recommendation_record(values, document)
        if record:
            records.append(record)
    return records


def _cumt_recommendation_record(values: list[str], document: dict[str, Any]) -> dict[str, Any] | None:
    if len(values) < 8:
        return None
    sequence, person_name, college_code, college_name, major_code, major_name, score, admission_type = values[:8]
    if not sequence.isdigit():
        return None
    if not _looks_like_chinese_name(person_name):
        return None
    if not (_looks_like_code(college_code, 3) and _looks_like_code(major_code, 6)):
        return None
    if not college_name or _looks_like_score_or_metric(college_name):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None
    if not _looks_like_score_or_metric(score):
        return None
    if admission_type not in {"硕士", "直博生"}:
        return None

    return _build_record(
        document,
        {
            "person_name": person_name,
            "college": _clean_text(f"{college_code} {college_name}"),
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "ranking": sequence,
            "remarks": _clean_text(" ".join([f"score {score}", admission_type, *values[8:]])),
        },
    )


def _records_from_xzmu_recommendation_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "西藏民族大学" not in document_text and "xzmu.edu.cn" not in document_text:
        return []
    if "推荐免试" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for values in rows:
        record = _xzmu_recommendation_record(values, document)
        if record:
            records.append(record)
    return records


def _xzmu_recommendation_record(values: list[str], document: dict[str, Any]) -> dict[str, Any] | None:
    if len(values) < 4:
        return None
    name_identity, major_code, major_name, score = values[:4]
    identity_match = re.fullmatch(r"([\u4e00-\u9fff·]{1,12})\s+([0-9Xx*]{10,})", name_identity or "")
    if not identity_match:
        return None
    person_name, identity = identity_match.groups()
    if not _looks_like_chinese_name(person_name):
        return None
    if not (_looks_like_masked_identifier(identity) and _looks_like_code(major_code, 6)):
        return None
    if not major_name or _looks_like_score_or_metric(major_name):
        return None
    if not _looks_like_score_or_metric(score):
        return None
    return _build_record(
        document,
        {
            "person_name": person_name,
            "student_id": identity,
            "admission_major": _clean_text(f"{major_code} {major_name}"),
            "remarks": f"score {score}",
        },
    )


def _records_from_njtech_adjustment_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    document_text = " ".join(
        [
            str(document.get("school_name") or ""),
            str(document.get("title") or ""),
            str(document.get("source_url") or ""),
        ]
    )
    if "南京工业大学" not in document_text and "njtech.edu.cn" not in document_text:
        return []
    if "拟录取" not in document_text:
        return []

    records: list[dict[str, Any]] = []
    for index, values in enumerate(rows):
        record = _njtech_adjustment_record(rows, index, document)
        if record:
            records.append(record)
    return records


def _njtech_adjustment_record(
    rows: list[list[str]],
    index: int,
    document: dict[str, Any],
) -> dict[str, Any] | None:
    values = rows[index]
    if len(values) < 8:
        return None
    sequence, student_id, person_name = values[:3]
    if not sequence.isdigit():
        return None
    if not _looks_like_identifier_only(student_id):
        return None
    if not _looks_like_chinese_name(person_name):
        return None

    college = ""
    major = ""
    category = ""
    study_mode = ""
    score_values: list[str] = []

    if len(values) >= 10 and _njtech_looks_like_major_code_name(values[4]):
        college = values[3]
        major = values[4]
        category, study_mode = values[5:7]
        score_values = values[7:10]
    elif len(values) >= 8 and _njtech_looks_like_major_code_name(values[4]):
        college = values[3]
        major = values[4]
        category, study_mode = values[5:7]
        score_values = [values[7], *_njtech_following_score_values(rows, index)]
    elif len(values) >= 9 and _njtech_looks_like_major_code_name(values[3]):
        college = _njtech_nearby_college(rows, index)
        major = values[3]
        category, study_mode = values[4:6]
        score_values = values[6:9]
    elif len(values) >= 8 and values[3] in {"定向", "非定向"}:
        nearby = _njtech_nearby_college_major(rows, index)
        if nearby is None:
            return None
        college, major = nearby
        category, study_mode = values[3:5]
        score_values = values[5:8]

    if not _njtech_valid_admission_values(college, major, category, study_mode, score_values):
        return None
    initial_score, reexam_score, total_score = score_values[:3]
    return _build_record(
        document,
        {
            "ranking": sequence,
            "student_id": student_id,
            "person_name": person_name,
            "college": college,
            "admission_major": major,
            "remarks": _clean_text(
                " ".join(
                    [
                        category,
                        study_mode,
                        f"initial_score {initial_score}",
                        f"reexam_score {reexam_score}",
                        f"total_score {total_score}",
                    ]
                )
            ),
        },
    )


def _njtech_valid_admission_values(
    college: str,
    major: str,
    category: str,
    study_mode: str,
    score_values: list[str],
) -> bool:
    if not college or college in {"南京工业大学", "定向", "非定向", "全日制", "非全日制"}:
        return False
    if not _njtech_looks_like_major_code_name(major):
        return False
    if category not in {"定向", "非定向"}:
        return False
    if study_mode not in {"全日制", "非全日制"}:
        return False
    return len(score_values) >= 3 and all(_looks_like_score_or_metric(value) for value in score_values[:3])


def _njtech_following_score_values(rows: list[list[str]], index: int) -> list[str]:
    for following in rows[index + 1 : index + 5]:
        if len(following) >= 2 and all(_looks_like_score_or_metric(value) for value in following[:2]):
            return following[:2]
    return []


def _njtech_looks_like_major_code_name(value: str) -> bool:
    return bool(re.fullmatch(r"\d{6}[\u4e00-\u9fffA-Za-z（）()·、]+", value or ""))


def _njtech_nearby_college_major(rows: list[list[str]], index: int) -> tuple[str, str] | None:
    for previous_index in range(index - 1, max(-1, index - 8), -1):
        text = _clean_text(" ".join(rows[previous_index]))
        if _njtech_ignore_context_text(text):
            continue
        match = re.search(r"(.+?学院)\s+(\d{6}.+)", text)
        if not match:
            continue
        college, major = match.groups()
        following = rows[index + 1] if index + 1 < len(rows) else []
        if len(following) >= 2 and following[0].startswith(("（", "(")) and not _looks_like_score_or_metric(following[1]):
            college = _clean_text(f"{college}{following[0]}")
            major = _clean_text(f"{major}{following[1]}")
        elif following:
            following_text = _clean_text(" ".join(following))
            continuation = re.fullmatch(r"([（(][^）)]+[）)])\s*(.+)", following_text)
            if continuation:
                college = _clean_text(f"{college}{continuation.group(1)}")
                major = _clean_text(f"{major}{continuation.group(2)}")
        return college, major
    return None


def _njtech_nearby_college(rows: list[list[str]], index: int) -> str:
    college = ""
    for previous_index in range(index - 1, max(-1, index - 8), -1):
        text = _clean_text(" ".join(rows[previous_index]))
        if _njtech_ignore_context_text(text):
            continue
        if re.search(r"\d{6}", text):
            continue
        if "学院" in text or re.fullmatch(r"[\u4e00-\u9fff（）()]+", text):
            college = text
            break
    following = rows[index + 1] if index + 1 < len(rows) else []
    if college and following and following[0].startswith(("（", "(")):
        college = _clean_text(f"{college}{following[0]}")
    return college


def _njtech_ignore_context_text(text: str) -> bool:
    if not text:
        return True
    if text in {"京", "南", "工", "业", "大", "学", "南京工业大学"}:
        return True
    if text.startswith("序号") or text.startswith("拟录取") or text.startswith("类别"):
        return True
    return False


def _looks_like_code(value: str, length: int) -> bool:
    return bool(re.fullmatch(rf"\d{{{length}}}", value or ""))


def _looks_like_masked_identifier(value: str) -> bool:
    return bool(re.fullmatch(r"[0-9Xx*]{10,}", value or "")) and "*" in (value or "")


def extract_pdf_text(path: Path) -> str:
    pdftotext_text = _extract_pdf_text_with_pdftotext(path)
    if pdftotext_text:
        return pdftotext_text

    pypdf_text = _extract_pdf_text_with_pypdf(path)
    if pypdf_text and not _looks_like_mojibake(pypdf_text):
        return pypdf_text
    return pypdf_text


def _extract_pdf_text_with_pypdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ModuleNotFoundError:
        return ""

    try:
        reader = PdfReader(str(path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception:
        return ""


def _extract_pdf_text_with_pdftotext(path: Path) -> str:
    try:
        completed = subprocess.run(
            ["pdftotext", "-layout", str(path), "-"],
            capture_output=True,
            text=False,
            timeout=30,
            check=False,
        )
    except (FileNotFoundError, OSError, subprocess.SubprocessError):
        return ""
    if completed.returncode != 0:
        return ""
    return _decode_pdftotext_output(completed.stdout).strip()


def _decode_pdftotext_output(output: bytes | str) -> str:
    if isinstance(output, str):
        return output
    if not output:
        return ""

    candidates = []
    for encoding in ("utf-8", "gb18030", "gbk"):
        text = output.decode(encoding, errors="replace")
        replacement_count = text.count("\ufffd")
        cjk_count = len(re.findall(r"[\u4e00-\u9fff]", text))
        candidates.append((_looks_like_mojibake(text), replacement_count, -cjk_count, text))
    return min(candidates, key=lambda candidate: candidate[:3])[3]


def _looks_like_mojibake(text: str) -> bool:
    if not text:
        return False
    replacement_ratio = text.count("\ufffd") / max(len(text), 1)
    has_cjk = bool(re.search(r"[\u4e00-\u9fff]", text))
    return replacement_ratio > 0.01 and not has_cjk


def crawl_seed_documents(
    seeds: list[dict[str, Any]],
    raw_dir: Path,
    processed_dir: Path,
    logs_dir: Path,
    fetcher: Fetcher | None = None,
    sleeper: Sleeper = time.sleep,
    delay_seconds: float = 0.8,
    timeout_seconds: float = 20,
    max_pages: int = 100,
    max_depth: int = 1,
    resume: bool = False,
    progress: bool = False,
) -> dict[str, Any]:
    raw_dir.mkdir(parents=True, exist_ok=True)
    processed_dir.mkdir(parents=True, exist_ok=True)
    logs_dir.mkdir(parents=True, exist_ok=True)

    documents_path = processed_dir / "documents.jsonl"
    records_path = processed_dir / "records.jsonl"
    failures_path = logs_dir / "graduate_outcome_failures.jsonl"
    seen_urls = read_document_urls(documents_path) if resume else set()
    queued: list[tuple[dict[str, Any], str, int, str, str | None]] = []
    link_metadata: dict[str, LinkCandidate] = {}

    for seed in seeds:
        start_url = str(seed.get("start_url") or "").strip()
        if start_url:
            queued.append((seed, start_url, 0, _clean_text(seed.get("discovery_title")), None))

    fetched = 0
    documents_written = 0
    records_written = 0
    failed = 0
    visited: set[str] = set()

    while queued and fetched < max_pages:
        seed, url, depth, link_title, referer = queued.pop(0)
        if url in visited:
            continue
        visited.add(url)
        if url in seen_urls:
            continue

        try:
            response = fetch_with_retries(
                url,
                timeout_seconds=timeout_seconds,
                fetcher=fetcher,
                referer=referer,
            )
            fetched += 1
            raw_path = _write_raw_document(raw_dir, response)
            html_text = _decode_response_text(response)
            title = _document_title(response, html_text, fallback=link_title)
            link_hint = link_metadata.get(url)
            classification = classify_document(title, url, html_text[:5000])
            if link_hint and classification["document_type"] == "unknown":
                classification = {
                    "document_type": link_hint.document_type,
                    "matched_keywords": link_hint.matched_keywords,
                }
            if (
                classification["document_type"] == "unknown"
                and seed.get("document_type")
                and not _is_homepage_redirect_from_detail(seed.get("start_url"), response.url)
            ):
                classification = {
                    "document_type": _clean_text(seed.get("document_type")),
                    "matched_keywords": [],
                }

            document = _build_document_record(
                seed=seed,
                response=response,
                raw_path=raw_path,
                title=title,
                classification=classification,
            )
            if document["year"] is None and html_text:
                document["year"] = _extract_year_from_html(html_text)
            if document["year"] is None and link_hint and link_hint.year_hint:
                document["year"] = link_hint.year_hint
            if document["year"] is None and seed.get("year"):
                document["year"] = _to_int(seed.get("year"))
            if _is_attachment_url(response.url):
                inherited_year = link_hint.year_hint if link_hint else _to_int(seed.get("year"))
                if inherited_year:
                    document["year"] = inherited_year
            records = _parse_response_records(response, raw_path, document, html_text)
            document["parse_status"] = _parse_status(response, records)
            document["record_count"] = len(records)

            _append_jsonl(documents_path, document)
            documents_written += 1

            for record in records:
                _append_jsonl(records_path, record)
                records_written += 1

            if _is_html_response(response) and depth < max_depth:
                allowed_domains = {urlparse(seed["start_url"]).netloc}
                for link in extract_candidate_links(
                    html_text,
                    response.url,
                    allowed_domains=allowed_domains,
                    include_all_attachments=document["document_type"] != "unknown",
                ):
                    if link.url not in visited:
                        if (
                            link.document_type == "unknown"
                            and document["document_type"] != "unknown"
                        ):
                            link = LinkCandidate(
                                url=link.url,
                                text=link.text,
                                link_kind=link.link_kind,
                                document_type=document["document_type"],
                                matched_keywords=document["matched_keywords"],
                                year_hint=document.get("year"),
                            )
                        elif link.document_type != "unknown":
                            link = LinkCandidate(
                                url=link.url,
                                text=link.text,
                                link_kind=link.link_kind,
                                document_type=link.document_type,
                                matched_keywords=link.matched_keywords,
                                year_hint=document.get("year"),
                            )
                        link_metadata[link.url] = link
                        queued.append((seed, link.url, depth + 1, link.text, response.url))

            if progress:
                print(f"fetched {url} records={len(records)}", flush=True)
            if delay_seconds > 0:
                sleeper(delay_seconds)
        except Exception as exc:
            failed += 1
            _append_jsonl(
                failures_path,
                {
                    "captured_at": _now_iso(),
                    "school_name": seed.get("school_name") or "",
                    "url": url,
                    "error": str(exc),
                    "error_type": type(exc).__name__,
                    "status_code": getattr(exc, "status_code", None),
                },
            )

    write_records_csv(records_path, processed_dir / "records.csv")
    return {
        "requested_seeds": len(seeds),
        "fetched": fetched,
        "documents_written": documents_written,
        "records_written": records_written,
        "failed": failed,
        "documents_path": str(documents_path),
        "records_path": str(records_path),
        "csv_path": str(processed_dir / "records.csv"),
        "failures_path": str(failures_path),
    }


def write_records_csv(jsonl_path: Path, csv_path: Path) -> int:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    rows = []
    if jsonl_path.exists():
        with jsonl_path.open("r", encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))

    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=RECORD_CSV_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _csv_value(row.get(field)) for field in RECORD_CSV_FIELDS})
    return len(rows)


def build_discovery_tasks(
    schools: list[dict[str, Any]],
    years: list[int],
) -> list[dict[str, Any]]:
    tasks: list[dict[str, Any]] = []
    for school in schools:
        school_name = _school_name_from_row(school)
        if not school_name:
            continue
        school_id = str(school.get("id") or school.get("school_id") or "")
        province = str(school.get("province") or "")
        level = str(school.get("level") or "")
        tags = _tags_from_row(school)
        eligibility_hint = _eligibility_hint(level, tags)
        preferred_domains = _preferred_domains_for_school(school_name)

        for year in years:
            for source in DISCOVERY_SOURCE_TYPES:
                tasks.append(
                    {
                        "school_id": school_id,
                        "school_name": school_name,
                        "province": province,
                        "level": level,
                        "year": year,
                        "source_type": source["source_type"],
                        "document_type": source["document_type"],
                        "eligibility_hint": eligibility_hint,
                        "search_query": source["query_template"].format(
                            school_name=school_name,
                            year=year,
                        ),
                        "preferred_domains": ";".join(preferred_domains),
                        "status": "pending",
                        "found_url": "",
                        "notes": "",
                    }
                )
    return tasks


def write_discovery_tasks_csv(tasks: list[dict[str, Any]], output_path: Path) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=DISCOVERY_TASK_CSV_FIELDS)
        writer.writeheader()
        for task in tasks:
            writer.writerow(
                {field: _csv_value(task.get(field)) for field in DISCOVERY_TASK_CSV_FIELDS}
            )
    return len(tasks)


def parse_bing_search_results(
    html: str,
    search_query: str,
    limit: int = 10,
    captured_at: str | None = None,
) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, Any]] = []
    captured = captured_at or _now_iso()

    for item in soup.select("li.b_algo"):
        anchor = item.select_one("h2 a") or item.select_one("a")
        if not anchor:
            continue
        url = _clean_text(anchor.get("href"))
        title = _clean_text(anchor.get_text(" ", strip=True))
        if not url or not title:
            continue
        snippet_node = item.select_one(".b_caption p") or item.select_one("p")
        snippet = _clean_text(snippet_node.get_text(" ", strip=True) if snippet_node else "")
        rows.append(
            {
                "search_query": search_query,
                "result_rank": len(rows) + 1,
                "result_title": title,
                "result_url": url,
                "result_snippet": snippet,
                "provider": "bing",
                "captured_at": captured,
            }
        )
        if len(rows) >= limit:
            break
    return rows


def parse_bing_rss_search_results(
    rss_text: str,
    search_query: str,
    limit: int = 10,
    captured_at: str | None = None,
) -> list[dict[str, Any]]:
    captured = captured_at or _now_iso()
    try:
        root = ET.fromstring(rss_text)
    except ET.ParseError:
        return []

    rows: list[dict[str, Any]] = []
    for item in root.findall(".//item"):
        title = _clean_text(_xml_child_text(item, "title"))
        url = _clean_text(_xml_child_text(item, "link"))
        snippet = _clean_text(BeautifulSoup(_xml_child_text(item, "description"), "html.parser").get_text(" ", strip=True))
        if not title or not url:
            continue
        rows.append(
            {
                "search_query": search_query,
                "result_rank": len(rows) + 1,
                "result_title": title,
                "result_url": url,
                "result_snippet": snippet,
                "provider": "bing-rss",
                "captured_at": captured,
            }
        )
        if len(rows) >= limit:
            break
    return rows


def parse_duckduckgo_search_results(
    html: str,
    search_query: str,
    limit: int = 10,
    captured_at: str | None = None,
) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, Any]] = []
    captured = captured_at or _now_iso()

    for result in soup.select(".result"):
        anchor = result.select_one(".result__a") or result.select_one("a")
        if not anchor:
            continue
        title = _clean_text(anchor.get_text(" ", strip=True))
        url = _unwrap_duckduckgo_url(_clean_text(anchor.get("href")))
        if not title or not url:
            continue
        snippet_node = result.select_one(".result__snippet")
        snippet = _clean_text(snippet_node.get_text(" ", strip=True) if snippet_node else "")
        rows.append(
            {
                "search_query": search_query,
                "result_rank": len(rows) + 1,
                "result_title": title,
                "result_url": url,
                "result_snippet": snippet,
                "provider": "duckduckgo-html",
                "captured_at": captured,
            }
        )
        if len(rows) >= limit:
            break
    return rows


def parse_chsi_school_index(html: str, base_url: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for anchor in soup.find_all("a"):
        href = _clean_text(anchor.get("href"))
        if "schoolInfo--schId-" not in href:
            continue
        url = urljoin(base_url, href)
        match = re.search(r"schId-(\d+)", href)
        if not match or url in seen:
            continue
        name = _clean_text(anchor.get_text(" ", strip=True))
        if not name:
            continue
        seen.add(url)
        rows.append(
            {
                "chsi_school_name": name,
                "chsi_sch_id": match.group(1),
                "chsi_school_url": url,
            }
        )
    return rows


def parse_chsi_school_info_bulletin_urls(html: str, base_url: str) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    urls: list[str] = []
    seen: set[str] = set()
    for anchor in soup.find_all("a"):
        href = _clean_text(anchor.get("href"))
        if "listBulletin--schId-" not in href:
            continue
        url = urljoin(base_url, href)
        if url not in seen:
            seen.add(url)
            urls.append(url)
    return urls


def parse_chsi_bulletin_seed_rows(
    html: str,
    base_url: str,
    school_name: str,
) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    seeds: list[dict[str, Any]] = []
    seen: set[str] = set()

    for anchor in soup.find_all("a"):
        href = _clean_text(anchor.get("href"))
        if "viewBulletin--" not in href and "infoId-" not in href:
            continue
        title = _clean_text(anchor.get_text(" ", strip=True))
        classification = classify_document(title, href, "")
        if classification["document_type"] == "unknown":
            continue
        if not _is_record_like_bulletin_title(title, classification["document_type"]):
            continue
        url = urljoin(base_url, href)
        if url in seen:
            continue
        seen.add(url)
        seeds.append(
            {
                "school_name": school_name,
                "source_type": _source_type_from_document_type(classification["document_type"]),
                "start_url": url,
                "year": _extract_year(title) or "",
                "document_type": classification["document_type"],
                "discovery_query": "chsi_bulletin",
                "discovery_title": title,
                "discovery_rank": len(seeds) + 1,
            }
        )
    return seeds


def parse_school_site_index_rows(
    html: str,
    base_url: str,
    school_rows: list[dict[str, Any]],
    recommended_only: bool = False,
) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    schools_by_name = {
        _school_name_from_row(row): row
        for row in school_rows
        if _school_name_from_row(row)
    }
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()

    for anchor in soup.find_all("a"):
        text = _clean_text(anchor.get_text(" ", strip=True))
        school = schools_by_name.get(text)
        if not school or text in seen:
            continue

        href = _clean_text(anchor.get("href"))
        if not href or href.startswith("#") or href.lower().startswith("javascript:"):
            continue
        try:
            url = urljoin(base_url, href)
            parsed = urlparse(url)
        except ValueError:
            continue
        if parsed.scheme not in {"http", "https"}:
            continue

        tags = _tags_from_row(school)
        eligibility_hint = _eligibility_hint(str(school.get("level") or ""), tags)
        if recommended_only and eligibility_hint != "recommended":
            continue

        seen.add(text)
        rows.append(
            {
                "school_id": str(school.get("id") or school.get("school_id") or ""),
                "school_name": text,
                "province": str(school.get("province") or ""),
                "level": str(school.get("level") or ""),
                "eligibility_hint": eligibility_hint,
                "official_url": url,
                "matched_link_text": text,
                "source_url": base_url,
                "source_rank": len(rows) + 1,
            }
        )
    return rows


def parse_official_site_portal_links(
    html: str,
    base_url: str,
    max_links: int = 8,
) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    base_root = _registrable_domain(urlparse(base_url).netloc)
    links: list[str] = []
    seen: set[str] = set()

    for anchor in soup.find_all("a"):
        href = _clean_text(anchor.get("href"))
        if not href or href.startswith("#") or href.lower().startswith("javascript:"):
            continue
        text = _clean_text(anchor.get_text(" ", strip=True))
        try:
            url = urljoin(base_url, href)
            parsed = urlparse(url)
        except ValueError:
            continue
        if parsed.scheme not in {"http", "https"}:
            continue
        if _registrable_domain(parsed.netloc) != base_root:
            continue

        haystack = f"{text} {unquote(url)}"
        if not any(keyword in haystack for keyword in OFFICIAL_SITE_PORTAL_KEYWORDS):
            continue
        if url in seen:
            continue
        seen.add(url)
        links.append(url)
        if len(links) >= max_links:
            break
    return links


def official_site_candidate_portal_urls(
    official_url: str,
    max_urls: int = 16,
) -> list[str]:
    parsed = urlparse(official_url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return []
    root = _registrable_domain(parsed.netloc)
    if not root:
        return []

    candidates: list[str] = []
    for subdomain in ("dean", "jwc", "jwb", "bks", "bk", "jw", "undergraduate"):
        candidates.append(f"https://{subdomain}.{root}/")

    base = f"{parsed.scheme}://{parsed.netloc}"
    for path in ("/jwc/", "/jwb/", "/jw/", "/bks/", "/bkjx/", "/undergraduate/", "/dean/"):
        candidates.append(urljoin(base, path))

    seen: set[str] = set()
    deduped: list[str] = []
    for url in candidates:
        if url == official_url or url in seen:
            continue
        seen.add(url)
        deduped.append(url)
        if len(deduped) >= max_urls:
            break
    return deduped


def parse_official_site_seed_rows(
    html: str,
    base_url: str,
    school_name: str,
) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    base_host = urlparse(base_url).netloc
    seeds: list[dict[str, Any]] = []
    seen: set[str] = set()

    for anchor in soup.find_all("a"):
        href = _clean_text(anchor.get("href"))
        if not href or href.startswith("#") or href.lower().startswith("javascript:"):
            continue
        try:
            url = urljoin(base_url, href)
            parsed = urlparse(url)
        except ValueError:
            continue
        if parsed.scheme not in {"http", "https"} or parsed.netloc != base_host:
            continue
        title = _clean_text(anchor.get_text(" ", strip=True))
        classification = classify_document(title, url, "")
        document_type = classification["document_type"]
        if document_type != "recommendation_exemption_list":
            continue
        if not _is_record_like_bulletin_title(f"{title} {unquote(url)}", document_type):
            continue
        if url in seen:
            continue
        seen.add(url)
        seeds.append(
            {
                "school_name": school_name,
                "source_type": "recommendation_exemption",
                "start_url": url,
                "year": _extract_year(title + " " + url) or "",
                "document_type": "recommendation_exemption_list",
                "discovery_query": "official_site_link",
                "discovery_title": title,
                "discovery_rank": len(seeds) + 1,
            }
        )
    return seeds


def collect_chsi_school_index(
    output_path: Path,
    fetcher: Fetcher | None = None,
    sleeper: Sleeper = time.sleep,
    delay_seconds: float = 1.0,
    start_page: int = 0,
    max_pages: int = 50,
    page_size: int = 20,
    timeout_seconds: float = 20,
    base_url: str = "https://yzst.chsi.com.cn/sch/",
) -> dict[str, Any]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    actual_fetcher = fetcher or fetch_url
    rows_by_url: dict[str, dict[str, Any]] = {}
    failed = 0
    failures_path = output_path.with_suffix(".failures.jsonl")
    failures_path.parent.mkdir(parents=True, exist_ok=True)
    failures_path.write_text("", encoding="utf-8")

    for page in range(start_page, start_page + max_pages):
        start = page * page_size
        url = base_url if start == 0 else f"{base_url}?start={start}"
        try:
            response = fetch_with_retries(
                url,
                timeout_seconds=timeout_seconds,
                fetcher=actual_fetcher,
                max_retries=1,
                sleeper=sleeper,
            )
            for row in parse_chsi_school_index(_decode_response_text(response), response.url):
                rows_by_url[row["chsi_school_url"]] = row
            if delay_seconds > 0:
                sleeper(delay_seconds)
        except Exception as exc:
            failed += 1
            _append_jsonl(
                failures_path,
                {
                    "captured_at": _now_iso(),
                    "url": url,
                    "error": str(exc),
                    "error_type": type(exc).__name__,
                    "status_code": getattr(exc, "status_code", None),
                },
            )

    rows = list(rows_by_url.values())
    _write_dict_rows_csv(rows, output_path, CHSI_SCHOOL_CSV_FIELDS)
    return {
        "pages_requested": max_pages,
        "start_page": start_page,
        "schools_written": len(rows),
        "failed": failed,
        "output_path": str(output_path),
        "failures_path": str(failures_path),
    }


def collect_school_site_index(
    school_rows: list[dict[str, Any]],
    output_path: Path,
    source_url: str = "https://laosheng.top/fuwu/yuanxiao",
    fetcher: Fetcher | None = None,
    recommended_only: bool = False,
    timeout_seconds: float = 20,
) -> dict[str, Any]:
    actual_fetcher = fetcher or fetch_url
    response = fetch_with_retries(
        source_url,
        timeout_seconds=timeout_seconds,
        fetcher=actual_fetcher,
        max_retries=1,
    )
    rows = parse_school_site_index_rows(
        _decode_response_text(response),
        response.url,
        school_rows,
        recommended_only=recommended_only,
    )
    _write_dict_rows_csv(rows, output_path, SCHOOL_SITE_CSV_FIELDS)
    return {
        "schools_considered": len(school_rows),
        "recommended_only": recommended_only,
        "sites_written": len(rows),
        "source_url": response.url,
        "output_path": str(output_path),
    }


def collect_chsi_bulletin_seeds(
    school_rows: list[dict[str, Any]],
    output_path: Path,
    fetcher: Fetcher | None = None,
    sleeper: Sleeper = time.sleep,
    delay_seconds: float = 1.0,
    max_schools: int | None = None,
    start_index: int = 0,
    max_bulletin_lists_per_school: int = 6,
    timeout_seconds: float = 20,
    workers: int = 1,
) -> dict[str, Any]:
    selected_schools = slice_seed_rows(school_rows, start_index, max_schools)
    actual_fetcher = fetcher or fetch_url
    seeds: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []
    failures_path = output_path.with_suffix(".failures.jsonl")
    failures_path.parent.mkdir(parents=True, exist_ok=True)
    failures_path.write_text("", encoding="utf-8")

    if workers <= 1:
        for school in selected_schools:
            result = _collect_chsi_bulletin_seeds_for_school(
                school,
                fetcher=actual_fetcher,
                sleeper=sleeper,
                delay_seconds=delay_seconds,
                max_bulletin_lists_per_school=max_bulletin_lists_per_school,
                timeout_seconds=timeout_seconds,
            )
            seeds.extend(result["seeds"])
            failures.extend(result["failures"])
    else:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_map = {
                executor.submit(
                    _collect_chsi_bulletin_seeds_for_school,
                    school,
                    actual_fetcher,
                    sleeper,
                    delay_seconds,
                    max_bulletin_lists_per_school,
                    timeout_seconds,
                ): school
                for school in selected_schools
            }
            for future in as_completed(future_map):
                result = future.result()
                seeds.extend(result["seeds"])
                failures.extend(result["failures"])

    for failure in failures:
        _append_jsonl(failures_path, failure)

    deduped = _dedupe_seed_rows(seeds)
    write_seed_rows_csv(deduped, output_path)
    return {
        "selected_schools": len(selected_schools),
        "seeds_written": len(deduped),
        "failed": len(failures),
        "output_path": str(output_path),
        "failures_path": str(failures_path),
    }


def collect_official_site_seeds(
    site_rows: list[dict[str, Any]],
    output_path: Path,
    fetcher: Fetcher | None = None,
    sleeper: Sleeper = time.sleep,
    delay_seconds: float = 1.0,
    start_index: int = 0,
    max_sites: int | None = None,
    timeout_seconds: float = 20,
    portal_depth: int = 1,
    max_portal_pages_per_site: int = 8,
    probe_candidate_portals: bool = False,
    max_candidate_portal_pages_per_site: int = 8,
    workers: int = 1,
) -> dict[str, Any]:
    selected_sites = slice_seed_rows(site_rows, start_index, max_sites)
    actual_fetcher = fetcher or fetch_url
    seeds: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []
    failures_path = output_path.with_suffix(".failures.jsonl")
    failures_path.parent.mkdir(parents=True, exist_ok=True)
    failures_path.write_text("", encoding="utf-8")

    def collect_site(site: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        site_seeds: list[dict[str, Any]] = []
        site_failures: list[dict[str, Any]] = []
        school_name = _school_name_from_row(site)
        start_url = _official_site_url_from_row(site)
        if not school_name or not start_url:
            site_failures.append(
                {
                    "captured_at": _now_iso(),
                    "school_name": school_name,
                    "url": start_url,
                    "error": "Missing school_name or start_url.",
                    "error_type": "ValueError",
                    "status_code": None,
                },
            )
            return site_seeds, site_failures
        queued: list[tuple[str, int]] = [(start_url, 0)]
        if probe_candidate_portals:
            for candidate_url in official_site_candidate_portal_urls(
                start_url,
                max_urls=max_candidate_portal_pages_per_site,
            ):
                queued.append((candidate_url, 1))
        visited: set[str] = set()
        portal_pages_queued = 0
        while queued:
            url, depth = queued.pop(0)
            if url in visited:
                continue
            visited.add(url)
            try:
                response = fetch_with_retries(
                    url,
                    timeout_seconds=timeout_seconds,
                    fetcher=actual_fetcher,
                    max_retries=1,
                    sleeper=sleeper,
                )
                html = _decode_response_text(response)
                site_seeds.extend(parse_official_site_seed_rows(html, response.url, school_name))
                if depth < portal_depth:
                    for portal_url in parse_official_site_portal_links(
                        html,
                        response.url,
                        max_links=max_portal_pages_per_site,
                    ):
                        if portal_url in visited:
                            continue
                        if portal_pages_queued >= max_portal_pages_per_site:
                            break
                        queued.append((portal_url, depth + 1))
                        portal_pages_queued += 1
                if delay_seconds > 0:
                    sleeper(delay_seconds)
            except Exception as exc:
                site_failures.append(
                    {
                        "captured_at": _now_iso(),
                        "school_name": school_name,
                        "url": url,
                        "error": str(exc),
                        "error_type": type(exc).__name__,
                        "status_code": getattr(exc, "status_code", None),
                    },
                )
        return site_seeds, site_failures

    worker_count = max(1, int(workers or 1))
    if worker_count == 1 or len(selected_sites) <= 1:
        for site in selected_sites:
            site_seeds, site_failures = collect_site(site)
            seeds.extend(site_seeds)
            failures.extend(site_failures)
    else:
        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            futures = [executor.submit(collect_site, site) for site in selected_sites]
            for future in as_completed(futures):
                site_seeds, site_failures = future.result()
                seeds.extend(site_seeds)
                failures.extend(site_failures)

    for failure in failures:
        _append_jsonl(failures_path, failure)

    deduped = _dedupe_seed_rows(seeds)
    write_seed_rows_csv(deduped, output_path)
    return {
        "selected_sites": len(selected_sites),
        "seeds_written": len(deduped),
        "failed": len(failures),
        "output_path": str(output_path),
        "failures_path": str(failures_path),
    }


def collect_search_results(
    tasks: list[dict[str, Any]],
    output_path: Path,
    fetcher: Fetcher | None = None,
    sleeper: Sleeper = time.sleep,
    delay_seconds: float = 1.5,
    start_index: int = 0,
    max_tasks: int | None = None,
    results_per_query: int = 5,
    resume: bool = True,
    timeout_seconds: float = 20,
    raw_debug_dir: Path | None = None,
    provider: str = "bing-rss",
) -> dict[str, Any]:
    selected_tasks = slice_seed_rows(tasks, start_index=start_index, max_seeds=max_tasks)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    completed_queries = _read_completed_search_queries(output_path) if resume else set()
    if not output_path.exists() or not resume:
        _write_search_result_header(output_path)

    actual_fetcher = fetcher or fetch_url
    queried = 0
    skipped_existing = 0
    results_written = 0
    failed = 0
    failures_path = output_path.with_suffix(".failures.jsonl")

    for task in selected_tasks:
        query = _clean_text(task.get("search_query"))
        if not query:
            continue
        if query in completed_queries:
            skipped_existing += 1
            continue
        try:
            url = _build_search_url(query, count=max(results_per_query, 1), provider=provider)
            response = fetch_with_retries(
                url,
                timeout_seconds=timeout_seconds,
                fetcher=actual_fetcher,
                max_retries=1,
                retry_base_seconds=2,
                sleeper=sleeper,
            )
            if raw_debug_dir is not None:
                _write_search_debug_html(raw_debug_dir, query, response.content)
            text = _decode_response_text(response)
            rows = _parse_search_results(text, query, results_per_query, provider)
            _append_search_result_rows(output_path, rows)
            queried += 1
            results_written += len(rows)
            if delay_seconds > 0:
                sleeper(delay_seconds)
        except Exception as exc:
            failed += 1
            _append_jsonl(
                failures_path,
                {
                    "captured_at": _now_iso(),
                    "search_query": query,
                    "error": str(exc),
                    "error_type": type(exc).__name__,
                    "status_code": getattr(exc, "status_code", None),
                },
            )

    return {
        "selected_tasks": len(selected_tasks),
        "queried": queried,
        "skipped_existing": skipped_existing,
        "results_written": results_written,
        "failed": failed,
        "output_path": str(output_path),
        "failures_path": str(failures_path),
    }


def select_seed_rows_from_search_results(
    tasks: list[dict[str, Any]],
    search_results: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    tasks_by_query: dict[str, dict[str, Any]] = {
        _clean_text(task.get("search_query")): task
        for task in tasks
        if _clean_text(task.get("search_query"))
    }
    seen_urls: set[str] = set()
    seeds: list[dict[str, Any]] = []
    rank_by_query: dict[str, int] = {}

    for result in search_results:
        query = _clean_text(result.get("search_query") or result.get("query"))
        task = tasks_by_query.get(query)
        if not task:
            continue
        rank_by_query[query] = rank_by_query.get(query, 0) + 1
        url = _clean_text(
            result.get("result_url")
            or result.get("url")
            or result.get("link")
            or result.get("found_url")
        )
        if not url or url in seen_urls:
            continue
        title = _clean_text(result.get("result_title") or result.get("title"))
        snippet = _clean_text(result.get("result_snippet") or result.get("snippet"))
        if not _is_acceptable_discovery_url(url):
            continue
        if not _search_result_matches_task(task, title, url, snippet):
            continue

        seen_urls.add(url)
        seeds.append(
            {
                "school_name": task.get("school_name") or "",
                "source_type": task.get("source_type") or "",
                "start_url": url,
                "year": task.get("year") or "",
                "document_type": task.get("document_type") or "",
                "discovery_query": query,
                "discovery_title": title,
                "discovery_rank": rank_by_query[query],
            }
        )
    return seeds


def write_seed_rows_csv(seeds: list[dict[str, Any]], output_path: Path) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SEED_CSV_FIELDS)
        writer.writeheader()
        for seed in seeds:
            writer.writerow({field: _csv_value(seed.get(field)) for field in SEED_CSV_FIELDS})
    return len(seeds)


def read_search_results_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_chsi_school_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def clean_records_to_outputs(
    records_jsonl_path: Path,
    clean_csv_path: Path,
    summary_csv_path: Path,
) -> dict[str, int]:
    return _clean_record_rows_to_outputs(
        list(_iter_jsonl(records_jsonl_path)),
        clean_csv_path,
        summary_csv_path,
    )


def merge_records_jsonl_to_outputs(
    records_jsonl_paths: list[Path],
    clean_csv_path: Path,
    summary_csv_path: Path,
) -> dict[str, int]:
    input_rows: list[dict[str, Any]] = []
    for records_jsonl_path in records_jsonl_paths:
        input_rows.extend(_iter_jsonl(records_jsonl_path))
    summary = _clean_record_rows_to_outputs(input_rows, clean_csv_path, summary_csv_path)
    return {"input_files": len(records_jsonl_paths), **summary}


def export_public_records_csv(clean_csv_path: Path, public_csv_path: Path) -> dict[str, int]:
    with clean_csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))

    public_csv_path.parent.mkdir(parents=True, exist_ok=True)
    with public_csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=PUBLIC_RECORD_CSV_FIELDS)
        writer.writeheader()
        for row in rows:
            public_row = {
                field: _csv_value(row.get(field))
                for field in PUBLIC_RECORD_CSV_FIELDS
                if field != "public_record_id"
            }
            public_row["public_record_id"] = _public_record_id(row)
            writer.writerow(public_row)

    return {
        "input_rows": len(rows),
        "public_rows": len(rows),
    }


def _clean_record_rows_to_outputs(
    input_rows: list[dict[str, Any]],
    clean_csv_path: Path,
    summary_csv_path: Path,
) -> dict[str, int]:
    by_key: dict[str, dict[str, Any]] = {}

    for index, record in enumerate(input_rows):
        indexed_record = dict(record)
        indexed_record["_input_row_index"] = index
        clean = _clean_record(indexed_record)
        if not (clean.get("school_name") and clean.get("year") and clean.get("source_url")):
            continue
        if not (clean.get("person_name") or clean.get("student_id")):
            continue
        if _looks_like_non_person_identity(clean):
            continue
        key = _record_dedupe_key(clean)
        for existing_key, existing in by_key.items():
            if _records_are_duplicate(clean, existing):
                key = existing_key
                break
        existing = by_key.get(key)
        if existing is None or clean["quality_score"] > existing["quality_score"]:
            by_key[key] = clean

    clean_rows = sorted(
        by_key.values(),
        key=lambda row: (
            str(row.get("school_name") or ""),
            str(row.get("year") or ""),
            str(row.get("document_type") or ""),
            bool(row.get("needs_review")),
            not bool(row.get("person_name")),
            str(row.get("person_name") or ""),
            str(row.get("source_url") or ""),
        ),
    )
    _write_clean_records_csv(clean_rows, clean_csv_path)
    summary_rows = _build_summary_rows(clean_rows)
    _write_summary_csv(summary_rows, summary_csv_path)
    return {
        "input_rows": len(input_rows),
        "clean_rows": len(clean_rows),
        "summary_rows": len(summary_rows),
    }


def read_seed_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = []
        for row in csv.DictReader(handle):
            if not row.get("start_url"):
                continue
            cleaned = {key: value for key, value in row.items() if key is not None}
            cleaned["school_name"] = cleaned.get("school_name") or ""
            cleaned["source_type"] = cleaned.get("source_type") or ""
            cleaned["start_url"] = cleaned.get("start_url") or ""
            rows.append(cleaned)
        return rows


def read_official_site_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = []
        for row in csv.DictReader(handle):
            cleaned = {key: value for key, value in row.items() if key is not None}
            if not _school_name_from_row(cleaned) or not _official_site_url_from_row(cleaned):
                continue
            rows.append(cleaned)
        return rows


def slice_seed_rows(
    seeds: list[dict[str, Any]],
    start_index: int = 0,
    max_seeds: int | None = None,
) -> list[dict[str, Any]]:
    if start_index < 0:
        raise ValueError("start_index must be non-negative.")
    if max_seeds is not None and max_seeds < 0:
        raise ValueError("max_seeds must be non-negative.")
    end_index = None if max_seeds is None else start_index + max_seeds
    return seeds[start_index:end_index]


def read_discovery_tasks_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_school_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _collect_chsi_bulletin_seeds_for_school(
    school: dict[str, Any],
    fetcher: Fetcher,
    sleeper: Sleeper,
    delay_seconds: float,
    max_bulletin_lists_per_school: int,
    timeout_seconds: float,
) -> dict[str, list[dict[str, Any]]]:
    school_name = _clean_text(school.get("chsi_school_name") or school.get("school_name"))
    school_url = _clean_text(school.get("chsi_school_url") or school.get("school_url"))
    if not school_name or not school_url:
        return {"seeds": [], "failures": []}

    seeds: list[dict[str, Any]] = []
    try:
        response = fetch_with_retries(
            school_url,
            timeout_seconds=timeout_seconds,
            fetcher=fetcher,
            max_retries=1,
            sleeper=sleeper,
        )
        list_urls = parse_chsi_school_info_bulletin_urls(
            _decode_response_text(response),
            response.url,
        )[:max_bulletin_lists_per_school]
        for list_url in list_urls:
            list_response = fetch_with_retries(
                list_url,
                timeout_seconds=timeout_seconds,
                fetcher=fetcher,
                max_retries=1,
                sleeper=sleeper,
            )
            seeds.extend(
                parse_chsi_bulletin_seed_rows(
                    _decode_response_text(list_response),
                    list_response.url,
                    school_name=school_name,
                )
            )
            if delay_seconds > 0:
                sleeper(delay_seconds)
        return {"seeds": seeds, "failures": []}
    except Exception as exc:
        return {
            "seeds": [],
            "failures": [
                {
                    "captured_at": _now_iso(),
                    "school_name": school_name,
                    "url": school_url,
                    "error": str(exc),
                    "error_type": type(exc).__name__,
                    "status_code": getattr(exc, "status_code", None),
                }
            ],
        }


def read_document_urls(path: Path) -> set[str]:
    urls: set[str] = set()
    if not path.exists():
        return urls
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            source_url = record.get("source_url")
            if source_url:
                urls.add(source_url)
    return urls


def fetch_with_retries(
    url: str,
    timeout_seconds: float = 20,
    fetcher: Fetcher | None = None,
    max_retries: int = 2,
    retry_base_seconds: float = 2,
    sleeper: Sleeper = time.sleep,
    referer: str | None = None,
) -> FetchResponse:
    last_error: Exception | None = None

    for attempt in range(max_retries + 1):
        try:
            if fetcher is None:
                return fetch_url(url, timeout_seconds, referer=referer)
            return fetcher(url, timeout_seconds)
        except FetchError as exc:
            last_error = exc
            if exc.status_code in BLOCKING_STATUS_CODES:
                raise
            if exc.status_code and exc.status_code not in RETRYABLE_STATUS_CODES:
                raise
        except (URLError, TimeoutError) as exc:
            last_error = exc

        if attempt < max_retries:
            sleeper(retry_base_seconds * (attempt + 1))

    if last_error:
        raise last_error
    raise FetchError("Unknown fetch failure.")


def fetch_url(url: str, timeout_seconds: float = 20, referer: str | None = None) -> FetchResponse:
    request = Request(_request_url(url), headers=_request_headers(referer=referer))
    try:
        with _open_url_with_ssl_fallback(request, url, timeout_seconds, referer=referer) as response:
            content = response.read()
            content_type = response.headers.get("Content-Type", "")
            final_url = response.geturl()
            status_code = getattr(response, "status", 200)
    except HTTPError as exc:
        if exc.code in BLOCKING_STATUS_CODES:
            curl_response = _fetch_url_with_curl_fallbacks(
                url,
                timeout_seconds,
                include_minimal=True,
                include_default=False,
                referer=referer,
            )
            if curl_response:
                return curl_response
        else:
            try:
                return _fetch_url_with_curl(url, timeout_seconds, referer=referer)
            except FetchError:
                pass
        raise FetchError(f"HTTP {exc.code} while fetching {url}", exc.code) from exc
    except URLError as exc:
        if _should_retry_with_curl(url, exc):
            curl_response = _fetch_url_with_curl_fallbacks(
                url,
                timeout_seconds,
                include_minimal=True,
                referer=referer,
            )
            if curl_response:
                return curl_response
        raise FetchError(f"Network error while fetching {url}: {exc}") from exc

    return FetchResponse(
        url=final_url,
        status_code=status_code,
        content_type=content_type,
        content=content,
    )


def _fetch_url_with_curl_fallbacks(
    url: str,
    timeout_seconds: float,
    include_minimal: bool = False,
    include_default: bool = True,
    referer: str | None = None,
) -> FetchResponse | None:
    attempts: list[dict[str, bool]] = [{}] if include_default else []
    if include_minimal:
        attempts.extend(
            [
                {"include_referer": False},
                {"include_headers": False, "include_ssl_no_revoke": False},
            ]
        )
    for curl_kwargs in attempts:
        try:
            return _fetch_url_with_curl(url, timeout_seconds, referer=referer, **curl_kwargs)
        except FetchError:
            continue
    return None


def _fetch_url_with_curl(
    url: str,
    timeout_seconds: float,
    include_referer: bool = True,
    include_headers: bool = True,
    include_ssl_no_revoke: bool = True,
    referer: str | None = None,
) -> FetchResponse:
    headers = _request_headers(referer=referer)
    command = [
        "curl.exe" if os.name == "nt" else "curl",
        "-L",
        "--fail",
        "--silent",
        "--show-error",
        "--noproxy",
        "*",
        "--max-time",
        f"{timeout_seconds:g}",
        _request_url(url),
    ]
    if include_headers:
        command[-1:-1] = [
            "-A",
            headers["User-Agent"],
            "-H",
            f"Accept: {headers['Accept']}",
            "-H",
            f"Accept-Language: {headers['Accept-Language']}",
        ]
    if include_headers and include_referer:
        command[-1:-1] = ["-H", f"Referer: {headers['Referer']}"]
    if os.name == "nt" and include_ssl_no_revoke:
        command.insert(5, "--ssl-no-revoke")
    try:
        completed = subprocess.run(
            command,
            capture_output=True,
            timeout=timeout_seconds + 5,
            check=False,
        )
    except (FileNotFoundError, OSError, subprocess.SubprocessError) as exc:
        raise FetchError(f"Curl fallback failed while fetching {url}: {exc}") from exc

    if completed.returncode != 0:
        stderr = completed.stderr.decode("utf-8", errors="replace").strip()
        raise FetchError(f"Curl fallback failed while fetching {url}: {stderr}")

    return FetchResponse(
        url=url,
        status_code=200,
        content_type="",
        content=completed.stdout,
    )


def _should_retry_with_curl(url: str, exc: URLError) -> bool:
    if urlparse(url).scheme not in {"http", "https"}:
        return False
    return True


def _open_url_with_ssl_fallback(
    request: Request,
    original_url: str,
    timeout_seconds: float,
    referer: str | None = None,
):
    try:
        return urlopen(request, timeout=timeout_seconds)
    except URLError as exc:
        if not _should_retry_with_unverified_ssl(original_url, exc):
            raise
        try:
            return urlopen(
                request,
                timeout=timeout_seconds,
                context=ssl._create_unverified_context(),
            )
        except URLError as ssl_retry_exc:
            if not _should_retry_with_plain_http(original_url, ssl_retry_exc):
                raise
            http_url = urlparse(original_url)._replace(scheme="http").geturl()
            http_request = Request(_request_url(http_url), headers=_request_headers(referer=referer))
            return urlopen(http_request, timeout=timeout_seconds)


def _should_retry_with_unverified_ssl(url: str, exc: URLError) -> bool:
    if urlparse(url).scheme != "https":
        return False
    reason = getattr(exc, "reason", None)
    if isinstance(reason, ssl.SSLError):
        return True
    text = str(exc).lower()
    return "ssl" in text and (
        "certificate" in text
        or "handshake" in text
        or "eof" in text
        or "wrong version number" in text
    )


def _should_retry_with_plain_http(url: str, exc: URLError) -> bool:
    if urlparse(url).scheme != "https":
        return False
    return _should_retry_with_unverified_ssl(url, exc)


def _request_url(url: str) -> str:
    parsed = urlparse(url)
    return parsed._replace(
        path=quote(parsed.path, safe="/%"),
        query=quote(parsed.query, safe="=&%/:+,"),
    ).geturl()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Crawl public recommendation-exemption and admission lists."
    )
    parser.add_argument("--seed-csv", type=Path)
    parser.add_argument(
        "--university-csv",
        type=Path,
        default=Path("data/seeds/rysxai_universities.csv"),
        help="School list CSV used when building all-school discovery tasks.",
    )
    parser.add_argument(
        "--build-discovery-tasks",
        type=Path,
        help="Write all-school search/discovery task CSV, then exit.",
    )
    parser.add_argument(
        "--discovery-tasks-csv",
        type=Path,
        help="Discovery task CSV used with --search-results-csv.",
    )
    parser.add_argument(
        "--search-results-csv",
        type=Path,
        help="CSV of search results with search_query/result_url/result_title/result_snippet columns.",
    )
    parser.add_argument(
        "--collect-search-results",
        type=Path,
        help="Run search discovery for discovery tasks and write search results CSV.",
    )
    parser.add_argument(
        "--collect-chsi-school-index",
        type=Path,
        help="Crawl CHSI graduate-school index and write school index CSV.",
    )
    parser.add_argument(
        "--collect-chsi-bulletin-seeds",
        type=Path,
        help="Crawl CHSI school bulletin pages and write relevant seed CSV.",
    )
    parser.add_argument(
        "--collect-school-site-index",
        type=Path,
        help="Fetch a public school website directory and write official school site CSV.",
    )
    parser.add_argument(
        "--school-site-source-url",
        default="https://laosheng.top/fuwu/yuanxiao",
        help="Source URL for the public school website directory.",
    )
    parser.add_argument(
        "--recommended-only-sites",
        action="store_true",
        help="Only write school sites whose local tags indicate graduate recommendation eligibility.",
    )
    parser.add_argument(
        "--collect-official-site-seeds",
        type=Path,
        help="Crawl known official school site pages and write recommendation-exemption seed CSV.",
    )
    parser.add_argument(
        "--official-site-csv",
        type=Path,
        help="CSV with school_name and start_url/official_url/homepage_url columns.",
    )
    parser.add_argument("--chsi-school-csv", type=Path)
    parser.add_argument("--chsi-max-pages", type=int, default=50)
    parser.add_argument("--chsi-start-page", type=int, default=0)
    parser.add_argument("--chsi-page-size", type=int, default=20)
    parser.add_argument("--chsi-start-index", type=int, default=0)
    parser.add_argument("--chsi-max-schools", type=int)
    parser.add_argument("--chsi-delay-seconds", type=float, default=1.0)
    parser.add_argument("--chsi-max-bulletin-lists-per-school", type=int, default=6)
    parser.add_argument("--chsi-workers", type=int, default=1)
    parser.add_argument(
        "--build-seed-csv",
        type=Path,
        help="Write crawl seed CSV selected from discovery/search results, then exit.",
    )
    parser.add_argument(
        "--years",
        default="2026,2025,2024",
        help="Comma-separated years for discovery tasks.",
    )
    parser.add_argument(
        "--clean-records",
        action="store_true",
        help="Clean processed records JSONL into analysis-ready CSVs, then exit.",
    )
    parser.add_argument(
        "--records-jsonl",
        type=Path,
        default=Path("data/processed/graduate_outcomes/records.jsonl"),
    )
    parser.add_argument(
        "--merge-records-jsonl",
        type=Path,
        nargs="+",
        help="Merge one or more processed records JSONL files into analysis-ready CSVs, then exit.",
    )
    parser.add_argument(
        "--export-public-clean",
        action="store_true",
        help="Export a de-identified public CSV from --clean-csv, then exit.",
    )
    parser.add_argument(
        "--public-csv",
        type=Path,
        default=Path("data/processed/graduate_outcomes/records_public.csv"),
    )
    parser.add_argument(
        "--clean-csv",
        type=Path,
        default=Path("data/processed/graduate_outcomes/records_clean.csv"),
    )
    parser.add_argument(
        "--summary-csv",
        type=Path,
        default=Path("data/processed/graduate_outcomes/school_year_summary.csv"),
    )
    parser.add_argument("--raw-dir", type=Path, default=Path("data/raw/graduate_outcomes"))
    parser.add_argument(
        "--processed-dir",
        type=Path,
        default=Path("data/processed/graduate_outcomes"),
    )
    parser.add_argument("--logs-dir", type=Path, default=Path("data/logs/graduate_outcomes"))
    parser.add_argument("--delay-seconds", type=float, default=0.8)
    parser.add_argument("--timeout-seconds", type=float, default=20)
    parser.add_argument("--max-pages", type=int, default=100)
    parser.add_argument("--max-depth", type=int, default=1)
    parser.add_argument(
        "--official-site-portal-depth",
        type=int,
        default=1,
        help="How many portal-link hops to follow when discovering official-site seeds.",
    )
    parser.add_argument(
        "--official-site-max-portal-pages",
        type=int,
        default=8,
        help="Maximum teaching/undergraduate portal pages to fetch per official site.",
    )
    parser.add_argument(
        "--official-site-probe-candidate-portals",
        action="store_true",
        help="Probe common teaching/undergraduate subdomains and paths for each official site.",
    )
    parser.add_argument(
        "--official-site-max-candidate-portals",
        type=int,
        default=8,
        help="Maximum generated teaching/undergraduate candidate portal URLs to fetch per official site.",
    )
    parser.add_argument(
        "--official-site-workers",
        type=int,
        default=1,
        help="Number of official school sites to probe concurrently.",
    )
    parser.add_argument("--start-index", type=int, default=0)
    parser.add_argument("--max-seeds", type=int)
    parser.add_argument("--search-start-index", type=int, default=0)
    parser.add_argument("--search-max-tasks", type=int)
    parser.add_argument("--search-delay-seconds", type=float, default=1.5)
    parser.add_argument("--results-per-query", type=int, default=5)
    parser.add_argument("--search-raw-dir", type=Path)
    parser.add_argument(
        "--search-provider",
        choices=["bing-rss", "bing-html", "duckduckgo-html"],
        default="bing-rss",
    )
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("--progress", action="store_true")
    args = parser.parse_args(argv)

    if args.build_discovery_tasks:
        years = _parse_years(args.years)
        schools = read_school_csv(args.university_csv)
        tasks = build_discovery_tasks(schools, years)
        count = write_discovery_tasks_csv(tasks, args.build_discovery_tasks)
        print(
            json.dumps(
                {
                    "schools": len(schools),
                    "years": years,
                    "tasks": count,
                    "output_path": str(args.build_discovery_tasks),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0

    if args.collect_chsi_school_index:
        summary = collect_chsi_school_index(
            args.collect_chsi_school_index,
            delay_seconds=args.chsi_delay_seconds,
            start_page=args.chsi_start_page,
            max_pages=args.chsi_max_pages,
            page_size=args.chsi_page_size,
            timeout_seconds=args.timeout_seconds,
        )
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    if args.collect_school_site_index:
        summary = collect_school_site_index(
            read_school_csv(args.university_csv),
            args.collect_school_site_index,
            source_url=args.school_site_source_url,
            recommended_only=args.recommended_only_sites,
            timeout_seconds=args.timeout_seconds,
        )
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    if args.collect_chsi_bulletin_seeds:
        if not args.chsi_school_csv:
            parser.error("--collect-chsi-bulletin-seeds requires --chsi-school-csv.")
        summary = collect_chsi_bulletin_seeds(
            read_chsi_school_csv(args.chsi_school_csv),
            args.collect_chsi_bulletin_seeds,
            delay_seconds=args.chsi_delay_seconds,
            start_index=args.chsi_start_index,
            max_schools=args.chsi_max_schools,
            max_bulletin_lists_per_school=args.chsi_max_bulletin_lists_per_school,
            timeout_seconds=args.timeout_seconds,
            workers=args.chsi_workers,
        )
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    if args.collect_official_site_seeds:
        if not args.official_site_csv:
            parser.error("--collect-official-site-seeds requires --official-site-csv.")
        summary = collect_official_site_seeds(
            read_official_site_csv(args.official_site_csv),
            args.collect_official_site_seeds,
            delay_seconds=args.delay_seconds,
            start_index=args.start_index,
            max_sites=args.max_seeds,
            timeout_seconds=args.timeout_seconds,
            portal_depth=args.official_site_portal_depth,
            max_portal_pages_per_site=args.official_site_max_portal_pages,
            probe_candidate_portals=args.official_site_probe_candidate_portals,
            max_candidate_portal_pages_per_site=args.official_site_max_candidate_portals,
            workers=args.official_site_workers,
        )
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    if args.collect_search_results:
        if not args.discovery_tasks_csv:
            parser.error("--collect-search-results requires --discovery-tasks-csv.")
        summary = collect_search_results(
            read_discovery_tasks_csv(args.discovery_tasks_csv),
            output_path=args.collect_search_results,
            delay_seconds=args.search_delay_seconds,
            start_index=args.search_start_index,
            max_tasks=args.search_max_tasks,
            results_per_query=args.results_per_query,
            resume=args.resume,
            timeout_seconds=args.timeout_seconds,
            raw_debug_dir=args.search_raw_dir,
            provider=args.search_provider,
        )
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    if args.build_seed_csv:
        if not args.discovery_tasks_csv or not args.search_results_csv:
            parser.error("--build-seed-csv requires --discovery-tasks-csv and --search-results-csv.")
        seeds = select_seed_rows_from_search_results(
            read_discovery_tasks_csv(args.discovery_tasks_csv),
            read_search_results_csv(args.search_results_csv),
        )
        count = write_seed_rows_csv(seeds, args.build_seed_csv)
        print(
            json.dumps(
                {"seeds": count, "output_path": str(args.build_seed_csv)},
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0

    if args.clean_records:
        summary = clean_records_to_outputs(
            args.records_jsonl,
            args.clean_csv,
            args.summary_csv,
        )
        print(json.dumps(summary | {"clean_csv": str(args.clean_csv), "summary_csv": str(args.summary_csv)}, ensure_ascii=False, indent=2))
        return 0

    if args.merge_records_jsonl:
        summary = merge_records_jsonl_to_outputs(
            args.merge_records_jsonl,
            args.clean_csv,
            args.summary_csv,
        )
        print(
            json.dumps(
                summary
                | {
                    "records_jsonl": [str(path) for path in args.merge_records_jsonl],
                    "clean_csv": str(args.clean_csv),
                    "summary_csv": str(args.summary_csv),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0

    if args.export_public_clean:
        summary = export_public_records_csv(args.clean_csv, args.public_csv)
        print(
            json.dumps(
                summary
                | {
                    "clean_csv": str(args.clean_csv),
                    "public_csv": str(args.public_csv),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0

    if not args.seed_csv:
        parser.error("Provide --seed-csv, --build-discovery-tasks, --clean-records, --merge-records-jsonl, --export-public-clean, --collect-school-site-index, or --collect-official-site-seeds.")

    all_seeds = read_seed_csv(args.seed_csv)
    seeds = slice_seed_rows(all_seeds, args.start_index, args.max_seeds)
    summary = crawl_seed_documents(
        seeds,
        raw_dir=args.raw_dir,
        processed_dir=args.processed_dir,
        logs_dir=args.logs_dir,
        delay_seconds=args.delay_seconds,
        timeout_seconds=args.timeout_seconds,
        max_pages=args.max_pages,
        max_depth=args.max_depth,
        resume=args.resume,
        progress=args.progress,
    )
    summary["total_seed_rows"] = len(all_seeds)
    summary["start_index"] = args.start_index
    summary["selected_seed_rows"] = len(seeds)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def _parse_years(value: str) -> list[int]:
    years = []
    for part in value.split(","):
        part = part.strip()
        if not part:
            continue
        years.append(int(part))
    return years


def _school_name_from_row(row: dict[str, Any]) -> str:
    return _clean_text(row.get("school_name") or row.get("name") or row.get("学校名称"))


def _official_site_url_from_row(row: dict[str, Any]) -> str:
    return _clean_text(
        row.get("start_url")
        or row.get("official_url")
        or row.get("homepage_url")
        or row.get("site_url")
        or row.get("url")
        or row.get("学校网址")
        or row.get("官网")
    )


def _registrable_domain(host: str) -> str:
    clean = (host or "").lower().split(":", 1)[0].strip(".")
    parts = [part for part in clean.split(".") if part]
    if len(parts) >= 3 and parts[-2:] == ["edu", "cn"]:
        return ".".join(parts[-3:])
    if len(parts) >= 2:
        return ".".join(parts[-2:])
    return clean


def _tags_from_row(row: dict[str, Any]) -> list[str]:
    value = row.get("tags") or row.get("tags_json") or []
    if isinstance(value, list):
        return [_clean_text(item) for item in value if _clean_text(item)]
    if not isinstance(value, str):
        return []
    try:
        decoded = json.loads(value)
    except json.JSONDecodeError:
        return [part.strip() for part in value.split(",") if part.strip()]
    if isinstance(decoded, list):
        return [_clean_text(item) for item in decoded if _clean_text(item)]
    return []


def _eligibility_hint(level: str, tags: list[str]) -> str:
    tag_text = " ".join(tags)
    if "研究生院" in tag_text or "保研" in tag_text:
        return "recommended"
    if "本科" in level:
        return "normal"
    return "low_priority"


def _preferred_domains_for_school(school_name: str) -> list[str]:
    compact_name = re.sub(r"(大学|学院|学校)$", "", school_name)
    return [
        f"{school_name} 研究生院",
        f"{school_name} 研究生招生网",
        f"{school_name} 教务处",
        f"{compact_name} 研招网" if compact_name != school_name else f"{school_name} 研招网",
    ]


def _source_type_from_document_type(document_type: str) -> str:
    if document_type == "recommendation_exemption_list":
        return "recommendation_exemption"
    if document_type == "incoming_recommendation_admission_list":
        return "incoming_recommendation"
    if document_type == "postgraduate_admission_list":
        return "postgraduate_admission"
    return "unknown"


def _is_record_like_bulletin_title(title: str, document_type: str) -> bool:
    if any(word in title for word in ("办法", "简章", "章程", "指南", "说明")):
        return False
    if document_type == "postgraduate_admission_list":
        return any(word in title for word in ("拟录取", "录取名单", "名单", "结果", "公示"))
    if document_type == "incoming_recommendation_admission_list":
        return any(word in title for word in ("拟录取", "名单", "公示", "结果"))
    if document_type == "recommendation_exemption_list":
        return any(word in title for word in ("名单", "公示", "资格", "结果"))
    return False


def _dedupe_seed_rows(seeds: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_url: dict[str, dict[str, Any]] = {}
    for seed in seeds:
        url = _clean_text(seed.get("start_url"))
        if url and url not in by_url:
            by_url[url] = seed
    return list(by_url.values())


def _write_dict_rows_csv(
    rows: list[dict[str, Any]],
    output_path: Path,
    fieldnames: list[str],
) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _csv_value(row.get(field)) for field in fieldnames})
    return len(rows)


def _build_search_url(query: str, count: int = 10, provider: str = "bing-rss") -> str:
    if provider == "duckduckgo-html":
        return "https://duckduckgo.com/html/?" + urlencode({"q": query})
    if provider == "bing-html":
        return _build_bing_search_url(query, count=count)
    return _build_bing_rss_search_url(query, count=count)


def _build_bing_search_url(query: str, count: int = 10) -> str:
    return "https://www.bing.com/search?" + urlencode(
        {
            "q": query,
            "count": count,
            "mkt": "zh-CN",
            "setlang": "zh-CN",
        }
    )


def _build_bing_rss_search_url(query: str, count: int = 10) -> str:
    return "https://www.bing.com/search?" + urlencode(
        {
            "q": query,
            "count": count,
            "format": "rss",
            "mkt": "zh-CN",
            "setlang": "zh-CN",
        }
    )


def _parse_search_results(
    text: str,
    query: str,
    limit: int,
    provider: str,
) -> list[dict[str, Any]]:
    if provider == "bing-rss":
        rows = parse_bing_rss_search_results(text, query, limit=limit)
        if rows:
            return rows
    if provider == "duckduckgo-html":
        return parse_duckduckgo_search_results(text, query, limit=limit)
    return parse_bing_search_results(text, query, limit=limit)


def _unwrap_duckduckgo_url(url: str) -> str:
    if not url:
        return ""
    parsed = urlparse(url)
    if parsed.path == "/l/":
        params = parse_qs(parsed.query)
        values = params.get("uddg")
        if values:
            return values[0]
    if parsed.scheme in {"http", "https"}:
        return url
    return urljoin("https://duckduckgo.com", url)


def _xml_child_text(item: ET.Element, tag: str) -> str:
    child = item.find(tag)
    if child is None or child.text is None:
        return ""
    return child.text


def _write_search_result_header(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SEARCH_RESULT_CSV_FIELDS)
        writer.writeheader()


def _append_search_result_rows(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not path.exists()
    with path.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SEARCH_RESULT_CSV_FIELDS)
        if write_header:
            writer.writeheader()
        for row in rows:
            writer.writerow(
                {field: _csv_value(row.get(field)) for field in SEARCH_RESULT_CSV_FIELDS}
            )


def _read_completed_search_queries(path: Path) -> set[str]:
    if not path.exists():
        return set()
    queries = set()
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            query = _clean_text(row.get("search_query"))
            if query:
                queries.add(query)
    return queries


def _write_search_debug_html(raw_dir: Path, query: str, content: bytes) -> Path:
    raw_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha1(query.encode("utf-8")).hexdigest()[:12]
    output_path = raw_dir / f"{digest}.html"
    output_path.write_bytes(content)
    return output_path


def _is_acceptable_discovery_url(url: str) -> bool:
    try:
        parsed = urlparse(url)
    except ValueError:
        return False
    if parsed.scheme not in {"http", "https"}:
        return False
    host = parsed.netloc.lower()
    if not host:
        return False
    return host.endswith(".edu.cn") or host.endswith(".edu") or ".edu.cn" in host


def _search_result_matches_task(
    task: dict[str, Any],
    title: str,
    url: str,
    snippet: str,
) -> bool:
    school_name = _clean_text(task.get("school_name"))
    haystack = f"{title} {unquote(url)} {snippet}"
    classification = classify_document(title, url, snippet)
    document_type = task.get("document_type")
    if classification["document_type"] == "unknown":
        return False
    if document_type and classification["document_type"] != document_type:
        if not (
            document_type == "incoming_recommendation_admission_list"
            and classification["document_type"] == "recommendation_exemption_list"
        ):
            return False
    if school_name and school_name not in haystack:
        compact_school = re.sub(r"(大学|学院|学校)$", "", school_name)
        if compact_school and compact_school not in haystack:
            return False
    return True


def _iter_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return rows


def _clean_record(record: dict[str, Any]) -> dict[str, Any]:
    base = {
        "school_name": _clean_text(record.get("school_name")),
        "year": _to_int(record.get("year")),
        "document_type": _clean_text(record.get("document_type") or "unknown"),
        "route": _clean_text(record.get("route") or "unknown"),
        "person_name": _clean_text(record.get("person_name")),
        "student_id": _clean_text(record.get("student_id")),
        "undergraduate_school": _clean_text(record.get("undergraduate_school")),
        "undergraduate_major": _clean_text(record.get("undergraduate_major")),
        "college": _clean_text(record.get("college")),
        "major": _clean_text(record.get("major")),
        "admission_major": _clean_text(record.get("admission_major")),
        "ranking": _clean_text(record.get("ranking")),
        "remarks": _clean_text(record.get("remarks")),
        "source_url": _clean_text(record.get("source_url")),
        "title": _clean_text(record.get("title")),
        "needs_review": bool(record.get("needs_review")),
        "_input_row_index": record.get("_input_row_index"),
        "_extra_quality_flags": [],
    }
    _repair_shifted_sequence_name(base)
    _move_plan_text_out_of_student_id(base)
    _clear_header_text_in_value_fields(base)
    _move_org_text_out_of_major(base)
    quality_flags = _quality_flags(base)
    base["quality_flags"] = ";".join(quality_flags)
    base["quality_score"] = _quality_score(base, quality_flags)
    base["person_name_masked"] = _mask_name(base["person_name"])
    base["student_id_masked"] = _mask_identifier(base["student_id"])
    base["record_id"] = _record_id(base)
    return base


def _move_org_text_out_of_major(record: dict[str, Any]) -> None:
    major = str(record.get("major") or "")
    if not major or record.get("admission_major"):
        return
    if not re.search(r"(大学|学院|研究院|研究生院|中心|学部|系|所)$", major):
        return
    if not record.get("college"):
        record["college"] = major
    record["major"] = ""


def _move_plan_text_out_of_student_id(record: dict[str, Any]) -> None:
    student_id = str(record.get("student_id") or "")
    match = re.fullmatch(
        r"\s*([0-9Xx*]{10,})\s+"
        r"(少干计划|少数民族骨干计划|骨干计划|退役大学生士兵计划|大学生士兵计划|士兵计划|专项计划|单列计划)\s*",
        student_id,
    )
    if not match:
        return

    record["student_id"] = match.group(1)
    plan_text = match.group(2)
    remarks = str(record.get("remarks") or "")
    if plan_text not in remarks:
        record["remarks"] = _clean_text(f"{remarks} {plan_text}")


def _repair_shifted_sequence_name(record: dict[str, Any]) -> None:
    person_name = str(record.get("person_name") or "")
    student_id = str(record.get("student_id") or "")
    admission_major = str(record.get("admission_major") or "")
    if _repair_direction_prefixed_name(record):
        return
    id_and_name_in_major = re.fullmatch(r"\s*([0-9Xx*]{6,})\s+([\u4e00-\u9fff·]{2,5})\s*", admission_major)
    if _looks_like_score_or_metric(person_name) and _looks_like_score_or_metric(student_id) and id_and_name_in_major:
        record["person_name"] = id_and_name_in_major.group(2)
        record["student_id"] = id_and_name_in_major.group(1)
        record["admission_major"] = ""
        extra_flags = record.setdefault("_extra_quality_flags", [])
        if "score_columns_shifted" not in extra_flags:
            extra_flags.append("score_columns_shifted")
        return
    id_gender_and_name = re.fullmatch(r"\s*([0-9Xx*]{6,})(?:\s+([男女]))?\s*", person_name)
    if id_gender_and_name and _looks_like_chinese_name(student_id):
        record["person_name"] = student_id
        record["student_id"] = id_gender_and_name.group(1)
        gender = id_gender_and_name.group(2)
        if gender:
            remarks = str(record.get("remarks") or "")
            if gender not in remarks:
                record["remarks"] = _clean_text(f"{remarks} {gender}")
        return
    id_and_name = re.fullmatch(r"\s*([0-9Xx*]{6,})\s+([\u4e00-\u9fff·]{2,5})\s*", student_id)
    if _looks_like_header_or_org_name(person_name) and id_and_name:
        record["college"] = person_name
        record["student_id"] = id_and_name.group(1)
        record["person_name"] = id_and_name.group(2)
        return
    if (
        _looks_like_identifier_only(person_name)
        and not student_id
        and _looks_like_chinese_name(admission_major)
    ):
        record["student_id"] = person_name
        record["person_name"] = admission_major
        record["admission_major"] = ""
        return
    if not person_name.isdigit() or not _looks_like_chinese_name(student_id):
        if person_name.isdigit() and _looks_like_identifier_only(student_id):
            record["person_name"] = ""
            record["needs_review"] = True
        return
    if not record.get("ranking"):
        record["ranking"] = person_name
    record["person_name"] = student_id
    record["student_id"] = ""


def _repair_direction_prefixed_name(record: dict[str, Any]) -> bool:
    person_name = str(record.get("person_name") or "")
    student_id = str(record.get("student_id") or "")
    direction_match = re.fullmatch(r"(\d{2}【[^】]+】)\s*(.*)", person_name)
    if direction_match:
        direction = direction_match.group(1)
        suffix = direction_match.group(2).strip()
        if _looks_like_chinese_name(suffix) and _looks_like_identifier_only(student_id):
            if not record.get("admission_major"):
                record["admission_major"] = direction
            record["person_name"] = suffix
            return True
        if not suffix and _looks_like_chinese_name(student_id):
            if not record.get("admission_major"):
                record["admission_major"] = direction
            record["person_name"] = student_id
            record["student_id"] = ""
            return True

    if person_name.startswith(tuple(f"{index:02d}【" for index in range(100))) and _looks_like_identifier_only(student_id):
        suffix = _trailing_chinese_name(person_name)
        if suffix:
            prefix = person_name[: -len(suffix)].strip()
            if prefix and not record.get("admission_major"):
                record["admission_major"] = prefix
            record["person_name"] = suffix
            return True
    return False


def _trailing_chinese_name(value: str) -> str:
    common_surnames = "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳鲍史唐费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮卞齐康伍余元卜顾孟平黄和穆萧尹"
    for length in (2, 3, 4):
        candidate = value[-length:]
        if _looks_like_chinese_name(candidate) and candidate[0] in common_surnames:
            return candidate
    suffix_match = re.search(r"([\u4e00-\u9fff]{2,4})$", value)
    return suffix_match.group(1) if suffix_match else ""


def _looks_like_chinese_name(value: str) -> bool:
    return bool(re.fullmatch(r"[\u4e00-\u9fff]{2,4}(?:·[\u4e00-\u9fff]{1,4})?", value or ""))


def _write_clean_records_csv(rows: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CLEAN_RECORD_CSV_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {field: _csv_value(row.get(field)) for field in CLEAN_RECORD_CSV_FIELDS}
            )


def _truthy_csv_value(value: Any) -> bool:
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y"}
    return bool(value)


def _build_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, Any, str, str], dict[str, Any]] = {}
    for row in rows:
        key = (
            row.get("school_name") or "",
            row.get("year") or "",
            row.get("document_type") or "unknown",
            row.get("route") or "unknown",
        )
        group = groups.setdefault(
            key,
            {
                "school_name": key[0],
                "year": key[1],
                "document_type": key[2],
                "route": key[3],
                "record_count": 0,
                "unique_people": set(),
                "needs_review_count": 0,
                "with_undergraduate_school_count": 0,
                "with_admission_major_count": 0,
                "source_documents": set(),
            },
        )
        group["record_count"] += 1
        if row.get("person_name") or row.get("student_id"):
            group["unique_people"].add((row.get("person_name"), row.get("student_id")))
        if _truthy_csv_value(row.get("needs_review")):
            group["needs_review_count"] += 1
        if row.get("undergraduate_school"):
            group["with_undergraduate_school_count"] += 1
        if row.get("admission_major"):
            group["with_admission_major_count"] += 1
        if row.get("source_url"):
            group["source_documents"].add(row.get("source_url"))

    summary_rows = []
    for group in groups.values():
        summary_rows.append(
            {
                "school_name": group["school_name"],
                "year": group["year"],
                "document_type": group["document_type"],
                "route": group["route"],
                "record_count": group["record_count"],
                "unique_person_count": len(group["unique_people"]),
                "needs_review_count": group["needs_review_count"],
                "with_undergraduate_school_count": group["with_undergraduate_school_count"],
                "with_admission_major_count": group["with_admission_major_count"],
                "source_document_count": len(group["source_documents"]),
            }
        )
    return sorted(
        summary_rows,
        key=lambda row: (
            str(row["school_name"]),
            str(row["year"]),
            str(row["document_type"]),
            str(row["route"]),
        ),
    )


def _write_summary_csv(rows: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SUMMARY_CSV_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _quality_flags(record: dict[str, Any]) -> list[str]:
    flags = []
    flags.extend(record.get("_extra_quality_flags") or [])
    if not record.get("person_name"):
        flags.append("missing_person_name")
    if not record.get("student_id"):
        flags.append("missing_student_id")
    if not record.get("undergraduate_school"):
        flags.append("missing_undergraduate_school")
    if not record.get("admission_major") and not record.get("major"):
        flags.append("missing_major")
    if record.get("needs_review"):
        flags.append("needs_review")
    return flags


def _quality_score(record: dict[str, Any], flags: list[str]) -> int:
    score = 100
    score -= 12 * len(flags)
    if record.get("source_url"):
        score += 5
    return max(0, min(100, score))


def _record_id(record: dict[str, Any]) -> str:
    parts = [
        str(record.get("school_name") or ""),
        str(record.get("year") or ""),
        str(record.get("document_type") or ""),
        str(record.get("person_name") or ""),
        str(record.get("student_id") or ""),
        str(record.get("undergraduate_school") or ""),
        str(record.get("admission_major") or record.get("major") or ""),
        str(record.get("source_url") or ""),
    ]
    if not record.get("student_id") and record.get("ranking"):
        parts.append(f"rank:{record.get('ranking')}")
    if _is_masked_identity_without_identifier(record):
        parts.append(str(record.get("_input_row_index") or ""))
    identity = "|".join(parts)
    return hashlib.sha1(identity.encode("utf-8")).hexdigest()


def _record_dedupe_key(record: dict[str, Any]) -> str:
    if not (record.get("person_name") or record.get("student_id")):
        return str(record.get("record_id") or "")
    if _is_masked_identity_without_identifier(record):
        return str(record.get("record_id") or "")
    person_name = str(record.get("person_name") or "")
    student_id = str(record.get("student_id") or "")
    identity_value = (
        f"{person_name}|{student_id}" if person_name and student_id else person_name or student_id
    )
    if not student_id and record.get("ranking"):
        identity_value = f"{identity_value}|rank:{record.get('ranking')}"
    identity = "|".join(
        [
            str(record.get("school_name") or ""),
            str(record.get("year") or ""),
            str(record.get("document_type") or ""),
            str(record.get("route") or ""),
            identity_value,
            str(record.get("college") or ""),
            str(record.get("admission_major") or record.get("major") or ""),
        ]
    )
    return hashlib.sha1(identity.encode("utf-8")).hexdigest()


def _records_are_duplicate(left: dict[str, Any], right: dict[str, Any]) -> bool:
    for field in ("school_name", "year", "document_type", "route"):
        if str(left.get(field) or "") != str(right.get(field) or ""):
            return False
    if _is_masked_identity_without_identifier(left) or _is_masked_identity_without_identifier(right):
        return False

    left_name = str(left.get("person_name") or "")
    right_name = str(right.get("person_name") or "")
    left_student_id = str(left.get("student_id") or "")
    right_student_id = str(right.get("student_id") or "")
    left_ranking = str(left.get("ranking") or "")
    right_ranking = str(right.get("ranking") or "")
    if left_student_id and right_student_id and left_student_id != right_student_id:
        return False
    if not (left_student_id or right_student_id) and left_ranking and right_ranking and left_ranking != right_ranking:
        return False
    if "*" in left_name or "*" in right_name:
        if left_name != right_name:
            return False
        if left_student_id or right_student_id:
            if not (left_student_id and right_student_id and left_student_id == right_student_id):
                return False
    left_identity = left_name or left_student_id
    right_identity = right_name or right_student_id
    if not left_identity or left_identity != right_identity:
        return False

    return _compatible_optional_field(left, right, "college") and _compatible_optional_major(left, right)


def _compatible_optional_field(left: dict[str, Any], right: dict[str, Any], field: str) -> bool:
    left_value = str(left.get(field) or "")
    right_value = str(right.get(field) or "")
    return not left_value or not right_value or left_value == right_value


def _compatible_optional_major(left: dict[str, Any], right: dict[str, Any]) -> bool:
    left_value = str(left.get("admission_major") or left.get("major") or "")
    right_value = str(right.get("admission_major") or right.get("major") or "")
    return not left_value or not right_value or left_value == right_value


def _is_masked_identity_without_identifier(record: dict[str, Any]) -> bool:
    return "*" in str(record.get("person_name") or "") and not str(record.get("student_id") or "")


def _looks_like_masked_person_name(value: str) -> bool:
    return bool(re.fullmatch(r"[\u4e00-\u9fff·*]{2,8}", value or ""))


def _looks_like_non_person_identity(record: dict[str, Any]) -> bool:
    name = str(record.get("person_name") or "")
    if not name:
        return not bool(record.get("student_id"))
    if name.endswith("老师"):
        return True
    status_only_major = str(record.get("admission_major") or record.get("major") or "")
    no_supporting_identity_context = (
        not record.get("student_id")
        and not any(
            record.get(field)
            for field in (
                "college",
                "undergraduate_school",
                "undergraduate_major",
                "ranking",
            )
        )
    )
    if no_supporting_identity_context and status_only_major in {
        "一审",
        "二审",
        "三审",
        "初审",
        "复审",
        "终审",
        "各位 推免生",
        "DATE",
        "辅修",
        "位置导航",
        "邮编",
        "友情链接",
    }:
        return True
    if no_supporting_identity_context and len(name) == 1:
        return True
    if no_supporting_identity_context and re.fullmatch(
        "[一二三]等(?:战功|功)|嘉奖|立功",
        name,
    ):
        return True
    if no_supporting_identity_context and not status_only_major and not record.get("remarks"):
        if re.fullmatch(r"\d{1,3}(?:\.\d+)?分", name):
            return True
        if re.fullmatch(r"加试[\u4e00-\u9fff0-9]+[:：]\d{1,3}(?:\.\d+)?", name):
            return True
        if re.fullmatch(r"[（(].{1,30}[）)]", name):
            return True
        if name in {"结果"}:
            return True
        source_url = str(record.get("source_url") or "").lower()
        if (
            record.get("needs_review")
            and record.get("document_type") == "postgraduate_admission_list"
            and "hebust.edu.cn/docs/" in source_url
        ):
            return True
        if (
            record.get("needs_review")
            and record.get("document_type") == "postgraduate_admission_list"
            and "grad.qdu.edu.cn/__local/" in source_url
        ):
            return True
    if (
        status_only_major in {"拟录取", "待录取", "录取"}
        and not record.get("student_id")
        and not any(
            record.get(field)
            for field in (
                "college",
                "undergraduate_school",
                "undergraduate_major",
                "ranking",
            )
        )
    ):
        return True
    if "关闭窗口" in name or name.startswith("当前位置"):
        return True
    if name in {
        "快速通道",
        "通知公告",
        "详细信息",
        "合格",
        "友情链接",
        "常用链接",
        "夜晚模式",
        "白天模式",
        "日间模式",
        "深色模式",
        "浅色模式",
        "联系人",
        "联系方式",
        "联系电话",
        "第一页",
        "上一页",
        "下一页",
        "最后一页",
        "末页",
        "下载中心",
        "主题教育",
        "党务公开",
        "党校在线",
        "党风廉政",
        "创制中心",
        "研究中心",
        "语合中心",
        "资料下载",
    }:
        return True
    if name in {
        "\u4e0b\u8f7d\u4e13\u533a",
        "\u4f18\u79c0\u8bba\u6587",
        "\u515a\u5efa\u601d\u653f",
        "\u521b\u65b0\u521b\u4e1a",
        "\u535a\u58eb\u62db\u751f",
        "\u57f9\u517b\u65b9\u6848",
        "\u5b66\u4f4d\u6388\u4e88",
        "\u5b66\u4f4d\u6587\u4ef6",
        "\u5b66\u7c4d\u7ba1\u7406",
        "\u5bfc\u5e08\u57f9\u8bad",
        "\u5bfc\u5e08\u5de5\u4f5c",
        "\u5bfc\u5e08\u8bc4\u4f18",
        "\u5bfc\u5e08\u9074\u9009",
        "\u5bfc\u5e08\u98ce\u91c7",
        "\u5de5\u4f5c\u52a8\u6001",
        "\u5e38\u7528\u4e0b\u8f7d",
        "\u62db\u751f\u7b80\u7ae0",
        "\u6559\u52a1\u7ba1\u7406",
        "\u6821\u56ed\u6d3b\u52a8",
        "\u76f8\u5173\u6587\u4ef6",
        "\u7814\u7a76\u751f\u4f1a",
        "\u7855\u58eb\u62db\u751f",
        "\u793e\u4f1a\u5b9e\u8df5",
        "\u79d1\u7814\u5de5\u4f5c",
        "\u884c\u4e3a\u7ba1\u7406",
    }:
        return True
    if name in {"\u975e\u5168\u65e5", "\u975e\u5168\u65e5\u5236", "\u5168\u65e5\u5236", "\u5168\u65e5"}:
        return True
    if name in {"\u603b\u5206", "\u6210\u7ee9", "\u521d\u8bd5", "\u590d\u8bd5", "\u8bed\u542c\u529b", "\u6027\u522b"}:
        return True
    if name in {"男", "女"}:
        return True
    if re.fullmatch(r"[男女]\s*\d{1,4}", name) and not record.get("student_id"):
        return True
    if _looks_like_score_or_metric(name):
        return True
    student_id = str(record.get("student_id") or "")
    student_id_is_identifier = _looks_like_identifier_only(student_id)
    if (
        not student_id
        and re.fullmatch(r"[（(][^（）()]{1,30}[）)]", name)
        and _looks_like_non_person_label(name)
    ):
        return True
    has_program_context = any(
        record.get(field)
        for field in (
            "college",
            "major",
            "admission_major",
            "undergraduate_school",
            "undergraduate_major",
            "ranking",
        )
    )
    compact_context = re.sub(
        r"\s+",
        "",
        " ".join(
            str(record.get(field) or "")
            for field in (
                "college",
                "major",
                "admission_major",
                "undergraduate_school",
                "undergraduate_major",
                "remarks",
            )
        ),
    )
    if (
        not student_id
        and not record.get("ranking")
        and not record.get("undergraduate_school")
        and not record.get("undergraduate_major")
        and not record.get("remarks")
        and re.fullmatch(r"\d{3}", str(record.get("college") or ""))
        and re.search(r"(学院|中心|系|部)", str(record.get("major") or ""))
        and re.fullmatch(r"\d{4}[0-9A-Za-z]{2}", str(record.get("admission_major") or ""))
    ):
        return True
    if not student_id and re.search(r"(联系人|联系方式|联系电话|电话|邮箱)", compact_context):
        return True
    if (
        has_program_context
        and _is_masked_identity_without_identifier(record)
        and _looks_like_masked_person_name(name)
    ):
        return False
    if (
        student_id
        and not student_id_is_identifier
        and not has_program_context
        and not re.fullmatch(r"[0-9Xx*]{6,}", student_id)
    ):
        return True
    if (
        student_id in {"院", "学院"}
        and not record.get("college")
        and not record.get("admission_major")
        and not record.get("major")
    ):
        return True
    if (
        not student_id_is_identifier
        and name in {"工程", "系统", "智能化", "化装置"}
        and not record.get("college")
        and not record.get("admission_major")
        and not record.get("major")
        and re.search(r"(学院|大学|工程|院)", student_id)
    ):
        return True
    if len(name) > 8 and not student_id_is_identifier and not _looks_like_dotted_person_name(name):
        return True
    if name.isdigit():
        if not student_id or _looks_like_score_or_metric(student_id):
            return True
        if not student_id_is_identifier and not _looks_like_chinese_name(student_id):
            return True
    if (
        student_id_is_identifier
        and (_looks_like_chinese_name(name) or _looks_like_dotted_person_name(name) or "*" in name)
        and not _looks_like_non_person_label(name)
    ):
        return False
    if (
        not student_id
        and record.get("ranking")
        and has_program_context
        and _looks_like_chinese_name(name)
        and not re.search(r"(学院|专业|学校|学科|备注|排名|序号|姓名|考生|录取|接收|申请|毕业)", name)
    ):
        return False
    if _looks_like_header_or_org_name(name):
        return True
    if (
        not record.get("student_id")
        and _looks_like_admission_notice_section(name)
        and str(record.get("admission_major") or record.get("major") or "") in {"招收", "接收"}
    ):
        return True
    if _looks_like_non_person_label(name):
        student_id = str(record.get("student_id") or "")
        if (
            not student_id
            or _looks_like_score_or_metric(student_id)
            or _looks_like_non_person_label(student_id)
        ):
            return True
    if (
        name.startswith("方向")
        and not record.get("student_id")
        and re.search(r"(学院|大学|学校|中心|部门)", str(record.get("admission_major") or record.get("major") or ""))
    ):
        return True
    if record.get("student_id"):
        if (
            not student_id_is_identifier
            and not _looks_like_chinese_name(name)
            and not _looks_like_dotted_person_name(name)
            and not re.fullmatch(r"[A-Za-z][A-Za-z .'-]{1,40}", name)
            and "*" not in name
        ):
            return True
        return False
    if name in {"其他", "准确", "有效", "机械", "在机械", "修复"}:
        return True
    if str(record.get("admission_major") or record.get("major") or "") == "接收" and not record.get("college"):
        return True
    return bool(re.search(r"(学院|专业|学校|学科|备注|排名|序号|姓名|考生|录取|接收|申请|毕业)", name))


def _looks_like_header_or_org_name(value: str) -> bool:
    return bool(
        re.search(
            r"(学院|大学|科学系|姓名|[^姓]名|考生|录取|专业|学校|学科|成绩|排名|序号|申请|毕业|备注|专项|权重|代码|编号)",
            value or "",
        )
    )


def _looks_like_non_person_label(value: str) -> bool:
    text = value or ""
    if text in {
        "滨海校区",
        "研究生院",
        "学业指导",
        "教师发展",
        "办公网",
        "图书馆",
        "学号",
        "教务处",
        "财务系统",
        "会计",
        "合作院校",
        "工商管理学",
        "管理科学与工程",
        "方向",
        "院系所",
        "支教团",
        "硕士",
        "直博生",
        "天开杯",
        "主要",
        "创新能力",
        "合作交流",
        "学生工作",
        "就业创业",
        "研究生",
        "科研项目",
        "素养",
        "能力",
        "后台管理",
        "特此通知",
    }:
        return True
    if re.search(r"(定向|非定向|全日制|非全日制|类别|类型|计划|方向)", text):
        return True
    if re.search(r"(面试|笔试|口试|总分|初试|复试|成绩|分数)", text):
        return True
    if re.search("(\u4e1a\u52a1\u8bfe\u4e00|\u4e1a\u52a1\u8bfe\u4e8c)", text):
        return True
    if re.search(r"(业务一|业务二|外语|英语|思想政|顺序号|管理类综|综合能力|合能力|政治|科目)", text):
        return True
    if re.search(r"[×xX]\s*\d+\s*[％%]", text):
        return True
    compact = re.sub(r"\s+", "", text)
    if re.search(r"(推荐人|被荐人|请您|报考导师|材料审核|以下请|公正而详尽|了解进行说明|^推荐$)", text):
        return True
    if re.search(r"(联系人|联系方式|联系电话|电话|邮箱)", compact):
        return True
    if re.search(r"(公示|名单|姓名|学院|专业|学校|学科|备注|排名|序号|考生|录取|接收|申请|毕业|院代码|代码|类别|类型|学习方式|成绩|分数|总分|初试|复试|合计|本科|推免|编号)", text):
        return True
    return bool(re.search(r"[（(][^）)]*方向[^）)]*[）)]|^（?[^）)]*方向）?$", text))


def _looks_like_dotted_person_name(value: str) -> bool:
    text = value or ""
    return ("·" in text or "&middot;" in text.lower()) and bool(
        re.fullmatch(r"[\u4e00-\u9fffA-Za-z&;.\s·]+", text)
    )


def _looks_like_score_or_metric(value: str) -> bool:
    return bool(re.fullmatch(r"\d{1,3}(?:\.\d+)?", (value or "").strip()))


def _looks_like_admission_notice_section(value: str) -> bool:
    return bool(
        re.fullmatch(
            r"(其他事项|奖学金|招收条件|招收程序|联系方式|诚实守信|遵纪守法)",
            value or "",
        )
    )


def _public_record_id(row: dict[str, Any]) -> str:
    identity = "|".join(
        [
            str(row.get("record_id") or ""),
            str(row.get("school_name") or ""),
            str(row.get("year") or ""),
            str(row.get("source_url") or ""),
        ]
    )
    return hashlib.sha1(identity.encode("utf-8")).hexdigest()


def _mask_name(value: str) -> str:
    if not value:
        return ""
    if len(value) <= 1:
        return "*"
    return value[0] + "*" * (len(value) - 1)


def _mask_identifier(value: str) -> str:
    if not value:
        return ""
    if len(value) <= 4:
        return "*" * len(value)
    return value[:4] + "*" * max(2, len(value) - 4)


def _records_from_table(rows: list[list[str]], document: dict[str, Any]) -> list[dict[str, Any]]:
    cleaned_rows = [[_clean_text(value) for value in row] for row in rows]
    gdufe_records = _records_from_gdufe_postgraduate_pdf_rows(cleaned_rows, document)
    if gdufe_records:
        return gdufe_records
    header_index, field_map = _find_header_row(cleaned_rows)
    if header_index is None:
        return []

    header = cleaned_rows[header_index]
    records: list[dict[str, Any]] = []
    for values in cleaned_rows[header_index + 1 :]:
        if not any(values):
            continue
        values = _align_values_to_header(header, values)
        normalized = {
            field: values[index] if index < len(values) else ""
            for index, field in field_map.items()
        }
        record = _build_record(document, normalized)
        if record:
            records.append(record)
    return records


def _records_from_recommendation_admission_status_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    header_index = _recommendation_admission_status_header_index(rows)
    if header_index is None:
        return []

    header = [_clean_text(value) for value in rows[header_index]]
    normalized_header = [_normalize_header(value) for value in header]
    sequence_index = _header_column_index(normalized_header, ["序号"])
    name_index = _header_column_index(normalized_header, ["姓名", "考生姓名"])
    student_id_index = _header_column_index(normalized_header, ["身份证号", "身份证号码", "证件号码", "证件号"])
    college_index = _header_column_index(normalized_header, ["报考学院", "拟录取学院", "录取学院", "学院"])
    major_index = _header_column_index(normalized_header, ["报考专业", "拟录取专业", "录取专业", "专业"])
    degree_type_index = _header_column_index(normalized_header, ["学位类型", "学位类别"])
    study_mode_index = _header_column_index(normalized_header, ["学习方式", "学习形式"])
    retest_score_index = _header_column_index(normalized_header, ["复试成绩", "考核成绩"])
    status_index = _header_column_index(normalized_header, ["录取状态", "拟录取状态", "录取结果"])
    if name_index is None or major_index is None:
        return []

    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        values = [_clean_text(value) for value in row]
        if not any(values):
            continue
        values = _align_values_to_header(header, values)
        person_name = values[name_index] if name_index < len(values) else ""
        admission_major = values[major_index] if major_index < len(values) else ""
        if not _looks_like_chinese_name(person_name) or not admission_major:
            continue

        status = values[status_index] if status_index is not None and status_index < len(values) else ""
        if status and not re.search(r"(拟录取|待录取|录取)", status):
            continue
        remarks = []
        for label, index in [
            ("学位类型", degree_type_index),
            ("学习方式", study_mode_index),
            ("复试成绩", retest_score_index),
            ("录取状态", status_index),
        ]:
            if index is not None and index < len(values) and values[index]:
                remarks.append(f"{label} {values[index]}")

        record = _build_record(
            document,
            {
                "person_name": person_name,
                "student_id": values[student_id_index] if student_id_index is not None and student_id_index < len(values) else "",
                "college": values[college_index] if college_index is not None and college_index < len(values) else "",
                "admission_major": admission_major,
                "ranking": values[sequence_index] if sequence_index is not None and sequence_index < len(values) else "",
                "remarks": "; ".join(remarks),
            },
        )
        if record:
            records.append(record)
    return records


def _recommendation_admission_status_header_index(rows: list[list[str]]) -> int | None:
    for index, values in enumerate(rows[:20]):
        normalized = {_normalize_header(value) for value in values}
        if (
            "姓名" in normalized
            and any(value in normalized for value in {"身份证号", "身份证号码", "证件号码", "证件号"})
            and any(value in normalized for value in {"报考专业", "拟录取专业", "录取专业"})
            and any(value in normalized for value in {"录取状态", "拟录取状态", "录取结果"})
        ):
            return index
    return None


def _records_from_gdufe_postgraduate_pdf_rows(
    rows: list[list[str]],
    document: dict[str, Any],
) -> list[dict[str, Any]]:
    if document.get("document_type") != "postgraduate_admission_list":
        return []
    title = str(document.get("title") or "")
    source_url = str(document.get("source_url") or "")
    school_name = str(document.get("school_name") or "")
    if not (
        "广东财经大学" in school_name
        or "广东财经大学" in title
        or "gdufe.edu.cn" in source_url
    ):
        return []

    header_index = None
    for index, row in enumerate(rows[:20]):
        header_text = " ".join(row)
        if all(term in header_text for term in ("姓名", "考生编号", "专业代码", "专业名称")) and "拟录取" in header_text:
            header_index = index
            break
    if header_index is None:
        return []

    records: list[dict[str, Any]] = []
    pending_college = ""
    index = header_index + 1
    while index < len(rows):
        row = [_clean_text(value) for value in rows[index]]
        if not any(row):
            index += 1
            continue

        if _gdufe_row_starts_candidate(row):
            suffix = ""
            if index + 1 < len(rows) and _gdufe_single_college_fragment(rows[index + 1]):
                suffix = _clean_text(rows[index + 1][0])
            record = _gdufe_postgraduate_record(row, pending_college, suffix, document)
            pending_college = ""
            if record:
                records.append(record)
            index += 2 if suffix else 1
            continue

        if _gdufe_single_college_fragment(row):
            pending_college = _clean_text(f"{pending_college}{row[0]}")
        index += 1

    return records


def _gdufe_row_starts_candidate(row: list[str]) -> bool:
    return (
        len(row) >= 12
        and bool(re.fullmatch(r"\d+", row[0] or ""))
        and _looks_like_chinese_name(row[1] or "")
        and _looks_like_identifier_only(row[2] or "")
    )


def _gdufe_single_college_fragment(row: list[str]) -> bool:
    if len(row) != 1:
        return False
    value = _clean_text(row[0])
    if not value or _row_looks_like_header_continuation({"person_name": value}):
        return False
    return bool(re.search(r"(学院|研究院|中心|学部|院[）)]?$)", value))


def _gdufe_postgraduate_record(
    values: list[str],
    pending_college: str,
    suffix: str,
    document: dict[str, Any],
) -> dict[str, Any] | None:
    if _gdufe_looks_like_major_code(values[3] if len(values) > 3 else ""):
        college = _clean_text(f"{pending_college}{suffix}")
        major_code_index = 3
    elif len(values) > 4 and _gdufe_looks_like_major_code(values[4]):
        college = values[3]
        major_code_index = 4
    else:
        return None

    major_name_index = major_code_index + 1
    status_index = _gdufe_status_index(values)
    if status_index is None:
        return None

    status = values[status_index]
    if "拟录取" not in status:
        return None
    total_score = _gdufe_total_score(values, major_name_index, status_index)

    return _build_record(
        document,
        {
            "person_name": values[1],
            "student_id": values[2],
            "college": college,
            "major": values[major_code_index],
            "admission_major": values[major_name_index] if len(values) > major_name_index else "",
            "remarks": total_score,
        },
    )


def _gdufe_looks_like_major_code(value: str) -> bool:
    return bool(re.fullmatch(r"\d{6}", value or ""))


def _gdufe_status_index(values: list[str]) -> int | None:
    for index, value in enumerate(values):
        if re.search(r"(拟录取|候选|不予录取)", value or ""):
            return index
    return None


def _gdufe_total_score(values: list[str], major_name_index: int, status_index: int) -> str:
    score_values = [
        (index, value)
        for index, value in enumerate(values[major_name_index + 1 : status_index], start=major_name_index + 1)
        if _looks_like_score_or_metric(value)
    ]
    if not score_values:
        return ""
    if len(score_values) >= 2:
        last_index, last_value = score_values[-1]
        previous_value = score_values[-2][1]
        tail_after_last_score = values[last_index + 1 : status_index]
        if (
            re.fullmatch(r"\d{1,2}", last_value)
            and any(value in {"全日制", "非全日制"} or "调剂" in value or "一志愿" in value for value in tail_after_last_score)
        ):
            return previous_value
    return score_values[-1][1]


def _align_values_to_header(header: list[str], values: list[str]) -> list[str]:
    if len(values) < len(header):
        values = _insert_missing_optional_header_values(header, values)
    if len(values) <= len(header):
        return values

    aligned: list[str] = []
    value_index = 0
    for header_index, header_value in enumerate(header):
        if value_index >= len(values):
            aligned.append("")
            continue

        remaining_headers = len(header) - header_index - 1
        can_consume_pair = len(values) - (value_index + 2) >= remaining_headers
        normalized_header = _normalize_header(header_value)
        if can_consume_pair and _header_combines_code_and_name(normalized_header):
            aligned.append(_clean_text(f"{values[value_index]} {values[value_index + 1]}"))
            value_index += 2
        else:
            aligned.append(values[value_index])
            value_index += 1

    return aligned


def _insert_missing_optional_header_values(header: list[str], values: list[str]) -> list[str]:
    aligned = list(values)
    for index, header_value in enumerate(header[:-1]):
        if index >= len(aligned):
            break
        if not _header_is_optional_direction_group(_normalize_header(header_value)):
            continue
        if _field_for_header(header[index + 1]) != "student_id":
            continue
        if _looks_like_identifier_only(aligned[index]):
            aligned.insert(index, "")
            break
    return aligned


def _header_is_optional_direction_group(normalized_header: str) -> bool:
    return "研究方向" in normalized_header or "方向组" in normalized_header


def _header_combines_code_and_name(normalized_header: str) -> bool:
    return "代码及名称" in normalized_header or "代码名称" in normalized_header


def _find_header_row(
    rows: list[list[str]],
) -> tuple[int | None, dict[int, str]]:
    for index, row in enumerate(rows[:20]):
        field_map: dict[int, str] = {}
        for column_index, value in enumerate(row):
            field = _field_for_header(value)
            if field and field not in field_map.values():
                field_map[column_index] = field
        if len(field_map) >= 2 and (
            "person_name" in field_map.values()
            or "undergraduate_school" in field_map.values()
            or "admission_major" in field_map.values()
        ):
            return index, field_map
    return None, {}


def _field_for_header(value: str) -> str | None:
    normalized = _normalize_header(value)
    if not normalized:
        return None

    if "毕业单位" in normalized or "本科毕业单位" in normalized:
        return "undergraduate_school"
    if "专业" in normalized and _header_combines_code_and_name(normalized):
        return "admission_major"
    if "专业代码" in normalized and not any(
        term in normalized for term in ("本科专业", "毕业专业", "所学专业")
    ):
        return "major"
    if "专业名称" in normalized and any(
        term in normalized for term in ("拟录取", "录取", "接收", "报考", "申请")
    ):
        return "admission_major"

    best_field = None
    best_len = -1
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_normalized = _normalize_header(alias)
            if alias_normalized and alias_normalized in normalized and len(alias_normalized) > best_len:
                best_field = field
                best_len = len(alias_normalized)
    return best_field


def _build_record(
    document: dict[str, Any],
    normalized: dict[str, str],
) -> dict[str, Any] | None:
    compact = {key: _clean_text(value) for key, value in normalized.items()}
    _normalize_compact_identity_fields(compact)
    _clear_header_text_in_value_fields(compact)
    if _row_looks_like_header_continuation(compact):
        return None
    if not any(compact.values()):
        return None
    if not (
        compact.get("person_name")
        or compact.get("undergraduate_school")
        or compact.get("admission_major")
        or compact.get("major")
    ):
        return None
    if _looks_like_identifier_only(compact.get("person_name") or ""):
        return None
    if _looks_like_score_or_metric(compact.get("person_name") or ""):
        return None
    if not compact.get("person_name") and compact.get("student_id") and not _looks_like_identifier_only(compact.get("student_id") or ""):
        return None

    route = _infer_route(document.get("document_type") or "", compact)
    needs_review = not bool(
        compact.get("person_name")
        and (
            compact.get("admission_major")
            or compact.get("major")
            or compact.get("undergraduate_school")
        )
    )
    return {
        "schema_version": RECORD_SCHEMA_VERSION,
        "school_name": document.get("school_name") or "",
        "year": document.get("year"),
        "document_type": document.get("document_type") or "unknown",
        "route": route,
        "person_name": compact.get("person_name") or "",
        "student_id": compact.get("student_id") or "",
        "undergraduate_school": compact.get("undergraduate_school") or "",
        "undergraduate_major": compact.get("undergraduate_major") or "",
        "college": compact.get("college") or "",
        "major": compact.get("major") or "",
        "admission_major": compact.get("admission_major") or "",
        "ranking": compact.get("ranking") or "",
        "remarks": compact.get("remarks") or "",
        "source_url": document.get("source_url") or "",
        "title": document.get("title") or "",
        "needs_review": needs_review,
        "raw_row_json": json.dumps(compact, ensure_ascii=False, separators=(",", ":")),
    }


def _clear_header_text_in_value_fields(row: dict[str, str]) -> None:
    if row.get("ranking") in {"排名", "专业", "专业排名", "综合排名"}:
        row["ranking"] = ""


def _normalize_compact_identity_fields(row: dict[str, str]) -> None:
    person_name = row.get("person_name") or ""
    if _looks_like_major_code_name(person_name):
        if not row.get("admission_major") or row.get("admission_major") in {"无", "无专项计划"}:
            row["admission_major"] = person_name
        row["person_name"] = ""
    student_id = row.get("student_id") or ""
    match = re.fullmatch(r"([0-9Xx*]{10,})\s+(.+)", student_id)
    if match:
        row["student_id"] = match.group(1)
        trailing_value = _clean_text(match.group(2))
        if trailing_value and (
            not row.get("admission_major")
            or _looks_like_enrollment_category(row.get("admission_major") or "")
        ):
            if row.get("admission_major") and not row.get("remarks"):
                row["remarks"] = row["admission_major"]
            row["admission_major"] = trailing_value
    if row.get("student_id"):
        return
    for field in ("college", "major", "admission_major"):
        value = row.get(field) or ""
        match = re.fullmatch(r"([0-9Xx*]{10,})\s+(.+)", value)
        if not match:
            continue
        row["student_id"] = match.group(1)
        row[field] = _clean_text(match.group(2))
        return


def _looks_like_enrollment_category(value: str) -> bool:
    text = _clean_text(value)
    return text in {"定向", "非定向", "定向就业", "非定向就业", "全日制", "非全日制"}


def _looks_like_major_code_name(value: str) -> bool:
    return bool(re.fullmatch(r"\d{4}[0-9A-Z]{2}\s*[|｜ ]\s*[\u4e00-\u9fffA-Za-z（）()·、]+", value or ""))


def _row_looks_like_header_continuation(row: dict[str, str]) -> bool:
    values = {value for value in row.values() if value}
    if not values:
        return False
    header_terms = {
        "\u521d\u8bd5",
        "\u590d\u8bd5",
        "\u4e13\u4e1a",
        "\u7efc\u5408",
        "\u7efc\u5408\u6210",
        "\u6210\u7ee9",
        "\u57fa\u7840",
        "\u80fd\u529b",
        "\u5916\u8bed\u53e3",
        "\u8bed\u542c\u529b",
        "\u603b\u5206",
        "\u7ee9",
        "\u59d3\u540d",
        "\u8003\u751f\u7f16\u53f7",
        "\u5b66\u9662\u540d\u79f0",
        "\u62df\u5f55\u53d6\u4e13\u4e1a\u540d\u79f0",
        "\u62df\u5f55\u53d6\u9662\u7cfb\u6240\u540d\u79f0",
        "\u8bc1\u4ef6\u53f7\u7801",
        "\u4ee3\u7801",
        "\u5b66\u4e60\u65b9\u5f0f",
        "\u7c7b\u578b",
        "\u7c7b\u522b",
        "\u5907\u6ce8",
        "\u4e1a\u4ee3\u7801",
        "\u5206\u5236)",
        "\u5236)",
    }
    if values <= header_terms:
        return True
    name = row.get("person_name") or ""
    student_id = row.get("student_id") or ""
    if name in header_terms and student_id and not _looks_like_identifier_only(student_id):
        return True
    return False


def _infer_route(document_type: str, row: dict[str, str]) -> str:
    row_text = " ".join(row.values())
    if "推免" in row_text or "推荐免试" in row_text:
        return "recommendation_exemption"
    if document_type in {
        "recommendation_exemption_list",
        "incoming_recommendation_admission_list",
    }:
        return "recommendation_exemption"
    if document_type == "postgraduate_admission_list":
        return "postgraduate_exam_or_admission"
    return "unknown"


def _looks_like_identifier_only(value: str) -> bool:
    compact = re.sub(r"\s+", "", value or "")
    if len(compact) < 10:
        return False
    return bool(re.fullmatch(r"[0-9Xx*]+", compact))


def _parse_response_records(
    response: FetchResponse,
    raw_path: Path,
    document: dict[str, Any],
    html_text: str,
) -> list[dict[str, Any]]:
    if document.get("document_type") == "unknown":
        return []
    suffix = raw_path.suffix.lower()
    if _is_html_response(response):
        return parse_html_records(html_text, document)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_excel_records(raw_path, document)
    if suffix == ".xls":
        return parse_legacy_xls_records(raw_path, document)
    if suffix == ".docx":
        return parse_docx_records(raw_path, document)
    if suffix == ".pdf":
        return parse_pdf_records(raw_path, document)
    return []


def _is_homepage_redirect_from_detail(start_url: Any, final_url: str) -> bool:
    if not start_url or not final_url:
        return False
    try:
        start = urlparse(str(start_url))
        final = urlparse(final_url)
    except ValueError:
        return False
    if not start.scheme or not start.netloc or not final.scheme or not final.netloc:
        return False
    return (not _url_path_is_homepage(start.path)) and _url_path_is_homepage(final.path)


def _url_path_is_homepage(path: str) -> bool:
    normalized = (path or "/").rstrip("/").lower()
    return normalized in {"", "/index", "/index.htm", "/index.html", "/index.shtml", "/index.php"}


def _parse_status(response: FetchResponse, records: list[dict[str, Any]]) -> str:
    if records:
        return "parsed"
    if _is_html_response(response) or _is_attachment_url(response.url):
        return "parsed_no_records"
    return "indexed"


def _build_document_record(
    seed: dict[str, Any],
    response: FetchResponse,
    raw_path: Path,
    title: str,
    classification: dict[str, Any],
) -> dict[str, Any]:
    content_hash = hashlib.sha256(response.content).hexdigest()
    return {
        "schema_version": DOCUMENT_SCHEMA_VERSION,
        "captured_at": _now_iso(),
        "school_name": seed.get("school_name") or "",
        "source_type": seed.get("source_type") or "",
        "source_url": response.url,
        "start_url": seed.get("start_url") or "",
        "title": title,
        "year": _extract_year(title + " " + response.url),
        "document_type": classification.get("document_type") or "unknown",
        "matched_keywords": classification.get("matched_keywords") or [],
        "content_type": response.content_type,
        "content_length": len(response.content),
        "content_hash": content_hash,
        "raw_path": str(raw_path),
        "parse_status": "indexed",
        "record_count": 0,
    }


def _write_raw_document(raw_dir: Path, response: FetchResponse) -> Path:
    parsed = urlparse(response.url)
    host = _safe_segment(parsed.netloc or "unknown-host")
    suffix = _response_suffix(response)
    digest = hashlib.sha256(response.content).hexdigest()[:16]
    filename = f"{digest}{suffix}"
    output_path = raw_dir / host / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(response.content)
    return output_path


def _document_title(response: FetchResponse, text: str, fallback: str = "") -> str:
    if _is_html_response(response):
        soup = BeautifulSoup(text, "html.parser")
        if soup.title and soup.title.get_text(strip=True):
            return _clean_text(soup.title.get_text(" ", strip=True))
        heading = soup.find(["h1", "h2", "h3"])
        if heading:
            return _clean_text(heading.get_text(" ", strip=True))
    return _clean_text(fallback) or Path(urlparse(response.url).path).name


def _decode_response_text(response: FetchResponse) -> str:
    charset = _charset_from_content_type(response.content_type)
    for encoding in [charset, "utf-8", "gb18030"]:
        if not encoding:
            continue
        try:
            return response.content.decode(encoding)
        except (LookupError, UnicodeDecodeError):
            continue
    return response.content.decode("utf-8", errors="replace")


def _charset_from_content_type(content_type: str) -> str | None:
    match = re.search(r"charset=([\w.-]+)", content_type or "", flags=re.I)
    return match.group(1) if match else None


def _is_html_response(response: FetchResponse) -> bool:
    content_type = (response.content_type or "").lower()
    if content_type:
        return "html" in content_type
    return _url_suffix(response.url) in {"", ".htm", ".html", ".shtml"}


def _is_attachment_url(url: str) -> bool:
    return _url_suffix(url) in ATTACHMENT_SUFFIXES


def _response_suffix(response: FetchResponse) -> str:
    url_suffix = _url_suffix(response.url)
    content_suffix = _suffix_from_content_type(response.content_type)
    magic_suffix = _suffix_from_content_bytes(response.content)
    if magic_suffix in ATTACHMENT_SUFFIXES:
        return magic_suffix
    if content_suffix in ATTACHMENT_SUFFIXES:
        return content_suffix
    if url_suffix in ATTACHMENT_SUFFIXES or url_suffix in {".htm", ".html", ".shtml"}:
        return url_suffix
    if content_suffix:
        return content_suffix
    if url_suffix and url_suffix not in {".do", ".jsp", ".php", ".asp", ".aspx"}:
        return url_suffix
    return ".html" if _is_html_response(response) else ".bin"


def _suffix_from_content_type(content_type: str) -> str:
    clean = (content_type or "").split(";", 1)[0].strip().lower()
    if not clean:
        return ""
    if "spreadsheetml" in clean:
        return ".xlsx"
    if clean in {"application/vnd.ms-excel", "application/msexcel"} or "ms-excel" in clean:
        return ".xls"
    if "wordprocessingml" in clean:
        return ".docx"
    if clean == "application/msword" or "msword" in clean:
        return ".doc"
    if "pdf" in clean:
        return ".pdf"
    if "html" in clean:
        return ".html"
    return ""


def _suffix_from_content_bytes(content: bytes) -> str:
    if content.startswith(b"%PDF-"):
        return ".pdf"
    if content.startswith(b"PK") and b"xl/workbook" in content[:2000000]:
        return ".xlsx"
    return ""


def _url_suffix(url: str) -> str:
    parsed = urlparse(url)
    path = parsed.path.lower()
    suffix = Path(path).suffix
    if suffix not in ATTACHMENT_SUFFIXES:
        query_suffix = _attachment_suffix_from_query(parsed.query)
        if query_suffix:
            return query_suffix
    return suffix


def _attachment_suffix_from_query(query: str) -> str:
    for values in parse_qs(query, keep_blank_values=True).values():
        for value in values:
            clean_value = unquote(value).lower()
            if clean_value in ATTACHMENT_SUFFIXES:
                return clean_value
            suffix = Path(clean_value).suffix
            if suffix in ATTACHMENT_SUFFIXES:
                return suffix
    return ""


def _rows_from_text_lines(text: str) -> list[list[str]]:
    rows = []
    for line in text.splitlines():
        raw = str(line).strip()
        if not _clean_text(raw):
            continue
        if "\t" in raw:
            rows.append([_clean_text(part) for part in raw.split("\t") if _clean_text(part)])
        else:
            rows.append([_clean_text(part) for part in re.split(r"[ \u3000]{2,}", raw) if _clean_text(part)])
    return rows


def _normalize_header(value: str) -> str:
    return re.sub(r"[\s　:：()（）\[\]【】,，.。/\\]+", "", _clean_text(value))


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return re.sub(r"\s+", " ", text).strip()


def _to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _extract_year(text: str) -> int | None:
    for match in re.finditer(r"(20\d{2})", text):
        year = int(match.group(1))
        if 2000 <= year <= 2030:
            return year
    return None


def _extract_year_from_html(html: str) -> int | None:
    soup = BeautifulSoup(html, "html.parser")
    for node in soup(["script", "style", "noscript"]):
        node.decompose()

    context_keywords = ["拟录取", "录取名单", "硕士研究生", "推免", "推荐免试", "公示"]
    for node in soup.find_all(["h1", "h2", "h3", "p", "td", "th", "span"]):
        text = _clean_text(node.get_text(" ", strip=True))
        if any(keyword in text for keyword in context_keywords):
            year = _extract_year(text)
            if year is not None:
                return year

    return _extract_year(_clean_text(soup.get_text(" ", strip=True)))


def _request_headers(referer: str | None = None) -> dict[str, str]:
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/125.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        "Referer": referer or "https://gaokao.chsi.com.cn/",
        "Upgrade-Insecure-Requests": "1",
    }


def _append_jsonl(path: Path, record: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def _safe_segment(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("._") or "unknown"


def _csv_value(value: Any) -> Any:
    if isinstance(value, bool):
        return "true" if value else "false"
    if value is None:
        return ""
    return value


def _unique_texts(values: list[str]) -> list[str]:
    result = []
    seen = set()
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def _now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


if __name__ == "__main__":
    raise SystemExit(main())
