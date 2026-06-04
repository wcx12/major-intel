from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from scripts import graduate_outcome_crawler as crawler
except ImportError:  # pragma: no cover - direct script execution path
    import graduate_outcome_crawler as crawler


XLSX_PATH = Path(
    "data/raw/graduate_outcomes_official_site_websearch_web_20260601_batch249_jnu_doctor_admission/yz.jnu.edu.cn/e4b19b8e5a05e00b.xlsx"
)
OUT_DIR = Path(
    "data/processed/graduate_outcomes_official_site_websearch_web_20260601_batch249_jnu_doctor_admission_curated"
)

SCHOOL_NAME = "暨南大学"
YEAR = 2026
SOURCE_URL = (
    "https://yz.jnu.edu.cn/_upload/article/files/99/b1/1fb265f14ab3a02802f00c1b0d79/"
    "c7235e9d-63f1-4efc-8781-7d1090e91a28.xlsx"
)
TITLE = "关于公布2026年第二批次拟录取博士研究生名单的通知"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _remarks(parts: dict[str, str]) -> str:
    labels = (
        ("院系代码", parts.get("college_code", "")),
        ("录取类别", parts.get("admission_category", "")),
        ("考试方式", parts.get("exam_method", "")),
        ("初试总分", parts.get("initial_score", "")),
        ("复试总分", parts.get("retest_score", "")),
        ("总成绩", parts.get("total_score", "")),
        ("拟录取研究方向名称", parts.get("direction", "")),
        ("备注", parts.get("note", "")),
    )
    return "; ".join(f"{label}: {value}" for label, value in labels if value)


def _record_from_row(values: list[str]) -> dict[str, Any] | None:
    if len(values) < 14 or not values[0].isdigit():
        return None
    college_code = values[1]
    student_id = values[2]
    person_name = values[3]
    college = values[4]
    major_code = values[5]
    major_name = values[6]
    if not (student_id and person_name and college and major_code and major_name):
        return None

    record = {
        "school_name": SCHOOL_NAME,
        "year": YEAR,
        "document_type": "postgraduate_admission_list",
        "route": "postgraduate_exam_or_admission",
        "person_name": person_name,
        "student_id": student_id,
        "undergraduate_school": "",
        "undergraduate_major": "",
        "college": college,
        "major": major_code,
        "admission_major": major_name,
        "ranking": values[0],
        "remarks": _remarks(
            {
                "college_code": college_code,
                "admission_category": values[7],
                "exam_method": values[8],
                "initial_score": values[9],
                "retest_score": values[10],
                "total_score": values[11],
                "direction": values[12],
                "note": values[13],
            }
        ),
        "source_url": SOURCE_URL,
        "title": TITLE,
        "needs_review": False,
    }
    return crawler._clean_record(record)


def curate_records(xlsx_path: Path | None = None) -> list[dict[str, Any]]:
    path = xlsx_path or XLSX_PATH
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        records: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(values_only=True):
            values = [_cell_text(value) for value in row]
            record = _record_from_row(values)
            if record:
                records.append(record)
        return records
    finally:
        workbook.close()


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = curate_records()
    output = OUT_DIR / "records_clean_curated.csv"
    crawler._write_clean_records_csv(rows, output)
    print({"records": len(rows), "output": str(output)})


if __name__ == "__main__":
    main()
