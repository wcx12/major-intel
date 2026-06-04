from __future__ import annotations

import hashlib
import re
from pathlib import Path
from typing import Any

from bs4 import BeautifulSoup
from openpyxl import load_workbook

try:
    from scripts import graduate_outcome_crawler as crawler
except ImportError:  # pragma: no cover - direct script execution path
    import graduate_outcome_crawler as crawler


RAW_DIR = Path("data/raw/graduate_outcomes_official_site_websearch_web_20260528_batch235_cuhk_sz")
OUT_DIR = Path(
    "data/processed/graduate_outcomes_official_site_websearch_web_20260528_batch235_cuhk_sz_curated"
)

SCHOOL_NAME = "香港中文大学（深圳）"
COLLEGE = "数据科学学院"
HTML_SOURCE_URL = "https://sds.cuhk.edu.cn/article/2036"
HTML_TITLE = "香港中文大学（深圳）数据科学学院2025年拟录取推荐免试硕士研究生名单公示"
XLSX_SOURCE_URL = (
    "https://sds.cuhk.edu.cn/sites/default/files/2025-11/"
    "%E9%A6%99%E6%B8%AF%E4%B8%AD%E6%96%87%E5%A4%A7%E5%AD%A6%EF%BC%88%E6%B7%B1%E5%9C%B3%EF%BC%89"
    "%E6%95%B0%E6%8D%AE%E7%A7%91%E5%AD%A6%E5%AD%A6%E9%99%A2%E7%9B%B4%E7%A1%95%E6%8B%9F%E5%BD%95"
    "%E5%8F%96%E5%90%8D%E5%8D%95%EF%BC%882026%E5%B9%B4%E7%A7%8B%E5%AD%A3%E5%85%A5%E5%AD%A6%EF%BC%89.xlsx"
)
XLSX_TITLE = "香港中文大学（深圳）数据科学学院直硕拟录取名单（2026年秋季入学）"

TRANSLATION = str.maketrans(
    {
        "⾦": "金",
        "⻜": "飞",
    }
)


def _clean_text(value: Any) -> str:
    text = str(value or "").replace("\ufeff", "").translate(TRANSLATION)
    return re.sub(r"\s+", " ", text).strip()


def _remarks(*parts: str) -> str:
    return "; ".join(part for part in (_clean_text(part) for part in parts) if part)


def _html_record(cells: list[str]) -> dict[str, Any]:
    ranking, student_id, person_name, undergraduate_major, admission_major = cells[:5]
    return crawler._clean_record(
        {
            "school_name": SCHOOL_NAME,
            "year": 2025,
            "document_type": "recommendation_exemption_list",
            "route": "recommendation_exemption",
            "person_name": person_name,
            "student_id": student_id,
            "undergraduate_major": undergraduate_major,
            "college": COLLEGE,
            "major": undergraduate_major,
            "admission_major": admission_major,
            "ranking": ranking,
            "source_url": HTML_SOURCE_URL,
            "title": HTML_TITLE,
            "needs_review": False,
            "remarks": _remarks(f"undergraduate_major {undergraduate_major}"),
        }
    )


def _xlsx_record(cells: list[str]) -> dict[str, Any]:
    ranking, person_name, undergraduate_major, admission_major = cells[:4]
    return crawler._clean_record(
        {
            "school_name": SCHOOL_NAME,
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": person_name,
            "undergraduate_major": undergraduate_major,
            "college": COLLEGE,
            "major": undergraduate_major,
            "admission_major": admission_major,
            "ranking": ranking,
            "source_url": XLSX_SOURCE_URL,
            "title": XLSX_TITLE,
            "needs_review": False,
            "remarks": _remarks(f"undergraduate_major {undergraduate_major}"),
        }
    )


def _has_header(cells: list[str], expected: list[str]) -> bool:
    return cells[: len(expected)] == expected


def _parse_html_table(path: Path) -> list[dict[str, Any]]:
    soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="replace"), "html.parser")
    rows: list[dict[str, Any]] = []
    for table in soup.find_all("table"):
        parsed_rows = [
            [_clean_text(cell.get_text(" ", strip=True)) for cell in tr.find_all(["td", "th"])]
            for tr in table.find_all("tr")
        ]
        if not parsed_rows or not _has_header(parsed_rows[0], ["序号", "学号", "姓名", "本科专业", "拟录取专业"]):
            continue
        for cells in parsed_rows[1:]:
            if len(cells) >= 5 and cells[0].isdigit() and cells[1]:
                rows.append(_html_record(cells))
    return rows


def _parse_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows: list[dict[str, Any]] = []
        header_seen = False
        for raw_cells in worksheet.iter_rows(values_only=True):
            cells = [_clean_text(value) for value in raw_cells[:4]]
            if not any(cells):
                continue
            if _has_header(cells, ["序号", "姓名", "本科专业", "拟录取专业"]):
                header_seen = True
                continue
            if header_seen and len(cells) >= 4 and cells[0].isdigit() and cells[1] and cells[3]:
                rows.append(_xlsx_record(cells))
        return rows
    finally:
        workbook.close()


def _dedup_paths(paths: list[Path]) -> list[Path]:
    seen_hashes: set[str] = set()
    unique: list[Path] = []
    for path in sorted(paths):
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        if digest in seen_hashes:
            continue
        seen_hashes.add(digest)
        unique.append(path)
    return unique


def curate_records(raw_dir: Path = RAW_DIR) -> list[dict[str, Any]]:
    html_rows: list[dict[str, Any]] = []
    for path in _dedup_paths(list(raw_dir.glob("**/*.html"))):
        text = path.read_text(encoding="utf-8", errors="replace")
        if HTML_TITLE in text:
            html_rows = _parse_html_table(path)
            if html_rows:
                break

    xlsx_rows: list[dict[str, Any]] = []
    for path in _dedup_paths(list(raw_dir.glob("**/*.xlsx"))):
        rows = _parse_xlsx(path)
        if rows:
            xlsx_rows = rows
            break

    return html_rows + xlsx_rows


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = curate_records()
    output = OUT_DIR / "records_clean_curated.csv"
    crawler._write_clean_records_csv(rows, output)
    print({"records": len(rows), "output": str(output)})


if __name__ == "__main__":
    main()
