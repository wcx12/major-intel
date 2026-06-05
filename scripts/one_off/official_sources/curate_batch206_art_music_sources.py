from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Iterable

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfReader

try:
    from scripts import graduate_outcome_crawler as crawler
except ImportError:  # pragma: no cover - direct script execution path
    import graduate_outcome_crawler as crawler


RAW_DIR = Path("data/raw/graduate_outcomes_official_site_websearch_web_20260527_batch206_art_music_sources")
OUT_DIR = Path("data/processed/graduate_outcomes_official_site_websearch_web_20260527_batch206_art_music_sources_curated")

LUMEI_XLSX = RAW_DIR / "www.lumei.edu.cn" / "79cd0fb1f2abecdf.xlsx"
CAA_HTML = RAW_DIR / "zb.caa.edu.cn" / "8e90fca12f1f8882.htm"
SCCM_PDF = RAW_DIR / "www.sccm.edu.cn" / "83e73544fc2f63fb.pdf"
JMU_XLSX = RAW_DIR / "arts.jmu.edu.cn" / "617dc24c4d992173.xlsx"

LUMEI_SOURCE_URL = "https://www.lumei.edu.cn/system/_content/download.jsp?urltype=news.DownloadAttachUrl&owner=2099302842&wbfileid=8D5779C52999BBF461947A172DD13D80"
CAA_SOURCE_URL = "https://zb.caa.edu.cn/info/1021/7101.htm"
SCCM_SOURCE_URL = "https://www.sccm.edu.cn/upload/202509/09/202509091749344443.pdf"
JMU_SOURCE_URL = "https://arts.jmu.edu.cn/system/_content/download.jsp?urltype=news.DownloadAttachUrl&owner=2039804140&wbfileid=242BE61F30C9C48A08D937A3A5A14ABD"

LUMEI_TITLE = "鲁迅美术学院2026年硕士研究生招生考试拟录取名单"
CAA_TITLE = "中国美术学院2026年港澳台硕士博士研究生拟录取名单公示"
SCCM_TITLE = "四川音乐学院2026年推免生拟推荐名单"
JMU_TITLE = "美术与设计学院拟推荐2026届免试硕士研究生名单（含备选名单）公示"


def _collapse_spaces(value: Any) -> str:
    return re.sub(r"\s+", " ", "" if value is None else str(value)).strip()


def _remarks(*parts: str) -> str:
    return "; ".join(part for part in (_collapse_spaces(part) for part in parts) if part)


def _format_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _collapse_spaces(value)


def parse_lumei_row(
    values: Iterable[Any],
    *,
    current_major: str,
    source_url: str = LUMEI_SOURCE_URL,
    title: str = LUMEI_TITLE,
) -> dict[str, Any] | None:
    row = list(values)
    if len(row) < 9 or not _collapse_spaces(row[0]).isdigit():
        return None
    ranking, name, student_id, direction, initial, interview, final, study_mode, admission_category, *rest = row
    remarks = _remarks(
        f"initial_score {_format_number(initial)}",
        f"interview_score {_format_number(interview)}",
        f"final_score {_format_number(final)}",
        f"study_mode {_collapse_spaces(study_mode)}",
        f"admission_category {_collapse_spaces(admission_category)}",
        f"note {_collapse_spaces(rest[0])}" if rest and _collapse_spaces(rest[0]) else "",
    )
    return crawler._clean_record(
        {
            "school_name": "鲁迅美术学院",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": _collapse_spaces(name),
            "student_id": _collapse_spaces(student_id),
            "major": current_major,
            "admission_major": _collapse_spaces(direction),
            "ranking": _collapse_spaces(ranking),
            "remarks": remarks,
            "source_url": source_url,
            "title": title,
            "needs_review": False,
        }
    )


def parse_caa_cells(cells: list[str], *, source_url: str = CAA_SOURCE_URL, title: str = CAA_TITLE) -> dict[str, Any] | None:
    if len(cells) < 12 or cells[0] == "申请类型":
        return None
    (
        application_type,
        student_id,
        name,
        college,
        major,
        direction,
        initial_total,
        professional_score,
        interview_score,
        reexam_total,
        composite_score,
        origin,
    ) = [_collapse_spaces(cell) for cell in cells[:12]]
    return crawler._clean_record(
        {
            "school_name": "中国美术学院",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": name,
            "student_id": student_id,
            "college": college,
            "major": major,
            "admission_major": _collapse_spaces(f"{major} {direction}"),
            "remarks": _remarks(
                f"application_type {application_type}",
                f"initial_total {initial_total}",
                f"professional_score {professional_score}",
                f"interview_score {interview_score}",
                f"reexam_total {reexam_total}",
                f"composite_score {composite_score}",
                f"origin {origin}",
            ),
            "source_url": source_url,
            "title": title,
            "needs_review": False,
        }
    )


def _extract_pdf_lines(pdf_path: Path) -> list[str]:
    lines: list[str] = []
    reader = PdfReader(str(pdf_path))
    for page in reader.pages:
        lines.extend((page.extract_text() or "").splitlines())
    return [_collapse_spaces(line) for line in lines if _collapse_spaces(line)]


def _sccm_chunks(lines: list[str]) -> list[list[str]]:
    chunks: list[list[str]] = []
    current: list[str] = []
    expected = 1
    for line in lines:
        if line == str(expected) or line.startswith(f"{expected} "):
            if current:
                chunks.append(current)
            current = [line]
            expected += 1
        elif current:
            current.append(line)
    if current:
        chunks.append(current)
    return chunks


def _parse_sccm_chunk(chunk: list[str], current_college: str) -> tuple[dict[str, Any], str]:
    text = " ".join(chunk)
    match = re.match(r"^(?P<seq>\d+)\s*(?P<body>.*)$", text)
    if not match:
        raise ValueError(f"Bad SCCM chunk: {chunk!r}")
    seq = match.group("seq")
    tokens = match.group("body").split()
    id_index = next((index for index, token in enumerate(tokens) if re.fullmatch(r"20\d{9}", token)), -1)
    if id_index < 1:
        raise ValueError(f"Missing SCCM student id: {chunk!r}")
    college_tokens = tokens[: id_index - 1]
    college = " ".join(college_tokens).strip() or current_college
    name = tokens[id_index - 1]
    student_id = tokens[id_index]
    gender = tokens[id_index + 1]
    tail = tokens[id_index + 2 :]
    if len(tail) < 9:
        raise ValueError(f"Short SCCM row: {chunk!r}")
    major_parts = tail[:-8]
    professional_count = ""
    if len(major_parts) > 1 and re.fullmatch(r"\d+", major_parts[-1]):
        professional_count = major_parts[-1]
        major_parts = major_parts[:-1]
    major = " ".join(major_parts)
    academic_score, academic_weighted, bonus_points, bonus_weighted, final_score, comprehensive_rank, political, violation = tail[-8:]
    record = crawler._clean_record(
        {
            "school_name": "四川音乐学院",
            "year": 2026,
            "document_type": "recommendation_exemption_list",
            "route": "recommendation_exemption",
            "person_name": name,
            "student_id": student_id,
            "college": college,
            "major": major,
            "ranking": seq,
            "remarks": _remarks(
                f"gender {gender}",
                f"professional_count {professional_count}" if professional_count else "",
                f"academic_score {academic_score}",
                f"academic_weighted {academic_weighted}",
                f"bonus_points {bonus_points}",
                f"bonus_weighted {bonus_weighted}",
                f"final_score {final_score}",
                f"comprehensive_rank {comprehensive_rank}",
                f"political {political}",
                f"violation {violation}",
            ),
            "source_url": SCCM_SOURCE_URL,
            "title": SCCM_TITLE,
            "needs_review": False,
        }
    )
    return record, college


def records_from_sccm_lines(lines: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_college = ""
    for chunk in _sccm_chunks([_collapse_spaces(line) for line in lines]):
        record, current_college = _parse_sccm_chunk(chunk, current_college)
        rows.append(record)
    return rows


def _curate_lumei(xlsx_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[dict[str, Any]] = []
    current_major = ""
    for values in worksheet.iter_rows(values_only=True):
        first = _collapse_spaces(values[0] if values else "")
        if not first:
            continue
        if first == "序号" or first.startswith("鲁迅美术学院"):
            continue
        if not first.isdigit():
            current_major = first
            continue
        record = parse_lumei_row(values, current_major=current_major)
        if record is None:
            raise ValueError(f"Unparsed LUMEI row: {values!r}")
        rows.append(record)
    return rows


def _curate_caa(html_path: Path) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html_path.read_text(encoding="utf-8", errors="replace"), "html.parser")
    rows: list[dict[str, Any]] = []
    for tr in soup.find_all("tr"):
        cells = [_collapse_spaces(cell.get_text(" ", strip=True)) for cell in tr.find_all(["td", "th"])]
        record = parse_caa_cells(cells)
        if record is not None:
            rows.append(record)
    return rows


def _curate_sccm(pdf_path: Path) -> list[dict[str, Any]]:
    return records_from_sccm_lines(_extract_pdf_lines(pdf_path))


def _curate_jmu(xlsx_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active
    header: list[str] | None = None
    rows: list[dict[str, Any]] = []
    for values in worksheet.iter_rows(values_only=True):
        first = _collapse_spaces(values[0] if values else "")
        if first == "序号":
            header = [_collapse_spaces(value) for value in values]
            continue
        if header is None or not first.isdigit():
            continue
        row = {header[index]: values[index] if index < len(values) else "" for index in range(len(header))}
        status = _collapse_spaces(row.get("备注"))
        if status != "正选":
            continue
        remarks = _remarks(
            f"status {status}",
            f"gender {_collapse_spaces(row.get('性别'))}",
            f"cet4 {_format_number(row.get('CET4成绩'))}",
            f"cet6 {_format_number(row.get('CET6成绩'))}",
            f"gpa {_format_number(row.get('平均学分绩点'))}",
            f"development_score {_format_number(row.get('全面发展测评成绩'))}",
            f"composite_score {_format_number(row.get('综合成绩'))}",
            f"major_rank {_format_number(row.get('综合排名'))}",
            f"major_count {_format_number(row.get('专业人数'))}",
        )
        rows.append(
            crawler._clean_record(
                {
                    "school_name": "集美大学",
                    "year": 2026,
                    "document_type": "recommendation_exemption_list",
                    "route": "recommendation_exemption",
                    "person_name": _collapse_spaces(row.get("姓名")),
                    "student_id": _collapse_spaces(row.get("学号")),
                    "college": _collapse_spaces(row.get("所在学院")),
                    "major": _collapse_spaces(row.get("所在专业")),
                    "ranking": _format_number(row.get("推荐排序")),
                    "remarks": remarks,
                    "source_url": JMU_SOURCE_URL,
                    "title": JMU_TITLE,
                    "needs_review": False,
                }
            )
        )
    return rows


def curate_records(
    *,
    lumei_xlsx: Path = LUMEI_XLSX,
    caa_html: Path = CAA_HTML,
    sccm_pdf: Path = SCCM_PDF,
    jmu_xlsx: Path = JMU_XLSX,
) -> list[dict[str, Any]]:
    rows = [*_curate_lumei(lumei_xlsx), *_curate_caa(caa_html), *_curate_sccm(sccm_pdf), *_curate_jmu(jmu_xlsx)]
    rows.sort(
        key=lambda row: (
            str(row.get("school_name") or ""),
            str(row.get("document_type") or ""),
            str(row.get("source_url") or ""),
            str(row.get("college") or ""),
            str(row.get("ranking") or ""),
            str(row.get("person_name") or ""),
            str(row.get("student_id") or ""),
        )
    )
    return rows


def main() -> None:
    rows = curate_records()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    crawler._write_clean_records_csv(rows, OUT_DIR / "records_clean_curated.csv")
    counts: dict[str, int] = {}
    for row in rows:
        school = str(row.get("school_name") or "")
        counts[school] = counts.get(school, 0) + 1
    notes = [
        "batch206_art_music_sources_curated: reparsed LUMEI XLSX, CAA HTML, SCCM PDF, and JMU XLSX.",
        "Dropped SCFAI page-navigation noise, SCCM page text noise, and JMU backup rows; SCFAI PDF attachments require a CAPTCHA bridge and were not bypassed.",
        f"rows={len(rows)}",
        f"counts={counts}",
    ]
    (OUT_DIR / "curation_notes.txt").write_text("\n".join(notes) + "\n", encoding="utf-8")
    print({"rows": len(rows), "output": str(OUT_DIR / "records_clean_curated.csv"), "counts": counts})


if __name__ == "__main__":
    main()
