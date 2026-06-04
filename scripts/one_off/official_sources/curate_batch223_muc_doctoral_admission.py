from __future__ import annotations

import json
import re
import sys
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from scripts import graduate_outcome_crawler as crawler
except ImportError:  # pragma: no cover - direct script execution path
    import graduate_outcome_crawler as crawler


PROCESSED_DIR = Path("data/processed/graduate_outcomes_official_site_websearch_web_20260528_batch223_muc_doctoral_admission")
OUT_DIR = Path(
    "data/processed/graduate_outcomes_official_site_websearch_web_20260528_batch223_muc_doctoral_admission_curated"
)

DEFAULT_DOCUMENTS_JSONL = PROCESSED_DIR / "documents.jsonl"


def _load_xlrd():
    try:
        import xlrd  # type: ignore[import-not-found]

        return xlrd
    except ImportError:
        local_package_dir = Path("tmp/xlrd_pkg")
        if local_package_dir.exists():
            sys.path.insert(0, str(local_package_dir.resolve()))
            import xlrd  # type: ignore[import-not-found]

            return xlrd
        raise RuntimeError(
            "xlrd is required to parse batch223 legacy .xls attachments. "
            "Install it with: python -m pip install xlrd==2.0.1 -t tmp/xlrd_pkg"
        )


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def _remarks(*parts: str) -> str:
    return "; ".join(part for part in (_clean_cell(part) for part in parts) if part)


def _read_documents(documents_jsonl: Path) -> list[dict[str, Any]]:
    docs_by_path: dict[str, dict[str, Any]] = {}
    with documents_jsonl.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            doc = json.loads(line)
            raw_path = doc.get("raw_path") or ""
            if raw_path and raw_path not in docs_by_path:
                docs_by_path[raw_path] = doc
    return list(docs_by_path.values())


def _iter_xlsx_rows(path: Path) -> Iterable[list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                yield [_clean_cell(cell) for cell in row]
    finally:
        workbook.close()


def _iter_legacy_xls_rows(path: Path) -> Iterable[list[str]]:
    xlrd = _load_xlrd()
    workbook = xlrd.open_workbook(str(path))
    for sheet in workbook.sheets():
        for row_index in range(sheet.nrows):
            yield [_clean_cell(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)]


def _iter_spreadsheet_rows(path: Path) -> Iterable[list[str]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _iter_xlsx_rows(path)
    if path.suffix.lower() in {".xls", ".bin"}:
        return _iter_legacy_xls_rows(path)
    return []


def _is_header_or_empty(values: list[str]) -> bool:
    compact = "".join(values)
    if not compact:
        return True
    return "姓名" in compact and ("报名号" in compact or "报考院系" in compact)


def _record_from_values(values: list[str], document: dict[str, Any]) -> dict[str, Any] | None:
    padded = values[:9] + [""] * max(0, 9 - len(values))
    person_name, student_id, college, admission_major, direction, special_plan, category, status, note = padded[:9]
    if _is_header_or_empty(padded) or not person_name or not student_id:
        return None
    if not re.fullmatch(r"B\d{8,}", student_id):
        return None
    if "不录取" in status or "拟不录取" in status:
        return None
    if "录取" not in status:
        return None

    remark_parts = [
        f"research_direction {direction}" if direction else "",
        f"special_plan {special_plan}" if special_plan else "",
        f"admission_category {category}" if category else "",
        f"admission_status {status}" if status else "",
        f"note {note}" if note else "",
    ]
    return crawler._clean_record(
        {
            "school_name": "中央民族大学",
            "year": 2026,
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": person_name,
            "student_id": student_id,
            "college": college,
            "admission_major": admission_major,
            "remarks": _remarks(*remark_parts),
            "source_url": document.get("source_url", ""),
            "title": document.get("title", ""),
            "needs_review": False,
        }
    )


def curate_records(*, documents_jsonl: Path = DEFAULT_DOCUMENTS_JSONL) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for document in _read_documents(documents_jsonl):
        raw_path = Path(document.get("raw_path") or "")
        if raw_path.suffix.lower() not in {".bin", ".xls", ".xlsx", ".xlsm"}:
            continue
        if raw_path.name.endswith(".converted.xlsx"):
            continue
        for values in _iter_spreadsheet_rows(raw_path):
            record = _record_from_values(values, document)
            if not record:
                continue
            key = (
                record.get("person_name", ""),
                record.get("student_id", ""),
                record.get("source_url", ""),
            )
            if key in seen:
                continue
            seen.add(key)
            rows.append(record)
    return sorted(rows, key=lambda row: (row["source_url"], row["student_id"], row["person_name"]))


def main() -> None:
    rows = curate_records()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    crawler._write_clean_records_csv(rows, OUT_DIR / "records_clean_curated.csv")
    crawler._write_summary_csv(crawler._build_summary_rows(rows), OUT_DIR / "school_year_summary_curated.csv")
    notes = [
        "batch223_muc_doctoral_admission_curated: parsed Central Minzu University 2026 doctoral admission attachments.",
        "Legacy .xls attachments are downloaded by the site as extensionless OLE .bin files; curated parser reads them with xlrd and reads .xlsx files with openpyxl.",
        f"rows={len(rows)}",
        "source=https://grs.muc.edu.cn/yjsyzsw/info/1069/5819.htm",
        "source=https://grs.muc.edu.cn/yjsyzsw/info/1069/5849.htm",
        "source=https://grs.muc.edu.cn/yjsyzsw/info/1069/5869.htm",
    ]
    (OUT_DIR / "curation_notes.txt").write_text("\n".join(notes) + "\n", encoding="utf-8")
    print({"rows": len(rows), "output": str(OUT_DIR / "records_clean_curated.csv")})


if __name__ == "__main__":
    main()
