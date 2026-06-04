from __future__ import annotations

import csv
import hashlib
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))


RAW_XLSX = (
    ROOT
    / "data/raw/official_recommendation_zcmu_2025_master_admission_current/"
    "zcmu_2025_master_admission_1760944454983762.xlsx"
)
OUTPUT_CSV = (
    ROOT
    / "data/processed/official_recommendation_zcmu_2025_master_admission_current/"
    "records_clean_curated.csv"
)
SOURCE_URL = "https://yjsgl.zcmu.edu.cn/storage/uploads/file/20251020/1760944454983762.xlsx"
SOURCE_PAGE = "https://yjsgl.zcmu.edu.cn/show/5810"
TITLE = "浙江中医药大学2025年硕士研究生拟录取名单公示"
CLEAN_RECORD_CSV_FIELDS = [
    "record_id",
    "school_name",
    "year",
    "document_type",
    "route",
    "person_name",
    "person_name_masked",
    "student_id",
    "student_id_masked",
    "undergraduate_school",
    "undergraduate_major",
    "college",
    "major",
    "admission_major",
    "ranking",
    "remarks",
    "source_url",
    "title",
    "needs_review",
    "quality_score",
    "quality_flags",
]


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def csv_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def mask_name(value: str) -> str:
    if not value:
        return ""
    if len(value) <= 1:
        return "*"
    return value[0] + "*" * (len(value) - 1)


def mask_identifier(value: str) -> str:
    if not value:
        return ""
    if len(value) <= 4:
        return "*" * len(value)
    return value[:4] + "*" * max(2, len(value) - 4)


def quality_score(record: dict[str, Any], flags: list[str]) -> int:
    score = 100
    score -= 12 * len(flags)
    if record.get("source_url"):
        score += 5
    return max(0, min(100, score))


def record_id(record: dict[str, Any]) -> str:
    parts = [
        str(record.get("school_name") or ""),
        str(record.get("year") or ""),
        str(record.get("document_type") or ""),
        str(record.get("person_name") or ""),
        str(record.get("student_id") or ""),
        str(record.get("undergraduate_school") or ""),
        str(record.get("admission_major") or record.get("major") or ""),
        str(record.get("source_url") or ""),
    ]
    if not record.get("student_id") and record.get("ranking"):
        parts.append(f"rank:{record.get('ranking')}")
    identity = "|".join(parts)
    return hashlib.sha1(identity.encode("utf-8")).hexdigest()


def build_remarks(row: dict[str, str]) -> str:
    parts = []
    if row["admission_major_type"]:
        parts.append(f"admission_major_type {row['admission_major_type']}")
    if row["initial_score"]:
        parts.append(f"initial_score {row['initial_score']}")
    if row["reexam_score"]:
        parts.append(f"reexam_score {row['reexam_score']}")
    if row["composite_score"]:
        parts.append(f"composite_score {row['composite_score']}")
    if row["note"]:
        parts.append(f"note {row['note']}")
    parts.extend(
        [
            "official_xlsx_download true",
            f"source_page {SOURCE_PAGE}",
            f"local_artifact {RAW_XLSX.relative_to(ROOT).as_posix()}",
        ]
    )
    return "; ".join(parts)


def parse_records() -> list[dict[str, str]]:
    workbook = load_workbook(RAW_XLSX, data_only=True, read_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    header = [text(value) for value in rows[0]]
    expected = [
        "姓名",
        "考生编号",
        "拟录取学院",
        "拟录取专业名称",
        "拟录取专业类型",
        "初试总分",
        "复试成绩",
        "总成绩",
        "备注",
    ]
    if header[: len(expected)] != expected:
        raise ValueError(f"unexpected ZCMU header: {header!r}")

    records: list[dict[str, str]] = []
    for index, values in enumerate(rows[1:], start=2):
        cells = [text(value) for value in values]
        if not any(cells):
            continue
        source = {
            "person_name": cells[0],
            "student_id": cells[1],
            "college": cells[2],
            "admission_major": cells[3],
            "admission_major_type": cells[4],
            "initial_score": cells[5],
            "reexam_score": cells[6],
            "composite_score": cells[7],
            "note": cells[8] if len(cells) > 8 else "",
        }
        record = {
            "school_name": "浙江中医药大学",
            "year": "2025",
            "document_type": "postgraduate_admission_list",
            "route": "postgraduate_exam_or_admission",
            "person_name": source["person_name"],
            "person_name_masked": mask_name(source["person_name"]),
            "student_id": source["student_id"],
            "student_id_masked": mask_identifier(source["student_id"]),
            "undergraduate_school": "",
            "undergraduate_major": "",
            "college": source["college"],
            "major": "",
            "admission_major": source["admission_major"],
            "ranking": "",
            "remarks": build_remarks(source),
            "source_url": SOURCE_URL,
            "title": TITLE,
            "needs_review": "false",
            "_input_row_index": index,
        }
        flags = ["missing_undergraduate_school"]
        record["quality_score"] = str(quality_score(record, flags))
        record["quality_flags"] = ";".join(flags)
        record["record_id"] = record_id(record)
        records.append({field: record.get(field, "") for field in CLEAN_RECORD_CSV_FIELDS})
    return records


def main() -> None:
    records = parse_records()
    if len(records) != 1715:
        raise ValueError(f"expected 1715 ZCMU records, got {len(records)}")

    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CLEAN_RECORD_CSV_FIELDS)
        writer.writeheader()
        for record in records:
            writer.writerow({field: csv_value(record.get(field)) for field in CLEAN_RECORD_CSV_FIELDS})
    print({"output": str(OUTPUT_CSV), "records": len(records)})


if __name__ == "__main__":
    main()
