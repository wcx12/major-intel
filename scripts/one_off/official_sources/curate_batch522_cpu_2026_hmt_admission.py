from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from scripts import graduate_outcome_crawler as crawler
except ImportError:  # pragma: no cover - direct script execution path
    import graduate_outcome_crawler as crawler


DEFAULT_PROCESSED_DIR = Path(
    "data/processed/official_site_recommendation_websearch_web_20260603_batch522_cpu_2026_hmt_admission"
)
DEFAULT_WORKBOOK = Path(
    "data/raw/official_site_recommendation_websearch_web_20260603_batch522_cpu_2026_hmt_admission/"
    "yjszs.cpu.edu.cn/0b4659e7f8b83ec7.xlsx"
)
DEFAULT_OUTPUT_CSV = DEFAULT_PROCESSED_DIR / "records_clean_curated.csv"
DEFAULT_SUMMARY_CSV = DEFAULT_PROCESSED_DIR / "school_year_summary_curated.csv"
DEFAULT_PUBLIC_CSV = DEFAULT_PROCESSED_DIR / "records_public_curated.csv"
SOURCE_URL = (
    "https://yjszs.cpu.edu.cn/_upload/article/files/55/3f/"
    "7d05258d4355aeb5fe79b3451f4d/2a98c5de-1544-438d-bc95-2f21903cdf69.xlsx"
)
TITLE = "附件：中国药科大学2026年面向香港、澳门、台湾地区招收研究生拟录取名单.xlsx"


def _clean_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def curate_workbook(workbook_path: Path, *, source_url: str = SOURCE_URL) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        header: list[str] | None = None
        for raw_row in worksheet.iter_rows(values_only=True):
            cells = [_clean_cell(cell) for cell in raw_row]
            if not any(cells):
                continue
            if "考生编号" in cells and "姓名" in cells and "专业名称" in cells:
                header = cells
                continue
            if header is None:
                continue

            row = dict(zip(header, cells))
            person_name = row.get("姓名", "")
            student_id = row.get("考生编号", "")
            major_name = row.get("专业名称", "")
            if not person_name or not student_id:
                continue

            remarks = [
                f"degree {row.get('拟攻读学位', '')}",
                f"college_code {row.get('院系所代码', '')}",
                f"major_code {row.get('专业代码', '')}",
                f"study_mode {row.get('学习方式', '')}",
                f"initial_score {row.get('初试成绩', '')}",
                f"reexam_score {row.get('复试成绩', '')}",
                f"total_score {row.get('录取总成绩', '')}",
            ]
            rows.append(
                {
                    "school_name": "中国药科大学",
                    "year": 2026,
                    "document_type": "postgraduate_admission_list",
                    "route": "postgraduate_exam_or_admission",
                    "person_name": person_name,
                    "student_id": student_id,
                    "undergraduate_school": "",
                    "undergraduate_major": "",
                    "college": row.get("院系所名称", ""),
                    "major": major_name,
                    "admission_major": major_name,
                    "ranking": "",
                    "remarks": "; ".join(part for part in remarks if not part.endswith(" ")),
                    "source_url": source_url,
                    "title": TITLE,
                    "needs_review": False,
                }
            )

    return rows


def write_outputs(
    *,
    workbook_path: Path = DEFAULT_WORKBOOK,
    output_csv: Path = DEFAULT_OUTPUT_CSV,
    summary_csv: Path = DEFAULT_SUMMARY_CSV,
    public_csv: Path = DEFAULT_PUBLIC_CSV,
) -> dict[str, int | str]:
    clean_rows = [crawler._clean_record(row) for row in curate_workbook(workbook_path)]
    crawler._write_clean_records_csv(clean_rows, output_csv)
    crawler._write_summary_csv(crawler._build_summary_rows(clean_rows), summary_csv)
    crawler.export_public_records_csv(output_csv, public_csv)
    return {
        "curated": len(clean_rows),
        "output_csv": str(output_csv),
        "summary_csv": str(summary_csv),
        "public_csv": str(public_csv),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT_CSV)
    parser.add_argument("--summary-csv", type=Path, default=DEFAULT_SUMMARY_CSV)
    parser.add_argument("--public-csv", type=Path, default=DEFAULT_PUBLIC_CSV)
    args = parser.parse_args()
    print(
        write_outputs(
            workbook_path=args.workbook,
            output_csv=args.output_csv,
            summary_csv=args.summary_csv,
            public_csv=args.public_csv,
        )
    )


if __name__ == "__main__":
    main()
